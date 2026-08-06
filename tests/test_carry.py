"""Carry-forward: rebuilding must not eat the user's book (D82).

The failure this guards against is silent and total — build() constructs a
fresh workbook from sample seeds and saves it over whatever was there — so
these tests care most about two things: that real inputs survive a round trip,
and that the sample-detector never mistakes a real book for a seeded one.
"""

from __future__ import annotations

import datetime as dt

import openpyxl
import pytest

from src.build_workbook import build, load_config, output_path
from src.carry import CarryError, holds_sample_data, read_inputs
from src.sheets_inputs import LR_COLS, lr_letter


@pytest.fixture(scope="module")
def cfg():
    return load_config("config/config.yaml")


@pytest.fixture(scope="module")
def seeded(tmp_path_factory, cfg):
    p = tmp_path_factory.mktemp("carry") / "seeded.xlsx"
    build(cfg, "Property").save(p)
    return p


def test_reads_every_input_table(seeded):
    got = read_inputs(seeded)
    assert got.lr_rows, "tbl_LR must carry"
    assert got.rate_rows, "the rate log must carry"
    assert got.mod_rows, "the mod log must carry"
    assert got.season_rows, "seasonality must carry"
    assert not got.missing, f"no defined name should be absent: {got.missing}"


def test_reads_control_scalars(seeded):
    got = read_inputs(seeded)
    assert got.term_months == 12
    assert got.season_on == "OFF"
    assert got.mod_master == "ON"
    assert got.trend_default == 0.0
    assert got.sel_bu and got.sel_state


def test_dates_come_back_as_dates_not_datetimes(seeded):
    got = read_inputs(seeded)
    effs = [r["eff"] for r in got.rate_rows]
    assert effs and all(isinstance(e, dt.date) for e in effs)
    assert not any(isinstance(e, dt.datetime) for e in effs), (
        "the engine keys on dt.date; a datetime compares unequal to it")


def test_blank_spare_rows_are_dropped(seeded, cfg):
    """Capacity is 240 rate rows; only the populated ones should carry."""
    got = read_inputs(seeded)
    assert len(got.rate_rows) < cfg.rate_log_rows
    assert all(r["bu"] and r["state"] for r in got.rate_rows)


def test_pristine_sample_is_recognised(seeded, cfg):
    assert holds_sample_data(seeded, cfg, "Property")


def test_one_edited_cell_makes_it_real_data(tmp_path, cfg, seeded):
    """The guard's whole job: an edited book must never read as sample."""
    p = tmp_path / "edited.xlsx"
    p.write_bytes(seeded.read_bytes())
    wb = openpyxl.load_workbook(p)
    ws = wb["Inputs"]
    lrp = lr_letter("lr_lrproj")                # NOT a literal: D107 moved it
    for r in range(1, 40):                      # first populated tbl_LR row
        if ws[f"A{r}"].value and isinstance(ws[f"{lrp}{r}"].value, float):
            ws[f"{lrp}{r}"] = 0.7777
            break
    wb.save(p)
    wb.close()
    assert not holds_sample_data(p, cfg, "Property")


def test_unreadable_source_reads_as_not_sample(tmp_path, cfg):
    """Refuse, don't guess: a file we cannot parse is never 'just the sample'."""
    p = tmp_path / "junk.xlsx"
    p.write_bytes(b"not a workbook")
    assert not holds_sample_data(p, cfg, "Property")
    with pytest.raises(CarryError):
        read_inputs(p)


def test_missing_source_raises(tmp_path):
    with pytest.raises(CarryError):
        read_inputs(tmp_path / "nope.xlsx")


def test_round_trip_preserves_edits(tmp_path, cfg, seeded):
    """Edit -> carry -> rebuild -> read: the edits must still be there."""
    p = tmp_path / "book.xlsx"
    p.write_bytes(seeded.read_bytes())
    wb = openpyxl.load_workbook(p)
    ws = wb["Inputs"]
    edited_key = None
    lrp = lr_letter("lr_lrproj")                # NOT a literal: D107 moved it
    for r in range(1, 40):
        if ws[f"A{r}"].value and isinstance(ws[f"{lrp}{r}"].value, float):
            ws[f"{lrp}{r}"] = 0.7777
            edited_key = f"{ws[f'A{r}'].value}|{ws[f'B{r}'].value}"
            break
    rl = wb["Rate Log"]
    last = max(r for r in range(1, 260) if rl[f"A{r}"].value and rl[f"C{r}"].value)
    bu, state = edited_key.split("|")
    for cell, val in ((f"A{last + 1}", bu), (f"B{last + 1}", state),
                      (f"C{last + 1}", dt.datetime(cfg.plan_year, 7, 1)),
                      (f"D{last + 1}", 0.0333), (f"E{last + 1}", "planned"),
                      (f"F{last + 1}", "N"), (f"G{last + 1}", 0.9),
                      (f"H{last + 1}", "my own filing")):
        rl[cell] = val
    wb["Control"]["C12"] = "ON"
    wb["Control"]["C15"] = 0.035
    wb.save(p)
    wb.close()

    carried = read_inputs(p)
    rebuilt = tmp_path / "rebuilt.xlsx"
    build(cfg, "Property", carried=carried).save(rebuilt)
    got = read_inputs(rebuilt)

    assert any(r["lr_proj"] == pytest.approx(0.7777) for r in got.lr_rows)
    assert any(r["comment"] == "my own filing" for r in got.rate_rows)
    # Every tbl_LR column has to survive the round trip, including the ones no
    # engine formula reads (D96). A reference column is the easiest to forget in
    # carry.py's explicit map, and losing it is silent — the rebuild just comes
    # back blank where the user had typed a number.
    for spec in LR_COLS:
        if spec.key is None:
            continue
        before = {(r["bu"], r["state"]): r[spec.key] for r in carried.lr_rows}
        after = {(r["bu"], r["state"]): r[spec.key] for r in got.lr_rows}
        assert after == before, f"{spec.name} did not survive carry-forward"
    assert len(got.rate_rows) == len(carried.rate_rows)
    assert len(got.mod_rows) == len(carried.mod_rows)
    assert len(got.lr_rows) == len(carried.lr_rows)
    assert got.season_on == "ON"
    assert got.trend_default == pytest.approx(0.035)


def test_carried_build_drops_the_sample_banner(tmp_path, cfg, seeded):
    """A SAMPLE banner over real rows is how a real row gets deleted."""
    carried = read_inputs(seeded)
    p = tmp_path / "carried.xlsx"
    build(cfg, "Property", carried=carried).save(p)
    wb = openpyxl.load_workbook(p)
    for sheet in ("Inputs", "Rate Log", "Mod Log"):
        assert wb[sheet]["A4"].value.startswith("CARRIED FORWARD"), sheet
    wb.close()


def test_carried_build_bakes_the_oracle_for_a_real_combo(tmp_path, cfg, seeded):
    """Oracle ties should describe the user's own book, not a retired sample."""
    carried = read_inputs(seeded)
    p = tmp_path / "carried2.xlsx"
    wb = build(cfg, "Property", carried=carried)
    wb.save(p)
    first = carried.lr_rows[0]
    got = openpyxl.load_workbook(p)
    assert got["Control"]["C7"].value == carried.sel_bu
    assert got["_oracle"]["A1"].value.startswith("_oracle")
    got.close()
    assert first["bu"] and first["state"]


def test_carry_refuses_when_rows_exceed_capacity(cfg, seeded):
    carried = read_inputs(seeded)
    fat = type(carried)(
        lr_rows=carried.lr_rows * 20, rate_rows=carried.rate_rows,
        mod_rows=carried.mod_rows, season_rows=carried.season_rows)
    with pytest.raises(ValueError, match="capacity"):
        build(cfg, "Property", carried=fat)


def test_carry_refuses_an_empty_roster(cfg, seeded):
    carried = read_inputs(seeded)
    empty = type(carried)(lr_rows=[], rate_rows=[], mod_rows=[], season_rows=[])
    with pytest.raises(ValueError, match="no populated"):
        build(cfg, "Property", carried=empty)
