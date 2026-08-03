"""Harvester contracts (D66): what it reads, and what it refuses.

The harvester is the book's only source of truth, so its failure modes matter
more than its happy path — a silently under-reported harvest would roll up
into a plausible, wrong book.

Run:  python -m pytest tests -q
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import openpyxl
import pytest

from src.build_workbook import Layout as L, load_config
from src.sheets_calc import BOOK_PUB, BOOK_SLOTS
from tools.harvest import (LAST_COL, REQUIRED, SCALARS, BookData, HarvestError,
                           harvest, read_lob)

CFG = load_config("config/config.yaml")


def _stub_workbook(path: Path, keys, *, calculated=True, version="9.9.9",
                   n_cols=LAST_COL, checks="ALL CHECKS PASS"):
    """A minimal stand-in for a generated workbook's published surface."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    rm = wb.create_sheet("Read Me")
    rm["B3"] = f"Plan year 2027  |  x  |  version {version}  |  built today"
    if checks is not None:
        wb.create_sheet("Checks")["C3"] = checks
    ws = wb.create_sheet("_calc")
    for i, key in enumerate(keys):
        r = L.CALC_RES_FIRST + i
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c, value=(1.0 if calculated else None))
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=SCALARS["state"], value=key.split("|")[1])
        ws.cell(row=r, column=SCALARS["bu"], value=key.split("|")[0])
    wb.save(path)
    return path


def test_reads_rows_and_provenance(tmp_path):
    p = _stub_workbook(tmp_path / "wb.xlsx", ["A|AZ", "A|CO", "B|AZ"])
    rows, src = read_lob(p, "Property")
    assert [r["key"] for r in rows] == ["A|AZ", "A|CO", "B|AZ"]
    assert all(r["lob"] == "Property" for r in rows)
    assert src.lob == "Property" and src.combos == 3 and src.version == "9.9.9"
    assert len(src.modified) == 16          # "YYYY-MM-DD HH:MM"
    # shapes: the monthly families and the rate-change slots
    assert len(rows[0]["delivered"]) == 12 and len(rows[0]["epw"]) == 12
    assert len(rows[0]["slots"]) == BOOK_SLOTS
    assert all(len(s) == 3 for s in rows[0]["slots"])


def test_skips_spare_rows(tmp_path):
    """Blank keys are the tbl_LR spares — skipped, not harvested as zeros."""
    p = tmp_path / "wb.xlsx"
    _stub_workbook(p, ["A|AZ"])
    wb = openpyxl.load_workbook(p)
    ws = wb["_calc"]
    for c in range(1, LAST_COL + 1):        # a fully blank spare row
        ws.cell(row=L.CALC_RES_FIRST + 1, column=c, value=None)
    wb.save(p)
    rows, src = read_lob(p, "Property")
    assert len(rows) == 1 and src.combos == 1


def test_refuses_uncalculated_workbook(tmp_path):
    p = _stub_workbook(tmp_path / "stale.xlsx", ["A|AZ"], calculated=False)
    with pytest.raises(HarvestError) as e:
        read_lob(p, "Property")
    assert "never" in str(e.value) and "A|AZ" in str(e.value)


def test_refuses_workbook_from_an_older_generator(tmp_path):
    """No BOOK_PUB columns at all — the schema probe must catch it."""
    p = _stub_workbook(tmp_path / "old.xlsx", ["A|AZ"],
                       n_cols=BOOK_PUB["ntaken"] - 1)
    with pytest.raises(HarvestError) as e:
        read_lob(p, "Property")
    assert "ntaken" in str(e.value)


def test_refuses_missing_file(tmp_path):
    with pytest.raises(HarvestError) as e:
        read_lob(tmp_path / "nope.xlsx", "Property")
    assert "not found" in str(e.value)


def test_refuses_duplicate_keys(tmp_path):
    p = _stub_workbook(tmp_path / "dupe.xlsx", ["A|AZ", "A|AZ"])
    with pytest.raises(HarvestError) as e:
        read_lob(p, "Property")
    assert "duplicate" in str(e.value) and "A|AZ" in str(e.value)


def test_refuses_empty_publication(tmp_path):
    p = _stub_workbook(tmp_path / "empty.xlsx", [])
    with pytest.raises(HarvestError):
        read_lob(p, "Property")


def test_refuses_a_source_whose_own_checks_are_failing(tmp_path):
    """A red LOB must not roll up into a green book (D83)."""
    p = _stub_workbook(tmp_path / "bad.xlsx", ["A|AZ"], checks="CHECKS FAILING: 2")
    with pytest.raises(HarvestError) as e:
        read_lob(p, "Property")
    assert "CHECKS FAILING" in str(e.value)


def test_allows_a_failing_source_when_asked_but_records_it(tmp_path):
    p = _stub_workbook(tmp_path / "bad.xlsx", ["A|AZ"], checks="CHECKS FAILING: 2")
    rows, src = read_lob(p, "Property", require_checks=False)
    assert rows and src.checks == "CHECKS FAILING: 2"
    assert not src.checks_ok, "the escape hatch must still mark the source bad"


def test_warnings_on_a_source_do_not_block_the_harvest(tmp_path):
    """Advisories are information, not breakage — the banner says PASS."""
    p = _stub_workbook(tmp_path / "warn.xlsx", ["A|AZ"],
                       checks="PASS WITH 3 WARNING(S)")
    rows, src = read_lob(p, "Property")
    assert rows and src.checks_ok


def test_refuses_a_source_with_no_checks_panel(tmp_path):
    p = _stub_workbook(tmp_path / "nopanel.xlsx", ["A|AZ"], checks=None)
    with pytest.raises(HarvestError) as e:
        read_lob(p, "Property")
    assert "no Checks status" in str(e.value)


def test_uncalculated_beats_checks_in_the_error_message(tmp_path):
    """Both gates fire; the more specific one must win (ordering contract)."""
    p = _stub_workbook(tmp_path / "both.xlsx", ["A|AZ"], calculated=False,
                       checks="CHECKS FAILING: 1")
    with pytest.raises(HarvestError) as e:
        read_lob(p, "Property")
    assert "never" in str(e.value)


def test_book_rosters_preserve_first_appearance_order():
    book = BookData(plan_year=2027, rows=[
        {"lob": "P", "bu": "MM", "state": "AZ"},
        {"lob": "P", "bu": "ABD", "state": "CO"},
        {"lob": "GL", "bu": "MM", "state": "AZ"},
    ])
    assert book.states == ["AZ", "CO"]          # de-duplicated, order kept
    assert book.business_units == ["ABD", "MM"]  # sorted


def test_harvest_of_the_shipped_book():
    """End to end against the real output/ workbooks (skips if not built)."""
    try:
        book = harvest(CFG, now=dt.datetime(2027, 1, 2, 3, 4))
    except HarvestError as e:
        pytest.skip(f"LOB workbooks not harvestable: {e}")
    assert book.as_of == "2027-01-02 03:04"
    assert len(book.sources) == len(CFG.output_lobs)
    assert len(book.rows) == sum(s.combos for s in book.sources)
    assert set(book.business_units) == set(CFG.business_units)
    assert book.states == list(CFG.states)
    for r in book.rows:
        assert all(r[f] is not None for f in REQUIRED)
        assert r["key"] == f"{r['bu']}|{r['state']}"


# ---------------------------------------------------------------------------
# D103: the long pivot dataset. The shape IS the safety property here — one
# weight per row is what makes a single calculated field correct everywhere.
# ---------------------------------------------------------------------------


def test_pivot_rows_are_long_and_self_weighting():
    from src.sheets_book import PIVOT_ANNUAL, PIVOT_MONTHLY, pivot_rows
    try:
        book = harvest(CFG, now=dt.datetime(2027, 1, 2, 3, 4))
    except HarvestError as e:
        pytest.skip(f"LOB workbooks not harvestable: {e}")
    rows = pivot_rows(book, CFG.plan_year)
    assert rows, "no pivot rows produced"
    # every row carries a POSITIVE weight — a zero would make a slice 0/0
    assert all(w > 0 for *_x, w, _wv in rows), "a row has no weight"
    # annual rows have no month; monthly rows all sit inside the plan year
    for *_d, cat, _m, month, _w, _wv in rows:
        if cat == "Monthly":
            assert month is not None and month.year == CFG.plan_year
        else:
            assert month is None
    measures = {m for *_d, _c, m, _mo, _w, _wv in rows}
    assert {m for m, _c, _s, _p in PIVOT_ANNUAL} <= measures
    assert {m for m, _l, _g in PIVOT_MONTHLY} <= measures


def test_pivot_monthly_weights_sum_to_each_combo_s_premium():
    """The /12 rescale: Weight means PREMIUM in every row, so a combo's twelve
    monthly weights must add back to the annual EP they were split from."""
    from src.sheets_book import pivot_rows
    try:
        book = harvest(CFG, now=dt.datetime(2027, 1, 2, 3, 4))
    except HarvestError as e:
        pytest.skip(f"LOB workbooks not harvestable: {e}")
    rows = pivot_rows(book, CFG.plan_year)
    ep, monthly = {}, {}
    for lob, bu, state, _cat, measure, _mo, w, _wv in rows:
        k = (lob, bu, state)
        if measure == "Plan LR":
            ep[k] = w
        elif measure == "Rate YoY":
            monthly[k] = monthly.get(k, 0.0) + w
    assert ep and monthly
    for k, e in ep.items():
        assert monthly[k] == pytest.approx(e, rel=1e-9), f"{k} monthly != annual EP"


def test_pivot_plan_lr_reproduces_the_ep_weighted_book():
    """SUM(Weighted value)/SUM(Weight) over the Plan LR rows must equal the
    EP-weighted plan LR of the harvest — the property every pivot slice relies
    on, checked at the one slice we can compute independently."""
    from src.sheets_book import pivot_rows
    try:
        book = harvest(CFG, now=dt.datetime(2027, 1, 2, 3, 4))
    except HarvestError as e:
        pytest.skip(f"LOB workbooks not harvestable: {e}")
    num = den = 0.0
    for _l, _b, _s, _c, measure, _mo, w, wv in pivot_rows(book, CFG.plan_year):
        if measure == "Plan LR":
            num, den = num + wv, den + w
    direct_n = sum((r["ep"] or 0.0) * (r["cylr_p"] or 0.0) for r in book.rows)
    direct_d = sum(r["ep"] or 0.0 for r in book.rows)
    assert num / den == pytest.approx(direct_n / direct_d, abs=1e-12)
    # and per LOB, since slicing is the whole point
    for lob in {r["lob"] for r in book.rows}:
        n = d = 0.0
        for l_, _b, _s, _c, measure, _mo, w, wv in pivot_rows(book, CFG.plan_year):
            if measure == "Plan LR" and l_ == lob:
                n, d = n + wv, d + w
        dn = sum((r["ep"] or 0.0) * (r["cylr_p"] or 0.0)
                 for r in book.rows if r["lob"] == lob)
        dd = sum(r["ep"] or 0.0 for r in book.rows if r["lob"] == lob)
        assert n / d == pytest.approx(dn / dd, abs=1e-12), lob
