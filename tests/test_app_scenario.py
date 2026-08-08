"""P1 ties: Scenario import -> compute == the workbook's own recalculated
values.

The app has ONE calculation path (sample_to_combo -> engine.run_bridge), so
these tests tie the MAPPING, not the math: importing the shipped Property
workbook and running the app's compute must reproduce the cy plan LRs that
Excel itself calculated into that file's hidden `_calc` results table —
values the release harness has already tied to the oracle. Three
implementations agree or something specific is wrong.

The scenario/compute layer is deliberately UI-free (no panel import), so
this file runs on the system interpreter too; the page-build test at the
bottom is venv-only.
"""
from __future__ import annotations

import sys
from pathlib import Path
from types import SimpleNamespace

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

WB = ROOT / "output" / "Plan_LR_Workbook_2027_Property.xlsx"


def _stub_session():
    """The attribute surface compute.results() actually uses — no panel."""
    return SimpleNamespace(
        page=SimpleNamespace(config=None, caches={}, data_version=0,
                             results={}),
        bus=SimpleNamespace(rev=0),
        ctx=SimpleNamespace(results_rev=0),
    )


@pytest.fixture(scope="module")
def scenario():
    from app import importers
    return importers.from_workbook(WB)


@pytest.fixture(scope="module")
def calc_cached():
    """{combo key: (cy_lr_p, cy_lr_p1)} straight from the shipped file's
    `_calc` results table — values EXCEL calculated, which the release
    harness ties to the oracle. Columns per tools/verify_workbook.py's
    phase-C map (A=key, N=CY LR P, O=CY LR P+1); the table starts at row 4."""
    import openpyxl
    wb = openpyxl.load_workbook(WB, data_only=True)
    ws = wb["_calc"]
    out = {}
    r = 4
    while True:
        key = ws[f"A{r}"].value
        if not key:
            break
        out[key] = (ws[f"N{r}"].value, ws[f"O{r}"].value)
        r += 1
    assert out, "no cached _calc rows — was the workbook recalculated?"
    return out


def test_import_shape(scenario):
    assert scenario.lob == "Property"
    assert scenario.plan_year == 2027
    assert scenario.term_months == 12
    keys = scenario.combo_keys()
    assert len(keys) == len(set(keys))
    assert len(keys) >= 3          # BU x state roster of the shipped sample
    assert all("|" in k for k in keys)


def test_app_ties_workbook_cached_values(scenario, calc_cached):
    from app import compute
    s = _stub_session()
    s.page.config = scenario
    res = compute.results(s)
    assert set(res) == set(calc_cached), (
        "app roster differs from the workbook's _calc table")
    bad = []
    for key, r in res.items():
        assert not compute.is_error(r), f"{key}: {r}"
        wp, wp1 = calc_cached[key]
        if abs(r.cy_lr_p - wp) > 1e-9 or abs(r.cy_lr_p1 - wp1) > 1e-9:
            bad.append((key, r.cy_lr_p, wp, r.cy_lr_p1, wp1))
    assert not bad, f"{len(bad)} combos off vs the workbook: {bad[:3]}"


def test_cache_hits_and_invalidation(scenario):
    from app import compute
    s = _stub_session()
    s.page.config = scenario
    first = compute.results(s)
    assert compute.results(s) is first          # same rev -> cached object
    rev0 = s.ctx.results_rev
    s.bus.rev += 1                              # an input changed
    second = compute.results(s)
    assert second is not first
    assert s.ctx.results_rev == rev0 + 1        # bumped exactly once


def test_mod_master_off_pins_every_a_mod(scenario):
    import dataclasses

    from app import compute
    s = _stub_session()
    s.page.config = dataclasses.replace(scenario, mod_master=False)
    res = compute.results(s)
    assert all(r.a_mod_p == 1.0 for r in res.values()
               if not compute.is_error(r))


@pytest.mark.skipif(
    __import__("importlib").util.find_spec("panel") is None,
    reason="panel not installed (system interpreter — app venv runs this)")
def test_combo_page_renders_live_bridge(scenario):
    """Headless page build: load the scenario, and the bridge card must
    render the selected combo's plan LR. Debounce degrades to synchronous
    passthrough headless, so a bus bump renders immediately."""
    from app.glue.format import fmt_pct
    from app.glue.session import PlanSession
    from app.pages import combo as combo_page

    session = PlanSession()
    page = combo_page.build(session)
    assert set(page) >= {"main", "sidebar"}
    session.replace_config(scenario)            # data_rev fires sync/render
    card = page["card"]                         # W2d put chips above it
    key = scenario.combo_keys()[0]
    assert "plw-bridge" in card.object
    assert key.split("|")[0] in card.object
    from app import compute
    s = _stub_session()
    s.page.config = scenario
    want = fmt_pct(compute.results(s)[key].cy_lr_p)
    assert want in card.object, f"card missing {want}"
