"""W4b: collecting the targets files back.

The contract: a file that came back untouched must diff to NOTHING (so a
real change is unmissable), a typed target must reach the book through
one undoable gesture, and everything the app cannot honestly apply — a
wrong plan year, a combo that left the roster, a cell holding 2.5 where a
fraction belongs — must be refused with a reason rather than guessed at.

openpyxl + engine only, so both interpreters run this.
"""
from __future__ import annotations

import sys
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

WB_PROP = ROOT / "output" / "Plan_LR_Workbook_2027_Property.xlsx"


class _Bus:
    def __init__(self):
        self.rev = 0

    def bump(self):
        self.rev += 1


def _sess(scns):
    page = SimpleNamespace(config=scns[0], data_version=0, results={},
                           caches={}, book={s.lob: s for s in scns})
    return SimpleNamespace(page=page, bus=_Bus(),
                           ctx=SimpleNamespace(results_rev=0))


@pytest.fixture
def scn():
    from app import importers
    return importers.from_workbook(WB_PROP)


@pytest.fixture
def built(scn, tmp_path):
    from app import targetbook
    return targetbook.build_targets_file(scn, out_dir=tmp_path)


def _type_targets(path, edits):
    """Type into the returned file the way a field actuary would: by
    defined name, one (BU|State) at a time."""
    from app import targetbook
    wb = load_workbook(path)
    for key, (p, p1) in edits.items():
        bu, state = key.split("|", 1)
        sfx = next(n[len("tgt_bu_"):] for n in wb.defined_names
                   if n.startswith("tgt_bu_")
                   and wb[list(wb.defined_names[n].destinations)[0][0]][
                       list(wb.defined_names[n].destinations)[0][1]
                       .replace("$", "")].value == bu)
        (sheet, ref), = wb.defined_names[f"tgt_state_{sfx}"].destinations
        ws = wb[sheet]
        states = [c.value for row in ws[ref.replace("$", "")] for c in row]
        r = targetbook.FIRST_ROW + states.index(state)
        ws[f"H{r}"] = p
        ws[f"I{r}"] = p1
    wb.save(path)
    return path


# ------------------------------------------------------------------- read

def test_scan_finds_targets_and_ignores_everything_else(built, tmp_path,
                                                        monkeypatch):
    """A folder of workbooks is normal; only files carrying the marker are
    ours. The standard directory is redirected here: the real one holds
    whatever the user last generated, and a test that reads it would pass
    or fail on that."""
    from app import targets_io
    import shutil
    monkeypatch.setattr(targets_io, "TARGETS_DIR", tmp_path)
    shutil.copy2(WB_PROP, tmp_path / "Plan_LR_Workbook_2027_Property.xlsx")
    (tmp_path / "notes.txt").write_text("not a workbook")
    found = targets_io.scan_targets(extra_dir=tmp_path)
    assert [p.name for p in found] == [built.name]
    with pytest.raises(targets_io.TargetsReadError, match="not a targets"):
        targets_io.read_targets(
            tmp_path / "Plan_LR_Workbook_2027_Property.xlsx")


def test_read_carries_identity_and_every_row(built, scn):
    from app import targets_io
    tf = targets_io.read_targets(built)
    assert tf.lob == "Property" and tf.plan_year == scn.plan_year
    assert tf.fmt == 1 and tf.generated and tf.fingerprint
    keys = {f"{r['bu']}|{r['state']}" for r in scn.lr_rows
            if r.get("bu") and r.get("state")}
    assert set(tf.rows) == keys
    assert not tf.problems


def test_an_untouched_file_diffs_to_nothing(built, scn):
    """The prefill exists for exactly this: if a returned file that nobody
    edited showed changes, a real edit would be lost in the noise."""
    from app import targets_io
    tf = targets_io.read_targets(built)
    changes, skipped = targets_io.diff_targets(tf, scn)
    assert changes == [] and skipped == []


# ------------------------------------------------------------------- diff

def test_typed_targets_read_back_and_speak_the_changelog(built, scn):
    from app import targets_io
    key = f"{scn.lr_rows[0]['bu']}|{scn.lr_rows[0]['state']}"
    _type_targets(built, {key: (0.025, 0.01)})
    tf = targets_io.read_targets(built)
    changes, _skipped = targets_io.diff_targets(tf, scn)
    assert [c.key for c in changes] == [key]
    c = changes[0]
    assert c.new_p == pytest.approx(0.025) and c.new_p1 == pytest.approx(0.01)
    sentences = targets_io.diff_sentences(tf, scn, changes)
    joined = " ".join(sentences)
    assert "Property" in joined and key in joined
    assert "2.5%" in joined                     # the changelog's own voice


def test_blank_clears_net_mode_and_p1_alone_warns(built, scn):
    """Blank is an instruction, not missing data — and a +1 target with no
    plan-year target does nothing, which the app says out loud."""
    from app import targets_io
    net = next((r for r in scn.lr_rows if r.get("netp") is not None), None)
    assert net is not None, "the sample book carries a net combo"
    k_net = f"{net['bu']}|{net['state']}"
    other = next(r for r in scn.lr_rows
                 if r.get("netp") is None and r.get("bu") and r.get("state"))
    k_other = f"{other['bu']}|{other['state']}"
    _type_targets(built, {k_net: (None, None), k_other: (None, 0.03)})
    tf = targets_io.read_targets(built)
    changes, _sk = targets_io.diff_targets(tf, scn)
    by = {c.key: c for c in changes}
    assert by[k_net].old_p is not None and by[k_net].new_p is None
    assert by[k_other].warn and "does nothing" in by[k_other].warn


def test_float_noise_is_not_a_change(built, scn):
    from app import targets_io
    row = next(r for r in scn.lr_rows if r.get("netp") is not None)
    key = f"{row['bu']}|{row['state']}"
    _type_targets(built, {key: (row["netp"] + 1e-15, row.get("netp1"))})
    tf = targets_io.read_targets(built)
    changes, _sk = targets_io.diff_targets(tf, scn)
    assert [c.key for c in changes] == []


def test_bad_cells_are_refused_with_a_reason(built, scn):
    """DV is defeated by a paste, so the bounds are re-checked here. 2.5
    where a fraction belongs is a typo, not a 250% rate change."""
    from app import targets_io
    keys = [f"{r['bu']}|{r['state']}" for r in scn.lr_rows[:2]]
    _type_targets(built, {keys[0]: (2.5, None), keys[1]: ("soon", None)})
    tf = targets_io.read_targets(built)
    assert len(tf.problems) == 2
    assert any("FRACTION" in m for m in tf.problems)
    assert any("not a number" in m for m in tf.problems)
    assert tf.rows[keys[0]][0] is None and tf.rows[keys[1]][0] is None
    # a percent STRING is a legitimate paste, though
    _type_targets(built, {keys[0]: ("3.5%", None)})
    tf2 = targets_io.read_targets(built)
    assert tf2.rows[keys[0]][0] == pytest.approx(0.035)


def test_stale_flags_but_the_targets_still_stand(built, scn):
    """A file built before an input change is stale — the LOSS RATIOS it
    displayed were computed from older inputs. The typed targets are still
    what the field asked for, so staleness warns and never blocks."""
    import dataclasses

    from app import targets_io
    tf = targets_io.read_targets(built)
    assert not targets_io.is_stale(tf, scn)
    rows = [dict(r) for r in scn.lr_rows]
    rows[0] = dict(rows[0], lr_proj=float(rows[0]["lr_proj"]) + 0.02)
    moved = dataclasses.replace(scn, lr_rows=rows)
    assert targets_io.is_stale(tf, moved)
    # ...and applying still works against the moved book
    sess = _sess([moved])
    key = f"{rows[1]['bu']}|{rows[1]['state']}"
    _type_targets(built, {key: (0.04, None)})
    tf2 = targets_io.read_targets(built)
    applied, _msgs = targets_io.apply_targets(sess, tf2, [key])
    assert applied == 1


# ------------------------------------------------------------------ apply

def test_apply_writes_bumps_once_and_undoes(built, scn):
    from app import targets_io
    from app.undo import UndoStack
    sess = _sess([scn])
    k1 = f"{scn.lr_rows[0]['bu']}|{scn.lr_rows[0]['state']}"
    k2 = f"{scn.lr_rows[1]['bu']}|{scn.lr_rows[1]['state']}"
    before = {k: (scn.row(k).get("netp"), scn.row(k).get("netp1"))
              for k in (k1, k2)}
    _type_targets(built, {k1: (0.02, 0.015), k2: (0.06, None)})
    tf = targets_io.read_targets(built)

    undo = UndoStack()
    rev0 = sess.bus.rev
    applied, msgs = targets_io.apply_targets(sess, tf, [k1], undo=undo)
    assert applied == 1 and not msgs
    assert sess.bus.rev == rev0 + 1                  # ONE bump per gesture
    assert scn.row(k1)["netp"] == pytest.approx(0.02)
    assert scn.row(k1)["netp1"] == pytest.approx(0.015)
    assert scn.row(k2).get("netp") == before[k2][0]  # not selected, untouched
    assert len(undo) == 1 and "Collect targets" in undo.peek()

    n, skipped = undo.pop_apply(sess)
    assert n == 1 and not skipped
    assert scn.row(k1).get("netp") == before[k1][0]
    assert scn.row(k1).get("netp1") == before[k1][1]


def test_apply_refuses_a_wrong_year_or_missing_line(built, scn):
    import dataclasses

    from app import targets_io
    tf = targets_io.read_targets(built)
    key = next(iter(tf.rows))

    sess = _sess([dataclasses.replace(scn, plan_year=scn.plan_year + 1)])
    applied, msgs = targets_io.apply_targets(sess, tf, [key])
    assert applied == 0 and "plan year" in msgs[0]

    empty = SimpleNamespace(
        page=SimpleNamespace(book={}, config=None, data_version=0,
                             results={}, caches={}),
        bus=_Bus(), ctx=SimpleNamespace(results_rev=0))
    applied, msgs = targets_io.apply_targets(empty, tf, [key])
    assert applied == 0 and "not loaded" in msgs[0]


def test_a_combo_that_left_the_book_is_skipped_not_invented(built, scn):
    import dataclasses

    from app import targets_io
    tf = targets_io.read_targets(built)
    gone_key = f"{scn.lr_rows[0]['bu']}|{scn.lr_rows[0]['state']}"
    trimmed = dataclasses.replace(scn, lr_rows=[dict(r)
                                                for r in scn.lr_rows[1:]])
    sess = _sess([trimmed])
    _type_targets(built, {gone_key: (0.05, None)})
    tf = targets_io.read_targets(built)
    changes, skipped = targets_io.diff_targets(tf, trimmed)
    assert gone_key not in [c.key for c in changes]
    assert any(gone_key in m and "no longer in the book" in m
               for m in skipped)
    applied, msgs = targets_io.apply_targets(sess, tf, [gone_key])
    assert applied == 0


# --------------------------------------------------------------- the page

@pytest.mark.skipif(
    __import__("importlib").util.find_spec("panel") is None,
    reason="panel not installed (system interpreter — app venv runs this)")
def test_scenarios_page_collects_a_returned_file(scn, tmp_path, monkeypatch):
    """The real motion, through the page: generate, someone types, scan,
    preview, apply the selected row, undo."""
    from app import targetbook, targets_io
    from app.glue.session import PlanSession
    from app.pages import scenarios as page_mod

    monkeypatch.setattr(targets_io, "TARGETS_DIR", tmp_path)
    session = PlanSession()
    session.replace_config(scn)
    page = page_mod.build(session)
    path = targetbook.build_targets_file(scn, out_dir=tmp_path)

    key = f"{scn.lr_rows[0]['bu']}|{scn.lr_rows[0]['state']}"
    _type_targets(path, {key: (0.075, None)})

    page["collect"]["scan"]()
    assert page["widgets"]["collect"].options, "the scan found nothing"
    page["widgets"]["collect"].value = str(path)
    assert key in page["widgets"]["rows"].options
    assert "1 row(s) changed" in page["flashes"]["preview"].object
    assert not page["widgets"]["apply"].disabled

    rev0 = session.bus.rev
    page["collect"]["apply"](None)
    assert session.page.book["Property"].row(key)["netp"] == \
        pytest.approx(0.075)
    assert session.bus.rev > rev0
    assert "Applied 1 row(s)" in page["flashes"]["collect"].object
    # the file now matches the book, so the preview says so
    assert "matches the book" in page["flashes"]["preview"].object

    page["widgets"]["coll_undo"].clicks += 1
    assert session.page.book["Property"].row(key).get("netp") != \
        pytest.approx(0.075)
