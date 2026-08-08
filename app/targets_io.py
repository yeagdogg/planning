"""Collecting the targets files back (W4b).

The other half of the round trip: field actuaries type net rate targets
into the small workbooks this app generated (``app/targetbook.py``), hand
them back, and this module reads them, shows exactly what changed in the
book's own vocabulary, and applies the rows the user approves.

Three rules the design turns on:

**Read the file, trust nothing.** Every harvested cell is a typed
constant, so the read never depends on a recalculation having happened —
but a paste defeats Excel's data validation, so bounds and types are
re-checked HERE. A cell holding ``2.5`` (meaning two and a half percent)
is refused with a reason, not applied as a 250% rate change.

**Blank is a value.** In this system a blank net target means "no net
selection — value this combo on its logged program", which is a different
answer from 0%. So a cleared cell is a real instruction and shows up in
the diff as a change; it is not "no data".

**Nothing lands without being shown.** The diff runs through
``scenarios_io.diff_books`` — the same sentences the scenario changelog
writes — so the preview and the audit trail speak one language. Apply
pushes ONE undo gesture per file and bumps the bus once.

A file built against different dates, mods or loss ratios (a moved
fingerprint) is called out as stale, but is NOT blocked: the targets a
person typed are still what they meant to ask for; only the loss ratios
they saw while typing were computed from older inputs, and this app
recomputes those from truth the moment the targets land.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from . import scenarios_io, targetbook
from .undo import entry

TARGETS_DIR = targetbook.TARGETS_DIR

#: the same bound the file's own data validation carries
_LO, _HI = -0.5, 0.5
#: below this two floats are the same number (0.025 written and read back)
_EPS = 1e-12


class TargetsReadError(RuntimeError):
    """This file is not a targets file we can read."""


@dataclass
class TargetsFile:
    path: Path
    lob: str
    plan_year: int
    fmt: int
    generated: str
    fingerprint: str
    rows: dict = field(default_factory=dict)     # "BU|State" -> (netp, netp1)
    problems: list = field(default_factory=list)


@dataclass
class RowChange:
    key: str
    old_p: float | None
    new_p: float | None
    old_p1: float | None
    new_p1: float | None
    warn: str = ""


# ------------------------------------------------------------------ read

def _cells(wb, name):
    """The defined-name read, in src/carry.py's idiom."""
    dn = wb.defined_names.get(name) if hasattr(wb.defined_names, "get") \
        else (wb.defined_names[name] if name in wb.defined_names else None)
    if dn is None:
        raise TargetsReadError(f"missing defined name {name!r}")
    out = []
    for sheet, ref in dn.destinations:
        ws = wb[sheet]
        got = ws[ref.replace("$", "")]
        if hasattr(got, "value"):               # single cell
            out.append(got)
        else:
            out.extend(c for row in got for c in row)
    return out


def _scalar(wb, name):
    got = _cells(wb, name)
    return got[0].value if got else None


def _clean(value, where: str, problems: list):
    """One harvested target cell -> a fraction, a blank, or a problem.

    Deliberately strict: Excel's validation is defeated by a paste, and a
    number outside the sane band is far more likely a typo (2.5 for 2.5%)
    than an intention."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.endswith("%"):
            try:
                return float(text[:-1].replace(",", "")) / 100.0
            except ValueError:
                pass
        problems.append(f"{where}: {value!r} is not a number — skipped")
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        problems.append(f"{where}: {value!r} is not a number — skipped")
        return None
    v = float(value)
    if not (_LO <= v <= _HI):
        problems.append(
            f"{where}: {v:g} is outside -50%..+50% — a target is a FRACTION "
            f"(2.5% is 0.025), so this looks like a typo; skipped")
        return None
    return v


def read_targets(path) -> TargetsFile:
    """Read one returned targets file. Values only — no recalculation
    needed, because everything this reads was typed, not computed."""
    path = Path(path)
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:                                  # noqa: BLE001
        raise TargetsReadError(f"{path.name}: cannot open — {e}") from e
    try:
        fmt = _scalar(wb, "tgt_Format")
    except TargetsReadError:
        raise TargetsReadError(
            f"{path.name}: not a targets file (no tgt_Format marker)")
    if int(fmt or 0) != targetbook.FORMAT_VERSION:
        raise TargetsReadError(
            f"{path.name}: targets format {fmt} — this app reads format "
            f"{targetbook.FORMAT_VERSION}. Regenerate the file.")

    tf = TargetsFile(
        path=path, lob=str(_scalar(wb, "tgt_LOB")),
        plan_year=int(_scalar(wb, "tgt_PlanYear") or 0),
        fmt=int(fmt), generated=str(_scalar(wb, "tgt_Generated") or ""),
        fingerprint=str(_scalar(wb, "tgt_Fingerprint") or ""))

    for name in sorted(wb.defined_names):
        if not name.startswith("tgt_state_"):
            continue
        sfx = name[len("tgt_state_"):]
        bu = _scalar(wb, f"tgt_bu_{sfx}")
        states = [c.value for c in _cells(wb, name)]
        netp = [c.value for c in _cells(wb, f"tgt_netp_{sfx}")]
        netp1 = [c.value for c in _cells(wb, f"tgt_netp1_{sfx}")]
        for st, a, b in zip(states, netp, netp1):
            if not st:
                continue
            key = f"{bu}|{st}"
            tf.rows[key] = (_clean(a, f"{key} net target", tf.problems),
                            _clean(b, f"{key} net target +1", tf.problems))
    if not tf.rows:
        raise TargetsReadError(f"{path.name}: no state rows found")
    return tf


def scan_targets(extra_dir=None) -> list:
    """Every readable targets file, newest first. A folder full of other
    workbooks is fine — anything without the marker is simply not ours."""
    seen, out = set(), []
    for d in (TARGETS_DIR, Path(extra_dir) if extra_dir else None):
        if d is None or not Path(d).is_dir():
            continue
        for p in sorted(Path(d).glob("*.xlsx")):
            if p.name.startswith("~$") or p.resolve() in seen:
                continue
            seen.add(p.resolve())
            try:
                read_targets(p)
            except TargetsReadError:
                continue
            out.append(p)
    return sorted(out, key=lambda p: p.stat().st_mtime, reverse=True)


# ------------------------------------------------------------------ diff

def is_stale(tf: TargetsFile, scn) -> bool:
    """True when the file was built against different INPUTS (dates, mods,
    loss ratios) than the book now holds. Targets themselves are excluded
    from the fingerprint — a changed target is the point, not staleness."""
    return bool(tf.fingerprint) and \
        tf.fingerprint != targetbook.fingerprint_line(scn)


def _same(a, b) -> bool:
    if a is None or b is None:
        return a is None and b is None       # blank vs a value IS a change
    return abs(float(a) - float(b)) <= _EPS


def diff_targets(tf: TargetsFile, scn) -> tuple:
    """(changes, skipped) — what this file asks the book to become.

    ``changes`` are structured for selection; the human sentences come
    from the scenario changelog's own differ, so the preview here and the
    audit trail there cannot drift apart."""
    changes, skipped = [], []
    by_key = {f"{r.get('bu')}|{r.get('state')}": r for r in scn.lr_rows
              if r.get("bu") and r.get("state")}
    for key, (new_p, new_p1) in tf.rows.items():
        row = by_key.get(key)
        if row is None:
            skipped.append(f"{key}: no longer in the book — skipped "
                           f"(regenerate the file to match)")
            continue
        old_p, old_p1 = row.get("netp"), row.get("netp1")
        if _same(old_p, new_p) and _same(old_p1, new_p1):
            continue
        warn = ("a +1 target with no plan-year target does nothing — the "
                "combo is not in net mode"
                if new_p is None and new_p1 is not None else "")
        changes.append(RowChange(key=key, old_p=old_p, new_p=new_p,
                                 old_p1=old_p1, new_p1=new_p1, warn=warn))
    return changes, skipped


def diff_sentences(tf: TargetsFile, scn, changes) -> list:
    """The changelog's own voice, by diffing the book against the book it
    WOULD become — no second formatter to drift."""
    old = scenarios_io.book_to_doc({tf.lob: scn}, "current")["lines"]
    new = {tf.lob: {**old[tf.lob],
                    "lr_rows": [dict(r) for r in old[tf.lob]["lr_rows"]]}}
    wanted = {c.key: c for c in changes}
    for r in new[tf.lob]["lr_rows"]:
        c = wanted.get(f"{r.get('bu')}|{r.get('state')}")
        if c is not None:
            r["netp"], r["netp1"] = c.new_p, c.new_p1
    return scenarios_io.diff_books(old, new)


# ----------------------------------------------------------------- apply

def apply_targets(session, tf: TargetsFile, keys, undo=None) -> tuple:
    """Write the selected rows' targets into the book. Returns
    (applied, [messages]). One undo gesture, one bus bump."""
    scn = session.page.book.get(tf.lob)
    if scn is None:
        return 0, [f"{tf.lob} is not loaded — load the line first"]
    if tf.plan_year != scn.plan_year:
        return 0, [f"{tf.path.name} is plan year {tf.plan_year}; the book's "
                   f"{tf.lob} is {scn.plan_year} — nothing applied"]
    changes, skipped = diff_targets(tf, scn)
    wanted = {c.key: c for c in changes if c.key in set(keys)}
    if not wanted:
        return 0, skipped or ["nothing selected"]

    by_key = {f"{r.get('bu')}|{r.get('state')}": r for r in scn.lr_rows
              if r.get("bu") and r.get("state")}
    entries, applied = [], 0
    for key, c in wanted.items():
        row = by_key.get(key)
        if row is None:                       # roster moved under the diff
            skipped.append(f"{key}: no longer in the book — skipped")
            continue
        entries.append(entry(tf.lob, "lr_rows", row,
                             {"netp": c.old_p, "netp1": c.old_p1},
                             after={"netp": c.new_p, "netp1": c.new_p1}))
        row["netp"], row["netp1"] = c.new_p, c.new_p1
        applied += 1
    if undo is not None and entries:
        undo.push(f"Collect targets {tf.path.name}: {applied} row(s)",
                  entries)
    if applied:
        session.bus.bump()
    return applied, skipped
