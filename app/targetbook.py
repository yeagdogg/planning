"""The TARGETS FILE — a small, live, net-target-only workbook (W4a).

The fleet workbook is the full model; this is the one page a field actuary
or product manager actually needs to make a selection. One file per line,
one tab per business unit, one row per state, and exactly three editable
cells per row: **Net rate target P**, **Net rate target P+1**, and a
filing-date dropdown for the suggested-change helper. Type a target and
the plan loss ratios move — in Excel, with no add-in, no macro, and no
recalculation trip back to the app.

WHY THAT IS EXACT, NOT AN APPROXIMATION. Freeze every date, mod action and
logged rate change, and the engine's plan LR is a closed form in the net
targets. The net path is ``q(k) = q(k-12) * (1 + x)`` on written cohorts
(engine.py net_index): cohorts written before Jan P are history constants,
cohorts written in P carry (1+x), cohorts written in P+1 carry (1+x)(1+x1).
Written weights and earning geometry do not depend on x at all. The bridge
divides by that earned level (A_rate = CRL / E_CY), so

    1 / cy_lr_p (x)      = c0 + c1*(1+x)
    1 / cy_lr_p1(x, x1)  = d0 + d1*(1+x) + d2*(1+x)*(1+x1)

are EXACTLY affine. This module does not re-derive any of that: it SAMPLES
``engine.run_bridge`` at four target pairs, solves for the five constants
in closed form, and then verifies the claim at a fifth, unrelated point. A
residual there is not a rounding wobble — it means the engine's functional
form changed — so it fails the BUILD rather than shipping a workbook that
lies. The constants are engine output; the Excel formula is arrangement.
One calculation path, exactly as everywhere else in this app.

The suggested-change helper needs no fitting at all: ``suggest_net_rate``
already publishes its own closed form, and ``net_delivery_components``
hands over the three constants behind it (c0, c1, sum_w) — baked once per
quarterly effective date.

WHAT IS DELIBERATELY NOT HERE. Dates are baked. A new filing date, a moved
mod action, a changed loss ratio: those change the constants, so they need
a regenerated file, and the cover says so in those words. This file is for
picking a target, not for rebuilding the model — that is what the fleet
workbook and the workbench are for.
"""
from __future__ import annotations

import datetime as dt
import hashlib
import json
import re
from dataclasses import dataclass, replace
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.workbook.defined_name import DefinedName

from src import engine
from src.xlstyle import (ALIGN_C, ALIGN_L, ALIGN_R, BORDER_THIN, F_INPUT,
                         F_SMALL_IT, FILL_GREY, FILL_REQ, FMT_EP, FMT_PCT,
                         FMT_PCT_SIGNED, NAVY, STEEL, add_dv,
                         assert_formulas_balanced, dv_decimal, formula,
                         header_row, jump, label, note, presentation_setup,
                         print_setup, prose, put, set_widths, title)

from . import compute

ROOT = Path(__file__).resolve().parents[1]
TARGETS_DIR = ROOT / "output" / "targets"

#: bumped only when the READ contract changes (targets_io checks it)
FORMAT_VERSION = 1

COVER = "Read Me"
FACTORS = "_factors"

#: the four filing dates the helper offers, as TEXT tokens. Deliberately
#: not Excel dates: a serial in a dropdown is a locale trap, and the token
#: is what a user says out loud ("we'd file 7/1").
DATE_TOKENS = ("1/1", "4/1", "7/1", "10/1")
DATE_MONTHS = (1, 4, 7, 10)

#: _factors column layout, exported so the tests read it the way the
#: workbook writes it (the tools/harvest.py SCALARS precedent).
FACTORS_COLS = {"key": 1, "bu": 2, "state": 3, "c0": 4, "c1": 5, "d0": 6,
                "d1": 7, "d2": 8, "prog_p": 9, "prog_p1": 10, "sum_w": 11,
                "a1": 12, "a2": 13, "a3": 14, "a4": 15,
                "b1": 16, "b2": 17, "b3": 18, "b4": 19,
                "resid_p": 20, "resid_p1": 21, "status": 22}

#: first data row on a BU tab / on _factors
FIRST_ROW = 6
FACT_FIRST = 3

# The fit's sample points. Any two distinct x values determine the P line;
# the (0, .10) point adds the P+1 restart leg; CHECK is never used to fit,
# only to falsify.
_X_A, _X_B, _X1_B = 0.0, 0.10, 0.10
_CHECK = (0.07, -0.03)
#: a structural break, not a rounding wobble — see the module docstring
_RESID_TOL = 1e-11


class TargetsBuildError(RuntimeError):
    """The build refused to ship a workbook that would compute wrongly."""


@dataclass
class LrFit:
    c0: float
    c1: float
    d0: float
    d1: float
    d2: float
    resid_p: float
    resid_p1: float

    def lr_p(self, x: float) -> float:
        return 1.0 / (self.c0 + self.c1 * (1.0 + x))

    def lr_p1(self, x: float, x1: float | None = None) -> float:
        x1 = x if x1 is None else x1            # the engine's carry rule
        u = 1.0 + x
        return 1.0 / (self.d0 + self.d1 * u + self.d2 * u * (1.0 + x1))


@dataclass
class HelperFit:
    sum_w: float
    a: list                                     # per DATE_TOKENS
    b: list

    def suggested(self, x: float, j: int) -> float | None:
        if not self.b[j]:
            return None                         # no exposure after that date
        return ((1.0 + x) * self.sum_w - self.a[j]) / self.b[j]


@dataclass
class ComboBake:
    key: str
    bu: str
    state: str
    row: dict
    ep: float | None = None
    hist: str = ""
    mods: str = ""
    prog_p: float | None = None
    prog_p1: float | None = None
    fit: LrFit | None = None
    helper: HelperFit | None = None
    error: str = ""

    @property
    def ok(self) -> bool:
        return self.error == ""


# --------------------------------------------------------------- fitting

def fit_lr(plan_year: int, ci) -> LrFit:
    """Extract the five constants by SAMPLING the engine, then falsify.

    Four runs pin the form; the fifth is spent trying to break it. Solving
    on 1/LR is what makes this closed-form rather than a regression: the
    reciprocal is affine, so the differences below ARE the coefficients.
    """
    def y(x, x1):
        # net_sel_p1 is ALWAYS explicit: a None there means "carry x"
        # (engine ComboInputs.net_x1) and would contaminate the grid
        res = engine.run_bridge(
            plan_year, replace(ci, net_sel_p=x, net_sel_p1=x1), "monthly")
        return 1.0 / res.cy_lr_p, 1.0 / res.cy_lr_p1

    ya_p, ya_p1 = y(_X_A, _X_A)                 # (0, 0)
    yb_p, yb_p1 = y(_X_B, _X_A)                 # (.10, 0)
    _yc_p, yc_p1 = y(_X_A, _X1_B)               # (0, .10)

    c1 = (yb_p - ya_p) / _X_B
    c0 = ya_p - c1 * (1.0 + _X_A)
    d2 = (yc_p1 - ya_p1) / _X1_B
    d1 = (yb_p1 - ya_p1) / _X_B - d2
    d0 = ya_p1 - d1 - d2

    fit = LrFit(c0=c0, c1=c1, d0=d0, d1=d1, d2=d2, resid_p=0.0, resid_p1=0.0)
    cx, cx1 = _CHECK
    want_p, want_p1 = y(cx, cx1)
    fit.resid_p = abs((c0 + c1 * (1.0 + cx)) - want_p)
    fit.resid_p1 = abs(
        (d0 + d1 * (1.0 + cx) + d2 * (1.0 + cx) * (1.0 + cx1)) - want_p1)
    return fit


def helper_fit(plan_year: int, ci) -> HelperFit:
    """The suggested-change constants, straight from the engine's own
    closed form (net_delivery_components). c0/c1/sum_w do not depend on the
    target, so one call per date serves every target the user will type."""
    net = replace(ci, net_sel_p=0.0, net_sel_p1=0.0)   # components needs one
    a, b, sum_w = [], [], 0.0
    for m in DATE_MONTHS:
        try:
            comp = engine.net_delivery_components(
                plan_year, net, dt.date(plan_year, m, 1))
        except ValueError:
            a.append(0.0)                       # no exposure after D, or the
            b.append(0.0)                       # date sits outside the year
            continue
        a.append(comp.c0)
        b.append(comp.c1)
        sum_w = comp.sum_w
    return HelperFit(sum_w=sum_w, a=a, b=b)


def _verify_helper(plan_year: int, ci, hf: HelperFit) -> None:
    """Belt and braces: the baked constants must reproduce the engine's own
    suggest_net_rate on the first date that has exposure."""
    for j, m in enumerate(DATE_MONTHS):
        if not hf.b[j]:
            continue
        x = 0.05
        want, _comp = engine.suggest_net_rate(
            plan_year, replace(ci, net_sel_p=x, net_sel_p1=x),
            dt.date(plan_year, m, 1))
        got = hf.suggested(x, j)
        if abs(got - want) > 1e-11 * max(1.0, abs(want)):
            raise TargetsBuildError(
                f"suggested-change constants disagree with the engine at "
                f"{DATE_TOKENS[j]}: baked {got!r} vs engine {want!r}")
        return


# ------------------------------------------------------------ the echoes

def _pct(v) -> str:
    return "—" if v is None else f"{v:+.1%}"


def history_text(scn, bu: str, state: str) -> str:
    """The rate-change chronology in the State Summary's own legend:
    T/P = taken/planned, percent = filed x achievement, * = not counted in
    the indication."""
    parts = []
    rows = [r for r in scn.rate_rows
            if r.get("bu") == bu and r.get("state") == state and r.get("eff")]
    for r in sorted(rows, key=lambda r: r["eff"]):
        filed = r.get("filed")
        ach = r.get("achievement")
        eff = filed if filed is None else filed * (1.0 if ach is None else ach)
        tok = "T" if r.get("status") == "taken" else "P"
        star = "" if str(r.get("considered", "") or "").upper().startswith("Y") \
            else "*"
        parts.append(f"{tok} {r['eff']:%m/%d/%y} {_pct(eff)}{star}")
    return " · ".join(parts) if parts else "no logged changes"


def mod_text(scn, row: dict, bu: str, state: str) -> str:
    """The mod path in one line: what the indication assumed, where the
    written mod stands now, and where it is projected to land."""
    def f3(v):
        return "—" if v is None else f"{v:.3f}"

    bits = [f"M_ind {f3(row.get('m_ind'))}"]
    asof = row.get("m0_asof")
    bits.append(f"M0 {f3(row.get('m0'))}"
                + (f" @ {asof:%m/%d/%y}" if hasattr(asof, "year") else ""))
    bits.append(f"→ M1 {f3(row.get('m1'))}")
    n = sum(1 for r in scn.mod_rows
            if r.get("bu") == bu and r.get("state") == state and r.get("eff"))
    if n:
        bits.append(f"{n} logged action(s)")
    off = (not scn.mod_master) or str(row.get("modadj") or "").upper() == "OFF"
    if off:
        bits.append("mod adj OFF")
    return " · ".join(bits)


# ------------------------------------------------------------ fingerprint

def fingerprint_line(scn) -> str:
    """12 hex of everything that DETERMINES the baked constants — and
    deliberately NOT netp/netp1. A returned file whose targets differ is
    the point of the exercise; a returned file built against different
    dates, mods or loss ratios is stale, and the collect step says so."""
    def clean(rows, drop=()):
        out = []
        for r in rows:
            item = {}
            for k in sorted(r):
                if k in drop:
                    continue
                v = r[k]
                if hasattr(v, "isoformat"):
                    v = v.isoformat()
                elif isinstance(v, float):
                    v = round(v, 12)
                item[k] = v
            out.append(item)
        return out

    payload = {
        "plan_year": scn.plan_year, "term": scn.term_months,
        "season_on": bool(scn.season_on), "mod_master": bool(scn.mod_master),
        "trend_default": round(float(scn.trend_default), 12),
        "lr": clean(scn.lr_rows, drop=("netp", "netp1")),
        "rate": clean(scn.rate_rows), "mod": clean(scn.mod_rows),
        "season": clean(scn.season_rows),
    }
    blob = json.dumps(payload, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(blob).hexdigest()[:12]


# ------------------------------------------------------------- the bake

def bake_combo(scn, row: dict, prog: dict | None = None) -> ComboBake:
    """Every constant one state row needs. Engine REJECTIONS (a row the
    bridge cannot value at all) dash the row and are listed on the cover;
    a broken functional form raises instead — see fit_lr."""
    bu, state = row.get("bu"), row.get("state")
    key = f"{bu}|{state}"
    ep = row.get("ep")
    bake = ComboBake(key=key, bu=bu, state=state, row=row,
                     ep=float(ep) if isinstance(ep, (int, float)) else None,
                     hist=history_text(scn, bu, state),
                     mods=mod_text(scn, row, bu, state))
    try:
        ci = compute.combo_inputs(scn, row)
    except Exception as e:                                  # noqa: BLE001
        bake.error = f"inputs rejected: {e}"
        return bake

    got = (prog or {}).get(key)
    try:
        if got is None:
            base = engine.run_bridge(
                scn.plan_year, replace(ci, net_sel_p=None, net_sel_p1=None),
                "monthly")
            got = (base.cy_lr_p, base.cy_lr_p1)
        bake.prog_p, bake.prog_p1 = float(got[0]), float(got[1])
        bake.fit = fit_lr(scn.plan_year, ci)
        bake.helper = helper_fit(scn.plan_year, ci)
        _verify_helper(scn.plan_year, ci, bake.helper)
    except TargetsBuildError:
        raise
    except Exception as e:                                  # noqa: BLE001
        bake.error = str(e)
        return bake

    fit = bake.fit
    if max(fit.resid_p, fit.resid_p1) > _RESID_TOL:
        raise TargetsBuildError(
            f"{scn.lob} {key}: the plan LR is no longer an exact function of "
            f"the net target (residual {max(fit.resid_p, fit.resid_p1):.3e} "
            f"> {_RESID_TOL:.0e}). The engine's net path changed shape — the "
            f"targets file would compute wrong numbers, so the build stops.")
    return bake


# ------------------------------------------------------- name registry

class _Names:
    """The defined-name registry, in the build_workbook.py idiom: Excel
    resolves names case-insensitively, so a case-only collision is a
    silently overwritten reference — catch it at build time."""

    def __init__(self) -> None:
        self._by_lower: dict = {}
        self._names: dict = {}

    def add(self, name: str, sheet: str, ref: str) -> None:
        low = name.lower()
        if low in self._by_lower:
            raise TargetsBuildError(
                f"defined name collision: {name!r} vs "
                f"{self._by_lower[low]!r} (Excel ignores case)")
        self._by_lower[low] = name
        q = f"'{sheet}'" if (" " in sheet or "-" in sheet) else sheet
        self._names[name] = f"{q}!{ref}"

    def flush(self, wb) -> None:
        for name, ref in sorted(self._names.items()):
            wb.defined_names.add(DefinedName(name, attr_text=ref))


def _sheet_name(bu: str, taken: set) -> str:
    """Excel's sheet-name rules, then collision suffixes. The TRUE business
    unit string rides in a named cell — never infer it back from this."""
    clean = re.sub(r"[\[\]:*?/\\]", "-", str(bu)).strip() or "BU"
    clean = clean[:31]
    base, n = clean, 2
    while clean.lower() in taken:
        suffix = f" ({n})"
        clean = base[:31 - len(suffix)] + suffix
        n += 1
    taken.add(clean.lower())
    return clean


def _suffix(bu: str, taken: set) -> str:
    """Defined-name suffix for a BU: identifier-safe, collision-checked."""
    s = re.sub(r"[^0-9A-Za-z_]", "_", str(bu)) or "BU"
    if s[0].isdigit():
        s = "_" + s
    base, n = s, 2
    while s.lower() in taken:
        s = f"{base}_{n}"
        n += 1
    taken.add(s.lower())
    return s


# ------------------------------------------------------------ the build

_HOWTO = [
    "This file asks you for one thing: the net rate target you want to "
    "deliver, by state. Type it in the blue cells (yellow fill) and the two "
    "plan loss ratios beside them update immediately — those are the real "
    "engine numbers, not a rule of thumb.",
    "Leave the target BLANK to see what the rate program already logged "
    "delivers on its own. Blank is not zero: a blank target means \"no net "
    "selection\", a 0.0% target means \"hold price flat\", and they are "
    "different answers. Leave the P+1 target blank to carry your P target "
    "into the following year.",
    "Targets are fractions the way Excel holds percents: type 2.5% or 0.025, "
    "both land the same. The filing-date dropdown is optional — pick a "
    "quarter and the sheet shows the rate change that would DELIVER your "
    "target from that date (that is the achieved change; gross it up by your "
    "expected achievement to get the ask you file).",
    "Everything else on the sheet is locked, because everything else is a "
    "consequence. Rate-change dates, mod actions and loss ratios are baked "
    "into this file: to move a DATE or change an input, the workbench has to "
    "regenerate it. That is a fast round trip — ask for a fresh file rather "
    "than editing around the lock.",
    "When you are done: save the file with the same name and hand it back. "
    "The workbench reads only your target columns, shows exactly what "
    "changed, and applies the rows you approve.",
]


def build_targets_wb(scn, prog: dict | None = None, now=None):
    """Build one line's targets workbook. Returns (Workbook, [ComboBake])."""
    stamp = (now or dt.datetime.now()).strftime("%Y-%m-%d %H:%M")
    fp = fingerprint_line(scn)

    rows = [r for r in scn.lr_rows if r.get("bu") and r.get("state")]
    if not rows:
        raise TargetsBuildError(f"{scn.lob}: no BU|State rows to write.")
    seen = set()
    for r in rows:
        k = f"{r['bu']}|{r['state']}"
        if k in seen:
            raise TargetsBuildError(
                f"{scn.lob}: duplicate combo {k} — one row per BU and state.")
        seen.add(k)

    by_bu: dict = {}
    bakes = []
    for r in rows:
        b = bake_combo(scn, r, prog=prog)
        bakes.append(b)
        by_bu.setdefault(b.bu, []).append(b)

    wb = Workbook()
    wb.remove(wb.active)
    names = _Names()
    cover = wb.create_sheet(COVER)

    # ---- the BU tabs (built first: the cover links to them) -------------
    sheet_taken: set = set()
    suffix_taken: set = set()
    tabs = []
    fact_row = FACT_FIRST
    for bu, group in by_bu.items():
        ws = wb.create_sheet(_sheet_name(bu, sheet_taken))
        sfx = _suffix(bu, suffix_taken)
        tabs.append((ws.title, bu))
        _write_bu_tab(ws, scn, bu, group, fact_row, names, sfx, stamp)
        fact_row += len(group)

    # ---- the hidden machine area ---------------------------------------
    fx = wb.create_sheet(FACTORS)
    _write_factors(fx, bakes)
    fx.sheet_state = "hidden"

    _write_cover(cover, scn, stamp, fp, tabs, bakes, names)

    names.add("tgt_Format", COVER, "$B$6")
    names.add("tgt_LOB", COVER, "$B$7")
    names.add("tgt_PlanYear", COVER, "$B$8")
    names.add("tgt_Generated", COVER, "$B$9")
    names.add("tgt_Fingerprint", COVER, "$B$10")
    names.flush(wb)

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    assert_formulas_balanced(wb)
    return wb, bakes


def _write_cover(ws, scn, stamp, fp, tabs, bakes, names) -> None:
    p = scn.plan_year
    set_widths(ws, {"A": 22, "B": 78})
    title(ws, "A1", f"{scn.lob} — net rate targets, plan year {p}",
          sub="Type a target; the plan loss ratios move. Everything else is "
              "a consequence of what the workbench already knows.")
    presentation_setup(ws, freeze="A5", tab_color=NAVY)
    print_setup(ws, landscape=False, title_text=f"{scn.lob} net rate targets")

    label(ws, "A6", "File format", bold=True)
    put(ws, "B6", FORMAT_VERSION, align=ALIGN_L)
    label(ws, "A7", "Line", bold=True)
    put(ws, "B7", scn.lob, align=ALIGN_L)
    label(ws, "A8", "Plan year", bold=True)
    put(ws, "B8", p, align=ALIGN_L)
    label(ws, "A9", "Generated", bold=True)
    put(ws, "B9", stamp, align=ALIGN_L)
    label(ws, "A10", "Inputs fingerprint", bold=True)
    put(ws, "B10", fp, align=ALIGN_L)

    r = 12
    label(ws, f"A{r}", "How this works", bold=True)
    for para in _HOWTO:
        prose(ws, f"B{r}", para, width=78)
        r += 1

    r += 1
    label(ws, f"A{r}", "Your tabs", bold=True)
    for sheet, bu in tabs:
        jump(ws, f"B{r}", f"'{sheet}'!A1", f"{bu}  →")
        r += 1

    bad = [b for b in bakes if not b.ok]
    if bad:
        r += 1
        label(ws, f"A{r}", "Not priced here", bold=True)
        prose(ws, f"B{r}", "These states are shown with — because the "
                           "workbench could not value them; they need a fix "
                           "in the workbench before a target means anything.",
              width=78)
        r += 1
        for b in bad:
            put(ws, f"A{r}", b.key, fnt=F_SMALL_IT)
            note(ws, f"B{r}", b.error)
            r += 1

    ws.protection.sheet = True


_HEADERS = ["State", "Plan EP (000s)", "Rate changes logged", "Mod path",
            "Plan LR — logged program", "Plan LR +1 — logged program",
            "Target LR", "NET RATE TARGET", "NET RATE TARGET +1",
            "PLAN LR", "PLAN LR +1", "File?", "Change to file", "Note"]
_WIDTHS = [7, 13, 46, 34, 12, 12, 10, 13, 13, 11, 11, 8, 13, 40]


def _write_bu_tab(ws, scn, bu, group, fact_first, names, sfx, stamp) -> None:
    p = scn.plan_year
    title(ws, "A1", f"{bu} — {scn.lob}",
          sub=f"Plan year {p}. Type in the blue cells only; generated {stamp}.")
    presentation_setup(ws, freeze="A6", tab_color=STEEL)
    print_setup(ws, title_text=f"{bu} — {scn.lob} net rate targets")
    hdr = list(_HEADERS)
    hdr[4] = f"Plan LR {p} — logged program"
    hdr[5] = f"Plan LR {p + 1} — logged program"
    hdr[9] = f"PLAN LR {p}"
    hdr[10] = f"PLAN LR {p + 1}"
    hdr[7] = f"NET RATE TARGET {p}"
    hdr[8] = f"NET RATE TARGET {p + 1}"
    header_row(ws, 5, 1, hdr, widths=_WIDTHS)

    for i, b in enumerate(group):
        r = FIRST_ROW + i
        f = fact_first + i
        ws.row_dimensions[r].height = 15
        put(ws, f"A{r}", b.state, align=ALIGN_L, border=BORDER_THIN)
        put(ws, f"B{r}", b.ep, fmt=FMT_EP, align=ALIGN_R, border=BORDER_THIN)
        put(ws, f"C{r}", b.hist, fnt=F_SMALL_IT, align=ALIGN_L,
            border=BORDER_THIN)
        put(ws, f"D{r}", b.mods, fnt=F_SMALL_IT, align=ALIGN_L,
            border=BORDER_THIN)
        put(ws, f"G{r}", b.row.get("target"), fmt=FMT_PCT, align=ALIGN_R,
            border=BORDER_THIN)

        if not b.ok:
            for c in ("E", "F", "J", "K", "M"):
                put(ws, f"{c}{r}", "—", align=ALIGN_C, border=BORDER_THIN)
            for c in ("H", "I", "L"):
                put(ws, f"{c}{r}", None, fill=FILL_GREY, border=BORDER_THIN)
            note(ws, f"N{r}", b.error)
            continue

        put(ws, f"E{r}", b.prog_p, fmt=FMT_PCT, align=ALIGN_R,
            border=BORDER_THIN)
        put(ws, f"F{r}", b.prog_p1, fmt=FMT_PCT, align=ALIGN_R,
            border=BORDER_THIN)
        # the three editable cells — prefilled with the book's current
        # targets so an untouched file round-trips to "nothing changed"
        for c, v in (("H", b.row.get("netp")), ("I", b.row.get("netp1"))):
            cell = put(ws, f"{c}{r}", v, fnt=F_INPUT, fmt=FMT_PCT_SIGNED,
                       fill=FILL_REQ, align=ALIGN_R, border=BORDER_THIN)
            cell.protection = Protection(locked=False)
        # TEXT format, not General: Excel parses "7/1" into a DATE the
        # moment it is typed or picked from the dropdown, and the MATCH
        # against the token row would then never hit (found by the live
        # Excel test — no python-level test can see this coercion).
        dcell = put(ws, f"L{r}", None, fnt=F_INPUT, fmt="@", fill=FILL_REQ,
                    align=ALIGN_C, border=BORDER_THIN)
        dcell.protection = Protection(locked=False)

        fr = f"{FACTORS}!"
        formula(ws, f"J{r}",
                f"=IF($H{r}=\"\",{fr}${_fc('prog_p')}${f},"
                f"1/({fr}${_fc('c0')}${f}+{fr}${_fc('c1')}${f}*(1+$H{r})))",
                fmt=FMT_PCT, border=BORDER_THIN, align=ALIGN_R, bold=True)
        formula(ws, f"K{r}",
                f"=IF($H{r}=\"\",{fr}${_fc('prog_p1')}${f},"
                f"1/({fr}${_fc('d0')}${f}"
                f"+{fr}${_fc('d1')}${f}*(1+$H{r})"
                f"+{fr}${_fc('d2')}${f}*(1+$H{r})"
                f"*(1+IF($I{r}=\"\",$H{r},$I{r}))))",
                fmt=FMT_PCT, border=BORDER_THIN, align=ALIGN_R, bold=True)

        # the helper: r* = ((1+x)*sum_w - a_d) / b_d, with a_d/b_d picked by
        # the date token. b_d = 0 means the engine had no exposure after
        # that date (or refused it) — say "n/a" rather than divide.
        m_tok = (f"MATCH($L{r},{fr}${_fc('a1')}$2:${_fc('a4')}$2,0)")
        a_d = f"INDEX({fr}${_fc('a1')}${f}:${_fc('a4')}${f},{m_tok})"
        b_d = f"INDEX({fr}${_fc('b1')}${f}:${_fc('b4')}${f},{m_tok})"
        formula(ws, f"M{r}",
                f"=IF(OR($H{r}=\"\",$L{r}=\"\"),\"\","
                f"IF({b_d}=0,\"n/a\","
                f"((1+$H{r})*{fr}${_fc('sum_w')}${f}-{a_d})/{b_d}))",
                fmt=FMT_PCT_SIGNED, border=BORDER_THIN, align=ALIGN_R)
        formula(ws, f"N{r}",
                f"=IF(AND($H{r}=\"\",$I{r}<>\"\"),"
                f"\"A +1 target alone does nothing — enter a {p} target too\","
                f"\"\")", border=BORDER_THIN)

    last = FIRST_ROW + len(group) - 1
    live = [b for b in group if b.ok]
    if live:
        rows_ref = [f"$H${FIRST_ROW}:$I${last}"]
        dv_decimal(ws, rows_ref, -0.5, 0.5,
                   error="Net rate targets are fractions: 2.5% or 0.025. "
                         "Anything outside -50%..+50% is a typo.")
        add_dv(ws, "list", [f"$L${FIRST_ROW}:$L${last}"],
               formula1='"' + ",".join(DATE_TOKENS) + '"',
               error="Pick one of the quarterly filing dates.")
    note(ws, f"A{last + 2}",
         "Blue cells are yours. Blank target = the logged program (the two "
         "grey plan-LR columns); blank +1 = your plan-year target carries. "
         "'Change to file' is the ACHIEVED change that delivers your target "
         "from the chosen date — gross it up by expected achievement.")

    # The TRUE business unit string, parked in a hidden column past the
    # exhibit: the tab NAME is sanitized for Excel's sheet rules and the
    # title is a composite, so neither can be read back as the key.
    put(ws, "Q1", bu, fnt=F_SMALL_IT)
    ws.column_dimensions["Q"].hidden = True
    names.add(f"tgt_bu_{sfx}", ws.title, "$Q$1")
    for nm, c in (("state", "A"), ("netp", "H"), ("netp1", "I"),
                  ("date", "L"), ("lrp", "J"), ("lrp1", "K"), ("sugg", "M")):
        names.add(f"tgt_{nm}_{sfx}", ws.title,
                  f"${c}${FIRST_ROW}:${c}${last}")
    ws.protection.sheet = True


def _fc(key: str) -> str:
    """_factors column letter for a schema key."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(FACTORS_COLS[key])


def _write_factors(ws, bakes) -> None:
    note(ws, "A1", "Machine area — the workbench regenerates this file; "
                   "editing these cells changes what the sheet computes.")
    inv = {v: k for k, v in FACTORS_COLS.items()}
    for c in range(1, max(FACTORS_COLS.values()) + 1):
        put(ws, f"{_col(c)}2", inv.get(c, ""), fnt=F_SMALL_IT)
    for j, tok in enumerate(DATE_TOKENS):        # L2:O2 — the MATCH row
        put(ws, f"{_col(FACTORS_COLS['a1'] + j)}2", tok, fnt=F_SMALL_IT)
    for i, b in enumerate(bakes):
        r = FACT_FIRST + i
        vals = {"key": b.key, "bu": b.bu, "state": b.state,
                "status": "ok" if b.ok else b.error}
        if b.ok:
            vals.update(c0=b.fit.c0, c1=b.fit.c1, d0=b.fit.d0, d1=b.fit.d1,
                        d2=b.fit.d2, prog_p=b.prog_p, prog_p1=b.prog_p1,
                        sum_w=b.helper.sum_w, resid_p=b.fit.resid_p,
                        resid_p1=b.fit.resid_p1)
            for j in range(len(DATE_TOKENS)):
                vals[f"a{j + 1}"] = b.helper.a[j]
                vals[f"b{j + 1}"] = b.helper.b[j]
        for k, v in vals.items():
            put(ws, f"{_col(FACTORS_COLS[k])}{r}", v)
    ws.protection.sheet = True


def _col(idx: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


# ---------------------------------------------------------------- files

def targets_filename(scn) -> str:
    """A stem the fleet loader can NEVER mistake for a workbook: a
    targets file read as a whole line would replace it with emptiness."""
    return (f"Plan_LR_Targets_{scn.plan_year}_"
            f"{str(scn.lob).replace(' ', '_')}.xlsx")


def build_targets_file(scn, out_dir=None, prog: dict | None = None,
                       now=None) -> Path:
    out = Path(out_dir) if out_dir else TARGETS_DIR
    out.mkdir(parents=True, exist_ok=True)
    wb, _bakes = build_targets_wb(scn, prog=prog, now=now)
    path = out / targets_filename(scn)
    wb.save(path)
    return path


def build_all(book: dict, out_dir=None, prog_by_lob: dict | None = None,
              now=None) -> list:
    """One targets file per loaded line, in book order."""
    paths = []
    for lob, scn in book.items():
        paths.append(build_targets_file(
            scn, out_dir=out_dir,
            prog=(prog_by_lob or {}).get(lob), now=now))
    return paths
