"""Sheet builders: Bridge, Portfolio, Scenarios, Solver, Attribution."""

from __future__ import annotations

from dataclasses import dataclass

import datetime as dt

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, Side

from .build_workbook import Ctx, Layout as L, SHEETS, mod_adj_on, slot_rank
from .xlstyle import (ALIGN_C, ALIGN_WRAP, BORDER_THIN, DOWN_BAR, FAIL_RED, FILL_GREY,
    FILL_NAVY, FILL_PANEL, FMT_DATE, FMT_DATE_S, FMT_EP_Z, FMT_GEN, FMT_IDX,
    FMT_IDX_Z, FMT_INT, FMT_MOD, FMT_PCT, FMT_PCT_SIGNED, FMT_PCT_Z, FMT_PTS_COL,
    F_HEADER, F_LABEL, F_SMALL_IT, GREY_DARK, NAVY, STEEL, STEEL_LIGHT, TOTAL_BAR,
    UP_BAR, WARN_AMBER, add_dv, chart_legend, col, dv_decimal, dv_plan_year_date,
    font, formula, header_row, input_cell, jump, label, link, nav_bar, note,
    presentation_setup, print_setup, prose, put, section, set_widths, title,
)

PTS_Z = '+0.00 "pts";-0.00 "pts";""'
PCT_SIGNED_Z = '+0.0%;-0.0%;""'


def _style_chart(chart, title_text, x_title=None, y_title=None, height=8.5, width=16):
    chart.title = title_text
    chart.style = 2
    chart.height = height
    chart.width = width
    # Excel's "plot visible cells only" default would blank any chart whose
    # staging rows are outline-hidden (openpyxl attr: visible_cells_only —
    # NOT plotVisOnly, which is silently ignored). Harness-guarded.
    chart.visible_cells_only = False
    if x_title:
        chart.x_axis.title = x_title
    if y_title:
        chart.y_axis.title = y_title
    # openpyxl omits <c:delete> when the attribute is None, and Excel treats
    # the omission as "axis deleted" on resave — every axis must be explicitly
    # visible unless deliberately hidden (DECISIONS.md D36).
    if chart.x_axis.delete is None:
        chart.x_axis.delete = False
    if chart.y_axis.delete is None:
        chart.y_axis.delete = False
    # D69, centralised: a legend left at openpyxl's default is drawn ON TOP of
    # the series. Callers that want no legend set it to None before styling and
    # are untouched; everyone else gets it moved out of the plot area here,
    # rather than each new chart having to remember (D71's lesson).
    chart_legend(chart)
    return chart


def _line_color(series, rgb, width_pt=2.25, dashed=False):
    series.graphicalProperties.line.solidFill = rgb
    series.graphicalProperties.line.width = int(width_pt * 12700)
    if dashed:
        series.graphicalProperties.line.dashStyle = "dash"
    series.marker = Marker(symbol="none")
    series.smooth = False


# ---------------------------------------------------------------------------
# Bridge
# ---------------------------------------------------------------------------


def build_bridge(ctx: Ctx):
    """Answer-first layout: the claim sentence and waterfall live above the fold
    with the chart at their right shoulder; the resolver block (the tbl_LR
    inputs that physically feed the engines via the nr_* names) is demoted to
    an outline-grouped audit block below. Harness binds Bridge ONLY via names,
    so the re-layout is free as long as every name is re-defined here."""
    ws = ctx.wb["Bridge"]
    p = ctx.p
    title(ws, "B2", f"Bridge — projected LR to calendar-year plan LR ({ctx.lob.name})",
          "The selected combo's answer, factor by factor. Expand the grouped rows at the "
          "bottom to audit exactly which inputs feed the engines.")

    label(ws, "E4", "Selected:")
    link(ws, "F4", "=nr_SelKey", bold=True)
    jump(ws, "H4", "Control!C7", "Change BU/state selection >")

    # ---- answer band ----
    formula(ws, "B5",
            '=IF(nr_SelOK,"Projected "&TEXT(nr_LRproj,"0.0%")&"  ->  CY "&nr_PlanYear&'
            '" plan "&TEXT(nr_CYLR_P,"0.0%")&"   ("&TEXT((nr_CYLR_P-nr_LRcur)*100,'
            '"+0.0;-0.0;0.0")&" pts vs current level)",'
            '"Select a BU and state that exist in tbl_LR to see the bridge.")',
            fmt=FMT_GEN)
    ws["B5"].font = font(NAVY, bold=True, size=14)
    ws["B5"].fill = FILL_PANEL
    for cc in range(3, 9):
        put(ws, ws.cell(row=5, column=cc).coordinate, None, fill=FILL_PANEL)
    put(ws, "B6", "Every factor below is calculated on the Rate Engine and Mod Engine "
                  "for this combo.", fnt=F_SMALL_IT)

    R = "$F${0}".format(L.BR_IN_FIRST - 2)   # br_selrow, on the resolver section row
    rows = [
        ("Business unit", "=nr_SelBU", FMT_GEN, None, None, True),
        ("State", "=nr_SelState", FMT_GEN, None, None, True),
        ("Projected LR (as input)", f"=IF({R}=0,0,N(INDEX(lr_lrproj,{R})))", FMT_PCT,
         "nr_LRproj", "Projected loss ratio from the indication, as entered", False),
        ("LR basis", f'=IF({R}=0,"current",INDEX(lr_basis,{R})&"")', FMT_GEN,
         "nr_Basis", "Basis of the input LR: current | proposed", False),
        ("Indication selected change (s)", f"=IF({R}=0,0,N(INDEX(lr_s,{R})))", FMT_PCT,
         "nr_SelS", "Selected/indicated change used for basis normalization", False),
        ("LR at current rate level",
         '=IF(nr_Basis="proposed",nr_LRproj*(1+nr_SelS),nr_LRproj)', FMT_PCT,
         "nr_LRcur", "Basis-normalized projected LR (§3.4.1)", False),
        ("Mod assumed in indication (M_ind)",
         f"=IF({R}=0,1,IF(N(INDEX(lr_mind,{R}))=0,1,INDEX(lr_mind,{R})))", FMT_MOD,
         "nr_MInd", "Average schedule mod assumed in the indication", False),
        ("Current avg written mod (M_0)",
         f"=IF({R}=0,1,IF(N(INDEX(lr_m0,{R}))=0,1,INDEX(lr_m0,{R})))",
         FMT_MOD, "nr_M0", "Current average written mod", False),
        ("Current mod as-of date (close of day)",
         f"=IF({R}=0,DATE(nr_PlanYear-1,10,1),IF(N(INDEX(lr_m0asof,{R}))=0,"
         f"DATE(nr_PlanYear-1,10,1),INDEX(lr_m0asof,{R})))", FMT_DATE,
         "nr_M0Asof", "As-of date of M_0 (anchored close-of-day, D2)", False),
        ("Projected mod, end of plan yr (M_1)",
         f"=IF({R}=0,1,IF(N(INDEX(lr_m1,{R}))=0,1,INDEX(lr_m1,{R})))",
         FMT_MOD, "nr_M1", "Projected average written mod at 12/31/P", False),
        ("Mod ~1 yr before as-of (M_prior; 0 = not given)",
         f"=IF({R}=0,0,N(INDEX(lr_mprior,{R})))", FMT_MOD,
         "nr_MPrior", "Average mod ~12 months before M_0; 0/blank = extrapolate backward", False),
        ("Projected mod, end plan yr+1 (M_2; 0 = use M_1)",
         f"=IF({R}=0,0,N(INDEX(lr_m2,{R})))", FMT_MOD,
         "nr_M2", "Projected written mod at 12/31/(P+1); 0/blank = M_1", False),
        ("Target loss ratio (profit provision; 0 = not set)",
         f"=IF({R}=0,0,N(INDEX(lr_target,{R})))", FMT_PCT,
         "nr_TargetLR", "Target loss ratio for this combo — reference only, no "
         "engine formula reads it (D96)", False),
        ("Net trend (P+1 view)",
         f"=IF({R}=0,N(nr_TrendDefault),IF(ISBLANK(INDEX(lr_trend,{R})),N(nr_TrendDefault),"
         f"INDEX(lr_trend,{R})))", FMT_PCT,
         "nr_Trend", "Effective P+1 net trend: tbl_LR value, or the Control default when blank",
         False),
        ("Other adjustment factor (A_other)",
         f"=IF({R}=0,1,IF(N(INDEX(lr_aother,{R}))=0,1,INDEX(lr_aother,{R})))", FMT_IDX,
         "nr_AOther", "Manual adjustment factor (label required when <> 1)", False),
        ("Reason for other adjustment", f'=IF({R}=0,"",INDEX(lr_aotherlbl,{R})&"")', FMT_GEN,
         "nr_AOtherLbl", "Reason for A_other", False),
        ("Mod adjustment (combo row)",
         f'=IF({R}=0,"ON",IF(INDEX(lr_modadj,{R})="OFF","OFF","ON"))', FMT_GEN,
         "nr_ModAdjRow", "Per-combo mod adjustment toggle from tbl_LR", False),
        ("Mod adjustment effective",
         '=IF(AND(nr_ModAdjMaster="ON",nr_ModAdjRow="ON"),"ON","OFF")', FMT_GEN,
         "nr_ModAdjEff", "TRUE state of the mod adjustment (master AND combo row)", False),
        ("Net rate selection active?",
         f"=IF({R}=0,FALSE,NOT(ISBLANK(INDEX(lr_netp,{R}))))", FMT_GEN,
         "nr_NetMode", "TRUE when this combo carries a net rate selection (D39)", False),
        ("Net selection x (P)", f"=IF({R}=0,0,N(INDEX(lr_netp,{R})))", FMT_PCT,
         "nr_NetSelP", "YoY combined net price target for cohorts written from 1/1/P", False),
        ("Net selection x (P+1; blank = carry)",
         f"=IF({R}=0,0,IF(ISBLANK(INDEX(lr_netp1,{R})),N(INDEX(lr_netp,{R})),"
         f"N(INDEX(lr_netp1,{R}))))", FMT_PCT,
         "nr_NetSelP1", "Net selection for P+1 cohorts (defaults to the P selection)", False),
    ]
    # ---- resolver block (demoted audit detail; outline-grouped) ----
    res_sec = L.BR_IN_FIRST - 2
    section(ws, res_sec, "B",
            "Selection resolver — the tbl_LR inputs feeding the engines (audit detail)")
    put(ws, f"D{res_sec + 1}",
        "These cells ARE the engine inputs (named nr_*); they resolve the Control "
        "selection against tbl_LR. Grey = machinery, not results.", fnt=F_SMALL_IT)
    label(ws, f"E{res_sec}", "row in tbl_LR")
    formula(ws, f"F{res_sec}", "=IF(nr_SelOK,MATCH(nr_SelKey,lr_key,0),0)", fmt=FMT_INT)
    ws[f"F{res_sec}"].font = font(GREY_DARK, size=9)
    ctx.define("br_selrow", "Bridge", f"$F${res_sec}",
               "tbl_LR row of the selected combo (0 = absent)")

    r = L.BR_IN_FIRST
    for lbl, f, fmt, name, desc, is_link in rows:
        label(ws, f"B{r}", lbl)
        c = (link if is_link else formula)(ws, f"C{r}", f, fmt=fmt, border=BORDER_THIN)
        c.font = font(GREY_DARK, size=10)
        if name:
            ctx.define(name, "Bridge", f"$C${r}", desc)
        r += 1
    # companion notes render the sentinel semantics as text WITHOUT touching the
    # numeric cells the engines consume.
    #
    # Keyed by the row's NAME, never by its position in `rows` (D98). These used
    # to be keyed by index, so inserting the target loss ratio at index 12 slid
    # every later note one row off its label: the A_other warning printed beside
    # "Net trend", and the net-selection note landed on "Mod adjustment
    # effective" — whose cell holds the TEXT "ON"/"OFF", making IF($C{r},...)
    # a permanent #VALUE!. It shipped in all six workbooks and the Checks panel
    # still said PASS, because a stray error in an advisory cell is exactly what
    # no check was watching. A name cannot drift when a row is inserted.
    companions = {
        "nr_MPrior": '=IF(N($C{r})=0,"blank -> backward anchor sits on the '
                     'M_0 -> M_1 line","")',
        "nr_M2": '=IF(N($C{r})=0,"blank -> M_1 carried flat beyond the plan year","")',
        "nr_AOther": '=IF(AND(nr_AOther<>1,nr_AOtherLbl=""),'
                     '"WARNING: A_other <> 1 requires a label","")',
        "nr_NetMode": '=IF($C{r},"NET SELECTION ACTIVE — supersedes planned rows '
                      'from 1/1","explicit rate program")',
    }
    row_of = {nm: L.BR_IN_FIRST + i
              for i, (_l, _f, _fm, nm, _d, _il) in enumerate(rows) if nm}
    for nm, tmpl in companions.items():
        rr_ = row_of[nm]        # KeyError here beats a silently misplaced note
        put(ws, f"D{rr_}", tmpl.format(r=rr_),
            fnt=font(FAIL_RED if nm == "nr_AOther" else GREY_DARK, size=9, italic=True))
    # The resolver block is fixed-origin and the chart staging block below it is
    # too, so the block can only grow into it. It has no slack left after the
    # target row, and silently overwriting the waterfall's data is not a failure
    # anyone would read as one.
    if L.BR_IN_FIRST + len(rows) > L.BR_CHART_DATA - 1:
        raise AssertionError(
            f"Bridge resolver block needs rows {L.BR_IN_FIRST}.."
            f"{L.BR_IN_FIRST + len(rows) - 1} but the waterfall chart data starts at "
            f"{L.BR_CHART_DATA - 1}. Raise Layout.BR_CHART_DATA (and the chart "
            f"ranges that follow it) before adding another resolver input.")
    for rr_ in range(res_sec + 1, L.BR_IN_FIRST + len(rows)):
        ws.row_dimensions[rr_].outlineLevel = 1
        ws.row_dimensions[rr_].hidden = True

    # ---- waterfall table ----
    section(ws, L.BR_WF_HDR - 1, "B", "The bridge — plan year and the indicative following year")
    header_row(ws, L.BR_WF_HDR, 2,
               ["Step", "Factor", "LR after", "Step chg", "Factor +1", "LR after +1",
                "Step chg +1"],
               widths=[30, 11, 12, 11, 11, 13, 11])
    # live year labels (D44)
    for cc, f in ((3, '="Factor "&nr_PlanYear'), (4, '="LR after ("&nr_PlanYear&")"'),
                  (5, '="Step ("&nr_PlanYear&")"'), (6, '="Factor "&(nr_PlanYear+1)'),
                  (7, '="LR after ("&(nr_PlanYear+1)&")"'),
                  (8, '="Step ("&(nr_PlanYear+1)&")"')):
        ws.cell(row=L.BR_WF_HDR, column=cc).value = f
    wf = L.BR_WF_FIRST
    steps = [
        ("Projected LR (as input)", None, "=nr_LRproj", None, "=nr_LRproj"),
        ("Convert to current rate level x(1+s)", '=IF(nr_Basis="proposed",1+nr_SelS,1)', None,
         '=IF(nr_Basis="proposed",1+nr_SelS,1)', None),
        ('="Net trend (applies to "&(nr_PlanYear+1)&" only)"', "=1", None, "=1+nr_Trend", None),
        ('=IF(nr_NetMode,"Rate + price earn-in (net)  A_net","Rate earn-in  A_rate")',
         "=nr_Arate_P", None, "=nr_Arate_P1", None),
        ('=IF(nr_NetMode,"Mod drift (merged into A_net)","Schedule mod drift  A_mod")',
         "=nr_Amod_P", None, "=nr_Amod_P1", None),
        ("Other  A_other", "=nr_AOther", None, "=nr_AOther", None),
    ]
    for i, (lbl, fac_p, lr_p, fac_p1, lr_p1) in enumerate(steps):
        rr = wf + i
        if lbl.startswith("="):
            formula(ws, f"B{rr}", lbl)
            ws[f"B{rr}"].font = F_LABEL
        else:
            label(ws, f"B{rr}", lbl)
        if fac_p:
            formula(ws, f"C{rr}", fac_p, fmt=FMT_IDX, align=ALIGN_C)
            formula(ws, f"D{rr}", f"=$D{rr - 1}*$C{rr}", fmt=FMT_PCT, align=ALIGN_C)
            formula(ws, f"E{rr}", f"=($D{rr}-$D{rr - 1})*100", fmt=PTS_Z, align=ALIGN_C)
            formula(ws, f"F{rr}", fac_p1, fmt=FMT_IDX, align=ALIGN_C)
            formula(ws, f"G{rr}", f"=$G{rr - 1}*$F{rr}", fmt=FMT_PCT, align=ALIGN_C)
            formula(ws, f"H{rr}", f"=($G{rr}-$G{rr - 1})*100", fmt=PTS_Z, align=ALIGN_C)
        else:
            formula(ws, f"D{rr}", lr_p, fmt=FMT_PCT, align=ALIGN_C)
            formula(ws, f"G{rr}", lr_p1, fmt=FMT_PCT, align=ALIGN_C)
    tot = wf + 6
    put(ws, f"B{tot}", f"CY plan loss ratio", fnt=font(NAVY, bold=True), fill=FILL_PANEL)
    formula(ws, f"D{tot}", f"=$D{tot - 1}", fmt=FMT_PCT, align=ALIGN_C, bold=True,
            fill=FILL_PANEL)
    formula(ws, f"E{tot}", f"=($D{tot}-$D{wf})*100", fmt=PTS_Z, align=ALIGN_C, fill=FILL_PANEL)
    formula(ws, f"G{tot}", f"=$G{tot - 1}", fmt=FMT_PCT, align=ALIGN_C, bold=True,
            fill=FILL_PANEL)
    formula(ws, f"H{tot}", f"=($G{tot}-$G{wf})*100", fmt=PTS_Z, align=ALIGN_C, fill=FILL_PANEL)
    ctx.define("nr_CYLR_P", "Bridge", f"$D${tot}", "CY plan loss ratio for the plan year")
    ctx.define("nr_CYLR_P1", "Bridge", f"$G${tot}",
               "Indicative CY LR for plan year + 1 (assumes no new indication)")
    # canonical P+1 caveat (other sheets point here rather than repeating it)
    prose(ws, f"B{tot + 2}",
          "The following-year column is indicative: it holds the indication fixed and "
          "applies the net trend input once (default 0.0% — a visible caveat, not a "
          "forecast).", size=9, width=34)
    ws[f"B{tot + 2}"].font = F_SMALL_IT

    # ---- against target (D96) ----
    # Reference only: nothing upstream reads it, and the gap is stated in points
    # rather than dressed up as a verdict — whether a combo "makes" its target
    # depends on expenses and mix this workbook does not carry.
    formula(ws, f"B{tot + 3}",
            '=IF(N(nr_TargetLR)=0,"",'
            '"Target loss ratio "&TEXT(nr_TargetLR,"0.0%")&"  —  the plan is "'
            '&TEXT(ABS(nr_CYLR_P-nr_TargetLR)*100,"0.00")&" pts "'
            '&IF(nr_CYLR_P>nr_TargetLR,"ABOVE","below")&" it.")')
    ws[f"B{tot + 3}"].font = font(NAVY, bold=True, size=11)
    ws[f"B{tot + 3}"].fill = FILL_PANEL
    for cc in range(3, 9):
        ws.cell(row=tot + 3, column=cc).fill = FILL_PANEL

    # ---- communication metrics (live year labels, D44) ----
    section(ws, tot + 4, "B", "Communication metrics")
    for i, (lbl, f, fmt) in enumerate([
        ('="Target loss ratio (profit provision)"', "=nr_TargetLR", FMT_PCT),
        ('="Plan LR vs target"',
         '=IF(N(nr_TargetLR)=0,"",(nr_CYLR_P-nr_TargetLR)*100)', FMT_PTS_COL),
        ('="CY earned rate chg vs indication level"', "=nr_EChgVsInd", PCT_SIGNED_Z),
        ('="Earned rate chg "&(nr_PlanYear-1)&" -> "&nr_PlanYear', "=nr_YoY_P", PCT_SIGNED_Z),
        ('="Earned rate chg "&nr_PlanYear&" -> "&(nr_PlanYear+1)&" (carryover + new actions)"',
         "=nr_YoY_P1", PCT_SIGNED_Z),
    ]):
        formula(ws, f"B{tot + 5 + i}", lbl)
        ws[f"B{tot + 5 + i}"].font = F_LABEL
        link(ws, f"D{tot + 5 + i}", f, fmt=fmt, align=ALIGN_C)

    # ---- rate changes for this combo (chronological slots via rl_seq) ----
    # the metrics block above now runs tot+5 .. tot+9 (two more since D96), so
    # the chronology starts two rows lower than it used to
    act = tot + 11
    section(ws, act, "B", "Rate changes for this combo — chronological")
    jump(ws, f"F{act}", "'Rate Log'!A1", "edit these on the Rate Log >", size=9)
    header_row(ws, act + 1, 2,
               ["#", "Effective", "Filed %", "Status", "In indication?", "Effective %"],
               widths=None, fill=FILL_GREY, fnt=font(GREY_DARK, bold=True, size=9))
    for j in range(1, 9):
        rr = act + 1 + j
        cnt = f"COUNTIFS(rl_key,nr_SelKey,rl_seq,{j})"
        put(ws, f"B{rr}", j, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
        formula(ws, f"C{rr}", f'=IF({cnt}=0,"",SUMIFS(rl_eff,rl_key,nr_SelKey,rl_seq,{j}))',
                fmt=FMT_DATE_S, align=ALIGN_C)
        formula(ws, f"D{rr}", f'=IF({cnt}=0,"",SUMIFS(rl_filed,rl_key,nr_SelKey,rl_seq,{j}))',
                fmt=FMT_PCT, align=ALIGN_C)
        formula(ws, f"E{rr}",
                f'=IF({cnt}=0,"",IF(COUNTIFS(rl_key,nr_SelKey,rl_seq,{j},rl_status,'
                f'"planned")>0,"planned","taken"))', align=ALIGN_C)
        formula(ws, f"F{rr}",
                f'=IF({cnt}=0,"",IF(COUNTIFS(rl_key,nr_SelKey,rl_seq,{j},rl_cons,"Y")>0,'
                f'"Y","N"))', align=ALIGN_C)
        formula(ws, f"G{rr}", f'=IF({cnt}=0,"",SUMIFS(rl_reff,rl_key,nr_SelKey,rl_seq,{j}))',
                fmt=FMT_PCT_SIGNED, align=ALIGN_C)
    formula(ws, f"B{act + 10}",
            '=IF(COUNTIFS(rl_key,nr_SelKey)>8,"Showing the first 8 of "&'
            'COUNTIFS(rl_key,nr_SelKey)&" changes - see the rate log for the rest.",'
            'IF(nr_NetMode,"Net selection active: planned rows on/after 1/1 of the plan '
            'year are superseded (D39).",""))')
    ws[f"B{act + 10}"].font = F_SMALL_IT
    # D65: the explicit-program counterfactual, shown only under a net selection
    formula(ws, f"B{act + 11}",
            '=IF(NOT(nr_NetMode),"",'
            '"If you booked the logged program instead of asserting the net target, '
            'CY "&nr_PlanYear&" plan LR would be "'
            '&TEXT(IF(br_selrow=0,0,INDEX(calc_cylr_prog,br_selrow)),"0.0%")&" — "'
            '&TEXT(IF(br_selrow=0,0,INDEX(calc_proggap,br_selrow)),"+0.00;-0.00")'
            '&" pts vs the asserted path. Net Delivery closes that gap state by state.")')
    ws[f"B{act + 11}"].font = font(NAVY, italic=True, size=9)

    # ---- waterfall chart data + chart (staging grouped-hidden below) ----
    cd = L.BR_CHART_DATA
    put(ws, f"B{cd - 1}", "Waterfall chart data (formulas — do not edit)", fnt=F_SMALL_IT)
    cats = [("Projected LR", f"$D${wf}", None, True),
            ("Basis", f"$D${wf}", f"$D${wf + 1}", False),
            ('=IF(nr_NetMode,"Net earn-in","Rate earn-in")',
             f"$D${wf + 2}", f"$D${wf + 3}", False),
            ('=IF(nr_NetMode,"Mod (in net)","Mod drift")',
             f"$D${wf + 3}", f"$D${wf + 4}", False),
            ("Other", f"$D${wf + 4}", f"$D${wf + 5}", False),
            ('="CY "&nr_PlanYear&" plan LR"', f"$D${tot}", None, True)]
    for j, h in enumerate(["Step", "base", "up", "down", "total"]):
        put(ws, f"{col(2 + j)}{cd}", h, fnt=font(GREY_DARK, size=9))
    for i, (lbl, a, b, is_total) in enumerate(cats):
        rr = cd + 1 + i
        # live category labels: the chart reads these cells, so net-mode and
        # plan-year relabeling reach the chart too (not just the table)
        if lbl.startswith("="):
            formula(ws, f"B{rr}", lbl)
        else:
            put(ws, f"B{rr}", lbl)
        ws[f"B{rr}"].font = font(GREY_DARK, size=9)
        if is_total:
            formula(ws, f"C{rr}", "=0", fmt=FMT_PCT)
            formula(ws, f"D{rr}", "=0", fmt=FMT_PCT)
            formula(ws, f"E{rr}", "=0", fmt=FMT_PCT)
            formula(ws, f"F{rr}", f"={a}", fmt=FMT_PCT)
        else:
            formula(ws, f"C{rr}", f"=MIN({a},{b})", fmt=FMT_PCT)
            formula(ws, f"D{rr}", f"=MAX(0,{b}-{a})", fmt=FMT_PCT)
            formula(ws, f"E{rr}", f"=MAX(0,{a}-{b})", fmt=FMT_PCT)
            formula(ws, f"F{rr}", "=0", fmt=FMT_PCT)
        for cL in "CDEF":
            ws[f"{cL}{rr}"].font = font(GREY_DARK, size=9)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 60
    data = Reference(ws, min_col=3, min_row=cd, max_col=6, max_row=cd + 6)
    cats_ref = Reference(ws, min_col=2, min_row=cd + 1, max_row=cd + 6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.noFill = True
    chart.series[1].graphicalProperties.solidFill = UP_BAR
    chart.series[2].graphicalProperties.solidFill = DOWN_BAR
    chart.series[3].graphicalProperties.solidFill = TOTAL_BAR
    chart.legend = None
    chart.y_axis.number_format = "0%"
    # Automatic axis (D99): a waterfall on a zero floor does compress the walk,
    # but a baked window cannot follow a paste and read worse in practice.
    _style_chart(chart, "Projected LR -> CY plan loss ratio (stacked-column waterfall)",
                 y_title="Loss ratio", height=9, width=15)
    ws.add_chart(chart, "J5")

    for rr in range(cd - 1, cd + 7):
        ws.row_dimensions[rr].outlineLevel = 1
        ws.row_dimensions[rr].hidden = True

    set_widths(ws, {"A": 2, "B": 34, "C": 13, "D": 12, "E": 11, "F": 12, "G": 13, "H": 11})
    presentation_setup(ws, gridlines_off=True, freeze="A8", tab_color=NAVY)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Portfolio
# ---------------------------------------------------------------------------


def build_portfolio(ctx: Ctx):
    ws = ctx.wb["Portfolio"]
    p = ctx.p
    title(ws, "A1", f"Portfolio — every BU x state ({ctx.lob.name})",
          "Computed simultaneously from the hidden _calc engine blocks (rows align 1:1 with tbl_LR). "
          "Heatmap: green = favorable vs projected.")
    nav_bar(ws, 3, 1, ["Control", "Inputs", "State Summary", "Bridge", "Checks"], step=2)
    header_row(ws, L.PF_HDR, 1,
               ["BU", "State", "Key", "Projected LR (as input)", "Basis",
                "Projected LR (current level)", "Plan LR",
                "Chg vs projected (pts)", "Rate earn-in (A_rate)", "Mod drift (A_mod)",
                "Earned rate chg vs indication", "Carryover", "Plan LR +1",
                "Adj plan EP (000s)", "EP weight",
                "Plan LR — program basis", "Program vs asserted (pts)"],
               widths=[9, 8, 4, 11, 10, 11, 11, 11, 10, 10, 13, 12, 11, 12, 8, 11, 12])
    ws.row_dimensions[L.PF_HDR].height = 42
    # live year labels (D44)
    ws.cell(row=L.PF_HDR, column=7).value = '="CY "&nr_PlanYear&" plan LR"'
    ws.cell(row=L.PF_HDR, column=12).value = '="Carryover into "&(nr_PlanYear+1)'
    ws.cell(row=L.PF_HDR, column=13).value = '="CY "&(nr_PlanYear+1)&" plan LR"'
    for i in range(L.LR_ROWS):
        n = i + 1
        r = L.PF_FIRST + i
        link(ws, f"A{r}", f'=IF(INDEX(lr_key,{n})="","",INDEX(lr_bu,{n}))')
        link(ws, f"B{r}", f'=IF(INDEX(lr_key,{n})="","",INDEX(lr_state,{n}))')
        link(ws, f"C{r}", f"=INDEX(lr_key,{n})")
        ws[f"C{r}"].font = font(GREY_DARK, size=8)
        # Spare rows return "" (never numeric 0): the min-anchored color scales
        # ignore text, so blank capacity rows can't distort the heatmap, and
        # the total rows below use SUM / comma-form SUMPRODUCT, which both
        # treat text as zero.
        link(ws, f"D{r}", f'=IF($C{r}="","",N(INDEX(lr_lrproj,{n})))',
             fmt=FMT_PCT_Z, align=ALIGN_C)
        link(ws, f"E{r}", f'=IF($C{r}="","",INDEX(lr_basis,{n})&"")', align=ALIGN_C)
        link(ws, f"F{r}", f'=IF($C{r}="","",INDEX(calc_lrcur,{n}))',
             fmt=FMT_PCT_Z, align=ALIGN_C)
        link(ws, f"G{r}", f'=IF($C{r}="","",INDEX(calc_cylr_p,{n}))',
             fmt=FMT_PCT_Z, align=ALIGN_C, bold=True)
        link(ws, f"H{r}",
             f'=IF($C{r}="","",(INDEX(calc_cylr_p,{n})-INDEX(calc_lrcur,{n}))*100)',
             fmt=PTS_Z, align=ALIGN_C)
        link(ws, f"I{r}", f'=IF($C{r}="","",INDEX(calc_arate_p,{n}))', fmt=FMT_IDX_Z,
             align=ALIGN_C)
        link(ws, f"J{r}", f'=IF($C{r}="","",INDEX(calc_amod_p,{n}))', fmt=FMT_IDX_Z,
             align=ALIGN_C)
        link(ws, f"K{r}", f'=IF($C{r}="","",INDEX(calc_echg,{n}))', fmt=PCT_SIGNED_Z,
             align=ALIGN_C)
        link(ws, f"L{r}", f'=IF($C{r}="","",INDEX(calc_carry,{n}))', fmt=PCT_SIGNED_Z,
             align=ALIGN_C)
        link(ws, f"M{r}", f'=IF($C{r}="","",INDEX(calc_cylr_p1,{n}))',
             fmt=FMT_PCT_Z, align=ALIGN_C)
        link(ws, f"N{r}", f'=IF($C{r}="","",N(INDEX(lr_ep,{n})))',
             fmt=FMT_EP_Z, align=ALIGN_C)
        formula(ws, f"O{r}", f'=IF($C{r}="","",IF(SUM($N${L.PF_FIRST}:$N${L.PF_LAST})=0,0,'
                             f"N($N{r})/SUM($N${L.PF_FIRST}:$N${L.PF_LAST})))",
                fmt=FMT_PCT_Z, align=ALIGN_C)
        ws[f"O{r}"].font = font(GREY_DARK, size=9)
        # D65: the explicit rate x mod counterfactual, and the gap a net
        # selection opens ("—" where there is no assertion to compare against)
        link(ws, f"P{r}", f'=IF($C{r}="","",INDEX(calc_cylr_prog,{n}))',
             fmt=FMT_PCT_Z, align=ALIGN_C)
        formula(ws, f"Q{r}",
                f'=IF($C{r}="","",IF(INDEX(calc_netmode,{n})=0,"—",'
                f"INDEX(calc_proggap,{n})))", fmt=FMT_PTS_COL, align=ALIGN_C)

    f0, f1 = L.PF_FIRST, L.PF_LAST
    wr, sr = L.PF_W_TOTAL, L.PF_S_TOTAL
    put(ws, f"B{wr}", "EP-weighted total *", fnt=font(NAVY, bold=True), fill=FILL_PANEL)
    put(ws, f"B{sr}", "Simple average", fnt=font(GREY_DARK, bold=True))
    epsum = f"SUM($N${f0}:$N${f1})"
    cnt = f'SUMPRODUCT(($C${f0}:$C${f1}<>"")*1)'
    for cL, fmt in (("F", FMT_PCT), ("G", FMT_PCT), ("K", PCT_SIGNED_Z), ("L", PCT_SIGNED_Z),
                    ("M", FMT_PCT)):
        formula(ws, f"{cL}{wr}",
                f'=IF({epsum}=0,"n/a",SUMPRODUCT({cL}${f0}:{cL}${f1},$N${f0}:$N${f1})/{epsum})',
                fmt=fmt, align=ALIGN_C, bold=(cL == "G"), fill=FILL_PANEL)
        formula(ws, f"{cL}{sr}",
                f'=IF({cnt}=0,"n/a",SUM({cL}${f0}:{cL}${f1})/{cnt})',
                fmt=fmt, align=ALIGN_C)
    formula(ws, f"H{wr}", f'=IF(ISNUMBER($G${wr}),($G${wr}-$F${wr})*100,"")', fmt=PTS_Z,
            align=ALIGN_C, fill=FILL_PANEL)
    formula(ws, f"N{wr}", f"={epsum}", fmt=FMT_EP_Z, align=ALIGN_C, fill=FILL_PANEL)
    note(ws, f"B{sr + 2}",
         "* Weighted by plan earned premium over the rows where EP is provided; rows without "
         "EP carry zero weight. The simple average covers every populated row equally.")

    # ---- ranking helpers (append-right; the col-7 header label never moves) ----
    put(ws, f"AA{L.PF_HDR}", "Rank (adverse)", fnt=font(GREY_DARK, bold=True, size=9),
        fill=FILL_GREY)
    put(ws, f"AB{L.PF_HDR}", "rk helper", fnt=font(GREY_DARK, size=8))
    put(ws, f"AC{L.PF_HDR}", "Contribution (pts)", fnt=font(GREY_DARK, bold=True, size=9),
        fill=FILL_GREY)
    for i in range(L.LR_ROWS):
        r = L.PF_FIRST + i
        formula(ws, f"AA{r}", f'=IF($C{r}="","",RANK($H{r},$H${f0}:$H${f1}))',
                fmt=FMT_INT, align=ALIGN_C)
        # epsilon tiebreak so LARGE/MATCH never lands on the same row twice.
        # The tiebreak is what makes it unreadable — General renders it as
        # 0.5000000371 — so it prints like the points column it is derived from.
        formula(ws, f"AB{r}", f'=IF($C{r}="","",$H{r}+ROW()/1000000000)',
                fmt=FMT_PTS_COL)
        formula(ws, f"AC{r}", f'=IF(OR($C{r}="",$O{r}=""),"",N($O{r})*$H{r})',
                fmt=FMT_PTS_COL, align=ALIGN_C)
        for cL in ("AA", "AB", "AC"):
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=9)

    # ---- Decision Board: the three answers a review meeting opens with ----
    section(ws, 4, "T", "Decision Board")
    put(ws, "T5", "Top 10 adverse movers (plan vs projected)", fnt=font(NAVY, bold=True,
                                                                        size=10))
    header_row(ws, 6, 20, ["Combo", "Plan LR", "Chg (pts)", "Contribution (pts)"],
               widths=[12, 10, 10, 13], fill=FILL_GREY,
               fnt=font(GREY_DARK, bold=True, size=9))
    for k in range(1, 11):
        r = 6 + k
        formula(ws, f"Y{r}",
                f"=IF(COUNT($AB${f0}:$AB${f1})<{k},0,"
                f"MATCH(LARGE($AB${f0}:$AB${f1},{k}),$AB${f0}:$AB${f1},0))", fmt=FMT_INT)
        ws[f"Y{r}"].font = font(GREY_DARK, size=8)
        formula(ws, f"T{r}", f'=IF($Y{r}=0,"",INDEX($C${f0}:$C${f1},$Y{r}))', align=ALIGN_C)
        formula(ws, f"U{r}", f'=IF($Y{r}=0,"",INDEX($G${f0}:$G${f1},$Y{r}))',
                fmt=FMT_PCT_Z, align=ALIGN_C)
        formula(ws, f"V{r}", f'=IF($Y{r}=0,"",INDEX($H${f0}:$H${f1},$Y{r}))',
                fmt=FMT_PTS_COL, align=ALIGN_C)
        formula(ws, f"W{r}", f'=IF($Y{r}=0,"",INDEX($AC${f0}:$AC${f1},$Y{r}))',
                fmt=FMT_PTS_COL, align=ALIGN_C, bold=True)
    put(ws, "T18",
        "Contribution = EP weight x change vs projected: which combos move the BOOK.",
        fnt=F_SMALL_IT)

    put(ws, "T20", "Portfolio bridge (EP-weighted)", fnt=font(NAVY, bold=True, size=10))
    pf_bridge = [
        ("Projected LR (current level)", '=IF(SUM(calc_ep)=0,"n/a",SUM(calc_w_lrcur)/SUM(calc_ep))',
         FMT_PCT, False),
        ("x  rate earn-in (avg factor)", '=IF(SUM(calc_ep)=0,"n/a",SUM(calc_w_arate)/SUM(calc_ep))',
         FMT_IDX, False),
        ("x  mod drift (avg factor)", '=IF(SUM(calc_ep)=0,"n/a",SUM(calc_w_amod)/SUM(calc_ep))',
         FMT_IDX, False),
        ("Mix / interaction (pts)",
         f'=IF(OR(SUM(calc_ep)=0,NOT(ISNUMBER($G${wr}))),"n/a",($G${wr}-$U$21*$U$22*$U$23)*100)',
         FMT_PTS_COL, False),
        ("CY plan LR — exact EP-weighted total", f"=$G${wr}", FMT_PCT, True),
    ]
    for i, (lbl, f, fmt, bold_) in enumerate(pf_bridge):
        r = 21 + i
        label(ws, f"T{r}", lbl, bold=bold_)
        formula(ws, f"U{r}", f, fmt=fmt, align=ALIGN_C, bold=bold_,
                fill=FILL_PANEL if bold_ else None)
    put(ws, "T27",
        "Factor averages don't compound exactly across a mixed book — the mix line is "
        "that honest residual.", fnt=F_SMALL_IT)
    jump(ws, "T28", f"'{SHEETS.PROGRAM_FLOW}'!A1",
         "Monthly flow by state and for the book > Program Flow", size=9)

    tor = BarChart()
    tor.type = "bar"
    tor.gapWidth = 40
    tor.add_data(Reference(ws, min_col=23, min_row=6, max_row=16), titles_from_data=True)
    tor.set_categories(Reference(ws, min_col=20, min_row=7, max_row=16))
    tor.series[0].graphicalProperties.solidFill = STEEL
    tor.legend = None
    tor.y_axis.number_format = "0.00"
    _style_chart(tor, "Who moves the book — EP-weighted contribution (pts)",
                 y_title="Contribution (pts)", height=9, width=11)
    ws.add_chart(tor, "T29")

    # The ranking machinery is collapsible: it sits right of the exhibit with no
    # explanation of its own, and only the Decision Board reads it. Grouped, not
    # hidden — the outline button says "there is more here" without asserting it
    # is interesting (same treatment as the Rate Engine's index columns).
    for cL in ("AA", "AB", "AC"):
        ws.column_dimensions[cL].outlineLevel = 1

    ws.auto_filter.ref = f"A{L.PF_HDR}:Q{f1}"
    put(ws, f"B{L.PF_S_TOTAL + 4}",
        "Filter freely; avoid SORTING this grid (rows are formulas aligned 1:1 with "
        "tbl_LR) — the Decision Board gives the sanctioned ranked view.", fnt=F_SMALL_IT)

    ws.conditional_formatting.add(
        f"G{f0}:G{f1}",
        ColorScaleRule(start_type="min", start_color="D6E8D5",
                       mid_type="percentile", mid_value=50, mid_color="FFF2CC",
                       end_type="max", end_color="F4B8B8"))
    ws.conditional_formatting.add(
        f"H{f0}:H{f1}",
        ColorScaleRule(start_type="min", start_color="D6E8D5",
                       mid_type="num", mid_value=0, mid_color="FFFFFF",
                       end_type="max", end_color="F4B8B8"))
    for cL in ("I", "J"):
        ws.conditional_formatting.add(
            f"{cL}{f0}:{cL}{f1}",
            ColorScaleRule(start_type="num", start_value=0.95, start_color="D6E8D5",
                           mid_type="num", mid_value=1.0, mid_color="FFFFFF",
                           end_type="num", end_value=1.05, end_color="F4B8B8"))
    presentation_setup(ws, gridlines_off=True, freeze=f"D{L.PF_FIRST}", tab_color=NAVY)
    print_setup(ws)


# ---------------------------------------------------------------------------
# State Summary — the flagship leadership exhibit (D38)
# ---------------------------------------------------------------------------

SS_CAPTION_FILL = "EDF1F5"

# The two long group captions, named so the Book's mirror uses the same words.
SS_G_BRIDGE = '="CY "&nr_PlanYear&" plan — the bridge, left to right"'
SS_G_HIST = ("Rate change history — chronological (single-BU view). "
             "T = taken, P = planned, * = not counted in the indication")
SS_G_P1 = '="CY "&(nr_PlanYear+1)&" indicative"'
SS_G_NET = "Net selection — asserted vs the logged program"


@dataclass(frozen=True)
class SsCol:
    """One State Summary column, declared once (D90).

    Before this, the exhibit was six parallel structures keyed by POSITION:
    the header list, the width list, the group spans, the metric map, the slot
    arithmetic (``6 + j * 3``), and every hard-coded letter in the bridge chain,
    the conditional formats, the Book's independent mirror and six harness
    reads. Moving one column meant editing all of them in lockstep, and a miss
    did not fail the build — it produced a correct-looking exhibit with one
    column of the wrong thing in it. Same lesson as ``LrCol`` (D72), applied to
    the flagship.
    """

    key: str                    # stable identity — what every consumer addresses
    header: str
    width: float
    group: str = ""             # two-tier caption; equal ADJACENT values span
    live: str | None = None     # formula header, for year-bearing labels (D44)
    level: int = 0              # Excel column outline level; 0 = always visible


def _ss_slots():
    for j in (1, 2, 3, 4):
        yield SsCol(f"chg{j}_date", f"Chg {j} date", 9, SS_G_HIST, level=SS_LV_HIST)
        yield SsCol(f"chg{j}_pct", f"Chg {j}", 7, SS_G_HIST, level=SS_LV_HIST)
        yield SsCol(f"chg{j}_tok", "T/P", 4, SS_G_HIST, level=SS_LV_HIST)


# The exhibit reads as a DERIVATION, not as a dashboard (D101, user direction):
# what we filed, what we assumed, what that earns — therefore this loss ratio.
# The inputs sit to the LEFT of the bridge because the bridge chain is itself a
# left-to-right product and its inputs belong on the same axis.
#
# D90 put the bridge first instead, to rescue a headline stranded at ~1,600px.
# That problem is real but reordering was the wrong cure: COLLAPSING the inputs
# solves it without breaking the argument's order. Two nested outline levels, so
# the toggles are independent — contiguous columns at one level would merge into
# a single group:
#   level 2  the rate-change chronology (12 columns), collapsed by default
#   level 1  the whole input region, so one more click leaves State | EP | Plan LR
# Plan LR lands at ~949px as shipped (unchanged), ~537px with inputs collapsed.
SS_LV_HIST = 2       # inner group: the chronology alone
SS_LV_INPUT = 1      # outer group: every input feeding the bridge

SS_COLS: tuple[SsCol, ...] = (
    SsCol("state", "State", 7),
    SsCol("ep", "Adj plan EP (000s)", 12, "Volume"),
    # ---- inputs, in the order the engine consumes them --------------------
    *_ss_slots(),
    SsCol("ntaken", "# taken", 7, SS_G_HIST, level=SS_LV_HIST),
    SsCol("nplanned", "# planned", 8, SS_G_HIST, level=SS_LV_HIST),
    SsCol("mind", "Mod in indication", 9, "Schedule mods", level=SS_LV_INPUT),
    SsCol("m0", "Current mod", 8.5, "Schedule mods", level=SS_LV_INPUT),
    SsCol("m1", "Proj. mod, plan-yr end", 9, "Schedule mods", level=SS_LV_INPUT),
    SsCol("crl", "Indication rate level", 9.5, "Engine levels", level=SS_LV_INPUT),
    SsCol("ecy", "Earned rate level", 9.5, "Engine levels",
          '=nr_PlanYear&" earned rate level"', level=SS_LV_INPUT),
    SsCol("mbar", "Earned mod", 9, "Engine levels", '=nr_PlanYear&" earned mod"',
          level=SS_LV_INPUT),
    # ---- the answer -------------------------------------------------------
    SsCol("lrcur", "Projected LR (current level)", 11, SS_G_BRIDGE),
    SsCol("arate", "x  Rate earn-in (A_rate)", 9.5, SS_G_BRIDGE),
    SsCol("amod", "x  Mod drift (A_mod)", 9.5, SS_G_BRIDGE),
    SsCol("aother", "x  Other adj (A_other)", 8.5, SS_G_BRIDGE),
    SsCol("planlr", "=  Plan LR", 10, SS_G_BRIDGE,
          '="=  CY "&nr_PlanYear&" plan LR"'),
    # the benchmark, immediately beside the answer (D96) — reference only, so it
    # sits in the bridge band without joining the product
    SsCol("target", "Target LR", 9, SS_G_BRIDGE),
    SsCol("mix", "Mix (pts)", 7, SS_G_BRIDGE),
    # ---- and the same answer for the following year, beside it (D101) -----
    SsCol("trend", "Net trend", 8.5, SS_G_P1, '="Net trend ("&(nr_PlanYear+1)&")"'),
    SsCol("arate1", "Rate earn-in +1", 9.5, SS_G_P1,
          '=(nr_PlanYear+1)&" rate earn-in"'),
    SsCol("amod1", "Mod drift +1", 9.5, SS_G_P1, '=(nr_PlanYear+1)&" mod drift"'),
    SsCol("planlr1", "Plan LR +1", 10, SS_G_P1,
          '="CY "&(nr_PlanYear+1)&" plan LR"'),
    SsCol("netsel", "Net rate sel", 8.5, SS_G_NET),
    SsCol("progbasis", "Plan LR — program basis", 11, SS_G_NET),
    SsCol("proggap", "Program vs asserted (pts)", 12, SS_G_NET),
)

SS_COL = {c.key: i for i, c in enumerate(SS_COLS, 1)}    # key -> 1-based column
SS_LAST = len(SS_COLS)
# helper columns live to the right of the exhibit proper
SS_HELP = {k: SS_LAST + 1 + i for i, k in
           enumerate(("pad", "crit", "cnt", "key", "netcnt", "mathchk"))}


def ss_c(key: str) -> int:
    """1-based column index of a State Summary column, by key."""
    return SS_COL[key]


def ss_l(key: str) -> str:
    """Column letter of a State Summary column, by key."""
    return col(SS_COL[key])


def ss_h(key: str) -> str:
    """Column letter of a State Summary HELPER column, by key."""
    return col(SS_HELP[key])


def ss_groups() -> list[tuple[int, int, str]]:
    """(first, last, caption) runs, derived from the declared order — so a
    reordered column takes its group span with it."""
    runs: list[list] = []
    for i, c in enumerate(SS_COLS, 1):
        if runs and runs[-1][2] == c.group:
            runs[-1][1] = i
        else:
            runs.append([i, i, c.group])
    return [tuple(r) for r in runs]


def ss_outline(ws, first_col: int = 1) -> None:
    """Apply the declared column outline levels (D101).

    Shared by the per-LOB exhibit and the Book's mirror so the two cannot drift
    into different collapse behaviour. ``first_col`` exists because a mirror may
    start the map at a different column.

    The chronology is hidden as well as grouped: an outline that ships expanded
    puts Plan LR back off the first screen, which is the whole problem D90 hit.
    Excel draws the +/- button to the RIGHT of a column group by default, which
    is where the bridge is — so the control sits against the thing it reveals.
    """
    for i, c in enumerate(SS_COLS, first_col):
        if c.level:
            dim = ws.column_dimensions[col(i)]
            dim.outlineLevel = c.level
            if c.level >= SS_LV_HIST:
                dim.hidden = True


def ss_group_starts() -> set[int]:
    """Columns that open a group — where the medium divider rule is drawn.
    The State/Volume boundary is deliberately not one: a state and its premium
    read as one label."""
    return {first for first, _last, cap in ss_groups()[2:] if cap}


def build_state_summary(ctx: Ctx):
    """One row per state: adjusted EP, mods, engine levels, the visible bridge
    through to the CY P plan LR, then the chronological rate-change history and
    the P+1 view — filtered by BU, with an 'All' view that combines business
    units on adjusted-EP weights.

    Every column position comes from :data:`SS_COLS`; nothing here counts
    columns by hand (D90)."""
    ws = ctx.wb["State Summary"]
    p = ctx.p
    states = ctx.cfg.states
    title(ws, "A1", f"State Summary — {ctx.lob.name}",
          "Plan-year view with the indicative following year. One row per state; choose a "
          "business unit or view all combined (adjusted-EP weighted). Every figure traces to "
          "the hidden _calc engine blocks and the Inputs tables.")
    nav_bar(ws, 3, 1, ["Control", "Inputs", "Portfolio", "Bridge", "Checks"], step=2)

    # ---- filter chip ----
    label(ws, "A4", "Business unit view", bold=True)
    c = input_cell(ws, "B4", "All")
    c.alignment = ALIGN_C
    c.border = Border(*(Side(style="medium", color=NAVY),) * 4)
    formula(ws, "C4",
            '="Showing: "&IF($B$4="All","all business units combined (adj-EP weighted)",$B$4)')
    ws["C4"].font = font(GREY_DARK, italic=True)
    ctx.define("nr_SumBU", "State Summary", "$B$4",
               "State Summary business-unit filter (All = EP-weighted across BUs)")
    # criteria helper: SUMIFS wildcard when 'All'
    CR, CN, KY, NC_, MC = (ss_h("crit"), ss_h("cnt"), ss_h("key"),
                           ss_h("netcnt"), ss_h("mathchk"))
    formula(ws, f"{CR}4", '=IF($B$4="All","*",$B$4)')
    ws[f"{CR}4"].font = font(GREY_DARK, size=8)
    add_dv(ws, "list", ["B4"], formula1="=lst_bu_all")

    cap_row, hdr, first = 6, 7, 8
    n_disp = len(states) + 3   # live-roster slots incl. headroom for new states (D42)
    last = first + n_disp - 1
    tot = last + 2
    crit = f"${CR}$4"

    # ---- two-tier header ----
    for c1, c2, text in ss_groups():
        for cc in range(c1, c2 + 1):
            cell = ws.cell(row=cap_row, column=cc, value=text if cc == c1 else None)
            cell.alignment = Alignment(horizontal="centerContinuous", vertical="center")
            cell.font = font(NAVY, bold=True, size=9)
            cell.fill = PatternFill_safe(SS_CAPTION_FILL)
    header_row(ws, hdr, 1, [c.header for c in SS_COLS],
               widths=[c.width for c in SS_COLS])
    ws.row_dimensions[hdr].height = 42
    # year-bearing headers are LIVE so a Control plan-year change relabels
    # the exhibit, not just the math (D44)
    for i, spec in enumerate(SS_COLS, 1):
        if spec.live:
            ws.cell(row=hdr, column=i).value = spec.live
    group_starts = ss_group_starts()
    med = Side(style="medium", color=STEEL)

    def wtd_body(row_, prod, plain, plain_dims=("calc_state", "calc_bu")):
        """EP-weighted mean, falling back to a simple average when EP is 0."""
        ps, pb = plain_dims
        return (f"IF($B{row_}>0,"
                f"SUMIFS({prod},calc_state,$A{row_},calc_bu,{crit})/$B{row_},"
                f"SUMIFS({plain},{ps},$A{row_},{pb},{crit})/${CN}{row_})")

    def wtd(row_, prod, plain, plain_dims=("calc_state", "calc_bu")):
        return f'=IF(${CN}{row_}=0,"",{wtd_body(row_, prod, plain, plain_dims)})'

    # the visible bridge chain: Projected LR x A_rate x A_mod x A_other
    CHAIN_P = tuple(f"${ss_l(k)}{{r}}"
                    for k in ("lrcur", "arate", "amod", "aother"))
    CHAIN_P1 = (f"${ss_l('lrcur')}{{r}}", f"(1+${ss_l('trend')}{{r}})",
                f"${ss_l('arate1')}{{r}}", f"${ss_l('amod1')}{{r}}",
                f"${ss_l('aother')}{{r}}")

    def chain(row_, parts):
        return "*".join(x.format(r=row_) for x in parts)

    metric_cols = [
        ("mind", "calc_w_mind", "lr_mind", ("lr_state", "lr_bu"), FMT_MOD),
        ("m0", "calc_w_m0", "lr_m0", ("lr_state", "lr_bu"), FMT_MOD),
        ("m1", "calc_w_m1", "lr_m1", ("lr_state", "lr_bu"), FMT_MOD),
        ("crl", "calc_w_crl", "calc_crl", None, FMT_IDX),
        ("ecy", "calc_w_ecy", "calc_ecy_p", None, FMT_IDX),
        ("mbar", "calc_w_mbar", "calc_mbar_p", None, FMT_MOD),
        ("lrcur", "calc_w_lrcur", "calc_lrcur", None, FMT_PCT),
        ("arate", "calc_w_arate", "calc_arate_p", None, FMT_IDX),
        ("amod", "calc_w_amod", "calc_amod_p", None, FMT_IDX),
        ("aother", "calc_w_aother", "calc_aother", None, FMT_IDX),
        ("target", "calc_w_target", "lr_target", ("lr_state", "lr_bu"), FMT_PCT),
        ("trend", "calc_w_trend", "calc_trend", None, FMT_PCT_SIGNED),
        ("arate1", "calc_w_arate1", "calc_arate_p1", None, FMT_IDX),
        ("amod1", "calc_w_amod1", "calc_amod_p1", None, FMT_IDX),
    ]

    for i in range(n_disp):
        r = first + i
        band = FILL_PANEL if i % 2 else None
        # live roster: the k-th distinct state currently in tbl_LR (D42)
        link(ws, f"A{r}", f"=_lists!$B${3 + i}", align=ALIGN_C, fill=band,
             border=BORDER_THIN, bold=True)
        formula(ws, f"{CN}{r}", f"=COUNTIFS(calc_state,$A{r},calc_bu,{crit})", fmt=FMT_INT)
        formula(ws, f"{KY}{r}", f'=IF($B$4="All","",$B$4&"|"&$A{r})')
        formula(ws, f"B{r}",
                f'=IF(${CN}{r}=0,"",SUMIFS(calc_ep,calc_state,$A{r},calc_bu,{crit}))',
                fmt="#,##0", align=ALIGN_C, fill=band, border=BORDER_THIN)
        for key, prod, plain, dims, fmt in metric_cols:
            f = wtd(r, prod, plain, dims) if dims else wtd(r, prod, plain)
            formula(ws, f"{ss_l(key)}{r}", f, fmt=fmt, align=ALIGN_C,
                    fill=band, border=BORDER_THIN)
        # rate-change slots: date, effective %, and the status token (D61) —
        # T/P for taken/planned, * when the row is outside the indication
        for j in range(1, 5):
            # the LAST four, not the first (D95): a combo with six filings keeps
            # its plan-year actions at ranks 5 and 6, and those are the ones a
            # planning exhibit exists to show
            sq = slot_rank(f"${KY}{r}", j)
            cnt = f"COUNTIFS(rl_key,${KY}{r},rl_seq,{sq})"
            formula(ws, f"{ss_l(f'chg{j}_date')}{r}",
                    f'=IF(${KY}{r}="","",IF({cnt}=0,"",'
                    f"SUMIFS(rl_eff,rl_key,${KY}{r},rl_seq,{sq})))",
                    fmt=FMT_DATE_S, align=ALIGN_C, fill=band, border=BORDER_THIN)
            formula(ws, f"{ss_l(f'chg{j}_pct')}{r}",
                    f'=IF(${KY}{r}="","",IF({cnt}=0,"",'
                    f"SUMIFS(rl_reff,rl_key,${KY}{r},rl_seq,{sq})))",
                    fmt=FMT_PCT_SIGNED, align=ALIGN_C, fill=band, border=BORDER_THIN)
            formula(ws, f"{ss_l(f'chg{j}_tok')}{r}",
                    f'=IF(${KY}{r}="","",IF({cnt}=0,"",'
                    f'IF(COUNTIFS(rl_key,${KY}{r},rl_seq,{sq},rl_status,"planned")>0,'
                    f'"P","T")&'
                    f'IF(COUNTIFS(rl_key,${KY}{r},rl_seq,{sq},rl_cons,"Y")>0,"","*")))',
                    align=ALIGN_C, fill=band, border=BORDER_THIN)
        # counts split by status — these work in the All view too, where the
        # per-change slots are necessarily blank (they are single-BU)
        for key, status in (("ntaken", "taken"), ("nplanned", "planned")):
            formula(ws, f"{ss_l(key)}{r}",
                    f'=IF(${CN}{r}=0,"",IF(${KY}{r}="",'
                    f'COUNTIFS(rl_state,$A{r},rl_status,"{status}",rl_eff,"<>"),'
                    f'COUNTIFS(rl_key,${KY}{r},rl_status,"{status}",rl_eff,"<>")))',
                    fmt=FMT_INT, align=ALIGN_C, fill=band, border=BORDER_THIN)
        # CY P and CY P+1 plan LR: on a single-combo row the bridge is visible
        # arithmetic across this row's own factor cells — no lookup. Where a row
        # genuinely combines combos, the EP-weighted engine value is exact and
        # the product of the displayed averages is not (mix).
        for key, parts, wprod, wplain in (
                ("planlr", CHAIN_P, "calc_w_cylr", "calc_cylr_p"),
                ("planlr1", CHAIN_P1, "calc_w_cylr1", "calc_cylr_p1")):
            formula(ws, f"{ss_l(key)}{r}",
                    f'=IF(${CN}{r}=0,"",IF(${CN}{r}=1,{chain(r, parts)},'
                    f"{wtd_body(r, wprod, wplain)}))",
                    fmt=FMT_PCT, align=ALIGN_C, fill=band, border=BORDER_THIN, bold=True)
        formula(ws, f"{ss_l('mix')}{r}",
                f'=IF(${CN}{r}<=1,"",(${ss_l("planlr")}{r}-{chain(r, CHAIN_P)})*100)',
                fmt=FMT_PTS_COL, align=ALIGN_C, fill=band, border=BORDER_THIN)
        formula(ws, f"{NC_}{r}", f"=SUMIFS(calc_netmode,calc_state,$A{r},calc_bu,{crit})",
                fmt=FMT_INT)
        formula(ws, f"{ss_l('netsel')}{r}",
                f'=IF(${CN}{r}=0,"",IF(${NC_}{r}=0,"—",'
                f"SUMIFS(calc_netx,calc_state,$A{r},calc_bu,{crit})/${NC_}{r}))",
                fmt=FMT_PCT_SIGNED, align=ALIGN_C, fill=band, border=BORDER_THIN)
        # D65: the same rows valued on the EXPLICIT rate x mod paths, and the
        # gap that opens only where a net target is asserted
        formula(ws, f"{ss_l('progbasis')}{r}",
                f'=IF(${CN}{r}=0,"",IF(${NC_}{r}=0,"—",'
                f"{wtd_body(r, 'calc_w_cylr_prog', 'calc_cylr_prog')}))",
                fmt=FMT_PCT, align=ALIGN_C, fill=band, border=BORDER_THIN)
        formula(ws, f"{ss_l('proggap')}{r}",
                f'=IF(${CN}{r}=0,"",IF(${NC_}{r}=0,"—",'
                f'(${ss_l("progbasis")}{r}-${ss_l("planlr")}{r})*100))',
                fmt=FMT_PTS_COL, align=ALIGN_C, fill=band, border=BORDER_THIN,
                bold=True)
        # audit helper: the visible product vs the engine's weighted value on
        # single-combo rows (0 elsewhere) — the Checks cross-tie, never circular
        # because the right-hand side is the engine, not this row's cells
        formula(ws, f"{MC}{r}",
                f"=IF(${CN}{r}<>1,0,"
                f"ABS({wtd_body(r, 'calc_w_cylr', 'calc_cylr_p')}-{chain(r, CHAIN_P)})"
                f"+ABS({wtd_body(r, 'calc_w_cylr1', 'calc_cylr_p1')}"
                f"-{chain(r, CHAIN_P1)}))", fmt=FMT_IDX)
        for cL in (ss_h("pad"), CR, CN, KY):
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=8)

    # ---- total band ----
    put(ws, f"A{tot}", "TOTAL", fnt=font(NAVY, bold=True), align=ALIGN_C,
        fill=FILL_STEEL_LIGHT_safe(), border=Border(top=med))
    formula(ws, f"B{tot}", f"=SUMIFS(calc_ep,calc_bu,{crit})", fmt="#,##0", align=ALIGN_C,
            bold=True, fill=FILL_STEEL_LIGHT_safe(), border=Border(top=med))
    for key, prod, plain, dims, fmt in metric_cols:
        formula(ws, f"{ss_l(key)}{tot}",
                f'=IF($B${tot}=0,"n/a",SUMIFS({prod},calc_bu,{crit})/$B${tot})',
                fmt=fmt, align=ALIGN_C,
                fill=FILL_STEEL_LIGHT_safe(), border=Border(top=med))
    # the book total always uses the exact EP-weighted engine value; the mix
    # cell shows what the displayed factors leave on the table
    for key, prod in (("planlr", "calc_w_cylr"), ("planlr1", "calc_w_cylr1")):
        formula(ws, f"{ss_l(key)}{tot}",
                f'=IF($B${tot}=0,"n/a",SUMIFS({prod},calc_bu,{crit})/$B${tot})',
                fmt=FMT_PCT, align=ALIGN_C, bold=True,
                fill=FILL_STEEL_LIGHT_safe(), border=Border(top=med))
    formula(ws, f"{ss_l('mix')}{tot}",
            f'=IF($B${tot}=0,"",(${ss_l("planlr")}{tot}-{chain(tot, CHAIN_P)})*100)',
            fmt=FMT_PTS_COL, align=ALIGN_C, fill=FILL_STEEL_LIGHT_safe(),
            border=Border(top=med))
    for key, status in (("ntaken", "taken"), ("nplanned", "planned")):
        formula(ws, f"{ss_l(key)}{tot}",
                f'=IF($B$4="All",COUNTIFS(rl_status,"{status}",rl_eff,"<>"),'
                f'COUNTIFS(rl_bu,$B$4,rl_status,"{status}",rl_eff,"<>"))',
                fmt=FMT_INT, align=ALIGN_C, fill=FILL_STEEL_LIGHT_safe(),
                border=Border(top=med))
    formula(ws, f"{ss_l('netsel')}{tot}",
            f'=IF(SUMIFS(calc_netmode,calc_bu,{crit})=0,"—",'
            f"SUMIFS(calc_netx,calc_bu,{crit})/SUMIFS(calc_netmode,calc_bu,{crit}))",
            fmt=FMT_PCT_SIGNED, align=ALIGN_C, fill=FILL_STEEL_LIGHT_safe(),
            border=Border(top=med))
    formula(ws, f"{ss_l('progbasis')}{tot}",
            f'=IF($B${tot}=0,"n/a",SUMIFS(calc_w_cylr_prog,calc_bu,{crit})/$B${tot})',
            fmt=FMT_PCT, align=ALIGN_C, fill=FILL_STEEL_LIGHT_safe(),
            border=Border(top=med))
    formula(ws, f"{ss_l('proggap')}{tot}",
            f'=IF($B${tot}=0,"",(${ss_l("progbasis")}{tot}-${ss_l("planlr")}{tot})*100)',
            fmt=FMT_PTS_COL, align=ALIGN_C, bold=True,
            fill=FILL_STEEL_LIGHT_safe(), border=Border(top=med))
    # the chronology band carries no total: four dated changes do not add up
    for j in range(1, 5):
        for part in ("date", "pct", "tok"):
            put(ws, f"{ss_l(f'chg{j}_{part}')}{tot}", None,
                fill=FILL_STEEL_LIGHT_safe(), border=Border(top=med))

    # group divider borders down the table
    for cc in group_starts:
        for r in range(cap_row, tot + 1):
            cell = ws.cell(row=r, column=cc)
            b = cell.border
            cell.border = Border(left=med, right=b.right, top=b.top, bottom=b.bottom)

    # conditional formatting: LR heatmaps + EP data bars
    for key in ("planlr", "planlr1"):
        cL = ss_l(key)
        ws.conditional_formatting.add(
            f"{cL}{first}:{cL}{last}",
            ColorScaleRule(start_type="min", start_color="D6E8D5",
                           mid_type="percentile", mid_value=50, mid_color="FFF2CC",
                           end_type="max", end_color="F4B8B8"))
    ws.conditional_formatting.add(
        f"B{first}:B{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color="8497B0", showValue=True))
    # planned changes read amber italic; taken stay plain (the token still
    # prints in black and white, so colour is reinforcement, not the message)
    for j in range(1, 5):
        cl = ss_l(f"chg{j}_tok")
        ws.conditional_formatting.add(
            f"{cl}{first}:{cl}{last}",
            FormulaRule(formula=[f'LEFT(${cl}{first},1)="P"'],
                        font=Font(name=ctx.cfg.font, size=9, bold=True,
                                  italic=True, color=WARN_AMBER)))

    # Short single-line footnotes (overflow display, never ribbon-wrapped);
    # the deep documentation lives on Read Me (glossary) and Walkthrough
    # (worked example) — this exhibit stays lean.
    notes = [
        "Weighted view: every metric is adjusted-plan-EP weighted across the business "
        "units in view (states with no EP fall back to a simple average).",
        "Plan LR is the product of the four factors to its left, multiplied on this row — "
        "on rows that combine business units those factors are EP-weighted averages that "
        "do not multiply exactly, so the exact weighted value is shown and Mix is the "
        "difference. Displayed factors are rounded; the cells carry full precision.",
        "Rate change history is chronological and single-BU: T = taken, P = planned, "
        "* = NOT counted in the indication; percents are filed % x achievement. The "
        "# taken / # planned counts work in every view; past four changes see the rate log.",
        "The following year is indicative — see the canonical caveat on the Bridge. "
        "It multiplies the same Projected LR and Other adj by (1 + net trend).",
        "Net sel: the net selection in force (— = explicit program); net combos carry the "
        "combined factor in A_rate with A_mod = 1.000.",
    ]
    for i, text in enumerate(notes):
        put(ws, f"A{tot + 2 + i}", text, fnt=F_SMALL_IT)

    d = tot + 2 + len(notes) + 2
    section(ws, d, "A", "How this exhibit is calculated")
    d += 1
    put(ws, f"A{d}",
        "Every figure below the headers is calculated, never typed. Each state row is the "
        "adjusted-EP-weighted combination of its BU x state engine results for the business "
        "units in view.", fnt=font(GREY_DARK, size=10))
    d += 2
    put(ws, f"A{d}", "THE BRIDGE IN ONE LINE      CY plan LR  =  Projected LR (current "
                     "level)  x  rate earn-in  x  mod drift  x  other adj      "
                     "— and those are literally the four columns to its left",
        fnt=font(NAVY, bold=True, size=11), fill=FILL_PANEL)
    for cc in range(2, 36):
        put(ws, ws.cell(row=d, column=cc).coordinate, None, fill=FILL_PANEL)
    d += 2
    jump(ws, f"A{d}", "'Walkthrough'!A1",
         "Worked example — every calculation for the selected combo, start to finish >")
    d += 1
    jump(ws, f"A{d}", "'Read Me'!A1",
         "Column definitions and the full glossary: Read Me >")
    d += 1
    jump(ws, f"A{d}", "'Net Delivery'!A1",
         "Net-target combos: how the target is delivered month by month — Net Delivery >")

    ctx.define("ss_mathchk", "State Summary", f"${MC}${first}:${MC}${last}",
               "Per-row |engine plan LR - the row's visible factor product| on "
               "single-combo rows, 0 elsewhere (D61)")
    for k, w in (("crit", 6), ("cnt", 6), ("key", 10), ("netcnt", 6), ("mathchk", 8)):
        ws.column_dimensions[ss_h(k)].width = w
    # six unlabelled helper columns immediately right of the flagship exhibit —
    # the first thing you meet scrolling right. Collapsible, so the exhibit ends
    # where it looks like it ends. (mathchk is ss_mathchk, which Checks reads: a
    # grouped column still calculates, it just isn't in the way.)
    for k in SS_HELP:
        ws.column_dimensions[ss_h(k)].outlineLevel = 1
    ss_outline(ws)
    presentation_setup(ws, gridlines_off=True, freeze=f"B{first}", tab_color=NAVY)
    print_setup(ws)


def PatternFill_safe(rgb):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=rgb)


def FILL_STEEL_LIGHT_safe():
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=STEEL_LIGHT)


# ---------------------------------------------------------------------------
# Scenarios
# ---------------------------------------------------------------------------


def build_scenarios(ctx: Ctx):
    ws = ctx.wb["Scenarios"]
    p = ctx.p
    title(ws, "B2", "Scenarios — selected BU x LOB",
          "Up to four what-if variations of the planned rate program and mod path. "
          "Blank lever = Base value. Engines live on the hidden _calc sheet.")
    label(ws, "B3", "Selected:")
    link(ws, "C3", "=nr_SelKey", bold=True)
    jump(ws, "D3", "Control!C7", "Change BU/state selection >")

    section(ws, 5, "B", "Scenario levers (blue = enter; blank = same as Base)")
    header_row(ws, 5, 3, ["S1", "S2", "S3", "S4"], widths=[10, 10, 10, 10])
    levers = [
        ("Extra filed rate on planned rows (+/- pts)", FMT_PCT, "sc_dpts",
         "Scenario lever: added to every planned row's filed %"),
        ("Shift planned effective dates (+/- months)", FMT_INT, "sc_shift",
         "Scenario lever: EDATE month shift applied to planned rows"),
        ("Achievement override (planned rows)", FMT_PCT, "sc_ach",
         "Scenario lever: replaces achievement % on planned rows"),
        ("Shift projected mod path M_1/M_2 (+/-)", FMT_MOD, "sc_dm1",
         "Scenario lever: added to M_1 and M_2 anchors"),
        ("Adjust net rate selection (+/- pts)", FMT_PCT, "sc_dnet",
         "Scenario lever: added to the net selection x for P and P+1 (D39)"),
    ]
    hints = ["e.g. 0.02 = +2.0 pts  (S1 ships seeded — clear to reset)",
             "e.g. 2 = two months later", "e.g. 0.80 = 80% achieved",
             "e.g. 0.02 = +2 mod pts on M_1 and M_2", "e.g. 0.01 = +1 pt net target"]
    for i, (lbl, fmt, name, desc) in enumerate(levers):
        r = 6 + i
        label(ws, f"B{r}", lbl)
        for cc in range(3, 7):
            # S1 ships with a demonstrative +2 pts so the tab teaches itself
            seed = 0.02 if (i == 0 and cc == 3) else None
            input_cell(ws, f"{col(cc)}{r}", seed, fmt=fmt, required=False)
        put(ws, f"G{r}", hints[i], fnt=font(GREY_DARK, size=9, italic=True))
        ctx.define(name, "Scenarios", f"$C${r}:$F${r}", desc)
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    for r, kind, lo, hi, msg in ((6, "decimal", "-0.5", "0.5", "Enter rate points as a decimal fraction in [-50%, +50%]."),
                                 (7, "whole", "-12", "12", "Shift is in whole months, -12..12."),
                                 (8, "decimal", "0", "1.5", "Achievement lies in [0%, 150%]."),
                                 (9, "decimal", "-0.5", "0.5", "Mod shift lies in [-0.5, +0.5]."),
                                 (10, "decimal", "-0.5", "0.5", "Net adjustment lies in [-50%, +50%].")):
        _dv_ = _DV(type=kind, operator="between", formula1=lo, formula2=hi,
                   allow_blank=True, showErrorMessage=True, error=msg)
        ws.add_data_validation(_dv_)
        _dv_.add(f"C{r}:F{r}")

    section(ws, 12, "B", "Results")
    put(ws, "B13", "Metric", fnt=F_HEADER, fill=FILL_NAVY)
    for j, h in enumerate(["Base", "S1", "S2", "S3", "S4"]):
        put(ws, f"{col(3 + j)}13", h, fnt=F_HEADER, fill=FILL_NAVY, align=ALIGN_C)
    ctx.define("sc_res_names", "Scenarios", "$C$13:$G$13", "Scenario result column headers")
    ctx.define("sc_res_cylr", "Scenarios", "$C$14:$G$14", "CY plan LR by scenario")

    metrics = ["Plan LR", "Change vs Base (pts)", "Rate earn-in (A_rate)",
               "Mod drift (A_mod)", "Earned rate level (E_CY)", "Carryover"]
    for i, m in enumerate(metrics):
        label(ws, f"B{14 + i}", m)
    # live year labels (D44)
    formula(ws, "B14", '="CY "&nr_PlanYear&" plan LR"')
    ws["B14"].font = F_LABEL
    formula(ws, "B19", '="Carryover into "&(nr_PlanYear+1)')
    ws["B19"].font = F_LABEL
    # Base column (green links to the live engines)
    for r, f, fmt in ((14, "=nr_CYLR_P", FMT_PCT), (15, "=($C$14-$C$14)*100", PTS_Z),
                      (16, "=nr_Arate_P", FMT_IDX), (17, "=nr_Amod_P", FMT_IDX),
                      (18, "=nr_ECY_P", FMT_IDX), (19, "=nr_YoY_P1", PCT_SIGNED_Z)):
        link(ws, f"C{r}", f, fmt=fmt, align=ALIGN_C, bold=(r == 14))
    for s in range(1, 5):
        cL = col(3 + s)
        res = ctx.lay_dyn["scenario_results"][s - 1]
        link(ws, f"{cL}16", f"=nr_CRLind/{res['e_p']}", fmt=FMT_IDX, align=ALIGN_C)
        link(ws, f"{cL}17",
             f'=IF(nr_NetMode,1,IF(AND(nr_ModAdjMaster="ON",nr_ModAdjRow="ON"),'
             f"nr_MInd/{res['mbar_p']},1))",
             fmt=FMT_IDX, align=ALIGN_C)
        formula(ws, f"{cL}14", f"=nr_LRcur*{cL}16*{cL}17*nr_AOther", fmt=FMT_PCT,
                align=ALIGN_C, bold=True)
        formula(ws, f"{cL}15", f"=({cL}14-$C$14)*100", fmt=PTS_Z, align=ALIGN_C)
        link(ws, f"{cL}18", f"={res['e_p']}", fmt=FMT_IDX, align=ALIGN_C)
        link(ws, f"{cL}19", f"={res['e_p1']}/{res['e_p']}-1", fmt=PCT_SIGNED_Z, align=ALIGN_C)
    note(ws, "B21", "Scenario CY LR = LR_current x A_rate' x A_mod' x A_other. CRL_ind is "
                    "unchanged (scenarios vary planned actions only, never the indication). "
                    "When the selected combo carries a net rate selection, only the D-net "
                    "lever moves results (planned-row levers are superseded, D39).")

    # chart the DELTAS (five near-identical LR levels are visually null; the
    # zero axis line is Base)
    chart = BarChart()
    chart.type = "col"
    chart.gapWidth = 60
    data = Reference(ws, min_col=4, min_row=15, max_col=7, max_row=15)
    cats = Reference(ws, min_col=4, min_row=13, max_col=7, max_row=13)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = STEEL
    chart.legend = None
    chart.y_axis.number_format = "0.00"
    _style_chart(chart, "Change vs Base (pts of loss ratio) — zero line = Base",
                 y_title="Change vs Base (pts)", height=8, width=13)
    ws.add_chart(chart, "B24")

    set_widths(ws, {"A": 2, "B": 40, "C": 11, "D": 11, "E": 11, "F": 11, "G": 11})
    presentation_setup(ws, gridlines_off=True, tab_color=STEEL_LIGHT)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Solver (E1)
# ---------------------------------------------------------------------------


def build_solver(ctx: Ctx):
    ws = ctx.wb["Solver"]
    p = ctx.p
    title(ws, "B2", "Solver — inverse plan (closed form, no iteration)",
          "Because the CY earned index is linear in a single unknown change r, the required "
          "rate solves in closed form.")
    # The Solver is the one tab with a LOCAL selection override: it gets used
    # standalone in filing conversations. An input cell cannot default via a
    # formula (typing destroys it), so the mechanism is Follow-Control? Y/N
    # plus override dropdowns; slv_key resolves the combo everything below uses.
    label(ws, "B3", "Solving for")
    link(ws, "C3", "=slv_key", bold=True)
    jump(ws, "E3", "Control!C7", "Change Control selection >")
    label(ws, "B4", "Follow the Control selection? (Y/N)")
    input_cell(ws, "C4", "Y")
    prose(ws, "D4",
          "Convention (D13): the Solver replaces the ENTIRE planned-rate program with a "
          "single action at the chosen date. Its base index W0 uses taken rows only.",
          size=9, width=30)
    ws["D4"].font = font(FAIL_RED, size=9, italic=True)
    label(ws, "B5", "Override BU | state (used when N)")
    input_cell(ws, "C5", ctx.we_row["bu"], required=False)
    input_cell(ws, "D5", ctx.we_row["state"], required=False)
    put(ws, "F4", "resolved key", fnt=font(GREY_DARK, size=8))
    formula(ws, "G4", '=IF(UPPER($C$4)="N",$C$5&"|"&$D$5,nr_SelKey)')
    ws["G4"].font = font(GREY_DARK, size=9)
    ctx.define("slv_key", "Solver", "$G$4",
               "Combo the Solver solves for (Control selection, or the local override)")
    put(ws, "F5", "resolved state", fnt=font(GREY_DARK, size=8))
    formula(ws, "G5", '=IF(UPPER($C$4)="N",$D$5,nr_SelState)')
    ws["G5"].font = font(GREY_DARK, size=9)
    ctx.define("slv_state", "Solver", "$G$5", "State of the combo the Solver solves for")
    for _rng, _f in (("C4", '"Y,N"'), ("C5", "=lst_bu"), ("D5", "=lst_state")):
        add_dv(ws, "list", [_rng], formula1=_f)
    # Every NUMBER typed on this tab is bounded too. A target loss ratio of 65
    # instead of 0.65 solves cleanly, prints a required change of several
    # thousand percent, and nothing on the sheet contradicts it.
    dv_decimal(ws, ["C7"], 0, 3,
               "Enter the target as a decimal fraction — 0.65 for a 65% loss ratio.")
    dv_decimal(ws, ["C9", "C47"], 0.1, 2,
               "Achievement is the fraction of a filing you expect to realise — "
               "0.7 for 70%, not 70.")
    dv_decimal(ws, ["C21", "C22", "G47"], 0, 1,
               "A reasonability bound is a positive decimal fraction.")
    dv_decimal(ws, ["C26"], -0.5, 2, "Enter a decimal fraction in [-50%, +200%].")
    dv_decimal(ws, ["G49"], -0.5, 0.5,
               "A mod action is a decimal fraction — mods live in [0.5, 1.5], so a "
               "step beyond ±50% is not a schedule mod.")
    dv_plan_year_date(ws, ["C8"],
                      "The effective date must lie inside the plan year — the solver "
                      "replaces the whole planned program with one action in P.")

    # ---- Mode A ----
    section(ws, 6, "B", "Mode A — required rate for a target CY LR")
    formula(ws, "B7", '="Target CY "&nr_PlanYear&" loss ratio"')
    ws["B7"].font = F_LABEL
    input_cell(ws, "C7", ctx.oracle_m.cy_lr_p, fmt=FMT_PCT)
    label(ws, "B8", "Effective date")
    input_cell(ws, "C8", dt.date(p, 4, 1), fmt=FMT_DATE)
    label(ws, "B9", "Achievement % assumed")
    input_cell(ws, "C9", 1.0, fmt=FMT_PCT)
    note(ws, "D7", "seeded with the sample book's own CY plan LR - the solver should return "
                   "the seeded planned change (+5.0% at 4/1) while sample data is intact")

    # ---- the answer, before the mechanics that produce it --------------------
    # Mode A's result was C17: seventh of nine rows in a list that opens with
    # cohort-month exposure sums. The number people come to this tab for should
    # not be something you find by reading down a column of intermediates.
    formula(ws, "B10",
            '=IF(COUNTIF(lr_key,slv_key)=0,"No row in tbl_LR for "&slv_key&" — the solver '
            'has nothing to solve.",'
            'IF($C$7="","Enter a target loss ratio above.",'
            'IF(NOT(ISNUMBER($C$17)),"No single rate change at this date reaches the '
            'target — nothing is left to earn. Try an earlier date, or Mode C.",'
            '"To reach "&TEXT($C$7,"0.0%")&" for "&slv_key&": FILE "'
            '&TEXT($C$18,"+0.00%;-0.00%")&" effective "&TEXT($C$8,"m/d/yy")'
            '&"   (earns "&TEXT($C$17,"+0.00%;-0.00%")&" at "&TEXT($C$9,"0%")'
            '&" achievement; "&TEXT($C$19,"0.0%")&" of the year is still ahead of that '
            'date).")))')
    ws["B10"].font = font(NAVY, bold=True, size=12)
    ws["B10"].fill = FILL_PANEL
    for cc in range(3, 5):        # C..D carry the band; E is the 3-wide gutter
        put(ws, ws.cell(row=10, column=cc).coordinate, None, fill=FILL_PANEL)

    helpers = [
        ("abs eff month", "=YEAR($C$8)*12+MONTH($C$8)-1", FMT_INT),
        ("month start", "=DATE(YEAR($C$8),MONTH($C$8),1)", FMT_DATE),
        ("month end", "=EOMONTH($C$8,0)", FMT_DATE),
        ("days in month", "=$G$9-$G$8+1", FMT_INT),
        ("days on/after eff", "=$G$9-$C$8+1", FMT_INT),
        ("eff month in cohort window?",
         "=AND($G$7>=(nr_PlanYear-2)*12,$G$7<=(nr_PlanYear+1)*12+11)", FMT_GEN),
        ("cohort row index (clamped)", "=MIN(48,MAX(1,$G$7-(nr_PlanYear-2)*12+1))", FMT_INT),
        ("first taken-change days (same month)", "=IF($G$12,INDEX(slv_dfirst,$G$13),0)", FMT_INT),
        ("split p (first-change rule, D4/D31)",
         "=IF($G$12,MIN(1,MAX($G$11,$G$14)/$G$10),1)", "0.000"),
        ("W0_pre at eff month", "=IF($G$12,INDEX(slv_wpre,$G$13),1)", FMT_IDX),
        ("W0_eom at eff month", "=IF($G$12,INDEX(slv_weom,$G$13),1)", FMT_IDX),
        ("w x ec at eff month", "=IF($G$12,INDEX(slv_w,$G$13)*INDEX(slv_ecp,$G$13),0)", FMT_IDX),
    ]
    for i, (lbl, f, fmt) in enumerate(helpers):
        r = 7 + i
        put(ws, f"F{r}", lbl, fnt=font(GREY_DARK, size=9))
        formula(ws, f"G{r}", f, fmt=fmt)
        ws[f"G{r}"].font = font(GREY_DARK, size=9)

    # local combo resolver: everything Mode A/B needs about the SOLVE combo,
    # served from the all-combo _calc results table (no new engine compute)
    locals_ = [
        ("combo row (_calc)", "slv_row",
         "=IF(COUNTIF(calc_key,slv_key)=0,0,MATCH(slv_key,calc_key,0))", FMT_INT,
         "Row of the solve combo in the _calc results table (0 = absent)"),
        ("CRL_ind (combo)", "slv_crl", "=IF(slv_row=0,1,INDEX(calc_crl,slv_row))", FMT_IDX,
         "Indication rate level of the solve combo"),
        ("LR at current level (combo)", "slv_lrcur",
         "=IF(slv_row=0,0,INDEX(calc_lrcur,slv_row))", FMT_PCT,
         "Basis-normalized projected LR of the solve combo"),
        ("A_mod in force (combo)", "slv_amod",
         "=IF(slv_row=0,1,INDEX(calc_amod_p,slv_row))", FMT_IDX,
         "Mod drift factor of the solve combo"),
        ("A_other (combo)", "slv_aother",
         "=IF(slv_row=0,1,INDEX(calc_aother,slv_row))", FMT_IDX,
         "Other adjustment factor of the solve combo"),
    ]
    for i, (lbl, name, f, fmt, desc) in enumerate(locals_):
        rr_ = 19 + i
        put(ws, f"F{rr_}", lbl, fnt=font(GREY_DARK, size=9))
        formula(ws, f"G{rr_}", f, fmt=fmt)
        ws[f"G{rr_}"].font = font(GREY_DARK, size=9)
        ctx.define(name, "Solver", f"$G${rr_}", desc)

    outputs = [
        ("Old-rate earned exposure before the date (C_pre, cohort-months)",
         "=SUMPRODUCT((slv_absmi<$G$7)*slv_w*slv_ecp*slv_widx)+$G$18*(1-$G$15)*$G$16", FMT_IDX,
         None),
        ("Earnable exposure on/after the date (C_post, cohort-months)",
         "=SUMPRODUCT((slv_absmi>$G$7)*slv_w*slv_ecp*slv_widx)+$G$18*$G$15*$G$17", FMT_IDX, None),
        ("Total CY earned exposure (D, cohort-months)", "=slv_den", FMT_IDX, None),
        ("Mod drift factor in force (A_mod)", "=slv_amod", FMT_IDX, None),
        ("Rate earn-in factor needed (A_rate)",
         "=IF(OR(slv_lrcur=0,$C$7=\"\"),0,$C$7/(slv_lrcur*$C$14*slv_aother))",
         FMT_IDX, None),
        ("Earned rate level needed (E_CY)", "=IF($C$15=0,0,slv_crl/$C$15)", FMT_IDX, None),
        ("REQUIRED CHANGE r",
         '=IF(OR($C$12=0,$C$16=0),"n/a",($C$16*$C$13-$C$11)/$C$12-1)', "0.000%", "nr_SolverR"),
        ("Filed-rate equivalent (r / achievement)",
         '=IF(OR($C$9=0,NOT(ISNUMBER($C$17))),"n/a",$C$17/$C$9)', "0.000%", None),
        ("Share of CY earned exposure on/after eff", "=$C$12/($C$11+$C$12)", FMT_PCT, None),
    ]
    for i, (lbl, f, fmt, name) in enumerate(outputs):
        r = 11 + i
        label(ws, f"B{r}", lbl, bold=(name == "nr_SolverR"))
        formula(ws, f"C{r}", f, fmt=fmt, border=BORDER_THIN, bold=(name == "nr_SolverR"),
                fill=FILL_PANEL if name == "nr_SolverR" else None)
        if name:
            ctx.define(name, "Solver", f"$C${r}", "Mode A required rate change")
    label(ws, "B21", "Reasonability bound |r|")
    input_cell(ws, "C21", ctx.cfg.solver_max_rate, fmt=FMT_PCT, required=False)
    ctx.define("nr_SolvMaxRate", "Solver", "$C$21", "Solver reasonability bound (config)")
    label(ws, "B22", "Minimum post-eff exposure share")
    input_cell(ws, "C22", ctx.cfg.solver_min_post_share, fmt=FMT_PCT, required=False)
    ctx.define("nr_SolvMinShare", "Solver", "$C$22", "Solver post-share warning threshold (config)")
    formula(ws, "D19",
            '=IF($C$19<nr_SolvMinShare,"WARNING: only "&TEXT($C$19,"0.0%")&" of CY exposure '
            'sits on/after this date - late-year targets are mathematically absurd","OK")',
            fmt=FMT_GEN)
    ws["D19"].font = font(FAIL_RED, size=9, italic=True)
    formula(ws, "D17",
            '=IF(ISNUMBER($C$17),IF(ABS($C$17)>nr_SolvMaxRate,"WARNING: exceeds the '
            'reasonability bound","OK"),"no solvable rate at this date")', fmt=FMT_GEN)
    ws["D17"].font = font(FAIL_RED, size=9, italic=True)

    # ---- Mode B ----
    section(ws, 25, "B",
            "Mode B — timing sensitivity: the FULL-YEAR plan LR under each possible "
            "start month for the same change")
    formula(ws, "B24",
            '=IF(slv_row=0,"",IF(INDEX(calc_netmode,slv_row)=1,"NOTE: this combo carries a '
            'NET RATE SELECTION - Solver results describe the explicit-program '
            'counterfactual (planned rows + classic A_mod), not the net path (D39).",""))')
    ws["B24"].font = font(FAIL_RED, size=9, italic=True)
    label(ws, "B26", "Rate change r to time")
    input_cell(ws, "C26", 0.05, fmt=FMT_PCT)
    note(ws, "D26",
         "Each row is a separate what-if, NOT the loss ratio moving through the year: the "
         "same increase made effective later earns in less during the plan year, so the "
         "full-year result lands higher. Mode A's target flags which start months still work.")
    header_row(ws, 28, 2, ["Effective month", "Effective", "C_pre", "C_post", "E_CY(P)",
                           "Full-yr plan LR", "Meets target?"],
               widths=[13, 11, 10, 10, 10, 12, 13])
    slv = ctx.lay_dyn["solver"]
    bf = slv["first"]
    for m in range(1, 13):
        r = 28 + m
        row_m = bf + 24 + (m - 1)          # cohort row of month m of P
        formula(ws, f"B{r}", f'=TEXT(DATE(nr_PlanYear,{m},1),"mmm yyyy")', align=ALIGN_C)
        formula(ws, f"C{r}", f"=DATE(nr_PlanYear,{m},1)", fmt=FMT_DATE, align=ALIGN_C)
        link(ws, f"D{r}", f"='_calc'!$T${row_m - 1}", fmt=FMT_IDX, align=ALIGN_C)
        link(ws, f"E{r}", f"=slv_total-'_calc'!$T${row_m}+'_calc'!$U${row_m}", fmt=FMT_IDX,
             align=ALIGN_C)
        formula(ws, f"F{r}", f"=($D{r}+(1+$C$26)*$E{r})/slv_den", fmt=FMT_IDX, align=ALIGN_C)
        formula(ws, f"G{r}",
                f"=IF(slv_lrcur=0,0,slv_lrcur*(slv_crl/$F{r})*$C$14*slv_aother)",
                fmt=FMT_PCT, align=ALIGN_C)
        formula(ws, f"H{r}", f'=IF($C$7="","-",IF($G{r}<=$C$7,"YES","no"))', align=ALIGN_C)
        formula(ws, f"I{r}", f'=IF($H{r}="YES",{m},0)', fmt=FMT_INT)
        ws[f"I{r}"].font = font(GREY_DARK, size=8)
        formula(ws, f"J{r}", f'=IF($C$7="","",$C$7)', fmt=FMT_PCT)
        ws[f"J{r}"].font = font(GREY_DARK, size=8)
    label(ws, "B42", "Latest start month that still meets the target", bold=True)
    formula(ws, "D42",
            '=IF(MAX($I$29:$I$40)=0,"No start month meets the target",'
            'TEXT(DATE(nr_PlanYear,MAX($I$29:$I$40),1),"mmm yyyy"))', bold=True,
            fill=FILL_PANEL)
    put(ws, "I28", "helper", fnt=font(GREY_DARK, size=8))
    put(ws, "J28", "target", fnt=font(GREY_DARK, size=8))

    lc = LineChart()
    data = Reference(ws, min_col=7, min_row=28, max_row=40)
    lc.add_data(data, titles_from_data=True)
    tgt = Reference(ws, min_col=10, min_row=28, max_row=40)
    lc.add_data(tgt, titles_from_data=True)
    cats = Reference(ws, min_col=2, min_row=29, max_row=40)
    lc.set_categories(cats)
    _line_color(lc.series[0], NAVY)
    _line_color(lc.series[1], FAIL_RED, width_pt=1.25, dashed=True)
    lc.legend = None
    lc.y_axis.number_format = "0.0%"
    _style_chart(lc,
                 "Rate lever: full-year plan LR by start month (dashed = target)",
                 y_title="Full-year plan LR", height=8, width=14)
    ws.add_chart(lc, "K28")

    # ---- Mode C: the same target, on the MOD lever (D73) ----
    section(ws, 45, "B",
            "Mode C — the same target on the PRICING lever: what dated mod action gets "
            "you there, and what it costs to wait")
    put(ws, "B46",
        "Earned mod is linear in a dated mod step exactly as the earned rate index is "
        "linear in a filed change, so this inverts in closed form too. Same target as "
        "Mode A (C7); the rate program is taken AS GIVEN here, including planned filings "
        "— the question is what pricing you still need on top of it. Planned mod actions "
        "are excluded, because they are what you are solving for.", fnt=F_SMALL_IT)
    label(ws, "B47", "Mod achievement % assumed")
    input_cell(ws, "C47", 0.70, fmt=FMT_PCT)
    note(ws, "D47", "mod is rarely fully realised — the DIRECTED column is what you have "
                    "to aim at to land the required step")
    # a mod step needs its OWN bound: the rate bound is far too loose here,
    # because mods live in [0.5, 1.5] and a 15% step already moves 0.85 to 0.98
    put(ws, "F47", "Reasonability bound |step|", fnt=font(GREY_DARK, size=9))
    input_cell(ws, "G47", 0.15, fmt=FMT_PCT, required=False)
    ctx.define("nr_SolvMaxMod", "Solver", "$G$47",
               "Mode C reasonability bound on a single mod step (mods are levels in "
               "[0.5, 1.5], so this is much tighter than the rate bound)")
    # D79: the chart times a FIXED action rather than plotting the solved one.
    # Mode A solves the rate size and Mode B times a fixed rate; this is the
    # same pair on the pricing lever. Plotting the SOLVED step's consequence
    # would draw a flat line on the target by construction.
    put(ws, "F49", "Mod action to time (directed)", fnt=font(GREY_DARK, size=9))
    input_cell(ws, "G49", 0.05, fmt=FMT_PCT)
    ctx.define("slv_modtime", "Solver", "$G$49",
               "Mode C: the DIRECTED mod action whose timing the chart traces (D79)")
    put(ws, "F50",
        "Directed, so the achievement applies: the chart traces what this one action does "
        "to the FULL-YEAR plan LR as its date slips. Where it crosses the target is the "
        "last month it still works.", fnt=F_SMALL_IT)
    lc_ = [
        ("Mod adjustment in force for this combo?",
         "=IF(COUNTIF(lr_key,slv_key)=0,0,IF("
         + mod_adj_on("MATCH(slv_key,lr_key,0)") + ",1,0))", FMT_INT, "slv_modon"),
        ("Mod assumed in indication (M_ind)",
         "=IF(slv_row=0,1,INDEX(lr_mind,slv_row))", FMT_MOD, "slv_mind"),
        ("Rate earn-in in force (A_rate)",
         "=IF(slv_row=0,1,INDEX(calc_arate_p,slv_row))", FMT_IDX, "slv_arate"),
        ("Earned mod needed (Mbar)",
         '=IF(OR(slv_lrcur=0,$C$7=""),0,'
         "slv_lrcur*slv_arate*slv_aother*slv_mind/$C$7)", FMT_MOD, "slv_mbarneed"),
    ]
    for i, (lbl, f, fmt, name) in enumerate(lc_):
        r = 48 + i
        label(ws, f"B{r}", lbl)
        formula(ws, f"C{r}", f, fmt=fmt, border=BORDER_THIN)
        ctx.define(name, "Solver", f"$C${r}", lbl)
    formula(ws, "F48",
            '=IF(slv_row=0,"",IF(slv_modon=0,'
            '"MOD ADJUSTMENT IS OFF for this combo — A_mod is pinned at 1.000 and no mod '
            'action can move the plan LR. Solve on rate instead.",'
            'IF(INDEX(calc_netmode,slv_row)=1,"NET SELECTION ACTIVE — the net path '
            'supersedes the mod leg from 1/1 (D39); these figures are the '
            'explicit-program counterfactual.","")))')
    ws["F48"].font = font(FAIL_RED, size=9, italic=True)

    header_row(ws, 52, 2,
               ["Effective month", "Effective", "K_pre", "K_post",
                "Required mod step", "Directed (step / achievement)",
                "Share left to earn", "Feasible?",
                "Earned mod at the timed action", "Full-yr plan LR"],
               widths=[13, 11, 10, 10, 14, 16, 13, 11, 13, 12])
    for m in range(1, 13):
        r = 52 + m
        row_m = bf + 24 + (m - 1)          # cohort row of month m of P
        formula(ws, f"B{r}", f'=TEXT(DATE(nr_PlanYear,{m},1),"mmm yyyy")', align=ALIGN_C)
        formula(ws, f"C{r}", f"=DATE(nr_PlanYear,{m},1)", fmt=FMT_DATE, align=ALIGN_C)
        # a 1st-of-month date makes the D31 split p = 1, so K_pre is the running
        # total through the previous cohort and K_post is everything after,
        # with the effective month itself valued at its END-of-month mod
        link(ws, f"D{r}", f"='_calc'!$AH${row_m - 1}", fmt=FMT_IDX, align=ALIGN_C)
        link(ws, f"E{r}", f"=slv_mtotal-'_calc'!$AH${row_m}+'_calc'!$AI${row_m}",
             fmt=FMT_IDX, align=ALIGN_C)
        formula(ws, f"F{r}",
                f'=IF(OR($E{r}=0,slv_mbarneed=0),"n/a",'
                f"(slv_mbarneed*slv_den-$D{r})/$E{r}-1)", fmt="0.000%", align=ALIGN_C)
        formula(ws, f"G{r}",
                f'=IF(OR(NOT(ISNUMBER($F{r})),$C$47=0),"n/a",$F{r}/$C$47)',
                fmt="0.000%", align=ALIGN_C, bold=True)
        formula(ws, f"H{r}", f"=$E{r}/($D{r}+$E{r})", fmt=FMT_PCT, align=ALIGN_C)
        formula(ws, f"I{r}",
                f'=IF(NOT(ISNUMBER($F{r})),"no",'
                f'IF(AND(ABS($F{r})<=nr_SolvMaxMod,$H{r}>=nr_SolvMinShare,'
                f'slv_modon=1),"YES","no"))', align=ALIGN_C)
        # D79: the SAME action, timed. Earned mod if a directed step of
        # slv_modtime (achieved at C47) lands in this month, and the full-year
        # plan LR it produces — the Mode B statistic, on the pricing lever.
        formula(ws, f"J{r}",
                f"=($D{r}+(1+slv_modtime*$C$47)*$E{r})/slv_den", fmt=FMT_MOD,
                align=ALIGN_C)
        formula(ws, f"K{r}",
                f'=IF(OR(slv_lrcur=0,$J{r}=0,slv_modon=0),"n/a",'
                f"slv_lrcur*slv_arate*(slv_mind/$J{r})*slv_aother)",
                fmt=FMT_PCT, align=ALIGN_C, bold=True)
        formula(ws, f"L{r}", f'=IF($I{r}="YES",{m},0)', fmt=FMT_INT)
        formula(ws, f"M{r}", f'=IF(ISNUMBER($K{r}),IF($K{r}<=$C$7,{m},0),0)', fmt=FMT_INT)
        formula(ws, f"N{r}", '=IF($C$7="","",$C$7)', fmt=FMT_PCT)
        for cL in "LMN":
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=8)
    put(ws, "L52", "helper", fnt=font(GREY_DARK, size=8))
    put(ws, "M52", "helper", fnt=font(GREY_DARK, size=8))
    put(ws, "N52", "target", fnt=font(GREY_DARK, size=8))
    label(ws, "B66", "Latest month the REQUIRED step is still inside the bound", bold=True)
    formula(ws, "E66",
            '=IF(MAX($L$53:$L$64)=0,"No month works on the mod lever alone",'
            'TEXT(DATE(nr_PlanYear,MAX($L$53:$L$64),1),"mmm yyyy"))', bold=True,
            fill=FILL_PANEL)
    label(ws, "B67", "Latest month the TIMED action above still meets the target",
          bold=True)
    formula(ws, "E67",
            '=IF(MAX($M$53:$M$64)=0,"This action never reaches the target — direct more",'
            'TEXT(DATE(nr_PlanYear,MAX($M$53:$M$64),1),"mmm yyyy"))', bold=True,
            fill=FILL_PANEL)
    put(ws, "B68",
        "Two different reads. The table SOLVES the step month by month — acting early "
        "costs less because more of the year is still there to earn it, and past the last "
        "feasible month the arithmetic still returns a number nobody could file. The "
        "chart TIMES one action you have chosen, in loss-ratio points, the way Mode B "
        "does for a filing.", fnt=F_SMALL_IT)

    # The required step is unbounded by construction — K_post goes to zero as
    # the year runs out, so a December solve reads +493% and the axis it forces
    # squashes every month anyone would act in onto the zero line. Charting the
    # LR CONSEQUENCE of a fixed action instead is bounded, monotonic, and reads
    # against the target exactly as Mode B does (D79).
    mc_ = LineChart()
    mc_.add_data(Reference(ws, min_col=11, min_row=52, max_row=64), titles_from_data=True)
    mc_.add_data(Reference(ws, min_col=14, min_row=52, max_row=64), titles_from_data=True)
    mc_.set_categories(Reference(ws, min_col=2, min_row=53, max_row=64))
    _line_color(mc_.series[0], NAVY)
    _line_color(mc_.series[1], FAIL_RED, width_pt=1.25, dashed=True)
    mc_.legend = None
    mc_.y_axis.number_format = "0.0%"
    _style_chart(mc_,
                 "Pricing lever: full-year plan LR by action month (dashed = target)",
                 y_title="Full-year plan LR", height=8, width=14)
    ws.add_chart(mc_, "P52")

    set_widths(ws, {"A": 2, "B": 38, "C": 12, "D": 30, "E": 3, "F": 30, "G": 11})
    presentation_setup(ws, gridlines_off=True, tab_color=STEEL)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Attribution (E2)
# ---------------------------------------------------------------------------


def build_attribution(ctx: Ctx):
    ws = ctx.wb["Attribution"]
    p = ctx.p
    title(ws, "B2", "Attribution — plan vs actual",
          "Sequential multiplicative decomposition of (actual - plan) CY LR. Order used: "
          "1) rate magnitude at planned dates, 2) rate timing at actual dates, 3) mod drift, "
          "4) loss-side residual. The decomposition is order-dependent (see Methodology).")
    label(ws, "B3", "Selected:")
    link(ws, "C3", "=nr_SelKey", bold=True)
    jump(ws, "D3", "Control!C7", "Change BU/state selection >")
    formula(ws, "E3",
            '=IF(nr_NetMode,"NOTE: this combo carries a NET RATE SELECTION - premium-side '
            'steps 1-3 are suspended (factors = 1) because rate and mods are merged in the '
            'net path; plan-vs-actual deviation flows to the residual (D39).","")')
    ws["E3"].font = font(FAIL_RED, size=9, italic=True)

    formula(ws, "B4",
            '=IF(N(att_actlr)=0,"AWAITING ACTUALS — enter the actual CY "&nr_PlanYear&'
            '" loss ratio below; the decomposition stays greyed until then.","")')
    ws["B4"].font = font("BF8F00", bold=True, size=10)
    # grey the decomposition while no actual LR is entered (visual state only)
    ws.conditional_formatting.add(
        "B25:E31",
        FormulaRule(formula=["N(att_actlr)=0"], font=font("BFBFBF")))

    section(ws, 5, "B", "Actuals — mod path and outcome (blank = as planned)")
    att_inputs = [
        ("Actual M_0", FMT_MOD, "att_m0", "Attribution: actual current written mod (blank = plan)"),
        ("Actual M_0 as-of", FMT_DATE, "att_asof", "Attribution: actual as-of date (blank = plan)"),
        ("Actual M_1 at 12/31/P", FMT_MOD, "att_m1", "Attribution: actual written mod at 12/31/P"),
        ("Actual M_2 (optional)", FMT_MOD, "att_m2", "Attribution: actual written mod at 12/31/(P+1)"),
        (f"ACTUAL CY {p} loss ratio", FMT_PCT, "att_actlr", "Attribution: actual CY LR outcome"),
    ]
    for i, (lbl, fmt, name, desc) in enumerate(att_inputs):
        r = 6 + i
        label(ws, f"B{r}", lbl, bold=(name == "att_actlr"))
        input_cell(ws, f"C{r}", None, fmt=fmt, required=(name == "att_actlr"))
        ctx.define(name, "Attribution", f"$C${r}", desc)
    note(ws, "D6", "The plan M_ind stays fixed - it is an indication property, not an actual (D31).")

    section(ws, 12, "B", "Planned rate actions — enter actuals (blank = as planned)")
    rank_rng = ctx.lay_dyn["att_rank"]
    header_row(ws, 13, 2, ["#", "Planned effective", "Planned eff chg",
                           "ACTUAL achieved chg", "ACTUAL effective"],
               widths=[5, 14, 13, 15, 14])
    for nslot in range(1, 9):
        r = 13 + nslot
        put(ws, f"B{r}", nslot, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
        link(ws, f"C{r}",
             f'=IF(COUNTIF({rank_rng},{nslot})=0,"",INDEX(rl_eff,MATCH({nslot},{rank_rng},0)))',
             fmt=FMT_DATE, align=ALIGN_C)
        link(ws, f"D{r}",
             f'=IF($C{r}="","",INDEX(rl_reff,MATCH({nslot},{rank_rng},0)))',
             fmt=FMT_PCT, align=ALIGN_C)
        input_cell(ws, f"E{r}", None, fmt=FMT_PCT, required=False)
        input_cell(ws, f"F{r}", None, fmt=FMT_DATE, required=False)
    ctx.define("att_actpct", "Attribution", "$E$14:$E$21",
               "Attribution: actual achieved total change per planned row (blank = plan)")
    ctx.define("att_actdate", "Attribution", "$F$14:$F$21",
               "Attribution: actual effective date per planned row (blank = plan)")
    # Actuals arrive late, typed under time pressure, and every one of them is
    # a DENOMINATOR in the decomposition — an unbounded typo here reappears as a
    # residual step and gets explained rather than caught. Same bounds tbl_LR
    # already applies to the plan-side twins of these cells.
    dv_decimal(ws, ["C6", "C8", "C9"], 0.5, 1.5,
               "Schedule mods must lie in [0.5, 1.5].")
    dv_decimal(ws, ["C10"], 0, 3,
               "Enter the actual loss ratio as a decimal fraction — 0.65 for 65%.")
    dv_decimal(ws, ["E14:E21"], -0.5, 2,
               "Enter the achieved change as a decimal fraction in [-50%, +200%].")
    # the as-of date may legitimately sit in the prior year; the effective dates
    # of plan-year actions may not
    dv_plan_year_date(ws, ["C7"], "The actual as-of date must fall within the plan "
                                  "year or the two before it.", back_years=2)
    dv_plan_year_date(ws, ["F14:F21"],
                      "An actual effective date belongs inside the plan year.")

    section(ws, 23, "B", "Decomposition")
    header_row(ws, 24, 2, ["Step", "Factor", "LR after", "Step"],
               widths=[30, 10, 10, 11])
    e_mag, e_time = ctx.lay_dyn["att_e_mag"], ctx.lay_dyn["att_e_time"]
    mbar_act = ctx.lay_dyn["att_mbar_act"]
    steps = [
        ("Plan CY LR", None, "=nr_CYLR_P"),
        ("1. Rate magnitude (achieved %, planned dates)",
         f"=IF(nr_NetMode,1,nr_ECY_P/{e_mag})", None),
        ("2. Rate timing (achieved %, actual dates)",
         f"=IF(nr_NetMode,1,{e_mag}/{e_time})", None),
        ("3. Mod drift (actual mod path)",
         f'=IF(nr_NetMode,1,IF(AND(nr_ModAdjMaster="ON",nr_ModAdjRow="ON"),'
         f"nr_MEarned_P/{mbar_act},1))", None),
        ("4. Loss-side residual", "=IF(N(att_actlr)=0,1,att_actlr/$D$28)", None),
        ("Actual CY LR", None, "=$D$29"),
    ]
    for i, (lbl, fac, lr_f) in enumerate(steps):
        r = 25 + i
        label(ws, f"B{r}", lbl, bold=(i in (0, 5)))
        if fac:
            formula(ws, f"C{r}", fac, fmt=FMT_IDX, align=ALIGN_C)
            formula(ws, f"D{r}", f"=$D{r - 1}*$C{r}", fmt=FMT_PCT, align=ALIGN_C)
            formula(ws, f"E{r}", f"=($D{r}-$D{r - 1})*100", fmt=PTS_Z, align=ALIGN_C)
        else:
            formula(ws, f"D{r}", lr_f, fmt=FMT_PCT, align=ALIGN_C, bold=True,
                    fill=FILL_PANEL if i == 5 else None)
    formula(ws, "B32",
            '=IF(N(att_actlr)=0,"Enter the actual CY LR above to complete the residual step.",'
            'IF(ABS($D$25*$C$26*$C$27*$C$28*$C$29-att_actlr)<0.0000001,'
            '"RECONCILES: factors multiply back to the actual CY LR.",'
            '"DOES NOT RECONCILE - check inputs."))', bold=True)
    ctx.define("att_reconciles", "Attribution", "$B$32", "Attribution reconciliation status")
    note(ws, "B34",
         "Steps 1-3 recompute the CY earned rate index / earned mod with actuals substituted "
         "sequentially; step 4 is the remainder. A different ordering would allocate the "
         "interaction terms differently (stated per §7 E2). Taken rows are assumed booked as "
         "logged; only planned rows are re-achieved.")

    # waterfall chart data + chart
    cd = 37
    put(ws, f"B{cd - 1}", "Waterfall chart data (formulas — do not edit)", fnt=F_SMALL_IT)
    cats = [("Plan CY LR", "$D$25", None, True),
            ("Rate magnitude", "$D$25", "$D$26", False),
            ("Rate timing", "$D$26", "$D$27", False),
            ("Mod drift", "$D$27", "$D$28", False),
            ("Loss residual", "$D$28", "$D$29", False),
            ("Actual CY LR", "$D$30", None, True)]
    for j, h in enumerate(["Step", "base", "up", "down", "total"]):
        put(ws, f"{col(2 + j)}{cd}", h, fnt=font(GREY_DARK, size=9))
    for i, (lbl, a, b, is_total) in enumerate(cats):
        rr = cd + 1 + i
        put(ws, f"B{rr}", lbl, fnt=font(GREY_DARK, size=9))
        if is_total:
            for cL, f in (("C", "=0"), ("D", "=0"), ("E", "=0"), ("F", f"={a}")):
                formula(ws, f"{cL}{rr}", f, fmt=FMT_PCT)
        else:
            formula(ws, f"C{rr}", f"=MIN({a},{b})", fmt=FMT_PCT)
            formula(ws, f"D{rr}", f"=MAX(0,{b}-{a})", fmt=FMT_PCT)
            formula(ws, f"E{rr}", f"=MAX(0,{a}-{b})", fmt=FMT_PCT)
            formula(ws, f"F{rr}", "=0", fmt=FMT_PCT)
        for cL in "CDEF":
            ws[f"{cL}{rr}"].font = font(GREY_DARK, size=9)
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 60
    data = Reference(ws, min_col=3, min_row=cd, max_col=6, max_row=cd + 6)
    cats_ref = Reference(ws, min_col=2, min_row=cd + 1, max_row=cd + 6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.noFill = True
    chart.series[1].graphicalProperties.solidFill = UP_BAR
    chart.series[2].graphicalProperties.solidFill = DOWN_BAR
    chart.series[3].graphicalProperties.solidFill = TOTAL_BAR
    chart.legend = None
    chart.y_axis.number_format = "0%"
    _style_chart(chart, "Plan -> actual CY LR attribution", y_title="Loss ratio",
                 height=8.5, width=14)
    ws.add_chart(chart, "H5")

    set_widths(ws, {"A": 2, "B": 34, "C": 14, "D": 15, "E": 15, "F": 14})
    presentation_setup(ws, gridlines_off=True, tab_color=STEEL)
    print_setup(ws)
