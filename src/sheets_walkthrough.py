"""Walkthrough sheet: the fully worked example, start to finish, LIVE for the
selected combo.

Grammar (one visual sentence per step): value | the formula in words | a
plain-English explanation in the wide prose column J | a jump to the source.
Everything references the same nr_* / calc_* / rl_* names the exhibits use —
this tab never recomputes the model EXCEPT where the recomputation IS the
pedagogy (the mid-month split, the assembly product), and those recomputations
are proven against the engine in the W9 reconciliation strip (plus a Checks
row and a harness tie on nr_WalkLR).
"""

from __future__ import annotations

from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font

from .build_workbook import Ctx, Layout as L, SHEETS
from .xlstyle import (
    ALIGN_C, BORDER_THIN, F_LABEL, F_SMALL_IT, FAIL_RED, FILL_GREEN, FILL_GREY,
    FILL_PANEL, FILL_RED, FMT_IDX, FMT_INT, FMT_MOD, FMT_PCT, GREY_DARK, NAVY,
    PASS_GREEN, STEEL, col, chart_legend, font, formula, header_row, jump, label, link,
    presentation_setup, print_setup, prose, put, quote_sheet, section, set_widths,
    title,
)

RE = quote_sheet(SHEETS.RATE_ENGINE)
ME = quote_sheet(SHEETS.MOD_ENGINE)
PCT_S = "+0.0%;-0.0%;0.0%"


def _expl(ws, r, text_or_formula, size=10):
    """Explanation in the wide prose column J (static text or live formula)."""
    if text_or_formula.startswith("="):
        formula(ws, f"J{r}", text_or_formula)
        ws[f"J{r}"].font = font(GREY_DARK, size=size, italic=True)
    else:
        prose(ws, f"J{r}", text_or_formula, size=size, width=85)
        ws[f"J{r}"].font = font(GREY_DARK, size=size, italic=True)


def build_walkthrough(ctx: Ctx):
    ws = ctx.wb[SHEETS.WALKTHROUGH]
    p = ctx.p
    cf, cl = L.RE_COH_FIRST, L.RE_COH_LAST

    # RE cohort-block column ranges (INDEX targets for the specimen cohorts)
    rng = {c: f"{RE}!${c}${cf}:${c}${cl}"
           for c in ("B", "D", "H", "I", "J", "K", "L", "M", "N", "P", "Q", "R")}

    # ---- W1: header band + the claim ----
    title(ws, "B2", "Walkthrough — every calculation for one combo, start to finish",
          "Reads top to bottom like the method itself: inputs -> rate level -> earning -> "
          "mod path -> assembly. Live for the Control selection; nothing here is typed.")
    label(ws, "B4", "Selected:")
    link(ws, "C4", "=nr_SelKey", bold=True)
    jump(ws, "E4", "Control!C7", "Change BU/state selection >")
    put(ws, "H4", "synced to Control", fnt=F_SMALL_IT)
    formula(ws, "B5",
            '=IF(nr_SelOK,"Your "&TEXT(nr_LRproj,"0.0%")&" projected loss ratio becomes a "'
            '&TEXT(nr_CYLR_P,"0.0%")&" CY "&nr_PlanYear&" plan loss ratio  ("'
            '&TEXT((nr_CYLR_P-nr_LRcur)*100,"+0.0;-0.0;0.0")&" pts vs current level).",'
            '"Select a BU and state that exist in tbl_LR.")')
    ws["B5"].font = font(NAVY, bold=True, size=13)
    ws["B5"].fill = FILL_PANEL
    for cc in range(3, 11):
        put(ws, ws.cell(row=5, column=cc).coordinate, None, fill=FILL_PANEL)
    formula(ws, "B6",
            '="THE BRIDGE IN ONE LINE:   "&TEXT(nr_LRcur,"0.0%")&"  x  "'
            '&TEXT(nr_Arate_P,"0.0000")&" rate earn-in  x  "&TEXT(nr_Amod_P,"0.0000")'
            '&" mod drift  x  "&TEXT(nr_AOther,"0.0000")&" other  =  "'
            '&TEXT(nr_CYLR_P,"0.0%")')
    ws["B6"].font = font(GREY_DARK, size=10, italic=True)

    # ---- W2: what you gave us ----
    r = 8
    section(ws, r, "B", "1.  What you gave us (tbl_LR, resolved for this combo)")
    _expl(ws, r, "Each row links the live input; 'edit >' opens the exact Inputs cell.")
    r += 1
    w2 = [
        ("Projected loss ratio", "=nr_LRproj", FMT_PCT, "C",
         '=IF(nr_Basis="proposed","entered at the PROPOSED level — step 2 converts it '
         'to current level","already at the current rate level — no conversion needed")'),
        ("LR basis", "=nr_Basis", None, "D",
         "current = today's rates; proposed = includes the indicated change s."),
        ("Indication selected change (s)", "=nr_SelS", FMT_PCT, "E",
         "Only used when the basis is 'proposed'."),
        ("Mod assumed in indication (M_ind)", "=nr_MInd", FMT_MOD, "F",
         "The average schedule mod the indication priced against — the mod benchmark."),
        ("Current avg written mod (M_0)", "=nr_M0", FMT_MOD, "G",
         "Where the written mod stands at the as-of date."),
        ("M_0 as-of date", "=nr_M0Asof", "m/d/yyyy", "H",
         "Anchored close-of-day: 'as of 9/30' means the 10/1 boundary."),
        ("Projected mod, plan-yr end (M_1)", "=nr_M1", FMT_MOD, "I",
         "Where you project the written mod by 12/31 of the plan year."),
        ("Mod ~1 yr before as-of (M_prior)", "=nr_MPrior", FMT_MOD, "J",
         '=IF(N(nr_MPrior)=0,"not provided — the path extends the M_0 -> M_1 slope '
         'backward","anchors the path 12 months before the as-of date")'),
        ("Projected mod, next-yr end (M_2)", "=nr_M2", FMT_MOD, "K",
         '=IF(N(nr_M2)=0,"not provided — M_1 carries flat beyond the plan year",'
         '"anchors the path at the end of the following year")'),
        ("Net trend (next-year view)", "=nr_Trend", FMT_PCT, "M",
         "Applied once, to the following year only (inherits the Control default when "
         "the tbl_LR cell is blank)."),
        ("Other adjustment (A_other)", "=nr_AOther", FMT_IDX, "N",
         '=IF(nr_AOtherLbl="","no manual adjustment","reason: "&nr_AOtherLbl)'),
        ("Mod adjustment effective?", "=nr_ModAdjEff", None, "P",
         "Master toggle (Control) AND the combo's own toggle must both be ON."),
        ("Net rate selection", "=IF(nr_NetMode,nr_NetSelP,0)", FMT_PCT, "Q",
         '=IF(nr_NetMode,"ACTIVE: from 1/1 the combined rate x price renews this % above '
         'the year-ago cohort — planned rows and the mod projection are superseded",'
         '"off — the explicit rate program below drives the year")'),
    ]
    lr0 = L.LR_FIRST - 1
    for lbl, f, fmt, colL, expl in w2:
        label(ws, f"B{r}", lbl)
        link(ws, f"C{r}", f, fmt=fmt or "General", border=BORDER_THIN, align=ALIGN_C)
        formula(ws, f"I{r}",
                f'=IF(br_selrow=0,"",HYPERLINK("#Inputs!{colL}"&(br_selrow+{lr0}),"edit >"))')
        ws[f"I{r}"].font = Font(name="Calibri", color=NAVY, size=9, underline="single")
        _expl(ws, r, expl, size=9)
        r += 1

    # ---- W3: where your rate level stands (CRL_ind) ----
    r += 1
    section(ws, r, "B", "2.  Where the indication's rate level stands (CRL_ind)")
    _expl(ws, r, "Chronological rate changes for this combo. The cumulative column "
                 "multiplies (1 + change) over the rows flagged 'in indication' — that "
                 "product is CRL_ind, the level the indication assumed FULLY EARNED.")
    r += 1
    header_row(ws, r, 2, ["#", "Effective", "Filed %", "Status", "In indication?",
                          "Effective %", "Cumulative level (considered rows)"],
               widths=None, fill=FILL_GREY, fnt=font(GREY_DARK, bold=True, size=9))
    r += 1
    slots0 = r
    for j in range(1, 9):
        cnt = f"COUNTIFS(rl_key,nr_SelKey,rl_seq,{j})"
        put(ws, f"B{r}", j, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
        formula(ws, f"C{r}", f'=IF({cnt}=0,"",SUMIFS(rl_eff,rl_key,nr_SelKey,rl_seq,{j}))',
                fmt="m/d/yy", align=ALIGN_C)
        formula(ws, f"D{r}", f'=IF({cnt}=0,"",SUMIFS(rl_filed,rl_key,nr_SelKey,rl_seq,{j}))',
                fmt=FMT_PCT, align=ALIGN_C)
        formula(ws, f"E{r}",
                f'=IF({cnt}=0,"",IF(COUNTIFS(rl_key,nr_SelKey,rl_seq,{j},rl_status,'
                f'"planned")>0,"planned","taken"))', align=ALIGN_C)
        formula(ws, f"F{r}",
                f'=IF({cnt}=0,"",IF(COUNTIFS(rl_key,nr_SelKey,rl_seq,{j},rl_cons,"Y")>0,'
                f'"Y","N"))', align=ALIGN_C)
        formula(ws, f"G{r}", f'=IF({cnt}=0,"",SUMIFS(rl_reff,rl_key,nr_SelKey,rl_seq,{j}))',
                fmt=PCT_S, align=ALIGN_C)
        formula(ws, f"H{r}",
                f'=IF({cnt}=0,"",EXP(SUMPRODUCT((rl_key=nr_SelKey)*(rl_cons="Y")'
                f"*(rl_seq<={j})*rl_ln1p)))", fmt=FMT_IDX, align=ALIGN_C)
        r += 1
    label(ws, f"F{r}", "CRL_ind", bold=True)
    link(ws, f"H{r}", "=nr_CRLind", fmt=FMT_IDX, align=ALIGN_C, bold=True, fill=FILL_PANEL)
    _expl(ws, r, "Computed as EXP(SUMPRODUCT(LN(1 + change))) over considered rows — the "
                 "exact product identity, blank-row safe (Methodology 4).", size=9)
    jump(ws, f"I{r}", f"{quote_sheet(SHEETS.RATE_LOG)}!A1", "rate log >", size=9)
    r += 2

    # ---- W4: three cohorts under the microscope ----
    section(ws, r, "B", "3.  How the calendar year earns it — three cohorts under the "
                        "microscope")
    _expl(ws, r, "The engine writes one cohort of policies per month (48 months of them) "
                 "and earns each across its term. Three specimens: A carries last year's "
                 "rates into the plan year; B catches a mid-month change; C straddles into "
                 "the following year.")
    r += 1
    hdr_r = r
    # specimen cohort indices (grey helpers in H): A carryover, B first cohort
    # with an in-month change (fallback: Jan P), C straddles into P+1
    idx_cells = {}
    for k, (lab, f) in enumerate([
        ("A", "=12+13-ROUND(nr_TermMonths/2,0)"),
        ("B", f'=IF(COUNTIF({rng["K"]},">0")=0,25,'
              f'MATCH(1,INDEX(({rng["K"]}>0)*1,0),0))'),
        ("C", "=24+13-ROUND(nr_TermMonths/2,0)"),
    ]):
        cc = col(3 + k)
        put(ws, f"H{hdr_r + 1 + k}", f"idx {lab}", fnt=font(GREY_DARK, size=8))
        formula(ws, f"I{hdr_r + 1 + k}", f, fmt=FMT_INT)
        ws[f"I{hdr_r + 1 + k}"].font = font(GREY_DARK, size=8)
        idx_cells[lab] = f"$I${hdr_r + 1 + k}"
    header_row(ws, hdr_r, 2,
               ["", "Cohort A - carryover", "Cohort B - mid-month chg", "Cohort C - straddles"],
               widths=None, fill=FILL_GREY, fnt=font(GREY_DARK, bold=True, size=9))
    r += 1
    ia, ib, ic = idx_cells["A"], idx_cells["B"], idx_cells["C"]
    w4 = [
        ("Written month", "B", "mmm yyyy",
         "When this cohort of policies goes on the books."),
        ("Weight w", "H", "0.00",
         "Written-exposure weight (seasonality-adjusted when active)."),
        ("Rate index before the month (W_pre)", "I", FMT_IDX,
         "Cumulative level from changes effective before the month."),
        ("Rate index at month end (W_post)", "J", FMT_IDX,
         "Cumulative level including in-month changes."),
        ("# changes in the month", "K", FMT_INT,
         "Cohort B catches one mid-month — see the split below."),
        ("Days on/after the first change", "L", FMT_INT,
         "Of the days in the month, how many fall on/after the change date."),
        ("Split p = days after / days in month", "M", "0.000",
         "The day-weighted share of the month's writings at the NEW level."),
        ("Blended written index W = (1-p) x W_pre + p x W_post", "N", FMT_IDX,
         "What this cohort actually goes on the books at."),
        ("Share earned in the prior year", "P", "0.0000",
         "Fractions of this cohort's premium falling in each calendar year;"),
        ("Share earned in the PLAN year", "Q", "0.0000",
         "they follow the mid-month rule C(o) = MIN(1, MAX(0, (o + 0.5) / T))"),
        ("Share earned in the following year", "R", "0.0000",
         "and always sum to 1 across the window."),
    ]
    for lbl, colL, fmt, expl in w4:
        label(ws, f"B{r}", lbl)
        for cc, idx in (("C", ia), ("D", ib), ("E", ic)):
            formula(ws, f"{cc}{r}", f"=INDEX({rng[colL]},{idx})", fmt=fmt, align=ALIGN_C)
        _expl(ws, r, expl, size=9)
        r += 1
    label(ws, f"B{r}", "Check: recompute the blend by hand", bold=True)
    for cc, idx in (("C", ia), ("D", ib), ("E", ic)):
        formula(ws, f"{cc}{r}",
                f"=(1-INDEX({rng['M']},{idx}))*INDEX({rng['I']},{idx})"
                f"+INDEX({rng['M']},{idx})*INDEX({rng['J']},{idx})",
                fmt=FMT_IDX, align=ALIGN_C)
        ws[f"{cc}{r}"].fill = FILL_PANEL
    _expl(ws, r, "Identical to the engine's W above — the split rule recomputed from its "
                 "ingredients (cohorts without an in-month change blend trivially).", size=9)
    jump(ws, f"I{r}", f"{RE}!A{L.RE_COH_HDR}", "all 48 cohorts >", size=9)
    r += 2

    # ---- W5: aggregate to the year ----
    section(ws, r, "B", "4.  Adding it up — the year's earned rate level and A_rate")
    _expl(ws, r, "E_CY is an aggregate ratio (never an average of monthly ratios): total "
                 "rate-weighted earned exposure over total earned exposure, cohort by "
                 "cohort, month by month.")
    r += 1
    mp0, mp1 = col(L.RE_MATRIX_COL + 12), col(L.RE_MATRIX_COL + 23)
    w5 = [
        ("Numerator: SUM of w x e x W over the plan year",
         f"=SUM({RE}!{mp0}${L.RE_ROW_NUM}:{mp1}${L.RE_ROW_NUM})", FMT_IDX,
         "Rate-weighted earned exposure across the plan year's 12 months."),
        ("Denominator: SUM of w x e",
         f"=SUM({RE}!{mp0}${L.RE_ROW_DEN}:{mp1}${L.RE_ROW_DEN})", FMT_IDX,
         "Total earned exposure in the plan year."),
        ("Earned rate level  E_CY(P)", "=nr_ECY_P", FMT_IDX,
         '="Numerator / denominator — the level CY "&nr_PlanYear&" premium actually '
         'carries."'),
        ("Rate earn-in  A_rate = CRL_ind / E_CY", "=nr_Arate_P", FMT_IDX,
         '=IF(nr_NetMode,"under the net selection this is the COMBINED rate x price '
         'factor A_net (mod drift shows 1.000)",IF(nr_Arate_P>1,"above 1.000: the year '
         'earns LESS rate than the indication assumed — pressure UP on the plan LR",'
         '"at or below 1.000: extra rate earns in — pressure DOWN on the plan LR"))'),
    ]
    for lbl, f, fmt, expl in w5:
        label(ws, f"B{r}", lbl)
        (link if f.startswith("=n") else formula)(ws, f"C{r}", f, fmt=fmt, align=ALIGN_C,
                                                  border=BORDER_THIN,
                                                  bold=("A_rate" in lbl))
        _expl(ws, r, expl, size=9)
        r += 1
    r += 1

    # ---- W6: the mod path ----
    section(ws, r, "B", "5.  The schedule-mod path — the price lever, same earning logic")
    _expl(ws, r, "The written mod runs straight lines through your anchors; each cohort "
                 "samples it at mid-month, and the year averages those levels through the "
                 "SAME earning matrix (mods are levels that average, not changes that "
                 "compound).")
    r += 1
    anchors = [
        ("Backward anchor (as-of - 12 mo)", 6),
        ("M_0 at the as-of date", 7),
        ("M_1 at plan-year end", 8),
        ("M_2 at next-year end", 9),
    ]
    for lbl, arow in anchors:
        label(ws, f"B{r}", lbl)
        link(ws, f"C{r}", f"={ME}!$F${arow}-1", fmt="m/d/yy", align=ALIGN_C)
        link(ws, f"D{r}", f"={ME}!$G${arow}", fmt=FMT_MOD, align=ALIGN_C)
        r += 1
    label(ws, f"B{r}", "Path slope over the plan year (mod pts / month)")
    link(ws, f"C{r}", f"={ME}!$H$7*30.4375*100", fmt="0.00", align=ALIGN_C)
    _expl(ws, r, "How fast the written mod drifts; the identity below only holds for a "
                 "globally straight path.", size=9)
    r += 1
    label(ws, f"B{r}", "Earned mod, plan year")
    link(ws, f"C{r}", "=nr_MEarned_P", fmt=FMT_MOD, align=ALIGN_C, border=BORDER_THIN)
    _expl(ws, r, '=IF(me_identity_ok,"sanity check: equals the written mod at Jan 1 ("'
                 '&TEXT(me_identity,"0.000")&") — the uniform + annual + linear identity",'
                 '"the Jan-1 identity is suspended (seasonality, short term, or a kinked '
                 'path) — the earned average still governs")')
    r += 1
    label(ws, f"B{r}", "Mod drift  A_mod = M_ind / earned mod", bold=True)
    link(ws, f"C{r}", "=nr_Amod_P", fmt=FMT_IDX, align=ALIGN_C, border=BORDER_THIN,
         bold=True)
    _expl(ws, r, '=IF(nr_NetMode,"1.000 — merged into the net factor in step 4",'
                 '"the indication assumed "&TEXT(nr_MInd,"0.000")&"; the year earns "'
                 '&TEXT(nr_MEarned_P,"0.000")&" — the gap is a price change that earns '
                 'in exactly like rate")')
    jump(ws, f"I{r}", f"{ME}!A1", "mod engine >", size=9)
    r += 2

    # ---- W7: assembly ----
    section(ws, r, "B", "6.  Assembly — multiply the factors")
    r += 1
    asm0 = r
    asm = [
        ("Projected LR at current rate level", None, "=nr_LRcur",
         "Step 1's loss pick, basis-normalized."),
        ("x  rate earn-in  A_rate", "=nr_Arate_P", None, ""),
        ("x  mod drift  A_mod", "=nr_Amod_P", None, ""),
        ("x  other  A_other", "=nr_AOther", None, ""),
    ]
    for i, (lbl, fac, seed, expl) in enumerate(asm):
        label(ws, f"B{r}", lbl, bold=(i == 0))
        if seed:
            formula(ws, f"D{r}", seed, fmt=FMT_PCT, align=ALIGN_C)
        else:
            link(ws, f"C{r}", fac, fmt=FMT_IDX, align=ALIGN_C)
            formula(ws, f"D{r}", f"=$D{r - 1}*$C{r}", fmt=FMT_PCT, align=ALIGN_C)
        if expl:
            _expl(ws, r, expl, size=9)
        r += 1
    label(ws, f"B{r}", "RESULT:  CY plan loss ratio", bold=True)
    formula(ws, f"D{r}", f"=$D{r - 1}", fmt=FMT_PCT, align=ALIGN_C, bold=True,
            fill=FILL_PANEL, border=BORDER_THIN)
    ctx.define("nr_WalkLR", SHEETS.WALKTHROUGH, f"$D${r}",
               "Walkthrough assembled CY plan LR (must tie nr_CYLR_P; Checks-enforced)")
    _expl(ws, r, '="The same number the Bridge, Portfolio row, and State Summary show for "'
                 '&nr_SelKey&"."')
    r += 1
    label(ws, f"B{r}", "Next year (indicative)")
    formula(ws, f"D{r}", "=nr_LRcur*(1+nr_Trend)*nr_Arate_P1*nr_Amod_P1*nr_AOther",
            fmt=FMT_PCT, align=ALIGN_C)
    _expl(ws, r, "Adds one year of net trend and swaps in next year's earn-in factors — "
                 "indicative only (the canonical caveat lives on the Bridge).", size=9)
    r += 2

    # ---- W8: say it upward ----
    section(ws, r, "B", "7.  Say it upward — two sentences for your VP")
    r += 1
    formula(ws, f"B{r}",
            '="CY "&nr_PlanYear&" plan loss ratio for "&nr_SelKey&" is "'
            '&TEXT(nr_CYLR_P,"0.0%")&", versus "&TEXT(nr_LRcur,"0.0%")&" projected at '
            'current rate level ("&TEXT((nr_CYLR_P-nr_LRcur)*100,"+0.0;-0.0;0.0")&" pts)."')
    ws[f"B{r}"].font = font(NAVY, size=11)
    r += 1
    formula(ws, f"B{r}",
            '=IF(nr_NetMode,"Driver: the "&TEXT(nr_NetSelP,"+0.0%;-0.0%")&" net rate '
            'selection from 1/1 — rate and price together renew at that pace, so earn-in '
            'nets to "&TEXT((nr_Arate_P*nr_Amod_P-1)*100,"+0.0;-0.0;0.0")&" pts of LR.",'
            '"Drivers: rate earn-in contributes "&TEXT(nr_LRcur*(nr_Arate_P-1)*100,'
            '"+0.0;-0.0;0.0")&" pts and schedule-mod drift "&TEXT(nr_LRcur*nr_Arate_P*'
            '(nr_Amod_P-1)*100,"+0.0;-0.0;0.0")&" pts; the year earns "'
            '&TEXT(nr_EChgVsInd,"+0.0%;-0.0%;0.0%")&" versus the level the indication '
            'assumed.")')
    ws[f"B{r}"].font = font(NAVY, size=11)
    r += 2

    # ---- W9: trust it ----
    section(ws, r, "B", "8.  Trust it — this page reconciles to everything else")
    r += 1
    strip = [
        ("This walkthrough vs the Bridge",
         '=IF(ABS(nr_WalkLR-nr_CYLR_P)<0.000000001,"MATCHES","DIFFERS")',
         "Same factors, same product."),
        ("Bridge vs the hidden all-combo engine",
         '=IF(br_selrow=0,"n/a",IF(ABS(INDEX(calc_cylr_p,br_selrow)-nr_CYLR_P)'
         '<0.000000001,"MATCHES","DIFFERS"))',
         "The visible engine equals the _calc block Portfolio and State Summary read."),
        ("Workbook vs the independent Python oracle",
         '=IF(NOT(orc_fp),"n/a - sample data changed (regenerate to re-bake)",'
         'IF(ABS(nr_CYLR_P-orc_cylr_p)<0.000001,"MATCHES","DIFFERS"))',
         "Constants baked at build time by src/engine.py; scoped to the seeded sample."),
    ]
    for lbl, f, expl in strip:
        label(ws, f"B{r}", lbl)
        formula(ws, f"C{r}", f, align=ALIGN_C, bold=True, border=BORDER_THIN)
        for val, fill_, fcol in (("MATCHES", FILL_GREEN, PASS_GREEN),
                                 ("DIFFERS", FILL_RED, FAIL_RED)):
            ws.conditional_formatting.add(
                f"C{r}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill_,
                                    font=font(fcol, bold=True)))
        _expl(ws, r, expl, size=9)
        r += 1
    jump(ws, f"B{r}", "Checks!A1", "Full validation panel: Checks >", size=10)

    # ---- W6 mini chart: written vs earned mod path (reads the Flow table) ----
    from .sheets_report import CD0
    flow = ctx.wb[SHEETS.FLOW]
    mini = LineChart()
    mini.add_data(Reference(flow, min_col=5, min_row=CD0, max_row=CD0 + L.N_MONTHS),
                  titles_from_data=True)
    mini.add_data(Reference(flow, min_col=6, min_row=CD0, max_row=CD0 + L.N_MONTHS),
                  titles_from_data=True)
    mini.set_categories(Reference(flow, min_col=2, min_row=CD0 + 1,
                                  max_row=CD0 + L.N_MONTHS))
    for s, rgb in zip(mini.series, (STEEL, NAVY)):
        s.graphicalProperties.line.solidFill = rgb
        s.graphicalProperties.line.width = int(2.25 * 12700)
        s.marker = Marker(symbol="none")
        s.smooth = False
    mini.y_axis.number_format = "0.000"
    mini.title = "Written vs earned mod path"
    mini.style = 2
    mini.height = 6.5
    mini.width = 11
    mini.visible_cells_only = False
    mini.x_axis.delete = False
    mini.y_axis.delete = False
    chart_legend(mini)           # legend beside the plot, not on it (D69)
    ws.add_chart(mini, "K8")

    set_widths(ws, {"A": 2, "B": 42, "C": 13, "D": 13, "E": 13, "F": 13, "G": 12,
                    "H": 12, "I": 12, "J": 85})
    presentation_setup(ws, gridlines_off=True, freeze="A7", tab_color=NAVY)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:J{r + 1}"
    print_setup(ws, landscape=False)