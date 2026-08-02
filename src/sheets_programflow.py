"""Program Flow sheet (DECISIONS.md D59; two-year view D68).

The DESCRIPTIVE twin of Net Delivery: no target, no suggested change — what
the program AS LOGGED is delivering on renewals, state by state and month by
month, split into its rate and pricing (schedule-mod) legs:

  delivered(m) = [W(m)/W(m-12)] x [M_w(m)/M_w(m-12)]

Program basis: taken rows at filed %, planned rows at filed % x achievement —
net-mode supersession is deliberately NOT applied, so for net-target combos
the gap between this tab and the asserted (1+x) is the exhibit's point (Net
Delivery closes it). Everything reads the existing _calc combo blocks via the
blessed stride-INDEX — the raw written index (col N) and the written mod path
(col O) are program-basis even in net mode, and the block results rows carry
the w-weighted ratio averages for the plan year (H/I/J) and for the year now
flowing (M/N/O). Oracle: engine.program_flow_by_month; every displayed figure
ties at 1e-9.

D68 — two years side by side. The cohort blocks always spanned Jan P-2..Dec
P+1, so P-1 was computed and never shown. The tab now reads left to right as
"what is flowing now, and how it flows into the plan year":

    A          B..M                        N..Y
    State      Jan P-1 .. Dec P-1          Jan P .. Dec P
               steel band                  navy band

The SUMMARY table uses the same two column bands (prior averages at B/C/D,
the plan-year metrics at N..T) rather than packing itself into A..I. That is
what makes the prior year genuinely collapsible: columns B..M carry an outline
group, and collapsing them removes the prior year from the summary AND the
grids in one click, leaving a tab identical to the pre-D68 layout. Freeze is
at column B so the state roster stays put across all 24 months.
"""

from __future__ import annotations

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation

from .build_workbook import Ctx, Layout as L, SHEETS
from .sheets_calc import FLOW_PUB, PRIOR_PUB
from .xlstyle import (ALIGN_C, ALIGN_L, BORDER_THIN, FAIL_RED, FILL_NAVY, FILL_PANEL,
    FILL_STEEL, FMT_PTS_COL, F_LABEL, F_SMALL_IT, GREY_DARK, NAVY, col, font, formula,
    header_row, input_cell, jump, label, link, nav_bar, presentation_setup,
    print_setup, put, quote_sheet as _q, section, set_widths, title,
)

from .xlstyle import FMT_PCT_SIGNED as PCT_S   # one definition, seven readers

# ---- sheet geometry (module-level so the harness can address the exhibit) ----
PF_BAND_ROW = 7                         # the two-year band caption
PF_SUM_HDR = 8                          # summary table header row
PF_SUM_FIRST = 9                        # first roster row
PF_PRIOR_COL = 2                        # B  — Jan P-1, and the prior-year summary
PF_PLAN_COL = 14                        # N  — Jan P, and the plan-year summary
PF_HELP_COL = 27                        # AA — grey helpers, right of both years


def n_disp(cfg) -> int:
    """Display slots: live states + 3 roster spares (the _lists convention)."""
    return len(cfg.states) + 3


def pf_tot(cfg) -> int:
    return PF_SUM_FIRST + n_disp(cfg) + 1


def grid_starts(cfg) -> list[int]:
    """Section rows of the three grids (rate leg, mod leg, delivered). Each
    grid block: section, note, month header, nd roster rows, and one IN VIEW
    total row (D60)."""
    nd = n_disp(cfg)
    g1 = pf_tot(cfg) + 4
    g2 = g1 + 3 + nd + 1 + 3
    return [g1, g2, g2 + 3 + nd + 1 + 3]


def build_program_flow(ctx: Ctx):
    ws = ctx.wb[SHEETS.PROGRAM_FLOW]
    nd = n_disp(ctx.cfg)
    cbf, stride = L.CALC_BLOCK_FIRST, L.CALC_BLOCK_STRIDE
    # summary column letters, derived from the two band anchors so the whole
    # exhibit moves together if the geometry ever changes
    pRATE, pMOD, pDEL = (col(PF_PRIOR_COL + k) for k in range(3))
    cEP, cNET, cRATE, cMOD, cDEL, cDLT, cGAP = (col(PF_PLAN_COL + k) for k in range(7))
    hKEY, hBLK, hGAP = (col(PF_HELP_COL + k) for k in range(3))
    # Bounded _calc column references. INDEX over a WHOLE column evaluates
    # cheaply but makes Excel carry a dependency on all 1,048,576 rows, and
    # this tab holds thousands of them — D68 doubled the count and a full
    # rebuild slowed measurably. Bounding costs nothing and changes no
    # arithmetic: the range starts at row 1, so INDEX's offset is still the
    # absolute row number. The bound covers every combo block plus its
    # results row.
    calc_last = L.CALC_BLOCK_FIRST + L.LR_ROWS * L.CALC_BLOCK_STRIDE + 60

    title(ws, "A1", f"Program Flow — what the logged program delivers ({ctx.lob.name})",
          "Month-by-month YoY change on renewals from the rate changes and mod path AS "
          "LOGGED — taken rows plus planned rows at achievement. Two years side by side: "
          "what is flowing now, and how it flows into the plan year. The descriptive twin "
          "of Net Delivery: no target, no suggestion, just the flow you are on today. "
          "'All' = the EP-weighted book view (D60).")

    def pub(key, j=0, mp=FLOW_PUB):
        """A published results column as an absolute range (D60/D68)."""
        cL = col(mp[key] + j)
        return f"'_calc'!${cL}${L.CALC_RES_FIRST}:${cL}${L.CALC_RES_LAST}"

    nav_bar(ws, 3, 1, ["Control", "State Summary", "Net Delivery", "Flow Dashboard",
                       "Rate Log", "Checks"], step=2)

    # ---- controls ----
    label(ws, "A5", "Business unit in view", bold=True)
    c = input_cell(ws, "B5", ctx.we_row["bu"])
    c.alignment = ALIGN_C
    ctx.define("pf_BU", SHEETS.PROGRAM_FLOW, "$B$5",
               "Program Flow BU in view ('All' = the EP-weighted book view, D60)")
    dv_bu = DataValidation(type="list", formula1="=lst_bu_all", allow_blank=False,
                           showErrorMessage=True)
    ws.add_data_validation(dv_bu)
    dv_bu.add("B5")
    formula(ws, "A6",
            '=IF(pf_BU="All","BOOK VIEW: EP-weighted across BUs (D60). Legs are weighted '
            'means — rate x mod need not multiply to delivered exactly under mix; '
            'delivered is the exact statistic. The delta column is per-BU: pick one BU '
            'for targeting.",'
            'IF(nr_ModAdjMaster="OFF","NOTE: master mod toggle is OFF — mod legs show a '
            'dash and delivered = the rate leg alone.",'
            'IF(SUMPRODUCT(calc_netmode*(calc_bu=pf_BU))>0,"Net-target combo(s) in view: '
            'these grids show the program AS LOGGED — the gap vs the asserted net is the '
            'point. See Net Delivery for the prescriptive view.","")))')
    ws["A6"].font = font(FAIL_RED, size=9, italic=True)

    # ---- the two-year band caption (D68) --------------------------------
    # Painted across the full 12 columns of each year so the summary table's
    # ragged right edge reads as part of a band, and so the seam between the
    # years is unmistakable even before you read a header.
    def band(row_, c0, ncol, fill, text):
        for k in range(ncol):
            cell = ws.cell(row=row_, column=c0 + k)
            cell.fill = fill
            cell.font = font("FFFFFF", bold=True, size=10)
            cell.alignment = ALIGN_C
        # left-aligned so the caption spills rightwards across its own band
        # (fill alone does not block overflow) instead of over the state column
        first_ = ws.cell(row=row_, column=c0)
        first_.value = text
        first_.alignment = ALIGN_L

    band(PF_BAND_ROW, PF_PRIOR_COL, 12, FILL_STEEL,
         '="  FLOWING NOW  —  "&(nr_PlanYear-1)')
    band(PF_BAND_ROW, PF_PLAN_COL, 12, FILL_NAVY,
         '="  PLAN YEAR  —  "&nr_PlanYear')

    # ---- per-state summary table ----
    hdr, first = PF_SUM_HDR, PF_SUM_FIRST
    put(ws, f"A{hdr}", "State", fnt=font("FFFFFF", bold=True), fill=FILL_NAVY,
        align=ALIGN_C)
    header_row(ws, hdr, PF_PRIOR_COL,
               ["Avg YoY rate leg", "Avg YoY mod leg", "Avg YoY delivered net"],
               fill=FILL_STEEL)
    header_row(ws, hdr, PF_PLAN_COL,
               ["Adj plan EP (000s)", "Net target", "Avg YoY rate leg",
                "Avg YoY mod leg", "Avg YoY delivered net", "Δ vs net assertion",
                "Plan LR gap: program vs asserted (pts)"])
    ws.row_dimensions[hdr].height = 42
    for h, cL in (("key", hKEY), ("blk row", hBLK), ("|gap|", hGAP)):
        put(ws, f"{cL}{hdr}", h, fnt=font(GREY_DARK, size=8), align=ALIGN_C)
    for i in range(nd):
        r = first + i
        band_fill = FILL_PANEL if i % 2 else None
        link(ws, f"A{r}", f"=_lists!$B${3 + i}", align=ALIGN_C, fill=band_fill,
             border=BORDER_THIN, bold=True)
        # grey helpers: resolved key and combo-block top row (0 = not in view)
        formula(ws, f"{hKEY}{r}", f'=IF(OR($A{r}="",pf_BU="All"),"",pf_BU&"|"&$A{r})')
        formula(ws, f"{hBLK}{r}",
                f'=IF(${hKEY}{r}="",0,IF(COUNTIF(calc_key,${hKEY}{r})=0,0,'
                f"{cbf}+(MATCH(${hKEY}{r},calc_key,0)-1)*{stride}))", fmt="0")
        formula(ws, f"{hGAP}{r}",
                f"=IF(${hBLK}{r}=0,0,IF(INDEX('_calc'!$L$1:$L${calc_last},${hBLK}{r}),"
                f"ABS(INDEX('_calc'!$J$1:$J${calc_last},${hBLK}{r}+51)-1"
                f"-INDEX('_calc'!$M$1:$M${calc_last},${hBLK}{r})),0))",
                fmt="0.0000")
        for cL in (hKEY, hBLK, hGAP):
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=8)
        formula(ws, f"{cEP}{r}",
                f'=IF($A{r}="","",IF(pf_BU="All",SUMIFS(calc_ep,calc_state,$A{r}),'
                f"SUMIFS(calc_ep,calc_state,$A{r},calc_bu,pf_BU)))",
                fmt="#,##0;-#,##0;\"\"", align=ALIGN_C, fill=band_fill, border=BORDER_THIN)
        # each metric: =IF(blank,"",IF(All,<EP-weighted book>,<single-BU block read>))
        sg = f'IF(${hBLK}{r}=0,"—",'                  # single-BU inner guard
        stc = f"(calc_state=$A{r})"
        epd = f"SUMIFS(calc_ep,calc_state,$A{r})"

        def _all_avg(key, mp=FLOW_PUB):
            return (f'IF({epd}=0,"—",'
                    f"SUMPRODUCT({stc}*calc_ep*{pub(key, 0, mp)})/{epd}-1)")

        me_d = f"SUMPRODUCT({stc}*calc_ep*{pub('modeff')})"

        def _all_avg_mod(mp=FLOW_PUB):
            """Mod averages weight by EP x the mod-in-force flag, not EP alone."""
            return (f'IF({me_d}=0,"—",'
                    f"SUMPRODUCT({stc}*calc_ep*{pub('modeff')}*"
                    f"{pub('avg_mod', 0, mp)})/{me_d}-1)")

        all_net = (f'IF(SUMIFS(calc_netmode,calc_state,$A{r})=0,"—",'
                   f"SUMIFS(calc_netx,calc_state,$A{r})/"
                   f"SUMIFS(calc_netmode,calc_state,$A{r}))")
        one_net = sg + (f"IF(INDEX('_calc'!$L$1:$L${calc_last},${hBLK}{r}),"
                        f"INDEX('_calc'!$M$1:$M${calc_last},${hBLK}{r}),\"—\"))")
        # plan-year block reads (block results row H/I/J); prior-year block
        # reads M/N/O on the SAME row. Note INDEX('_calc'!$O$1:$O${calc_last}, blk) is the
        # block HEADER row's mod-in-force flag while INDEX(..., blk+51) is the
        # results row's prior-year delivered average — same column, different
        # rows, which is the established stride-INDEX idiom.
        one_rate = sg + f"INDEX('_calc'!$H$1:$H${calc_last},${hBLK}{r}+51)-1)"
        one_mod = sg + (f"IF(INDEX('_calc'!$O$1:$O${calc_last},${hBLK}{r})=0,\"—\","
                        f"INDEX('_calc'!$I$1:$I${calc_last},${hBLK}{r}+51)-1))")
        one_del = sg + f"INDEX('_calc'!$J$1:$J${calc_last},${hBLK}{r}+51)-1)"
        one_prate = sg + f"INDEX('_calc'!$M$1:$M${calc_last},${hBLK}{r}+51)-1)"
        one_pmod = sg + (f"IF(INDEX('_calc'!$O$1:$O${calc_last},${hBLK}{r})=0,\"—\","
                         f"INDEX('_calc'!$N$1:$N${calc_last},${hBLK}{r}+51)-1))")
        one_pdel = sg + f"INDEX('_calc'!$O$1:$O${calc_last},${hBLK}{r}+51)-1)"
        one_dlt = sg + (f"IF(INDEX('_calc'!$L$1:$L${calc_last},${hBLK}{r}),"
                        f"INDEX('_calc'!$J$1:$J${calc_last},${hBLK}{r}+51)-1"
                        f"-INDEX('_calc'!$M$1:$M${calc_last},${hBLK}{r}),\"—\"))")
        # D65: the same disagreement one level up — plan LR, not the legs.
        # The All branch averages over the state's NET combos; the single-BU
        # branch must gate on THAT combo's own net flag, not the state's.
        nm_st = f"SUMIFS(calc_netmode,calc_state,$A{r})"
        one_gap = (f'IF(${hBLK}{r}=0,"—",'
                   f'IF(INDEX(calc_netmode,MATCH(${hKEY}{r},calc_key,0))=0,"—",'
                   f"INDEX(calc_proggap,MATCH(${hKEY}{r},calc_key,0))))")
        all_gap = (f'IF({nm_st}=0,"—",'
                   f"SUMPRODUCT((calc_state=$A{r})*calc_netmode*calc_proggap)/{nm_st})")
        formula(ws, f"{cGAP}{r}",
                f'=IF($A{r}="","",IF(pf_BU="All",{all_gap},{one_gap}))',
                fmt=FMT_PTS_COL, align=ALIGN_C, fill=band_fill,
                border=BORDER_THIN)
        cells = (
            (pRATE, _all_avg("avg_rate", PRIOR_PUB), one_prate, False),
            (pMOD, _all_avg_mod(PRIOR_PUB), one_pmod, False),
            (pDEL, _all_avg("avg_del", PRIOR_PUB), one_pdel, True),
            (cNET, all_net, one_net, False),
            (cRATE, _all_avg("avg_rate"), one_rate, False),
            (cMOD, _all_avg_mod(), one_mod, False),
            (cDEL, _all_avg("avg_del"), one_del, True),
            (cDLT, '"—"', one_dlt, False),
        )
        for cc, allf, onef, bold_ in cells:
            formula(ws, f"{cc}{r}",
                    f'=IF($A{r}="","",IF(pf_BU="All",{allf},{onef}))',
                    fmt=PCT_S, align=ALIGN_C, fill=band_fill, border=BORDER_THIN,
                    bold=bold_)
    tot = pf_tot(ctx.cfg)
    put(ws, f"A{tot}", "TOTAL", fnt=font(NAVY, bold=True), align=ALIGN_C)
    formula(ws, f"{cEP}{tot}",
            '=IF(pf_BU="All",SUM(calc_ep),SUMIFS(calc_ep,calc_bu,pf_BU))',
            fmt="#,##0", align=ALIGN_C, bold=True)
    formula(ws, f"{cNET}{tot}",
            '="all states, "&IF(pf_BU="All","every business unit",pf_BU)&" — "'
            '&IF(pf_BU="All",SUMPRODUCT(calc_netmode),'
            'SUMPRODUCT(calc_netmode*(calc_bu=pf_BU)))&" net combo(s)"')
    ws[f"{cNET}{tot}"].font = F_LABEL
    ctx.define("pfd_states", SHEETS.PROGRAM_FLOW,
               f"$A${first}:$A${first + nd - 1}", "Program Flow roster state column")
    ctx.define("pfd_gap", SHEETS.PROGRAM_FLOW,
               f"${hGAP}${first}:${hGAP}${first + nd - 1}",
               "Abs gap: program-basis delivered avg vs the asserted net (0 = non-net)")
    put(ws, f"A{tot + 2}",
        "Averages are written-weighted means of the monthly YoY ratios over each year "
        "(the Net Delivery convention); under 'All' they are EP-weighted across BUs, "
        "where delivered is the exact statistic (D60). The steel block is the year now "
        "flowing — its year-ago base is the year before that — and it collapses to "
        "nothing with the outline button above column B. Deep dive on one combo: set it "
        "on Control, then see the Flow Dashboard's written legs and locked/planned split.",
        fnt=F_SMALL_IT)

    # ---- three state x month grids, two years wide (D68) ----
    g1, g2, g3 = grid_starts(ctx.cfg)
    secs = [
        (g1, "rate", "The rate leg — YoY written rate change on renewals",
         "History and planned filings at achievement, day-blended in their effective "
         "months. Steps mark anniversaries; a cliff is a filing finishing its year."),
        (g2, "mod", "The pricing leg — YoY written schedule-mod change on renewals",
         "Drift along the anchored mod path (dash = the mod adjustment is off for the "
         "combo). Slow-motion price change, same YoY lens."),
        (g3, "delivered", "Delivered net — rate x pricing, the YoY change customers "
         "actually see",
         "The product of the two legs. For net-target combos compare against the "
         "asserted net: the shortfall months are Net Delivery's to close."),
    ]
    roll = ("Bottom row: every state above collapsed onto one line, weighted by "
            "adjusted plan EP x written weight — the whole book when the picker is "
            "'All' (BOOK AVG), otherwise the selected business unit (BU AVG). It is "
            "an average of the state rows, not a total.")
    # (label, first column, cohort-row offsets num/den, published map, fill)
    years = (("prior", PF_PRIOR_COL, 15, 3, PRIOR_PUB, FILL_STEEL, "nr_PlanYear-1"),
             ("plan", PF_PLAN_COL, 27, 15, FLOW_PUB, FILL_NAVY, "nr_PlanYear"))
    for g0, key, head, note_txt in secs:
        section(ws, g0, "A", head)
        put(ws, f"A{g0 + 1}",
            f"{note_txt}  Left block = the year now flowing; right block = the plan "
            f"year.  {roll}", fnt=F_SMALL_IT)
        put(ws, f"A{g0 + 2}", "State", fnt=font("FFFFFF", bold=True), fill=FILL_NAVY,
            align=ALIGN_C)
        for _yr, c0, _on, _od, _mp, fill_, yexpr in years:
            for j in range(12):
                cell = ws.cell(row=g0 + 2, column=c0 + j)
                cell.value = f'=TEXT(DATE({yexpr},{j + 1},1),"mmm")&" "&({yexpr})'
                cell.font = font("FFFFFF", bold=True, size=9)
                cell.fill = fill_
                cell.alignment = ALIGN_C
        for i in range(nd):
            r = g0 + 3 + i
            rs = first + i
            link(ws, f"A{r}", f"=_lists!$B${3 + i}", align=ALIGN_C, bold=True)
            for _yr, c0, on_, od_, mp, _f, _y in years:
                for j in range(12):
                    nr_ = (f"INDEX('_calc'!$N$1:$N${calc_last},${hBLK}${rs}+{on_ + j})"
                           f"/INDEX('_calc'!$N$1:$N${calc_last},${hBLK}${rs}+{od_ + j})")
                    or_ = (f"INDEX('_calc'!$O$1:$O${calc_last},${hBLK}${rs}+{on_ + j})"
                           f"/INDEX('_calc'!$O$1:$O${calc_last},${hBLK}${rs}+{od_ + j})")
                    if key == "rate":
                        one = f'IF(${hBLK}${rs}=0,"",{nr_}-1)'
                    elif key == "mod":
                        one = (f'IF(${hBLK}${rs}=0,"",'
                               f"IF(INDEX('_calc'!$O$1:$O${calc_last},${hBLK}${rs})=0,\"—\","
                               f"{or_}-1))")
                    else:
                        one = (f'IF(${hBLK}${rs}=0,"",{nr_}*'
                               f"IF(INDEX('_calc'!$O$1:$O${calc_last},${hBLK}${rs})=1,{or_},1)-1)")
                    # All view (D60): EP x w weighted mean across the state's BUs
                    # (w cancels within a state — state-keyed seasonality). The
                    # epw family is the plan year's and serves both years: w is
                    # a function of calendar MONTH, so it repeats (D68).
                    sw = f"(calc_state=$A{r})*{pub('epw', j)}"
                    if key == "mod":
                        dn = f"SUMPRODUCT({sw}*{pub('modeff')})"
                        alv = (f'IF({dn}=0,"—",'
                               f"SUMPRODUCT({sw}*{pub('modeff')}*"
                               f"{pub('mod', j, mp)})/{dn})")
                    else:
                        dn = f"SUMPRODUCT({sw})"
                        alv = (f'IF({dn}=0,"",'
                               f"SUMPRODUCT({sw}*{pub(key, j, mp)})/{dn})")
                    f = f'=IF($A{r}="","",IF(pf_BU="All",{alv},{one}))'
                    cc = ws.cell(row=r, column=c0 + j)
                    formula(ws, cc.coordinate, f, fmt=PCT_S, align=ALIGN_C)
                    cc.font = font(GREY_DARK, size=9)
        # roll-up row (D60): the BU / book-level monthly flow line — every
        # state in view collapsed onto one line, EP x weight weighted.
        # View condition is purely multiplicative (no IF over arrays):
        # ((pf_BU="All")+(calc_bu=pf_BU)) is 1 for every combo under All,
        # else 1 only for the picked BU's combos.
        r_t = g0 + 3 + nd
        formula(ws, f"A{r_t}", '=IF(pf_BU="All","BOOK AVG","BU AVG")',
                align=ALIGN_C)
        ws[f"A{r_t}"].font = font(NAVY, bold=True, size=9)
        for _yr, c0, _on, _od, mp, _f, _y in years:
            for j in range(12):
                vw = f'((pf_BU="All")+(calc_bu=pf_BU))*{pub("epw", j)}'
                ext = f"*{pub('modeff')}" if key == "mod" else ""
                dn = f"SUMPRODUCT({vw}{ext})"
                f = (f'=IF({dn}=0,"{"—" if key == "mod" else ""}",'
                     f"SUMPRODUCT({vw}{ext}*{pub(key, j, mp)})/{dn})")
                cc = ws.cell(row=r_t, column=c0 + j)
                formula(ws, cc.coordinate, f, fmt=PCT_S, align=ALIGN_C)
                cc.font = font(NAVY, bold=True, size=9)
    # one colour scale PER YEAR: a single scale spanning both would normalise
    # over the boundary and flatten the contrast inside each year (D68)
    for c0 in (PF_PRIOR_COL, PF_PLAN_COL):
        ws.conditional_formatting.add(
            f"{col(c0)}{g3 + 3}:{col(c0 + 11)}{g3 + 2 + nd}",
            ColorScaleRule(start_type="min", start_color="D6E8D5",
                           mid_type="percentile", mid_value=50, mid_color="FFF2CC",
                           end_type="max", end_color="F4B8B8"))

    # ---- footer ----
    fn = g3 + 3 + nd + 2
    put(ws, f"A{fn}",
        "Program basis (D59): every Rate Log row enters exactly as logged — taken at "
        "filed %, planned at filed % x achievement. A net rate selection changes NOTHING "
        "here by design; this tab is the reality check against that assertion. Under "
        "'All' (and on IN VIEW rows) legs are EP x weight means — rate x mod need not "
        "multiply to delivered exactly under mix; delivered is exact (D60).",
        fnt=F_SMALL_IT)
    put(ws, f"A{fn + 1}",
        "Two years (D68): the left block is the year now flowing, measured against the "
        "year before it; the right block is the plan year measured against the year now "
        "flowing. So a filing appears in the left block while it is earning through, and "
        "in the right block only for the months before its anniversary — reading across "
        "the seam shows exactly how much of today's momentum is still there next year. "
        "Collapse the left block with the outline button above column B to get the "
        "plan-year-only view back.", fnt=F_SMALL_IT)
    jump(ws, f"A{fn + 2}", f"{_q(SHEETS.NET_DELIVERY)}!A1",
         "Net combos: see Net Delivery for the filing + pricing walk that closes the gap >")
    set_widths(ws, {"A": 10})
    for j in range(12):
        ws.column_dimensions[col(PF_PRIOR_COL + j)].width = 10
        ws.column_dimensions[col(PF_PLAN_COL + j)].width = 10
    # the prior year collapses as one outline group; Excel puts the +/- button
    # above the first column to its right (the plan-year January)
    for j in range(12):
        ws.column_dimensions[col(PF_PRIOR_COL + j)].outlineLevel = 1
    set_widths(ws, {hKEY: 14, hBLK: 7, hGAP: 8})
    presentation_setup(ws, gridlines_off=True, freeze=f"B{PF_SUM_FIRST}",
                       tab_color=NAVY)
    print_setup(ws)
