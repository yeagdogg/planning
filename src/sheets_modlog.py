"""Mod Log sheet: stepped schedule-mod actions (DECISIONS.md D70).

A mod action is a dated percent that stays in force until superseded — the
same mechanic as a rate filing, applied to the schedule mod. It is how the net
rate strategy is actually executed, and it makes the mod path something the
book EARNS rather than something the plan asserts.

Its OWN sheet, and its own ``ml_*`` names, deliberately. A ``kind`` column on
the Rate Log looked cheaper, but ~20 places read ``rl_ln1p`` / ``rl_seq`` /
``rl_status`` assuming every row is a rate row — CRL_ind among them — so a mod
row in that table would have to be excluded at each of them, and one missed
site produces a plausible wrong on-level index. A separate table makes that
contamination impossible by construction rather than by audit.

Layout mirrors the Rate Log so the two feel identical:
  A:G  paste block (no "in indication?" — M_ind carries that, per combo)
  I:R  engine helpers
  T:U  live same-row feedback

The live column T is the one this exhibit exists for: the share of CY P the
action actually earns. Type a date and watch it fall — a January action earns
half the plan year on an annual book, an October action barely 3%.
"""

from __future__ import annotations

from openpyxl.worksheet.table import Table

from .build_workbook import Ctx, Layout as L, SHEETS
from .sheets_inputs import TABLE_STYLE, _dv
from .xlstyle import (
    ALIGN_C, ALIGN_L, BORDER_THIN, F_SMALL_IT, FAIL_RED, FILL_GREY, FMT_DATE,
    FMT_IDX, FMT_INT, FMT_MOD, FMT_PCT, GREY_DARK, WARN_AMBER, font, formula,
    header_row,
    input_cell, nav_bar, presentation_setup, print_setup, put, quote_sheet,
    set_widths, title,
)

RE = quote_sheet(SHEETS.RATE_ENGINE)
from .xlstyle import FMT_PCT_SIGNED as PCT_S   # one definition, seven readers

ML_HELPER_HEADERS = ["Key", "m_eff", "ln(1+m)", "Eff month", "Days on/after",
                     "Earlier cnt", "Same cnt", "First?", "Dup month?", "Seq"]


def ml_headers() -> list[str]:
    return ["BU", "State", "Effective date", "Mod change %",
            "Status (taken / planned)", "Achievement % (planned)", "Comment"]


def build_mod_log(ctx: Ctx):
    ws = ctx.wb[SHEETS.MOD_LOG]
    p = ctx.p
    title(ws, "A1", f"Schedule-Mod Action Log — {ctx.lob.name}",
          "One row per mod action per BU x state: a dated percent change to the average "
          "schedule mod that stays in force until another action supersedes it. Positive = "
          "taking mod (raising price); negative = giving it away. Actions compound on the "
          "projected CURRENT-year-end mod (M_endPrior on Inputs) from 1/1 of the plan year "
          "— everything before that is carried by the drift path, so nothing is counted "
          "twice. Logging a combo's FIRST action also makes M_endPrior live: the drift path "
          "then lands on it at 12/31 of the current year instead of running on to M_1, so "
          "expect the current year's earned mod to move a little the first time.")
    nav_bar(ws, 3, 1, ["Control", "Inputs", "Rate Log", "Net Delivery", "Checks"], step=2)
    put(ws, "A4",
        ("CARRIED FORWARD: these are your own logged actions. Do not insert or delete "
         "rows — spare capacity is provided below."
         if ctx.carried is not None else
         "SAMPLE DATA: replace the populated rows with your own actions — these seeded "
         "rows attach to real combos and MOVE REAL PLAN LOSS RATIOS until you do. Do not "
         "insert or delete rows — spare capacity is provided below the sample. Leave this "
         "sheet EMPTY to keep the pre-D70 behaviour exactly: with no actions the mod "
         "follows the drift path (M_0 -> M_1 -> M_2) and every number is unchanged."),
        fnt=font(FAIL_RED, bold=True, italic=True))

    header_row(ws, L.ML_HDR, 1, ml_headers(), widths=[9, 8, 11, 11, 12, 13, 34],
               fill=FILL_GREY, fnt=font(GREY_DARK, bold=True, size=9))
    ws.row_dimensions[L.ML_HDR].height = 30
    header_row(ws, L.ML_HDR, 9, ML_HELPER_HEADERS,
               widths=[14, 9, 9, 11, 12, 10, 9, 7, 11, 6], fill=FILL_GREY,
               fnt=font(GREY_DARK, bold=True, size=9))
    put(ws, f"I{L.ML_HDR - 1}", "engine helpers (formulas — do not edit)", fnt=F_SMALL_IT)
    header_row(ws, L.ML_HDR, 20,
               [f"% of CY {p} this action earns (selected combo)", "Warnings"],
               widths=[16, 30], fill=FILL_GREY,
               fnt=font(GREY_DARK, bold=True, size=9))
    put(ws, f"T{L.ML_HDR - 1}", "live results (formulas — do not edit)", fnt=F_SMALL_IT)
    ws.column_dimensions["H"].width = 2
    ws.column_dimensions["S"].width = 2

    # share-of-year machinery, mirroring the Rate Log's column X: exact, via
    # the SELECTED combo's Rate Engine block (other combos' rows show blank)
    wrng = f"{RE}!$H${L.RE_COH_FIRST}:$H${L.RE_COH_LAST}"
    ecp = f"{RE}!$Q${L.RE_COH_FIRST}:$Q${L.RE_COH_LAST}"
    absmi = f"{RE}!$G${L.RE_COH_FIRST}:$G${L.RE_COH_LAST}"

    fr, lr_ = L.ML_FIRST, L.ML_LAST
    for i in range(L.ML_ROWS):
        r = fr + i
        row = ctx.mod_rows[i] if i < len(ctx.mod_rows) else None
        input_cell(ws, f"A{r}", row and row["bu"])
        input_cell(ws, f"B{r}", row and row["state"])
        input_cell(ws, f"C{r}", row and row["eff"], fmt=FMT_DATE)
        input_cell(ws, f"D{r}", row and row["chg"], fmt=PCT_S)
        input_cell(ws, f"E{r}", row and row["status"])
        input_cell(ws, f"F{r}", row and row["achievement"], fmt=FMT_PCT, required=False)
        input_cell(ws, f"G{r}", (row["comment"] if row else None), required=False)
        ws[f"G{r}"].alignment = ALIGN_L
        # helpers I:R — the same shapes the rate log uses, on the mod columns
        formula(ws, f"I{r}", f'=IF(OR($A{r}="",$B{r}=""),"",$A{r}&"|"&$B{r})',
                border=BORDER_THIN)
        formula(ws, f"J{r}",
                f'=IF($C{r}="","",$D{r}*IF($E{r}="planned",IF($F{r}="",1,$F{r}),1))',
                fmt=PCT_S)
        formula(ws, f"K{r}", f'=IF($C{r}="",0,IF(1+$J{r}>0,LN(1+$J{r}),0))', fmt=FMT_IDX)
        formula(ws, f"L{r}", f'=IF($C{r}="","",DATE(YEAR($C{r}),MONTH($C{r}),1))',
                fmt=FMT_DATE)
        formula(ws, f"M{r}", f'=IF($C{r}="",0,EOMONTH($C{r},0)-$C{r}+1)', fmt=FMT_INT)
        formula(ws, f"N{r}",
                f'=IF($C{r}="",0,COUNTIFS(ml_key,$I{r},ml_effmonth,$L{r},ml_eff,"<"&$C{r}))',
                fmt=FMT_INT)
        formula(ws, f"O{r}",
                f'=IF($C{r}="",0,COUNTIFS($I${fr}:$I{r},$I{r},$L${fr}:$L{r},$L{r},'
                f"$C${fr}:$C{r},$C{r}))", fmt=FMT_INT)
        formula(ws, f"P{r}", f"=IF($C{r}=\"\",0,IF(AND($N{r}=0,$O{r}=1),1,0))", fmt=FMT_INT)
        formula(ws, f"Q{r}",
                f'=IF($C{r}="",0,IF(COUNTIFS(ml_key,$I{r},ml_effmonth,$L{r})>1,1,0))',
                fmt=FMT_INT)
        formula(ws, f"R{r}",
                f'=IF($C{r}="","",COUNTIFS(ml_key,$I{r},ml_eff,"<"&$C{r})'
                f"+COUNTIFS($I${fr}:$I{r},$I{r},$C${fr}:$C{r},$C{r}))", fmt=FMT_INT)
        # first action among the TAKEN rows of this key + cohort month. The
        # plain first-flag (P) is status-blind, so a planned action earlier in
        # the month would steal it and the mod solve's taken-only base would
        # silently drop the taken action from that cohort — D58, on the mod leg.
        formula(ws, f"V{r}",
                f'=IF(OR($C{r}="",$E{r}<>"taken"),0,'
                f'IF(AND(COUNTIFS(ml_key,$I{r},ml_effmonth,$L{r},ml_eff,"<"&$C{r},'
                f'ml_status,"taken")=0,'
                f"COUNTIFS($I${fr}:$I{r},$I{r},$L${fr}:$L{r},$L{r},"
                f'$C${fr}:$C{r},$C{r},$E${fr}:$E{r},"taken")=1),1,0))',
                fmt=FMT_INT)
        # D77: where this combo's year-end written mod lands if every action
        # in its key holds. Reads the engine's own M_endPrior cell rather than
        # re-deriving the blank-input fallback, so a feasibility warning can
        # never disagree with the path it is warning about.
        mend = (f"INDEX('_calc'!$Q$1:$Q${L.CALC_BLOCK_FIRST + L.LR_ROWS * L.CALC_BLOCK_STRIDE},"
                f"{L.CALC_BLOCK_FIRST}+(MATCH($I{r},lr_key,0)-1)*{L.CALC_BLOCK_STRIDE})")
        formula(ws, f"W{r}",
                f'=IF(OR($C{r}="",COUNTIF(lr_key,$I{r})=0),0,{mend}'
                f'*EXP(SUMIFS(ml_ln1p,ml_key,$I{r},ml_eff,"<="&DATE(nr_PlanYear,12,31))))',
                fmt=FMT_MOD)
        for cL in "IJKLMNOPQRVW":
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=9)
        # live feedback: what share of the plan year this action actually earns
        formula(ws, f"T{r}",
                f'=IF(OR($C{r}="",$I{r}<>nr_SelKey),"",'
                f"(SUMPRODUCT(({absmi}>YEAR($C{r})*12+MONTH($C{r})-1)*{wrng},{ecp})"
                f"+(EOMONTH($C{r},0)-$C{r}+1)/DAY(EOMONTH($C{r},0))"
                f"*SUMPRODUCT(({absmi}=YEAR($C{r})*12+MONTH($C{r})-1)*{wrng},{ecp}))"
                f"/SUMPRODUCT({wrng},{ecp}))", fmt="0%", align=ALIGN_C)
        formula(ws, f"U{r}",
                f'=IF($C{r}="","",TRIM('
                f'IF($C{r}<DATE(nr_PlanYear,1,1),"before 1/1 of the plan year — mod history '
                f'belongs in M_0 / M_endPrior ","")'
                f'&IF(AND($E{r}="taken",$F{r}<>"",$F{r}<>1),"achievement ignored on a taken '
                f'row — restate the change ","")'
                f'&IF(AND($I{r}<>"",COUNTIF(lr_key,$I{r})=0),"key not in tbl_LR ","")'
                f'&IF($Q{r}=1,"another action shares this month ","")'
                # D39 x D70: a net selection supersedes the mod leg from 1/1/P,
                # so the action shapes Net Delivery but not the plan LR — the
                # "why did nothing move?" question, answered before it is asked
                f'&IF($I{r}="","",IF(COUNTIF(lr_key,$I{r})=0,"",'
                f'IF(ISBLANK(INDEX(lr_netp,MATCH($I{r},lr_key,0))),"",'
                f'"combo is on a net rate selection — the net path supersedes the mod leg '
                f'(D39), so this shapes Net Delivery, not the plan LR ")))'
                f'))')
        for cL in "TU":
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=9)
        ws[f"U{r}"].font = font(FAIL_RED, size=9, italic=True)

    tbl = Table(displayName="tbl_ModLog", ref=f"A{L.ML_HDR}:G{L.ML_LAST}")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)
    ml_names = {
        "ml_bu": ("A", "tbl_ModLog business unit"),
        "ml_state": ("B", "tbl_ModLog state"),
        "ml_eff": ("C", "Mod action effective date (start of day)"),
        "ml_chg": ("D", "Directed mod change (decimal fraction)"),
        "ml_status": ("E", "taken | planned"),
        "ml_ach": ("F", "Achievement % (planned rows; blank = 100%)"),
        "ml_comment": ("G", "Free-text comment (the sample-data tripwire reads it)"),
        "ml_key": ("I", "Helper: BU|State key"),
        "ml_meff": ("J", "Helper: effective change = directed x achievement (planned only)"),
        "ml_ln1p": ("K", "Helper: ln(1 + effective mod change); 0 for blank rows"),
        "ml_effmonth": ("L", "Helper: first day of the effective month"),
        "ml_daysafter": ("M", "Helper: days in effective month on/after the action"),
        "ml_first": ("P", "Helper: 1 for the first action in its BU|State cohort month"),
        "ml_dup": ("Q", "Helper: 1 when >1 action shares a BU|State cohort month"),
        "ml_seq": ("R", "Helper: chronological rank of the action within its BU|State key"),
        "ml_firsttaken": ("V", "Helper: 1 for the first TAKEN action in its key + cohort "
                               "month (the mod solve's status-aware split flag, D58/D73)"),
    }
    for name, (colL, desc) in ml_names.items():
        ctx.define(name, SHEETS.MOD_LOG, f"${colL}${fr}:${colL}${lr_}", desc)

    _dv(ws, "list", [f"A{fr}:A{lr_}"], formula1="=lst_bu")
    _dv(ws, "list", [f"B{fr}:B{lr_}"], formula1="=lst_state")
    _dv(ws, "list", [f"E{fr}:E{lr_}"], formula1="=lst_status")
    _dv(ws, "decimal", [f"D{fr}:D{lr_}"], operator="between", formula1="-0.5", formula2="1",
        error="Mod change must lie in [-50%, +100%] (decimal fraction).")
    _dv(ws, "decimal", [f"F{fr}:F{lr_}"], operator="between", formula1="0", formula2="1.5",
        error="Achievement must lie in [0%, 150%].")
    _dv(ws, "date", [f"C{fr}:C{lr_}"], operator="between",
        formula1="DATE(nr_PlanYear,1,1)", formula2="DATE(2100,12,31)",
        error="Mod actions are dated inside the plan year or later — history is carried "
              "by M_0 and the projected current-year-end mod (D70).")

    put(ws, f"A{L.ML_LAST + 2}",
        "Achievement matters more here than on a filing: directed mod is rarely fully "
        "realised, so a planned action enters at directed x achievement. Column T shows "
        "what share of the plan year the action actually earns for the combo selected on "
        "Control — the cost of a late effective date, in the one number that makes it "
        "obvious.", fnt=F_SMALL_IT)
    put(ws, f"V{L.ML_HDR}", "First taken?", fnt=font(GREY_DARK, bold=True, size=9),
        fill=FILL_GREY)
    put(ws, f"W{L.ML_HDR}", "Yr-end mod if all hold",
        fnt=font(GREY_DARK, bold=True, size=9), fill=FILL_GREY)
    ctx.define("ml_endplan", SHEETS.MOD_LOG, f"$W${fr}:$W${lr_}",
               "Helper: the combo's written mod at 12/31/P with every logged action in "
               "force (0 on blank rows) — the feasibility check's subject (D77)")
    set_widths(ws, {"T": 16, "U": 30, "V": 11, "W": 12})
    for cL in ("V", "W"):
        ws.column_dimensions[cL].outlineLevel = 1
    # amber marks an ENTRY sheet, like Inputs and Rate Log — this is the third
    # paste target, and steel (the engine/read-only colour) hid that (D80)
    presentation_setup(ws, gridlines_off=False, freeze=f"C{L.ML_FIRST}",
                       tab_color=WARN_AMBER)
    print_setup(ws)
