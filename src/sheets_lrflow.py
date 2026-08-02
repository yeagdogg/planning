"""LR Flow: the plan loss ratio month by month — trend against earn-in (D93).

Every other exhibit in this workbook reports the plan year as ONE number. That
is the right answer for a plan, and it hides the thing a pricing actuary is
actually managing: the loss ratio does not sit at 63.4% all year. It starts
high, because January renews against a rate level that has barely earned in,
and it falls as rate and price earn through — while trend pushes the other way
the whole time. The plan loss ratio is the average of a race.

This tab runs the race. Each month is the SAME bridge the headline uses,
evaluated at that month rather than averaged over the year::

    LR(m) = LR_current x T^((mm - 6.5)/12) x (CRL/E_m) x (M_ind/Mbar_m) x A_other

and the exhibit's one number is the BREAKEVEN TREND: the net trend at which
January and December land on the same loss ratio. Above it, trend is winning
and the year deteriorates as it runs; below it, earn-in is winning.

The weighted mean of the walk does not equal the CY headline, and this tab does
not force it to (D46: disclose the residual, never normalise it away). The gap
is a few basis points on an annual term and it decomposes exactly — see the
reconciliation strip at the bottom, and ``engine.lr_flow_by_month``.
"""

from __future__ import annotations

from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker

from .build_workbook import Ctx, Layout as L, SHEETS
from .engine import LR_FLOW_ANCHOR
from .xlstyle import (
    ALIGN_C, BORDER_THIN, FAIL_RED, FILL_NAVY, FILL_PANEL, FILL_STEEL,
    FMT_IDX, FMT_MOD, FMT_PCT, FMT_PTS_COL, F_LABEL, F_SMALL_IT, GREY_DARK,
    NAVY, STEEL, chart_legend, col, dv_decimal, font, formula, header_row,
    input_cell, jump, label, link, nav_bar, presentation_setup, print_setup,
    put, quote_sheet, section, set_widths, title, zoom_axis,
)

RE = quote_sheet(SHEETS.RATE_ENGINE)

# geometry (module-level so the harness can address the exhibit)
LF_HDR = 17                      # month-table header row
LF_FIRST = 18                    # Jan P
LF_MONTHS = 24                   # Jan P .. Dec P+1
LF_LAST = LF_FIRST + LF_MONTHS - 1
LF_REC = LF_LAST + 2             # reconciliation strip
# the Rate Engine's month columns start at Jan P-1, so the plan year is 12 in
LF_RE_OFFSET = 12


def _mcol(i: int) -> str:
    """Rate Engine column for the i-th month of the walk (0 = Jan P)."""
    return col(L.RE_MATRIX_COL + LF_RE_OFFSET + i)


def build_lr_flow(ctx: Ctx):
    ws = ctx.wb[SHEETS.LR_FLOW]
    p = ctx.p
    title(ws, "A1", f"LR Flow — the plan loss ratio month by month ({ctx.lob.name})",
          "The same bridge the headline uses, evaluated at each month instead of averaged "
          "over the year. Rate and price earn IN (the factors fall as the year runs); "
          "trend pushes the other way. Which one wins is a date.")
    nav_bar(ws, 3, 1, ["Control", "Bridge", "Walkthrough", "Flow Dashboard",
                       "Program Flow", "Checks"], step=2)

    label(ws, "A5", "Selected:")
    link(ws, "B5", "=nr_SelKey", bold=True)
    jump(ws, "D5", "Control!C7", "Change BU/state selection >")

    # ---- the race, in one sentence -----------------------------------------
    formula(ws, "A6",
            '=IF(NOT(nr_SelOK),"Select a BU and state that exist in tbl_LR.",'
            'IF(NOT(ISNUMBER(lrf_tstar)),'
            '"Rate and price do not move this year, so there is no race to run — '
            'the loss ratio walks with trend alone.",'
            '"At "&TEXT(nr_Trend,"0.0%")&" net trend the year "'
            '&IF(ABS(nr_Trend-lrf_tstar)<0.0005,"is FLAT: trend and earn-in cancel",'
            'IF(nr_Trend>lrf_tstar,"DETERIORATES from "&TEXT(lrf_jan,"0.0%")&" in January '
            'to "&TEXT(lrf_dec,"0.0%")&" in December — trend is winning",'
            '"IMPROVES from "&TEXT(lrf_jan,"0.0%")&" in January to "&TEXT(lrf_dec,"0.0%")'
            '&" in December — earn-in is winning"))'
            '&".  They break even at "&TEXT(lrf_tstar,"0.0%")&" trend."))')
    ws["A6"].font = font(NAVY, bold=True, size=12)
    for cc in range(1, 12):
        ws.cell(row=6, column=cc).fill = FILL_PANEL

    # ---- levers and the breakeven ------------------------------------------
    section(ws, 8, "A", "The race")
    rows = [
        ("Net trend in force (from tbl_LR, or the Control default)", "=nr_Trend",
         FMT_PCT, None,
         "Loss over premium, annual. This tab is the only place it touches the PLAN "
         "year — the headline applies it to next year only."),
        # C x D are the two EARN-IN factors (rate, mod); the trend factors at the
        # two ends are T^(-5.5/12) and T^(+5.5/12), so the whole solve collapses
        # to their ratio raised to 12/11. Reading E (the loss ratio) here instead
        # would fold the trend back in and solve for the wrong thing.
        ("Breakeven net trend — January and December land together",
         f'=IF(OR(NOT(nr_SelOK),$C{LF_FIRST}="",$C{LF_FIRST + 11}=""),"n/a",'
         f'IF(ABS($C{LF_FIRST}*$D{LF_FIRST}'
         f'/($C{LF_FIRST + 11}*$D{LF_FIRST + 11})-1)<0.0000000001,"n/a",'
         f'($C{LF_FIRST}*$D{LF_FIRST}'
         f'/($C{LF_FIRST + 11}*$D{LF_FIRST + 11}))^(12/11)-1))',
         FMT_PCT, "lrf_tstar",
         "Solved in closed form from January's and December's earn-in factors: the trend "
         "at which the two months' loss ratios are equal. Above it the year deteriorates "
         "as it runs; below it, it improves."),
        ("Headroom (trend in force vs breakeven)",
         f'=IF(NOT(ISNUMBER(lrf_tstar)),"n/a",(nr_Trend-lrf_tstar)*100)',
         FMT_PTS_COL, None,
         "Negative = earn-in is ahead. Positive = trend is ahead."),
        ("Target loss ratio to plot (optional)", None, FMT_PCT, "lrf_target",
         "Draws a reference line on the chart and fills the 'vs target' column. Blank = "
         "no line."),
    ]
    for i, (lbl, f, fmt, name, note_) in enumerate(rows):
        r = 9 + i
        label(ws, f"A{r}", lbl, bold=(name == "lrf_tstar"))
        if f is None:
            input_cell(ws, f"C{r}", None, fmt=fmt, required=False)
        else:
            formula(ws, f"C{r}", f, fmt=fmt, align=ALIGN_C, border=BORDER_THIN,
                    bold=(name == "lrf_tstar"),
                    fill=FILL_PANEL if name == "lrf_tstar" else None)
        if name:
            ctx.define(name, SHEETS.LR_FLOW, f"$C${r}", lbl)
        put(ws, f"E{r}", note_, fnt=F_SMALL_IT)
    dv_decimal(ws, [f"C{9 + 3}"], 0, 3,
               "Enter the target as a decimal fraction — 0.65 for a 65% loss ratio.")

    put(ws, "A13",
        f"Trend is anchored at 7/1 of the plan year (month {LR_FLOW_ANCHOR} of 12), which is "
        "the only anchor at which the plan-year walk averages to the headline AND next "
        "year's walk averages to the headline's single trend step. The mod leg is "
        "suspended under a net selection — the net path already combines rate and price.",
        fnt=F_SMALL_IT)
    formula(ws, "A14",
            '=IF(nr_NetMode,"NET SELECTION ACTIVE: the rate column carries the combined '
            'rate x price path and the mod column is pinned at 1.000 (D39).",'
            'IF(nr_ModAdjEff<>"ON","MOD ADJUSTMENT OFF: the mod column is pinned at 1.000 '
            '— the walk is rate and trend only.",""))')
    ws["A14"].font = font(FAIL_RED, size=9, italic=True)

    # ---- the walk -----------------------------------------------------------
    section(ws, 16, "A", "Month by month — the plan year and the one after it")
    header_row(ws, LF_HDR, 1,
               ["Month", "Trend factor", "Rate earn-in (A_rate)",
                "Mod earn-in (A_mod)", "Plan loss ratio", "vs target (pts)",
                "Earned exposure", "Premium weight"],
               widths=[11, 11, 13, 12, 12, 11, 12, 12])
    header_row(ws, LF_HDR, 9,
               ["E_m", "Mbar_m", "price", "delta", "LR x weight", "target"],
               widths=[10, 10, 9, 9, 12, 10], fill=FILL_STEEL,
               fnt=font("FFFFFF", bold=True, size=8))
    ws.row_dimensions[LF_HDR].height = 40

    mod_on = 'AND(nr_ModAdjEff="ON",NOT(nr_NetMode))'
    for i in range(LF_MONTHS):
        r = LF_FIRST + i
        mm = i + 1                                   # month number, Jan P = 1
        mc = _mcol(i)
        plan_year_row = i < 12
        band = None if plan_year_row else FILL_PANEL
        yexpr = "nr_PlanYear" if plan_year_row else "(nr_PlanYear+1)"
        formula(ws, f"A{r}", f'=TEXT(DATE({yexpr},{(i % 12) + 1},1),"mmm yyyy")',
                align=ALIGN_C, fill=band, border=BORDER_THIN)
        # the three engine reads for this month, kept together so a column shift
        # in the Rate Engine moves one line rather than five
        link(ws, f"I{r}", f"={RE}!${mc}${L.RE_ROW_EM}", fmt=FMT_IDX, fill=band)
        link(ws, f"J{r}", f"={RE}!${mc}${L.RE_ROW_MODM}", fmt=FMT_MOD, fill=band)
        link(ws, f"G{r}", f"={RE}!${mc}${L.RE_ROW_DEN}", fmt=FMT_IDX,
             align=ALIGN_C, fill=band, border=BORDER_THIN)
        dead = f'$G{r}=0'
        formula(ws, f"K{r}", f'=IF({dead},1,IF({mod_on},$J{r}/nr_MInd,1))', fmt=FMT_IDX,
                fill=band)
        formula(ws, f"L{r}", f"=({mm}-{LR_FLOW_ANCHOR})/12", fmt="0.0000", fill=band)
        formula(ws, f"B{r}", f"=(1+nr_Trend)^$L{r}", fmt=FMT_IDX, align=ALIGN_C,
                fill=band, border=BORDER_THIN)
        formula(ws, f"C{r}", f'=IF({dead},"",nr_CRLind/$I{r})', fmt=FMT_IDX,
                align=ALIGN_C, fill=band, border=BORDER_THIN)
        formula(ws, f"D{r}", f'=IF({dead},"",IF({mod_on},nr_MInd/$J{r},1))', fmt=FMT_IDX,
                align=ALIGN_C, fill=band, border=BORDER_THIN)
        # a month that earns nothing has no loss ratio — a gap, never a zero
        formula(ws, f"E{r}",
                f'=IF({dead},"—",nr_LRcur*$B{r}*$C{r}*$D{r}*nr_AOther)',
                fmt=FMT_PCT, align=ALIGN_C, fill=band, border=BORDER_THIN, bold=True)
        formula(ws, f"F{r}",
                f'=IF(OR({dead},N(lrf_target)=0),"",($E{r}-lrf_target)*100)',
                fmt=FMT_PTS_COL, align=ALIGN_C, fill=band, border=BORDER_THIN)
        formula(ws, f"H{r}", f'=IF({dead},0,$G{r}*$I{r}*$K{r})', fmt=FMT_IDX,
                align=ALIGN_C, fill=band, border=BORDER_THIN)
        formula(ws, f"M{r}", f'=IF({dead},0,$H{r}*$E{r})', fmt=FMT_IDX, fill=band)
        # the target as a flat series the chart can plot beside the walk
        formula(ws, f"N{r}", '=IF(N(lrf_target)=0,"",lrf_target)', fmt=FMT_PCT,
                fill=band)
        for cL in "IJKLMN":
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=8)

    # the two months the race is read between, named so the banner can say them
    ctx.define("lrf_jan", SHEETS.LR_FLOW, f"$E${LF_FIRST}",
               "LR Flow: the plan year's January loss ratio")
    ctx.define("lrf_dec", SHEETS.LR_FLOW, f"$E${LF_FIRST + 11}",
               "LR Flow: the plan year's December loss ratio")
    ctx.define("lrf_lr", SHEETS.LR_FLOW, f"$E${LF_FIRST}:$E${LF_LAST}",
               "LR Flow: the 24-month plan loss ratio walk (Jan P .. Dec P+1)")
    ctx.define("lrf_w", SHEETS.LR_FLOW, f"$H${LF_FIRST}:$H${LF_LAST}",
               "LR Flow: the monthly premium weights")
    ctx.define("lrf_delta_p", SHEETS.LR_FLOW, f"$L${LF_FIRST}:$L${LF_FIRST + 11}",
               "LR Flow: plan-year trend exponents (must sum to 0 — the 7/1 anchor)")
    ctx.define("lrf_delta_p1", SHEETS.LR_FLOW, f"$L${LF_FIRST + 12}:$L${LF_LAST}",
               "LR Flow: P+1 trend exponents (must sum to 12 — one whole trend step)")

    # ---- reconciliation: what the mean leaves on the table ------------------
    p12, p12e = LF_FIRST, LF_FIRST + 11               # plan year rows
    n12, n12e = LF_FIRST + 12, LF_LAST                # the year after
    section(ws, LF_REC, "A",
            "Reconciliation — the walk against the headline, and every basis point of "
            "the difference")
    rec = [
        ('="Premium-weighted mean, CY "&nr_PlanYear',
         f"=IF(SUM($H${p12}:$H${p12e})=0,0,"
         f"SUM($M${p12}:$M${p12e})/SUM($H${p12}:$H${p12e}))", FMT_PCT, "lrf_mean_p",
         "Weighted by earned exposure x the price in force — premium, not exposure. "
         "Exposure weights alone miss by five times as much."),
        ('="CY "&nr_PlanYear&" plan loss ratio (the headline)"', "=nr_CYLR_P",
         FMT_PCT, None, "What every other exhibit reports."),
        ("Residual (mean vs headline)",
         f'=IF(nr_CYLR_P=0,0,(lrf_mean_p/nr_CYLR_P-1)*10000)', "0.00\" bp\"",
         "lrf_resid_p",
         "Disclosed, not normalised (D46). It decomposes exactly into the three lines "
         "below — the identity is mean = headline x E[T^delta] / (1 + covariance)."),
        ("…of which: rate x price covariance",
         f"=IF(OR(SUM($G${p12}:$G${p12e})=0,SUMPRODUCT($G${p12}:$G${p12e},$I${p12}:$I${p12e})=0),0,"
         f"SUMPRODUCT($G${p12}:$G${p12e},$I${p12}:$I${p12e},$K${p12}:$K${p12e})"
         f"*SUM($G${p12}:$G${p12e})"
         f"/SUMPRODUCT($G${p12}:$G${p12e},$I${p12}:$I${p12e})"
         f"/SUMPRODUCT($G${p12}:$G${p12e},$K${p12}:$K${p12e})-1)*10000", "0.0000\" bp\"",
         "lrf_cov",
         "The headline divides by two SEPARATE averages (earned rate level, earned mod); "
         "the walk divides month by month by their PRODUCT. The gap is their covariance. "
         "Zero when there is no separate price leg."),
        ("…of which: trend centre-of-gravity tilt",
         f"=IF(SUM($G${p12}:$G${p12e})=0,0,"
         f"(1+nr_Trend)^(SUMPRODUCT($G${p12}:$G${p12e},$L${p12}:$L${p12e})"
         f"/SUM($G${p12}:$G${p12e}))-1)*10000", "0.0000\" bp\"", "lrf_tilt",
         "Zero for an annual term at any seasonality — earned exposure per month is "
         "uniform, so 7/1 IS the centre of gravity. A short term with lumpy writings "
         "moves it, and this is the exhibit saying so."),
        ("…of which: trend convexity (Jensen)",
         f"=IF(SUM($G${p12}:$G${p12e})=0,0,"
         f"(SUMPRODUCT($G${p12}:$G${p12e},$B${p12}:$B${p12e})/SUM($G${p12}:$G${p12e})"
         f"/(1+lrf_tilt/10000)-1)*10000)", "0.0000\" bp\"", "lrf_convex",
         "The average of a curve exceeds the curve of the average: E[T^d] > T^E[d]. "
         "Vanishes when trend is zero."),
        ('="Premium-weighted mean, CY "&(nr_PlanYear+1)',
         f"=IF(SUM($H${n12}:$H${n12e})=0,0,"
         f"SUM($M${n12}:$M${n12e})/SUM($H${n12}:$H${n12e}))", FMT_PCT, "lrf_mean_p1",
         "The following year is indicative — see the canonical caveat on the Bridge."),
        ('="CY "&(nr_PlanYear+1)&" plan loss ratio (the headline)"', "=nr_CYLR_P1",
         FMT_PCT, None, ""),
    ]
    for i, (lbl, f, fmt, name, note_) in enumerate(rec):
        r = LF_REC + 1 + i
        if lbl.startswith("="):
            formula(ws, f"A{r}", lbl)
            ws[f"A{r}"].font = F_LABEL
        else:
            label(ws, f"A{r}", lbl)
        bold = name in ("lrf_mean_p", "lrf_resid_p")
        formula(ws, f"C{r}", f, fmt=fmt, align=ALIGN_C, border=BORDER_THIN, bold=bold,
                fill=FILL_PANEL if bold else None)
        if name:
            ctx.define(name, SHEETS.LR_FLOW, f"$C${r}", lbl.strip('="'))
        if note_:
            put(ws, f"E{r}", note_, fnt=F_SMALL_IT)
    # the identity, as a residual of a residual: this is what makes the three
    # lines above an EXPLANATION rather than three plausible numbers
    r_id = LF_REC + 1 + len(rec)
    label(ws, f"A{r_id}", "Unexplained (identity check — must be zero)", bold=True)
    formula(ws, f"C{r_id}",
            "=IF(nr_CYLR_P=0,0,(lrf_mean_p-nr_CYLR_P*(1+lrf_tilt/10000)"
            "*(1+lrf_convex/10000)/(1+lrf_cov/10000))*10000)", '0.000000" bp"', align=ALIGN_C,
            border=BORDER_THIN, bold=True)
    ctx.define("lrf_ident", SHEETS.LR_FLOW, f"$C${r_id}",
               "LR Flow: mean minus its own decomposition (Checks-enforced zero)")
    put(ws, f"E{r_id}",
        "If this is ever non-zero the three lines above have stopped being an account "
        "of the residual and become three numbers that merely look like one.",
        fnt=F_SMALL_IT)

    # ---- chart --------------------------------------------------------------
    lc = LineChart()
    lc.add_data(Reference(ws, min_col=5, min_row=LF_HDR, max_row=LF_LAST),
                titles_from_data=True)
    lc.set_categories(Reference(ws, min_col=1, min_row=LF_FIRST, max_row=LF_LAST))
    s = lc.series[0]
    s.graphicalProperties.line.solidFill = NAVY
    s.graphicalProperties.line.width = int(2.5 * 12700)
    s.marker = Marker(symbol="none")
    s.smooth = False
    lc.y_axis.number_format = "0.0%"
    lc.y_axis.title = "Plan loss ratio"
    lc.x_axis.delete = False
    lc.y_axis.delete = False
    lc.title = "The race: plan loss ratio by month (plan year, then the year after)"
    lc.style = 2
    lc.height, lc.width = 9, 17
    lc.visible_cells_only = False
    # the target, flat and dashed, so you can read the month the walk crosses it
    lc.add_data(Reference(ws, min_col=14, min_row=LF_HDR, max_row=LF_LAST),
                titles_from_data=True)
    t = lc.series[1]
    t.graphicalProperties.line.solidFill = FAIL_RED
    t.graphicalProperties.line.width = int(1.25 * 12700)
    t.graphicalProperties.line.dashStyle = "dash"
    t.marker = Marker(symbol="none")
    t.smooth = False
    # framed on the walk itself, not on zero — this is a LEVELS chart and the
    # whole point is the slope (D91)
    lo, hi = ctx.lay_dyn["axis"]["lr"]
    zoom_axis(lc, lo, hi, step=0.05)
    chart_legend(lc)
    ws.add_chart(lc, "O5")

    for cL in "IJKLMN":
        ws.column_dimensions[cL].outlineLevel = 1
    set_widths(ws, {"A": 34, "B": 11, "C": 13, "D": 12, "E": 12, "F": 11,
                    "G": 12, "H": 12})
    presentation_setup(ws, gridlines_off=True, freeze=f"A{LF_FIRST}", tab_color=NAVY)
    print_setup(ws)
