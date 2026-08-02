"""Theme, number formats, and cell-writing helpers for the workbook generator.

Financial-model conventions (brief §8):
  blue font   = hardcoded inputs / levers
  black font  = formulas calculated on the same sheet
  green font  = links pulled from another sheet
  yellow fill = required input cells
"""

from __future__ import annotations

import math

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Palette (restrained: navy headers, steel accents, near-white panels)
# ---------------------------------------------------------------------------

NAVY = "1F3864"
STEEL = "8497B0"
STEEL_LIGHT = "D6DCE4"
PANEL = "F7F9FB"
PANEL_2 = "EDF1F5"
WHITE = "FFFFFF"
INPUT_BLUE = "0000FF"
LINK_GREEN = "008000"
BLACK = "000000"
GREY = "808080"
GREY_DARK = "595959"
YELLOW_REQ = "FFF7CC"
WARN_AMBER = "FFC000"
FAIL_RED = "C00000"
PASS_GREEN = "2E7D32"
RED_FILL = "F8CBCC"
GREEN_FILL = "D6E8D5"
AMBER_FILL = "FFE8A6"
GREY_FILL = "E7E6E6"
BAND_FILL = "DDEBF7"  # plan-year band / highlight
UP_BAR = "8497B0"  # waterfall increase
DOWN_BAR = "1F3864"  # waterfall decrease
TOTAL_BAR = "548235"  # waterfall endpoints

FONT_NAME = "Calibri"

# ---------------------------------------------------------------------------
# Number formats (brief §8)
# ---------------------------------------------------------------------------

FMT_PCT = "0.0%"
FMT_PCT2 = "0.00%"
FMT_PCT_SIGNED = "+0.0%;-0.0%;0.0%"   # a CHANGE in rate/price: the sign is the message
FMT_IDX = "0.0000"
FMT_FACTOR = "0.0000"
FMT_MOD = "0.000"
FMT_PTS = '0.00 "pts"'
FMT_PTS_SIGNED = '+0.00 "pts";-0.00 "pts";0.00 "pts"'
# ...and the same number where the COLUMN HEADER already says "(pts)", so the
# suffix would repeat itself on every row.
FMT_PTS_COL = "+0.00;-0.00;0.00"
FMT_INT = "0"
FMT_EP = "#,##0"
FMT_GEN = "General"

# One date vocabulary, three widths — anything else is an accident:
#   FMT_DATE     every date you TYPE, and every date in a column wide enough
#   FMT_DATE_S   dates inside the 10-wide exhibit grids, where the long form ###s
#   FMT_MONTH    month headers on cohort and flow grids
# FMT_DATETIME is the one deliberate oddity: the Rate Engine samples cohorts at
# mid-month, so its x-coordinate carries a time of day and must show it.
FMT_DATE = "mm/dd/yyyy"
FMT_DATE_S = "m/d/yy"
FMT_MONTH = "mmm-yy"
FMT_DATETIME = "m/d/yy h:mm"

# zero-hiding variants for lookup grids whose blank-row fallback is 0:
FMT_PCT_Z = '0.0%;-0.0%;""'
FMT_IDX_Z = '0.0000;-0.0000;""'
FMT_EP_Z = '#,##0;-#,##0;""'

# ---------------------------------------------------------------------------
# Fonts / fills / borders
# ---------------------------------------------------------------------------


def font(color=BLACK, bold=False, size=11, italic=False, name=FONT_NAME):
    return Font(name=name, color=color, bold=bold, size=size, italic=italic)


F_LABEL = font(GREY_DARK)
F_LABEL_B = font(GREY_DARK, bold=True)
F_INPUT = font(INPUT_BLUE)
F_FORMULA = font(BLACK)
F_LINK = font(LINK_GREEN)
F_HEADER = font(WHITE, bold=True)
F_TITLE = font(NAVY, bold=True, size=16)
F_SUB = font(GREY, size=10, italic=True)
F_KPI = font(NAVY, bold=True, size=18)
F_SMALL = font(GREY_DARK, size=9)
F_SMALL_IT = font(GREY_DARK, size=9, italic=True)

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_STEEL = PatternFill("solid", fgColor=STEEL)
FILL_STEEL_LIGHT = PatternFill("solid", fgColor=STEEL_LIGHT)
FILL_PANEL = PatternFill("solid", fgColor=PANEL)
FILL_PANEL_2 = PatternFill("solid", fgColor=PANEL_2)
FILL_REQ = PatternFill("solid", fgColor=YELLOW_REQ)
FILL_GREY = PatternFill("solid", fgColor=GREY_FILL)
FILL_BAND = PatternFill("solid", fgColor=BAND_FILL)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_AMBER = PatternFill("solid", fgColor=AMBER_FILL)

_thin = Side(style="thin", color=STEEL_LIGHT)
_med = Side(style="medium", color=STEEL)
_hair = Side(style="hair", color=STEEL_LIGHT)
BORDER_THIN = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_BOTTOM = Border(bottom=_med)
BORDER_TOP = Border(top=_med)
BORDER_UNDER_HDR = Border(bottom=Side(style="medium", color=NAVY))

ALIGN_L = Alignment(horizontal="left", vertical="center")
ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_R = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_TITLE = Alignment(horizontal="left", vertical="center")
ALIGN_CC = Alignment(horizontal="centerContinuous", vertical="center")

# ---------------------------------------------------------------------------
# Cell writers
# ---------------------------------------------------------------------------


def put(ws, addr, value, fnt=None, fmt=None, fill=None, align=None, border=None):
    """Low-level cell writer; returns the cell."""
    c = ws[addr]
    c.value = value
    if fnt is not None:
        c.font = fnt
    if fmt is not None:
        c.number_format = fmt
    if fill is not None:
        c.fill = fill
    if align is not None:
        c.alignment = align
    if border is not None:
        c.border = border
    return c


def label(ws, addr, text, bold=False, **kw):
    return put(ws, addr, text, fnt=F_LABEL_B if bold else F_LABEL, **kw)


def title(ws, addr, text, sub=None):
    put(ws, addr, text, fnt=F_TITLE, align=ALIGN_TITLE)
    if sub:
        row = int("".join(ch for ch in addr if ch.isdigit())) + 1
        col = "".join(ch for ch in addr if ch.isalpha())
        put(ws, f"{col}{row}", sub, fnt=F_SUB)


def input_cell(ws, addr, value, fmt=FMT_GEN, required=True, border=BORDER_THIN):
    """Blue font; yellow fill when required."""
    return put(
        ws, addr, value, fnt=F_INPUT, fmt=fmt,
        fill=FILL_REQ if required else None, border=border,
    )


def formula(ws, addr, f, fmt=FMT_GEN, fill=None, border=None, align=None, bold=False):
    """Black same-sheet formula."""
    fnt = font(BLACK, bold=bold)
    return put(ws, addr, f, fnt=fnt, fmt=fmt, fill=fill, border=border, align=align)


def link(ws, addr, f, fmt=FMT_GEN, fill=None, border=None, align=None, bold=False):
    """Green cross-sheet link formula."""
    fnt = font(LINK_GREEN, bold=bold)
    return put(ws, addr, f, fnt=fnt, fmt=fmt, fill=fill, border=border, align=align)


def header_row(ws, row, first_col, labels, widths=None, fill=FILL_NAVY, fnt=None):
    """Write a navy header band across columns."""
    fnt = fnt or F_HEADER
    for i, text in enumerate(labels):
        col = first_col + i
        c = ws.cell(row=row, column=col, value=text)
        c.font = fnt
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if widths and widths[i]:
            ws.column_dimensions[get_column_letter(col)].width = widths[i]


def section(ws, row, col_letter, text, span_note=None):
    """Bold navy section caption with a medium underline."""
    c = put(ws, f"{col_letter}{row}", text, fnt=font(NAVY, bold=True, size=12))
    c.border = BORDER_UNDER_HDR
    if span_note:
        put(ws, f"{col_letter}{row + 1}", span_note, fnt=F_SUB)
    return c


def note(ws, addr, text, italic=True):
    return put(ws, addr, text, fnt=F_SMALL_IT if italic else F_SMALL, align=ALIGN_WRAP)


def add_dv(ws, kind, ranges, blocking=True, **kw):
    """Attach one data validation to a set of ranges.

    Every typed cell in the workbook should carry a bound. The point is not to
    stop a determined user — DV is defeated by a paste, which is how the inputs
    arrive — but to catch the fat-finger that a formula would otherwise swallow
    silently: a target loss ratio of 65 instead of 0.65 returns a plausible
    number everywhere downstream, and nothing on the sheet says it is wrong.

    ``blocking=False`` makes the list a suggestion rather than a rule, for the
    places where typing a NEW value is legitimate (tbl_LR defines the roster).
    """
    dv = DataValidation(type=kind, allow_blank=True, showErrorMessage=blocking, **kw)
    ws.add_data_validation(dv)
    for r in ranges:
        dv.add(r)
    return dv


def dv_decimal(ws, ranges, lo, hi, error, blocking=True):
    """A numeric bound on typed cells (the commonest case)."""
    return add_dv(ws, "decimal", ranges, blocking=blocking, operator="between",
                  formula1=str(lo), formula2=str(hi), error=error)


def dv_plan_year_date(ws, ranges, error, back_years=0):
    """A date bound anchored on ``nr_PlanYear``, NOT on the build-time year.

    The plan year is an INPUT (Control!C6), and the sheets say so — "all
    calculations follow the plan year entered above". A validation baked with
    the year the file was generated quietly stops agreeing with that the moment
    anyone rolls the year forward, and rejects dates the engine accepts.
    """
    lo = f"DATE(nr_PlanYear-{back_years},1,1)" if back_years else "DATE(nr_PlanYear,1,1)"
    return add_dv(ws, "date", ranges, operator="between",
                  formula1=lo, formula2="DATE(nr_PlanYear,12,31)", error=error)


def text_height(text: str, width: float, size: int = 10) -> float:
    """Row height (pt) that shows the whole wrapped ``text`` in a column of the
    given Excel width. Excel skips auto-fit once a height is set explicitly, so
    an under-estimate clips the text invisibly — calibration errs tall.
    Excel caps row height at 409.5pt; prose needing more must move to a wider
    column, which the min() here makes a visible (not silent) failure."""
    chars_per_line = max(6.0, width * (11.0 / max(size, 6)) * 0.85)
    lines = sum(max(1, math.ceil(len(para) / chars_per_line))
                for para in str(text).split("\n"))
    return min(405.0, max(15.0, lines * (size + 4.0) + 6.0))


def prose(ws, addr, text, size=10, bold=False, color=None, width=None):
    """Paragraph text sized to its HOST column so nothing clips (vs note(),
    which leaves height to Excel's 409.5pt-capped auto-fit). Pass ``width``
    when the host column's width is set after the call."""
    col_letter = "".join(ch for ch in addr if ch.isalpha())
    row = int("".join(ch for ch in addr if ch.isdigit()))
    w = width or ws.column_dimensions[col_letter].width or 8.43
    c = put(ws, addr, text, fnt=font(color or GREY_DARK, bold=bold, size=size),
            align=ALIGN_WRAP)
    ws.row_dimensions[row].height = text_height(text, w, size=size)
    return c


def jump(ws, addr, target: str, text: str, size=10, bold=False):
    """In-workbook navigation link via the classic HYPERLINK function.

    ``target`` example: "'Rate Engine'!A1" or "Control!C7".
    """
    c = put(ws, addr, f'=HYPERLINK("#{target}","{text}")')
    c.font = Font(name=FONT_NAME, color=NAVY, size=size, bold=bold, underline="single")
    return c


def nav_bar(ws, row: int, first_col: int, sheets, label_text="Go to:", size=9, step=1):
    """One-row navigation bar: a small label followed by one jump per sheet.
    ``step=2`` skips every other column so link text can overflow on
    narrow-column sheets instead of clipping against its neighbor."""
    if label_text:
        put(ws, f"{get_column_letter(first_col)}{row}", label_text, fnt=F_SMALL)
        first_col += step
    for i, sheet in enumerate(sheets):
        jump(ws, f"{get_column_letter(first_col + i * step)}{row}",
             f"{quote_sheet(sheet)}!A1", sheet, size=size)


def col(idx: int) -> str:
    return get_column_letter(idx)


def rng(c1: int, r1: int, c2: int, r2: int, sheet: str | None = None, absolute=True) -> str:
    """A1-style range string, optionally sheet-qualified (quoted if needed)."""
    d = "$" if absolute else ""
    body = f"{d}{col(c1)}{d}{r1}:{d}{col(c2)}{d}{r2}"
    if sheet:
        return f"{quote_sheet(sheet)}!{body}"
    return body


def quote_sheet(name: str) -> str:
    return f"'{name}'" if (" " in name or "-" in name) else name


def set_widths(ws, spec: dict):
    """spec: {'A': 12, 'B': 9, ...} or {1: 12, ...}"""
    for k, w in spec.items():
        letter = k if isinstance(k, str) else get_column_letter(k)
        ws.column_dimensions[letter].width = w


STATUS_PASS = "ALL CHECKS PASS"


def status_banner_cf(ws, addr: str, size: int = 11):
    """Green / amber / red conditional formatting for a checks-status cell (D80).

    Three states, because a two-state banner had to call warnings "pass": the
    old rules were equal / notEqual against the pass string, so any number of
    WARNs stayed green at the one place a reader actually looks. The banner now
    reads "PASS WITH n WARNING(S)" in that state and this paints it amber.

    Rules are mutually exclusive (no stopIfTrue needed) and reference the cell
    absolutely, so the same three lines work for a status cell on any sheet.
    """
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils.cell import coordinate_from_string

    c_letter, c_row = coordinate_from_string(addr.replace("$", ""))
    a = f"${c_letter}${c_row}"
    for rule_f, fill, fcolor in (
            (f'={a}="{STATUS_PASS}"', FILL_GREEN, PASS_GREEN),
            (f'=AND(LEFT({a},4)="PASS",{a}<>"{STATUS_PASS}")', FILL_AMBER, "7F6000"),
            (f'=LEFT({a},4)<>"PASS"', FILL_RED, FAIL_RED)):
        ws.conditional_formatting.add(
            addr, FormulaRule(formula=[rule_f], fill=fill,
                              font=font(fcolor, bold=True, size=size)))
    return ws


def chart_legend(c, pos="b"):
    """Move the legend OUT of the plot area (D69).

    openpyxl leaves ``overlay`` as None, so ``<c:overlay>`` is omitted — and
    ECMA-376's default for a missing overlay is TRUE. Excel therefore
    materialises ``<c:overlay val="1"/>`` on resave and draws the legend ON TOP
    of the series. Exactly the shape of D36 (the omitted ``<c:delete>`` on an
    axis): an optional element left out acquires a hostile default, and the
    damage only appears after the Excel round-trip.

    Bottom by default: every chart here is a month or cohort series, so a
    bottom legend costs height that is cheap and keeps the full plot width.
    Note ``Legend`` has no ``position`` alias — the attribute is ``legendPos``.
    """
    if getattr(c, "legend", None) is None:
        return c
    c.legend.legendPos = pos
    c.legend.overlay = False
    return c


def unoverlay_titles(wb):
    """Force every chart and axis TITLE to reserve its own space (D71).

    The same omitted-element trap as the legend one line above, one level
    deeper: ``Title`` also carries a ``<c:overlay>``, openpyxl leaves it None,
    and Excel materialises ``val="1"`` on resave — so axis titles are drawn ON
    TOP of the plot area and read as unreadable text sitting behind the bars.

    Swept over the whole workbook at save time rather than fixed at each of the
    sixteen ``add_chart`` sites: this class of bug is one a new chart silently
    reintroduces, and a sweep cannot be forgotten. ``verify_workbook`` phase A
    holds the line by requiring ``<c:overlay val="0"/>`` inside every title.
    """
    n = 0
    for ws in wb.worksheets:
        for ch in getattr(ws, "_charts", []):
            # a combined chart (c1 += c2) keeps its parts, and each one can
            # carry its own axes — walk the whole family, not just the head
            parts = [ch] + list(getattr(ch, "_charts", []) or [])
            for part in parts:
                for holder in (part, getattr(part, "x_axis", None),
                               getattr(part, "y_axis", None),
                               getattr(part, "z_axis", None)):
                    t = getattr(holder, "title", None)
                    if t is not None and hasattr(t, "overlay"):
                        t.overlay = False
                        n += 1
    return n


def assert_formulas_balanced(wb):
    """Fail the BUILD on a formula whose parentheses don't close (D89).

    Excel does not report an unbalanced formula as an unbalanced formula. It
    refuses to open the file at all, and the failure surfaces several minutes
    later, from the recalculation step, as::

        Open method of Workbooks class failed

    — which names neither the sheet, the cell, nor the defect. These formulas
    are assembled from nested f-strings, so a stray closing paren is an ordinary
    typo; catching it here costs a fraction of a second and points at the cell.

    Quotes are tracked because Excel string literals legitimately contain
    parentheses (every "(pts)" caption, and every TEXT() format mask).
    """
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                depth = 0
                in_str = False
                for ch in v:
                    if ch == '"':
                        in_str = not in_str
                    elif not in_str:
                        depth += (ch == "(") - (ch == ")")
                        if depth < 0:
                            break
                if depth or in_str:
                    bad.append(f"  {ws.title}!{c.coordinate}: "
                               f"{'unterminated string' if in_str else f'depth {depth:+d}'}"
                               f"\n    {v[:160]}")
    if bad:
        raise ValueError(
            f"{len(bad)} malformed formula(s) — Excel would refuse to open this "
            f"file:\n" + "\n".join(bad))


def presentation_setup(ws, gridlines_off=True, zoom=100, freeze=None, tab_color=None):
    ws.sheet_view.showGridLines = not gridlines_off
    ws.sheet_view.zoomScale = zoom
    if freeze:
        ws.freeze_panes = freeze
    if tab_color:
        ws.sheet_properties.tabColor = tab_color


def print_setup(ws, landscape=True, title_text="CY Plan Loss Ratio Workbook"):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddHeader.center.text = title_text
    ws.oddHeader.center.font = f"{FONT_NAME},Bold"
    ws.oddHeader.center.size = 10
    ws.oddFooter.left.text = "&A"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.left.size = 9
    ws.oddFooter.right.size = 9
