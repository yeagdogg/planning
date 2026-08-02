"""Net Delivery sheet + hidden _netcalc engine blocks (DECISIONS.md D57).

Answers the PM question for combos on a net rate selection: given each
state's rate-change effective date and the current flow, how is the target
delivered month by month, and what filed change + pricing (mod) walk
satisfies it?

Math (oracle: engine.net_delivery_components / suggest_net_rate /
required_m1 / required_mod_step / net_program_plan_lr — every displayed
figure ties at 1e-9):
  delivered(m) = [W(m)/W(m-12)] x [M(m)/M(m-12)]
The rate leg's base is the LIVE rows (taken + planned before 1/1/P — D39,
not the Solver's taken-only D13). One new change r at the state's date D
enters affinely per month, W_new(m) = A(m) + B(m)*r, with the D31
first-in-month day-blend, so the suggested filed change and the required
year-end mod both solve in closed form.

The mod ask is published twice (D75). As a LEVEL — the year-end written mod
M_1' — which is the natural answer while the mod path is a drift line, and
goes n/a the moment the Mod Log pins the path. And as an ACTION — a dated
mod change at the filing's own date, grossed up by achievement — which
always answers, because the plan-year written mod is affine in (1 + c)
along the stepped path exactly as W_new is affine in r.

_netcalc geometry: a live-filtered log helper strip, then one 56-row-stride
block per roster state (21 + 3 spares) plus one band block for the tab's
local picker. Blocks go inert ("") for non-net combos, the 'All' view, and
absent states — this tab never fabricates.
"""

from __future__ import annotations

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation

from .build_workbook import Ctx, Layout as L, SHEETS, mod_adj_on
from .sheets_engine import (
    LogRefs, mod_drift_at_switch, mod_value_formula, write_cohort_block,
    write_mod_anchor_cells)
from .xlstyle import (
    ALIGN_C, ALIGN_L, BORDER_THIN, F_LABEL, F_SMALL_IT, FAIL_RED, FILL_GREY,
    FILL_NAVY, FILL_PANEL, FILL_STEEL, FMT_DATE, FMT_IDX, FMT_INT, FMT_MOD,
    FMT_PCT, GREY_DARK, NAVY,
    STEEL, WARN_AMBER, col, chart_legend, font, formula, header_row, input_cell, jump, label,
    link, nav_bar, presentation_setup, print_setup, prose, put, quote_sheet,
    section, set_widths, title,
)

ND = quote_sheet(SHEETS.NET_DELIVERY)
NC = SHEETS.NETCALC                     # "_netcalc" (no quoting needed)
RL = quote_sheet(SHEETS.RATE_LOG)
PCT_S = "+0.0%;-0.0%;0.0%"

# ---- _netcalc geometry (module-level so the harness can address blocks) ----
NDL_FIRST = 4                           # live-log strip first row
NC_STRIDE = 56
# D70 stepped-mod strip, parked right of everything the block already uses
# (W..Y, AB..AI on cohort rows; AK/AL on the band block) so nothing is
# overloaded and the whole mechanic reads as one contiguous appendix
NC_MOD_CNT_COL = 39                     # AM — mod actions logged for this combo
NC_MOD_BASE_COL = 40                    # AN — M_endPrior, what they compound on
NC_MOD_COL = 41                         # AO..AS — the five step-index helpers
# D75 mod-step solve: the mod ask as a dated ACTION rather than a level. Its
# columns sit past the D70 strip for the same reason that one sits past the
# rest — one column, one meaning, nothing overloaded.
NC_MSTEP_HDR_COL = 46                   # AT, AU — mod day-blend, header row only
NC_MSTEP_COL = 48                       # AV..AZ — M_w(m-12) stepped, A, B, h, diff
NC_MSTEP_ANCH_COL = 53                  # BA..BK — anchors with steps FORCED on

# ---- Net Delivery sheet geometry ----
# The deep-dive band leads the tab (D74): its chart is the one exhibit that
# makes the whole decomposition legible at a glance, so it opens the sheet
# instead of closing it. The band is FIXED height — it does not grow with the
# state roster — so the summary below it still starts at a constant row and
# the harness can keep importing ND_FIRST as a plain number.
ND_BAND_TOP = 8                         # "One combo under the microscope"
ND_HDR = 31                             # summary table header row
ND_FIRST = 32
# The band is hand-placed, so its height is hand-counted; build() asserts the
# two never collide rather than letting a longer band overwrite the summary.
GRID_GAP = 5                            # rows from the IN VIEW row to the first grid
                                        # section (the notes under the table live here)
ND_PRIOR_COL = 2                        # B  — Jan P-1 on the rate-leg grid (D68)
ND_PLAN_COL = 14                        # N  — Jan P on both grids
ND_FLAG_COL = 26                        # Z  — flags, parked right of both bands so
                                        #      the wide text is never clipped by a
                                        #      month column


def ndl_last():
    return NDL_FIRST + L.RL_ROWS - 1


def nc_block_first():
    return ndl_last() + 4


def block_top(i: int) -> int:
    """Top row of state block i (0-based roster order); the band block is
    index n_disp (one past the last roster slot)."""
    return nc_block_first() + i * NC_STRIDE


def n_disp(ctx: Ctx) -> int:
    return len(ctx.cfg.states) + 3


# ---------------------------------------------------------------------------
# _netcalc
# ---------------------------------------------------------------------------


def build_netcalc(ctx: Ctx):
    ws = ctx.wb[NC]
    put(ws, "A1",
        "_netcalc (hidden): net-delivery decomposition blocks (D57). Live-filtered log "
        "strip (taken + planned before 1/1/P), then one block per roster state plus the "
        "band block. Generated by src/sheets_netdelivery.py; do not edit.",
        fnt=font(GREY_DARK, bold=True, size=9))

    # ---- live-log helper strip -------------------------------------------
    for j, h in enumerate(["live?", "eff''", "effmo''", "ln''", "days''",
                           "ecnt", "scnt", "first''"]):
        put(ws, f"{col(1 + j)}{NDL_FIRST - 1}", h, fnt=font(GREY_DARK, size=8))
    r0 = NDL_FIRST
    rlast = ndl_last()
    for j in range(L.RL_ROWS):
        r = r0 + j
        ir = L.RL_FIRST + j
        formula(ws, f"A{r}",
                f"=IF({RL}!$C${ir}=\"\",0,IF(OR({RL}!$E${ir}=\"taken\","
                f"{RL}!$C${ir}<DATE(nr_PlanYear,1,1)),1,0))", fmt=FMT_INT)
        formula(ws, f"B{r}", f'=IF($A{r}=0,"",{RL}!$C${ir})', fmt="mm/dd/yyyy")
        formula(ws, f"C{r}", f'=IF($A{r}=0,"",DATE(YEAR($B{r}),MONTH($B{r}),1))',
                fmt="mm/dd/yyyy")
        formula(ws, f"D{r}", f"=IF($A{r}=0,0,{RL}!$L${ir})", fmt=FMT_IDX)
        formula(ws, f"E{r}", f"=IF($A{r}=0,0,{RL}!$N${ir})", fmt=FMT_INT)
        # first LIVE change in its key + cohort month (rl_first is
        # filter-unaware; recompute over the strip, scenario-log style)
        formula(ws, f"F{r}",
                f'=IF($A{r}=0,0,COUNTIFS(rl_key,{RL}!$J${ir},'
                f"$C${r0}:$C${rlast},$C{r},$B${r0}:$B${rlast},\"<\"&$B{r}))", fmt=FMT_INT)
        formula(ws, f"G{r}",
                f"=IF($A{r}=0,0,COUNTIFS({RL}!$J${L.RL_FIRST}:$J${ir},{RL}!$J${ir},"
                f"$C${r0}:$C{r},$C{r},$B${r0}:$B{r},$B{r}))", fmt=FMT_INT)
        formula(ws, f"H{r}", f"=IF($A{r}=0,0,IF(AND($F{r}=0,$G{r}=1),1,0))", fmt=FMT_INT)
        for cL in "ABCDEFGH":
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=8)
    strip = {
        "ndl_live": ("A", "1 for rows live under a net selection (D57)"),
        "ndl_eff": ("B", "Live-row effective date ('' when superseded)"),
        "ndl_effmonth": ("C", "Live-row effective month"),
        "ndl_ln": ("D", "Live-row ln(1 + effective change); 0 when superseded"),
        "ndl_days": ("E", "Live-row days on/after the change"),
        "ndl_first": ("H", "1 for the first LIVE change in its key + cohort month"),
    }
    for name, (cL, desc) in strip.items():
        ctx.define(name, NC, f"${cL}${r0}:${cL}${rlast}", desc)

    refs = None  # per-block LogRefs built below

    # ---- state blocks + the band block ------------------------------------
    nd = n_disp(ctx)
    for i in range(nd + 1):
        t = block_top(i)
        band = (i == nd)
        if band:
            put(ws, f"M{t - 1}", "band block (Net Delivery monthly deep-dive pickers)",
                fnt=font(GREY_DARK, size=8))
            formula(ws, f"A{t}", "=nd_BandState")
            formula(ws, f"B{t}",
                    '=IF(OR(nd_BandBU="",nd_BandState=""),"",nd_BandBU&"|"&nd_BandState)')
            # per-state date / filed override, looked up from the tab's rows.
            # Blank falls back to the PLANNED plan-year filing in the rate log
            # (W/X/Y below), and only then to the default date / suggestion (D63).
            nd_hit = "COUNTIF(ndd_states,nd_BandState)=0"
            eff_f = (f'=IF(OR($B{t}="",{nd_hit}),$X{t},'
                     f'IF(INDEX(ndd_dates,MATCH(nd_BandState,ndd_states,0))="",'
                     f"$X{t},INDEX(ndd_dates,MATCH(nd_BandState,ndd_states,0))))")
            ovr_f = (f'=IF(OR($B{t}="",{nd_hit}),$Y{t},'
                     f"IF(ISBLANK(INDEX(ndd_filed,MATCH(nd_BandState,ndd_states,0))),"
                     f"$Y{t},INDEX(ndd_filed,MATCH(nd_BandState,ndd_states,0))))")
        else:
            formula(ws, f"A{t}", f"=_lists!$B${3 + i}")
            formula(ws, f"B{t}", f'=IF(OR($A{t}="",nd_BU="All"),"",nd_BU&"|"&$A{t})')
            row_v = ND_FIRST + i
            eff_f = (f'=IF({ND}!$D${row_v}="",$X{t},{ND}!$D${row_v})')
            ovr_f = (f'=IF(ISBLANK({ND}!$E${row_v}),$Y{t},{ND}!$E${row_v})')
        # header cells: combo resolution + guarded inputs (the _calc pattern)
        formula(ws, f"C{t}", f'=IF($B{t}="",0,IF(COUNTIF(lr_key,$B{t})=0,0,'
                             f"MATCH($B{t},lr_key,0)))", fmt=FMT_INT)
        formula(ws, f"D{t}", f"=IF($C{t}=0,0,IF(ISBLANK(INDEX(lr_netp,$C{t})),0,1))",
                fmt=FMT_INT)
        formula(ws, f"E{t}", f"=IF($D{t}=0,0,N(INDEX(lr_netp,$C{t})))", fmt="0.0%")
        formula(ws, f"F{t}", f"=IF($C{t}=0,0,IF({mod_adj_on(f'$C{t}')},1,0))",
                fmt=FMT_INT)
        formula(ws, f"G{t}", f"=IF($C{t}=0,1,IF(N(INDEX(lr_mind,$C{t}))=0,1,"
                             f"INDEX(lr_mind,$C{t})))", fmt=FMT_MOD)
        formula(ws, f"H{t}", f"=IF($C{t}=0,1,IF(N(INDEX(lr_m0,$C{t}))=0,1,"
                             f"INDEX(lr_m0,$C{t})))", fmt=FMT_MOD)
        formula(ws, f"I{t}", f"=IF($C{t}=0,DATE(nr_PlanYear-1,10,1),"
                             f"IF(N(INDEX(lr_m0asof,$C{t}))=0,DATE(nr_PlanYear-1,10,1),"
                             f"INDEX(lr_m0asof,$C{t})))", fmt="mm/dd/yyyy")
        formula(ws, f"J{t}", f"=IF($C{t}=0,1,IF(N(INDEX(lr_m1,$C{t}))=0,1,"
                             f"INDEX(lr_m1,$C{t})))", fmt=FMT_MOD)
        formula(ws, f"K{t}", f"=IF($C{t}=0,0,N(INDEX(lr_mprior,$C{t})))", fmt=FMT_MOD)
        formula(ws, f"L{t}", f"=IF($C{t}=0,0,N(INDEX(lr_m2,$C{t})))", fmt=FMT_MOD)
        formula(ws, f"M{t}",
                f'=IF(OR($A{t}="",COUNTIF(se_state,$A{t})=0),0,MATCH($A{t},se_state,0))',
                fmt=FMT_INT)
        formula(ws, f"N{t}", f"=IF($M{t}=0,0,INDEX(se_sum,$M{t}))", fmt="0.00")
        # effective-date helpers (D31 blend at the earlier of D / first live);
        # O is written below, after the planned-filing pickup it falls back to
        formula(ws, f"P{t}", f"=IF(AND($O{t}>=DATE(nr_PlanYear,1,1),"
                             f"$O{t}<=DATE(nr_PlanYear,12,31)),1,0)", fmt=FMT_INT)
        formula(ws, f"Q{t}", f"=YEAR($O{t})*12+MONTH($O{t})-1", fmt=FMT_INT)
        formula(ws, f"R{t}", f"=DAY(EOMONTH($O{t},0))", fmt=FMT_INT)
        formula(ws, f"S{t}", f"=EOMONTH($O{t},0)-$O{t}+1", fmt=FMT_INT)
        formula(ws, f"T{t}", f"=SUMIFS(ndl_days,rl_key,$B{t},ndl_effmonth,"
                             f"DATE(YEAR($O{t}),MONTH($O{t}),1),ndl_first,1)", fmt=FMT_INT)
        formula(ws, f"U{t}", f"=IF($P{t}=1,MIN(1,MAX($S{t},$T{t})/$R{t}),1)", fmt="0.000")
        # the PLANNED plan-year filing already in the rate log (D63): count,
        # date, effective % (filed x achievement — what actually earns), and
        # the filed % for display. Written BEFORE V/O consume them by address.
        jan1, dec31 = "DATE(nr_PlanYear,1,1)", "DATE(nr_PlanYear,12,31)"
        formula(ws, f"W{t}",
                f'=IF($B{t}="",0,COUNTIFS(rl_key,$B{t},rl_status,"planned",'
                f'rl_eff,">="&{jan1},rl_eff,"<="&{dec31}))', fmt=FMT_INT)
        formula(ws, f"X{t}",
                f"=IF($W{t}=0,nd_DefaultDate,SUMIFS(rl_eff,rl_key,$B{t},rl_firstplanned,1))",
                fmt="mm/dd/yyyy")
        formula(ws, f"Y{t}",
                f'=IF($W{t}=0,"",SUMIFS(rl_reff,rl_key,$B{t},rl_firstplanned,1))',
                fmt="0.00%")
        formula(ws, f"Z{t}",
                f'=IF($W{t}=0,"",SUMIFS(rl_filed,rl_key,$B{t},rl_firstplanned,1))',
                fmt="0.00%")
        formula(ws, f"O{t}", eff_f, fmt="mm/dd/yyyy")
        formula(ws, f"V{t}", ovr_f, fmt="0.00%")
        # D70: how many mod actions this combo has logged, and the level they
        # compound on. The mod log is read WHOLE here, unlike the rate log: a
        # net selection supersedes planned FILINGS, but a logged mod action is
        # the pricing plan this tab exists to measure the target against.
        cnt_ref = f"${col(NC_MOD_CNT_COL)}${t}"
        base_ref = f"${col(NC_MOD_BASE_COL)}${t}"
        formula(ws, cnt_ref.replace("$", ""),
                f'=IF($B{t}="",0,COUNTIF(ml_key,$B{t}))', fmt=FMT_INT)
        drift_sw = mod_drift_at_switch(f"$H{t}", f"$I{t}", f"$J{t}", f"$K{t}")
        formula(ws, base_ref.replace("$", ""),
                f"=IF($C{t}=0,$H{t},IF(N(INDEX(lr_mendprior,$C{t}))=0,{drift_sw},"
                f"INDEX(lr_mendprior,$C{t})))", fmt=FMT_MOD)
        # D75: the day-blend for a mod step dated at the FILING's own date —
        # one filing normally carries both levers, so the tab does not ask for
        # a second date. The split lands at the earlier of D and the first
        # logged mod action in D's month, exactly as $T/$U do for the rate leg.
        mdays_ref = f"${col(NC_MSTEP_HDR_COL)}${t}"
        pmod_ref = f"${col(NC_MSTEP_HDR_COL + 1)}${t}"
        formula(ws, mdays_ref.replace("$", ""),
                f'=IF($B{t}="",0,SUMIFS(ml_daysafter,ml_key,$B{t},ml_effmonth,'
                f"DATE(YEAR($O{t}),MONTH($O{t}),1),ml_first,1))", fmt=FMT_INT)
        formula(ws, pmod_ref.replace("$", ""),
                f"=IF($P{t}=1,MIN(1,MAX($S{t},{mdays_ref})/$R{t}),1)", fmt="0.000")
        for cc in list("ABCDEFGHIJKLMNOPQRSTUVWXYZ") + [
                col(NC_MOD_CNT_COL), col(NC_MOD_BASE_COL),
                col(NC_MSTEP_HDR_COL), col(NC_MSTEP_HDR_COL + 1)]:
            ws[f"{cc}{t}"].font = font(GREY_DARK, size=8)

        anchors = write_mod_anchor_cells(
            ws, ctx, t + 1, 1, mind_ref=f"$G${t}", m0_ref=f"$H${t}",
            asof_ref=f"$I${t}", m1_ref=f"$J${t}", mprior_ref=f"$K${t}",
            m2_ref=f"$L${t}", steps_cond=f"{cnt_ref}>0", mend_ref=base_ref)
        # ... and the same path with steps FORCED on. Asking for a mod step
        # commits the combo to the stepped regime whether or not it is already
        # in it, so the D70 re-anchor (history runs to M_endPrior at
        # 12/31/(P-1), the log owns the plan year) is in force for the solve
        # even when the log is empty. Solving on the c = 0 member of that same
        # family is what makes the closed form exact rather than nearly right.
        anchors_s = write_mod_anchor_cells(
            ws, ctx, t + 1, NC_MSTEP_ANCH_COL, mind_ref=f"$G${t}",
            m0_ref=f"$H${t}", asof_ref=f"$I${t}", m1_ref=base_ref,
            mprior_ref=f"$K${t}", m2_ref=f"$L${t}", steps_cond="TRUE",
            mend_ref=base_ref)
        refs = LogRefs(key_cond=f"$B${t}", eff="ndl_eff", ln="ndl_ln",
                       effmonth="ndl_effmonth", first="ndl_first",
                       daysafter="ndl_days")
        write_cohort_block(ws, ctx, t + 3, refs, t_ref="nr_TermMonths",
                           srow_ref=f"$M${t}", ssum_ref=f"$N${t}",
                           anchors=anchors, header=True, header_row_at=t + 2,
                           mod_refs=LogRefs(key_cond=f"$B${t}", eff="ml_eff",
                                            ln="ml_ln1p", effmonth="ml_effmonth",
                                            first="ml_first", daysafter="ml_daysafter"),
                           mod_col=NC_MOD_COL, mod_base_ref=base_ref,
                           mod_count_ref=cnt_ref,
                           # D84: blank key = a spare state, or ANY state while
                           # the tab is on the All view. Both are states where
                           # the sweep provably returns the gate's constants,
                           # so this is free in the common view too.
                           idle_cond=f'$B${t}=""')

        p0, p1 = t + 3 + 24, t + 3 + 35        # plan-year cohort rows
        y0, y1 = t + 3 + 12, t + 3 + 23        # year-ago cohort rows
        rr = t + 51                            # results row
        # x-anchor cells (close-of-day serials) for the alpha/beta mod affine
        x_asof, x_end = f"$B${t + 1}", f"$C${t + 1}"
        m0v = f"$F${t + 1}"                    # M_0 value anchor (ModAnchors.m[1])
        for j in range(12):
            r = p0 + j
            y = r - 12
            # A(m), B(m): W_new(m) = A + B*r (affine; D31 blend in D's month)
            formula(ws, f"W{r}",
                    f"=IF($G{r}<>$Q${t},$N{r},(1-$U${t})*$I{r}+$U${t}*$J{r})", fmt=FMT_IDX)
            formula(ws, f"X{r}",
                    f"=IF($G{r}<$Q${t},0,IF($G{r}>$Q${t},$N{r},$U${t}*$J{r}))", fmt=FMT_IDX)
            formula(ws, f"Y{r}", f"=IF($F${t}=1,$O{r}/$O{y},1)", fmt=FMT_IDX)
            # alpha/beta: M(m) = alpha + beta * M_1 along the anchor path.
            # D70: with actions logged there is no free M_1 to solve for — the
            # log pins the path — so beta collapses to 0 and alpha becomes the
            # stepped level itself. Every downstream user stays correct: the
            # booked-program q (AE) reads the real path, and the implied-mod
            # solve (results H) divides by SUMPRODUCT(AD, AB) = 0 and reports
            # n/a, which is the honest answer rather than a fabricated one.
            formula(ws, f"AB{r}",
                    f"=IF({cnt_ref}>0,0,IF(AND(N($K${t})<>0,$F{r}<={x_asof}),0,"
                    f"($F{r}-{x_asof})/({x_end}-{x_asof})))", fmt="0.0000")
            formula(ws, f"AC{r}",
                    f"=IF(OR({cnt_ref}>0,AND(N($K${t})<>0,$F{r}<={x_asof})),$O{r},"
                    f"{m0v}*(1-$AB{r}))", fmt=FMT_MOD)
            # h(m) at the rate in force (results row F, forward reference)
            formula(ws, f"AD{r}",
                    f"=$H{r}*($W{r}+$X{r}*IF(ISNUMBER($F${rr}),$F${rr},0))"
                    f"/($N{y}*$O{y})", fmt=FMT_IDX)
            # share indicator s(m): 0 / p / 1
            formula(ws, f"AF{r}",
                    f"=IF($G{r}<$Q${t},0,IF($G{r}>$Q${t},1,$U${t}))", fmt="0.000")
            # the tab's grid values: rate leg and required pricing leg by month
            formula(ws, f"AG{r}",
                    f'=IF(OR($D${t}=0,$P${t}=0),"",'
                    f"($W{r}+$X{r}*IF(ISNUMBER($F${rr}),$F${rr},0))/$N{y}-1)", fmt=PCT_S)
            formula(ws, f"AH{r}",
                    f'=IF(OR($AG{r}="",$F${t}=0),"",(1+$E${t})/(1+$AG{r})-1)', fmt=PCT_S)
            formula(ws, f"AI{r}",
                    f'=IF($AH{r}="","",$O{y}*(1+$E${t})/(1+$AG{r}))', fmt=FMT_MOD)
        # ---- D75: the mod ask as a dated ACTION, not a level ----------------
        # Along the stepped path the plan-year written mod is affine in (1 + c),
        # M_w(m) = A(m) + B(m)*(1 + c), so the required step solves in one sum
        # the way the filing does. A/B mirror the rate leg's W/X exactly:
        # months before D carry no step, months after carry it whole, and D's
        # own month splits on the day-blend.
        c_pre = f"${col(NC_MOD_COL)}"        # mod index before the month
        c_eom = f"${col(NC_MOD_COL + 1)}"    # mod index at month end
        c_bl = f"${col(NC_MOD_COL + 4)}"     # blended step index
        sc = [col(NC_MSTEP_COL + i) for i in range(5)]   # AV..AZ
        for j in range(12):
            y, r = y0 + j, p0 + j
            formula(ws, f"{sc[0]}{y}", mod_value_formula(f"$F{y}", anchors_s),
                    fmt=FMT_MOD)
            formula(ws, f"{sc[4]}{y}",
                    f"=IF({cnt_ref}=0,0,ABS(${sc[0]}{y}-$O{y}))", fmt="0.0000000000")
            formula(ws, f"{sc[1]}{r}",
                    f"=IF($G{r}<$Q${t},{base_ref}*{c_bl}{r},"
                    f"IF($G{r}>$Q${t},0,{base_ref}*(1-{pmod_ref})*{c_pre}{r}))",
                    fmt=FMT_MOD)
            formula(ws, f"{sc[2]}{r}",
                    f"=IF($G{r}<$Q${t},0,IF($G{r}>$Q${t},{base_ref}*{c_bl}{r},"
                    f"{base_ref}*{pmod_ref}*{c_eom}{r}))", fmt=FMT_MOD)
            formula(ws, f"{sc[3]}{r}",
                    f"=$H{r}*($W{r}+$X{r}*IF(ISNUMBER($F${rr}),$F${rr},0))"
                    f"/($N{y}*${sc[0]}{y})", fmt=FMT_IDX)
            # agreement: outside D's own month these columns must reproduce the
            # block's own mod path whenever the log already drives it. IN D's
            # month they legitimately differ — a step there moves the two-point
            # split to the earlier of D and the first logged action, which
            # changes the blend even at c = 0 (D31).
            formula(ws, f"{sc[4]}{r}",
                    f"=IF(OR({cnt_ref}=0,$G{r}=$Q${t}),0,"
                    f"ABS(${sc[1]}{r}+${sc[2]}{r}-$O{r}))", fmt="0.0000000000")

        # program-q column over ALL rows (for the booked-program reconciliation)
        for k in range(48):
            r = t + 3 + k
            formula(ws, f"AE{r}",
                    f"=IF(OR($G{r}<nr_MStartP,$G{r}>=nr_MStartP1),"
                    f"IF($F${t}=1,$N{r}*$O{r}/$G${t},$N{r}),"
                    f"($W{r}+$X{r}*IF(ISNUMBER($F${rr}),$F${rr},0))"
                    f"*IF($F${t}=1,($AC{r}+$AB{r}*IF(ISNUMBER($H${rr}),$H${rr},"
                    f"$J${t}))/$G${t},1))", fmt=FMT_IDX)
        # D68: the rate leg one year back — the year now flowing. The tab's new
        # change is always dated INSIDE the plan year (the input carries a
        # date validation to it), so it can never touch a P-1 month: the prior
        # year is the live log measured against its own P-2 base, with no
        # affine term and nothing to solve.
        for j in range(12):
            r = t + 15 + j
            formula(ws, f"AG{r}", f'=IF($D${t}=0,"",$N{r}/$N{r - 12}-1)', fmt=PCT_S)
        for cc in ("W", "X", "Y", "AB", "AC", "AD", "AE", "AF", "AG", "AH",
                   "AI", *sc):
            for k in range(48):
                c = ws[f"{cc}{t + 3 + k}"]
                if c.value is not None:
                    c.font = font(GREY_DARK, size=8)

        # ---- results row ----
        hp = f"$H${p0}:$H${p1}"
        label(ws, f"A{rr}", "delivery results:")
        ws[f"A{rr}"].font = font(GREY_DARK, size=8)
        formula(ws, f"B{rr}", f"=SUM({hp})", fmt="0.00")
        formula(ws, f"C{rr}",
                f"=SUMPRODUCT({hp},$W${p0}:$W${p1}/$N${y0}:$N${y1},$Y${p0}:$Y${p1})",
                fmt=FMT_IDX)
        formula(ws, f"D{rr}",
                f"=SUMPRODUCT({hp},$X${p0}:$X${p1}/$N${y0}:$N${y1},$Y${p0}:$Y${p1})",
                fmt=FMT_IDX)
        formula(ws, f"E{rr}",
                f'=IF(OR($D${t}=0,$P${t}=0,$D{rr}=0),"n/a",'
                f"((1+$E${t})*$B{rr}-$C{rr})/$D{rr})", fmt="0.000%")
        formula(ws, f"F{rr}", f'=IF($V${t}<>"",$V${t},$E{rr})', fmt="0.000%")
        formula(ws, f"G{rr}", f'=IF(OR(NOT(ISNUMBER($F{rr})),nd_Ach=0),"n/a",'
                              f"$F{rr}/nd_Ach)", fmt="0.000%")
        formula(ws, f"H{rr}",
                f'=IF(OR(NOT(ISNUMBER($F{rr})),$F${t}=0,'
                f'SUMPRODUCT($AD${p0}:$AD${p1},$AB${p0}:$AB${p1})=0),"n/a",'
                f"((1+$E${t})*$B{rr}-SUMPRODUCT($AD${p0}:$AD${p1},$AC${p0}:$AC${p1}))"
                f"/SUMPRODUCT($AD${p0}:$AD${p1},$AB${p0}:$AB${p1}))", fmt=FMT_MOD)
        formula(ws, f"I{rr}",
                f'=IF(NOT(ISNUMBER($F{rr})),"n/a",SUMPRODUCT({hp},'
                f"($W${p0}:$W${p1}+$X${p0}:$X${p1}*$F{rr})/$N${y0}:$N${y1})/$B{rr}-1)",
                fmt=PCT_S)
        formula(ws, f"J{rr}", f'=IF(OR(NOT(ISNUMBER($I{rr})),$F${t}=0),"n/a",'
                              f"(1+$E${t})/(1+$I{rr})-1)", fmt=PCT_S)
        formula(ws, f"K{rr}",
                f'=IF(NOT(ISNUMBER($F{rr})),"n/a",($C{rr}+$D{rr}*$F{rr})/$B{rr}-1)',
                fmt=PCT_S)
        formula(ws, f"L{rr}",
                f'=IF(NOT(ISNUMBER($F{rr})),"n/a",'
                f"SUMPRODUCT({hp},$AF${p0}:$AF${p1})/$B{rr})", fmt=FMT_PCT)
        # D75: the same ask as a dated action — required step, what you must
        # direct to land it, the year-end mod it produces, and the share of
        # delivery it can still reach. AW/AX carry their own column footers.
        hsr = f"${sc[3]}${p0}:${sc[3]}${p1}"
        formula(ws, f"{sc[1]}{rr}",
                f"=SUMPRODUCT({hsr},${sc[1]}${p0}:${sc[1]}${p1})", fmt=FMT_IDX)
        formula(ws, f"{sc[2]}{rr}",
                f"=SUMPRODUCT({hsr},${sc[2]}${p0}:${sc[2]}${p1})", fmt=FMT_IDX)
        formula(ws, f"S{rr}",
                f'=IF(OR(NOT(ISNUMBER($F{rr})),$F${t}=0,${sc[2]}{rr}=0),"n/a",'
                f"((1+$E${t})*$B{rr}-${sc[1]}{rr})/${sc[2]}{rr}-1)", fmt="0.000%")
        formula(ws, f"T{rr}", f'=IF(OR(NOT(ISNUMBER($S{rr})),nd_Ach=0),"n/a",'
                              f"$S{rr}/nd_Ach)", fmt="0.000%")
        formula(ws, f"U{rr}", f'=IF(NOT(ISNUMBER($S{rr})),"n/a",'
                              f"{base_ref}*(1+$S{rr})*{c_eom}{p1})", fmt=FMT_MOD)
        formula(ws, f"V{rr}", f'=IF(NOT(ISNUMBER($S{rr})),"n/a",'
                              f"${sc[2]}{rr}/(${sc[1]}{rr}+${sc[2]}{rr}))", fmt=FMT_PCT)
        formula(ws, f"{sc[4]}{rr}",
                f"=MAX(${sc[4]}${y0}:${sc[4]}${y1},${sc[4]}${p0}:${sc[4]}${p1})",
                fmt="0.0000000000")
        # Flags: implied mod out of bounds / |r| beyond bound / thin share.
        # Every numeric test is NESTED inside its own ISNUMBER rather than
        # ANDed with it — AND does not short-circuit in Excel, so ABS("n/a")
        # is evaluated and poisons the whole string with #VALUE!. These cells
        # go "n/a" whenever the mod adjustment is off.
        num = lambda ref, test, tag: (f'IF(ISNUMBER({ref}),IF({test},"{tag} ",""),"")')
        formula(ws, f"M{rr}",
                f'=IF($D${t}=0,"",TRIM('
                + num(f"$H{rr}", f"OR($H{rr}<0.5,$H{rr}>1.5)", "MOD-BOUND")
                + "&" + num(f"$F{rr}", f"ABS($F{rr})>nr_SolvMaxRate", "RATE-BOUND")
                + "&" + num(f"$L{rr}", f"$L{rr}<nr_SolvMinShare", "THIN-SHARE")
                + "&" + num(f"$S{rr}", f"ABS($S{rr})>nr_SolvMaxMod", "MODSTEP-BOUND")
                + "&" + num(f"$U{rr}", f"OR($U{rr}<0.5,$U{rr}>1.5)", "MODEND-BOUND")
                + f'&IF($W${t}>1,"MULTI-PLANNED ","")'
                + f'&IF({cnt_ref}>0,"MOD-STEPS ","")'
                + f'&IF($F${t}=0,"RATE-ONLY","")))')
        # booked-program plan LR vs the asserted net-mode plan LR
        formula(ws, f"N{rr}",
                f'=IF(OR($D${t}=0,NOT(ISNUMBER($F{rr}))),"n/a",'
                f"INDEX(calc_lrcur,$C${t})*INDEX(calc_crl,$C${t})"
                f"/(SUMPRODUCT($H${t + 3}:$H${t + 50},$Q${t + 3}:$Q${t + 50},"
                f"$AE${t + 3}:$AE${t + 50})/SUMPRODUCT($H${t + 3}:$H${t + 50},"
                f"$Q${t + 3}:$Q${t + 50}))*INDEX(calc_aother,$C${t}))", fmt=FMT_PCT)
        formula(ws, f"O{rr}", f'=IF($D${t}=0,"n/a",INDEX(calc_cylr_p,$C${t}))',
                fmt=FMT_PCT)
        # solve residual (Checks): delivered at r* minus target, 0 when n/a
        formula(ws, f"P{rr}",
                f"=IF(NOT(ISNUMBER($E{rr})),0,"
                f"ABS(($C{rr}+$D{rr}*$E{rr})/$B{rr}-(1+$E${t})))", fmt="0.0000000000")
        for cc in ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M",
                   "N", "O", "P", "S", "T", "U", "V", sc[1], sc[2], sc[4]):
            ws[f"{cc}{rr}"].font = font(GREY_DARK, size=8)

    # names for the band block + the residual aggregate
    tb = block_top(nd)
    ctx.define("nd_band_top", NC, f"$A${tb}", "Band block anchor (state echo)")
    # cross-checks on the band combo (the one legitimate computed-row INDEX):
    # (a) live-filter year-ago N must equal the combo's full-log _calc block N
    #     (planned rows on/after 1/1/P can never enter P-1 cohorts — D57);
    # (b) the engine's own net path must satisfy T(m)/T(m-12) = 1 + x.
    calc_base = (f"{L.CALC_BLOCK_FIRST}+(MATCH($B${tb},calc_key,0)-1)"
                 f"*{L.CALC_BLOCK_STRIDE}")
    # D84: bound the INDEX ranges. Whole-column refs ('_calc'!$N:$N) make these
    # 24 cells depend on all 1,048,576 rows of two columns and re-dirty on any
    # block change — the exact pattern Program Flow already replaced (D68).
    cl = L.CALC_BLOCK_FIRST + L.LR_ROWS * L.CALC_BLOCK_STRIDE + 60
    rbb = tb + 51
    for j in range(12):
        r = tb + 3 + 12 + j            # year-ago cohort rows
        formula(ws, f"AK{r}",
                f"=IF($D${tb}=0,0,ABS($N{r}"
                f"-INDEX('_calc'!$N$1:$N{cl},{calc_base}+{15 + j})))", fmt="0.0000000000")
        rp = tb + 3 + 24 + j           # plan cohort rows
        formula(ws, f"AL{rp}",
                f"=IF(OR($D${tb}=0,$F${tb}=0),0,"
                f"ABS(INDEX('_calc'!$T$1:$T{cl},{calc_base}+{27 + j})"
                f"/INDEX('_calc'!$T$1:$T{cl},{calc_base}+{15 + j})-(1+$E${tb})))",
                fmt="0.0000000000")
        ws[f"AK{r}"].font = font(GREY_DARK, size=8)
        ws[f"AL{rp}"].font = font(GREY_DARK, size=8)
    formula(ws, f"Q{rbb}", f"=MAX($AK${tb + 15}:$AK${tb + 26})", fmt="0.0000000000")
    ctx.define("nd_YagoDiff", NC, f"$Q${rbb}",
               "Max |live-filter year-ago index - full-log _calc index| (band combo)")
    formula(ws, f"R{rbb}", f"=MAX($AL${tb + 27}:$AL${tb + 38})", fmt="0.0000000000")
    ctx.define("nd_NetSelfChk", NC, f"$R${rbb}",
               "Max |T(m)/T(m-12) - (1+x)| on the engine's own net path (band combo)")
    resid_cells = ",".join(f"$P${block_top(i) + 51}" for i in range(nd + 1))
    formula(ws, "J1", f"=MAX({resid_cells})", fmt="0.0000000000")
    ctx.define("nd_MaxResid", NC, "$J$1",
               "Largest closed-form solve residual across all delivery blocks")
    # D75: outside D's own month the mod-step columns must reproduce the
    # block's own mod path on any combo the log already drives. A non-zero
    # here means the forced-stepped anchors, M_endPrior, or a column offset
    # is wrong — exactly the failure modes that hide inside a plausible number.
    step_cells = ",".join(f"${col(NC_MSTEP_COL + 4)}${block_top(i) + 51}"
                          for i in range(nd + 1))
    formula(ws, "K1", f"=MAX({step_cells})", fmt="0.0000000000")
    ctx.define("nd_MaxStepDiff", NC, "$K$1",
               "Max |stepped mod columns - engine mod path| on already-stepped combos")
    set_widths(ws, {c: 10 for c in
                    ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L",
                     "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X",
                     "Y", "Z", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI")})
    set_widths(ws, {col(NC_MOD_CNT_COL + j): 10 for j in range(7)})   # AM..AS (D70)
    set_widths(ws, {col(NC_MSTEP_HDR_COL + j): 10 for j in range(7)})  # AT..AZ (D75)


# ---------------------------------------------------------------------------
# Net Delivery (visible)
# ---------------------------------------------------------------------------


def build_net_delivery(ctx: Ctx):
    ws = ctx.wb[SHEETS.NET_DELIVERY]
    p = ctx.p
    states = ctx.cfg.states
    nd = n_disp(ctx)
    title(ws, "A1", f"Net Delivery — how the net rate target gets delivered ({ctx.lob.name})",
          "For combos on a net rate selection: month-by-month renewals must average the "
          "target. Given each state's rate-change date and the current flow, this tab "
          "shows the rate change coming through and the pricing change that satisfies "
          "the rest. Oracle-tied like everything else.")
    nav_bar(ws, 3, 1, ["Control", "State Summary", "Rate Log", "Solver", "Checks"], step=2)

    # ---- controls ----
    # seed the view on the BU that actually carries a net selection in the
    # sample, so the tab demonstrates itself on first open
    net_rows = [r for r in ctx.lr_rows if r.get("netp") is not None]
    seed_bu = net_rows[0]["bu"] if net_rows else ctx.we_row["bu"]
    label(ws, "A5", "Business unit in view", bold=True)
    c = input_cell(ws, "B5", seed_bu)
    c.alignment = ALIGN_C
    ctx.define("nd_BU", SHEETS.NET_DELIVERY, "$B$5",
               "Net Delivery BU in view (delivery needs a single BU; All = roll-up only)")
    dv_bu = DataValidation(type="list", formula1="=lst_bu_all", allow_blank=False,
                           showErrorMessage=True)
    ws.add_data_validation(dv_bu)
    dv_bu.add("B5")
    label(ws, "D5", "Default effective date")
    input_cell(ws, "E5", None, fmt=FMT_DATE, required=False)
    ctx.define("nd_DefaultDateRaw", SHEETS.NET_DELIVERY, "$E$5",
               "Default rate-change date (blank = April 1 of the plan year)")
    formula(ws, "F5", '=IF(N(nd_DefaultDateRaw)=0,DATE(nr_PlanYear,4,1),nd_DefaultDateRaw)',
            fmt=FMT_DATE)
    ws["F5"].font = font(GREY_DARK, size=9)
    ctx.define("nd_DefaultDate", SHEETS.NET_DELIVERY, "$F$5",
               "Resolved default effective date for suggested changes")
    label(ws, "H5", "Achievement %")
    input_cell(ws, "I5", 1.0, fmt=FMT_PCT, required=False)
    ctx.define("nd_Ach", SHEETS.NET_DELIVERY, "$I$5",
               "Assumed achievement on the suggested filing (filed equivalent = r / this)")
    formula(ws, "A6",
            '=IF(nd_BU="All","Pick a single business unit to see delivery and '
            'suggestions — combining different rate histories across BUs would not '
            'mean anything.",'
            'IF(nr_ModAdjMaster="OFF","NOTE: master mod toggle is OFF — targets are '
            'interpreted RATE-ONLY (D39); the pricing columns show a dash.",'
            'IF(COUNT(sc_dnet)>0,"NOTE: a Scenarios delta-net lever is set; scenario '
            'levers do not affect this tab.","")))')
    ws["A6"].font = font(FAIL_RED, size=9, italic=True)

    # ---- per-state summary table ----
    hdr, first = ND_HDR, ND_FIRST
    headers = ["State", "Adj plan EP (000s)", "Net target", "Eff date (enter)",
               "Filed % (enter; blank = your planned filing, else suggested)",
               "Planned filing date (rate log)", "Planned filed %",
               "Suggested change",
               "Change in force", "Filed equivalent", "Avg YoY rate leg",
               "Avg YoY pricing needed", "Implied yr-end mod", "vs projected M_1 (pts)",
               "Required mod change at that date", "Yr-end mod it produces",
               "Share written on/after date"]
    header_row(ws, hdr, 1, headers)
    header_row(ws, hdr, ND_FLAG_COL, ["Flags"])
    ws.row_dimensions[hdr].height = 42
    dv_date = DataValidation(
        type="date", operator="between",
        formula1=f"DATE({p},1,1)", formula2=f"DATE({p},12,31)",
        allow_blank=True, showErrorMessage=True,
        error="The rate-change date must lie inside the plan year (prior-year "
              "vehicles belong in the Rate Log).")
    ws.add_data_validation(dv_date)
    dv_r = DataValidation(type="decimal", operator="between", formula1="-0.5",
                          formula2="2", allow_blank=True, showErrorMessage=True,
                          error="Enter a decimal fraction in [-50%, +200%].")
    ws.add_data_validation(dv_r)
    for i in range(nd):
        r = first + i
        t = block_top(i)
        rr = t + 51
        band_fill = FILL_PANEL if i % 2 else None
        link(ws, f"A{r}", f"=_lists!$B${3 + i}", align=ALIGN_C, fill=band_fill,
             border=BORDER_THIN, bold=True)
        formula(ws, f"B{r}",
                f'=IF($A{r}="","",IF(nd_BU="All",SUMIFS(calc_ep,calc_state,$A{r}),'
                f"SUMIFS(calc_ep,calc_state,$A{r},calc_bu,nd_BU)))",
                fmt="#,##0;-#,##0;\"\"", align=ALIGN_C, fill=band_fill, border=BORDER_THIN)
        formula(ws, f"C{r}",
                f'=IF($A{r}="","",IF(nd_BU="All",'
                f'IF(SUMIFS(calc_netmode,calc_state,$A{r})=0,"—",'
                f'SUMIFS(calc_netx,calc_state,$A{r})/SUMIFS(calc_netmode,calc_state,$A{r})),'
                f"IF('{NC}'!$D${t}=0,\"—\",'{NC}'!$E${t})))",
                fmt=PCT_S, align=ALIGN_C, fill=band_fill, border=BORDER_THIN)
        input_cell(ws, f"D{r}", None, fmt="m/d/yy", required=False)
        input_cell(ws, f"E{r}", None, fmt="0.0%", required=False)
        dv_date.add(f"D{r}")
        dv_r.add(f"E{r}")
        # the planned filing picked up from the rate log (blank when none)
        for cc, src, fmt in (("F", "X", "m/d/yy"), ("G", "Z", "+0.0%;-0.0%;0.0%")):
            formula(ws, f"{cc}{r}",
                    f'=IF(OR($A{r}="",nd_BU="All"),"",'
                    f"IF('{NC}'!$D${t}=0,\"—\",IF('{NC}'!$W${t}=0,\"none\","
                    f"'{NC}'!${src}${t})))",
                    fmt=fmt, align=ALIGN_C, fill=band_fill, border=BORDER_THIN)
        # M/N are the LEVEL answer and its gap; O/P are the same ask as an
        # ACTION (D75), which is the one that survives once the Mod Log drives
        # the path; Q is the share a change at that date can still reach.
        for cc, src, fmt in (("H", "E", "0.000%"), ("I", "F", "0.000%"),
                             ("J", "G", "0.000%"), ("K", "I", PCT_S),
                             ("L", "J", PCT_S), ("M", "H", FMT_MOD),
                             ("O", "S", "+0.000%;-0.000%"), ("P", "U", FMT_MOD),
                             ("Q", "L", FMT_PCT), (col(ND_FLAG_COL), "M", None)):
            formula(ws, f"{cc}{r}",
                    f'=IF(OR($A{r}="",nd_BU="All"),"",'
                    f"IF('{NC}'!$D${t}=0,\"—\",'{NC}'!${src}${rr}))",
                    fmt=fmt or "General", align=ALIGN_C, fill=band_fill,
                    border=BORDER_THIN, bold=(cc in ("I", "L", "O")))
        formula(ws, f"N{r}",
                f'=IF(OR($A{r}="",nd_BU="All"),"",IF(\'{NC}\'!$D${t}=0,"—",'
                f"IF(ISNUMBER('{NC}'!$H${rr}),"
                f"('{NC}'!$H${rr}-'{NC}'!$J${t})*100,\"n/a\")))",
                fmt='+0.0;-0.0;0.0', align=ALIGN_C, fill=band_fill, border=BORDER_THIN)
    tot = first + nd + 1
    put(ws, f"A{tot}", "IN VIEW", fnt=font(NAVY, bold=True), align=ALIGN_C)
    formula(ws, f"B{tot}", '=IF(nd_BU="All",SUM(calc_ep),SUMIFS(calc_ep,calc_bu,nd_BU))',
            fmt="#,##0", align=ALIGN_C, bold=True)
    formula(ws, f"C{tot}",
            '=IF(nd_BU="All",SUMPRODUCT(calc_netmode),'
            'SUMPRODUCT(calc_netmode*(calc_bu=nd_BU)))&" net combo(s)"')
    ws[f"C{tot}"].font = F_LABEL
    ctx.define("ndd_states", SHEETS.NET_DELIVERY,
               f"$A${first}:$A${first + nd - 1}", "Net Delivery roster state column")
    ctx.define("ndd_dates", SHEETS.NET_DELIVERY,
               f"$D${first}:$D${first + nd - 1}",
               "Per-state rate-change dates (blank = default)")
    ctx.define("ndd_filed", SHEETS.NET_DELIVERY,
               f"$E${first}:$E${first + nd - 1}",
               "Per-state filed % overrides (blank = the planned filing, else suggested)")
    ctx.define("ndd_flags", SHEETS.NET_DELIVERY,
               f"${col(ND_FLAG_COL)}${first}:${col(ND_FLAG_COL)}${first + nd - 1}",
               "Per-state feasibility flags (empty = clean; — = not on a net selection)")
    put(ws, f"A{tot + 2}",
        "Leave the two input columns blank to adopt the PLANNED filing already in the "
        "rate log for the plan year (shown beside them): its date and its effective % "
        "= filed x achievement. With no planned filing the tab falls back to the "
        "default date and the suggested change. Type in either column to override. "
        "A net selection supersedes planned rows in the engine (D39) — this tab is "
        "where you put one back to ask what pricing still has to carry.", fnt=F_SMALL_IT)
    put(ws, f"A{tot + 3}",
        "Only the FIRST planned filing inside the plan year is adopted (the model "
        "carries one change at one date); a state with more is flagged MULTI-PLANNED. "
        "The Solver answers a different question (CY plan LR under the explicit "
        "program, D13).", fnt=F_SMALL_IT)
    put(ws, f"A{tot + 4}",
        "MOD-STEPS means the combo already has actions in the Mod Log, so its pricing "
        "path is decided rather than open: the projected mod row below follows those "
        "actions, and the implied year-end MOD LEVEL reads n/a because the log — not "
        "this tab — now sets it. The required mod CHANGE beside it still answers, and "
        "is the column to read: it is the further action your logged plan still leaves "
        "for the target to find, dated at the same filing.", fnt=F_SMALL_IT)

    # ---- twin state x month grids, the rate leg two years wide (D68) ----
    # Same seam and the same colours as Program Flow — steel for the year now
    # flowing at B..M, navy for the plan year at N..Y — so the two tabs read
    # identically. Unlike Program Flow the prior block carries no outline
    # group: this tab's summary keeps its INPUT columns (D, E) under B..M, and
    # collapsing would hide them.
    g1 = tot + GRID_GAP
    section(ws, g1, "A", "How the target is delivered, month by month — the rate leg")
    put(ws, f"A{g1 + 1}",
        "YoY change on renewals from the rate side: history plus the change in force at "
        "each state's date. Steps mark anniversaries of past changes and the new filing. "
        "The steel block is the year now flowing — what the book is already delivering "
        "before this filing lands.", fnt=F_SMALL_IT)
    g2 = g1 + 3 + nd + 3
    section(ws, g2, "A", "…and the pricing change that satisfies the target")
    put(ws, f"A{g2 + 1}",
        "Required YoY pricing (schedule-mod) change on renewals so that rate x pricing "
        "averages the target. Red = the months where the burden bites. Plan year only: "
        "there is no target in a prior year, so a required walk is undefined there — the "
        "steel block is deliberately empty.", fnt=F_SMALL_IT)
    years = ((ND_PRIOR_COL, 15, FILL_STEEL, "nr_PlanYear-1"),
             (ND_PLAN_COL, 27, FILL_NAVY, "nr_PlanYear"))
    for g0, colL, two_year in ((g1 + 2, "AG", True), (g2 + 2, "AH", False)):
        put(ws, f"A{g0}", "State", fnt=font("FFFFFF", bold=True), fill=FILL_NAVY,
            align=ALIGN_C)
        for c0, off, fill_, yexpr in years:
            for j in range(12):
                cell = ws.cell(row=g0, column=c0 + j)
                cell.font = font("FFFFFF", bold=True, size=9)
                cell.fill = fill_
                cell.alignment = ALIGN_C
                if two_year or c0 == ND_PLAN_COL:
                    cell.value = (f'=TEXT(DATE({yexpr},{j + 1},1),"mmm")'
                                  f'&" "&({yexpr})')
            if not two_year and c0 == ND_PRIOR_COL:
                c = ws.cell(row=g0, column=c0)
                c.value = '="  no target in "&(nr_PlanYear-1)&" — nothing to solve"'
                c.alignment = ALIGN_L
        for i in range(nd):
            r = g0 + 1 + i
            t = block_top(i)
            link(ws, f"A{r}", f"=_lists!$B${3 + i}", align=ALIGN_C, bold=True)
            for c0, off, _f, _y in years:
                if not two_year and c0 == ND_PRIOR_COL:
                    continue
                for j in range(12):
                    cc = ws.cell(row=r, column=c0 + j)
                    formula(ws, cc.coordinate,
                            f"=IF(nd_BU=\"All\",\"\",'{NC}'!${colL}${t + off + j})",
                            fmt=PCT_S, align=ALIGN_C)
                    cc.font = font(GREY_DARK, size=9)
    ws.conditional_formatting.add(
        f"{col(ND_PLAN_COL)}{g2 + 3}:{col(ND_PLAN_COL + 11)}{g2 + 2 + nd}",
        ColorScaleRule(start_type="min", start_color="D6E8D5",
                       mid_type="percentile", mid_value=50, mid_color="FFF2CC",
                       end_type="max", end_color="F4B8B8"))

    # ---- monthly deep-dive band (own pickers) — leads the tab (D74) ----
    # Written here, after the grids, but placed ABOVE them: nothing about a
    # cell's position depends on the order the builder emits it in.
    b0 = ND_BAND_TOP
    tb = block_top(nd)
    rb = tb + 51
    section(ws, b0, "A", "One combo under the microscope")
    label(ws, f"A{b0 + 1}", "Business unit")
    input_cell(ws, f"B{b0 + 1}", seed_bu, required=False)
    ctx.define("nd_BandBU", SHEETS.NET_DELIVERY, f"$B${b0 + 1}",
               "Deep-dive band BU picker (independent of the table's BU filter)")
    label(ws, f"C{b0 + 1}", "State")
    net_state = (net_rows[0]["state"] if net_rows
                 else (states[8] if len(states) > 8 else states[0]))
    input_cell(ws, f"D{b0 + 1}", net_state, required=False)
    ctx.define("nd_BandState", SHEETS.NET_DELIVERY, f"$D${b0 + 1}",
               "Deep-dive band state picker")
    formula(ws, f"F{b0 + 1}",
            f"=IF('{NC}'!$D${tb}=0,\"— not on a net selection (explicit program; see "
            f"Solver / Scenarios) —\",\"net target \"&TEXT('{NC}'!$E${tb},\"+0.0%;-0.0%\")"
            f"&\"  |  change in force \"&IF(ISNUMBER('{NC}'!$F${rb}),"
            f"TEXT('{NC}'!$F${rb},\"+0.000%;-0.000%\"),\"n/a\")&\"  |  date \""
            f"&TEXT('{NC}'!$O${tb},\"m/d/yy\"))")
    ws[f"F{b0 + 1}"].font = font(NAVY, bold=True)
    dvb = DataValidation(type="list", formula1="=lst_bu", allow_blank=True,
                         showErrorMessage=False)
    ws.add_data_validation(dvb)
    dvb.add(f"B{b0 + 1}")
    dvs = DataValidation(type="list", formula1="=lst_state", allow_blank=True,
                         showErrorMessage=False)
    ws.add_data_validation(dvs)
    dvs.add(f"D{b0 + 1}")

    # the microscope band stays plan-year, and sits under the PLAN-year columns
    # so its months line up with the grid block above them (D68)
    bh = b0 + 3
    put(ws, f"A{bh}", "", fill=FILL_GREY)
    for j in range(12):
        cell = ws.cell(row=bh, column=ND_PLAN_COL + j)
        cell.value = f'=TEXT(DATE(nr_PlanYear,{j + 1},1),"mmm")&" "&nr_PlanYear'
        cell.font = font(GREY_DARK, bold=True, size=9)
        cell.fill = FILL_GREY
        cell.alignment = ALIGN_C
    band_rows = [
        ("Written weight w(m)", "H", "0.00", "t3"),
        ("YoY rate leg on renewals", "AG", PCT_S, "plan"),
        ("Required YoY pricing leg", "AH", PCT_S, "plan"),
        ("Required written mod level", "AI", FMT_MOD, "plan"),
        ("Projected written mod level", "O", FMT_MOD, "plan"),
    ]
    r = bh + 1
    for lbl, colL, fmt, kind in band_rows:
        label(ws, f"A{r}", lbl)
        for j in range(12):
            src = tb + 27 + j if kind == "plan" else tb + 27 + j
            formula(ws, ws.cell(row=r, column=ND_PLAN_COL + j).coordinate,
                    f"='{NC}'!${colL}${src}", fmt=fmt, align=ALIGN_C)
            ws.cell(row=r, column=ND_PLAN_COL + j).font = font(GREY_DARK, size=9)
        r += 1
    label(ws, f"A{r}", "Gap: required minus projected mod (pts)")
    for j in range(12):
        formula(ws, ws.cell(row=r, column=ND_PLAN_COL + j).coordinate,
                f"=IF('{NC}'!$AI${tb + 27 + j}=\"\",\"\","
                f"('{NC}'!$AI${tb + 27 + j}-'{NC}'!$O${tb + 27 + j})*100)",
                fmt="+0.0;-0.0;0.0", align=ALIGN_C)
        ws.cell(row=r, column=ND_PLAN_COL + j).font = font(FAIL_RED, size=9)
    r += 2
    recon = [
        ("Required mod CHANGE, dated at the filing above (D75)",
         f"=IF(ISNUMBER('{NC}'!$S${rb}),'{NC}'!$S${rb},\"n/a\")", "+0.000%;-0.000%"),
        ("...  what you must DIRECT at the achievement above",
         f"=IF(ISNUMBER('{NC}'!$T${rb}),'{NC}'!$T${rb},\"n/a\")", "+0.000%;-0.000%"),
        ("...  the written mod at 12/31 it produces",
         f"=IF(ISNUMBER('{NC}'!$U${rb}),'{NC}'!$U${rb},\"n/a\")", FMT_MOD),
        ("...  share of delivery a change at that date can still reach",
         f"=IF(ISNUMBER('{NC}'!$V${rb}),'{NC}'!$V${rb},\"n/a\")", FMT_PCT),
        ("Implied year-end written mod, LEVEL basis (n/a once the Mod Log drives it)",
         f"=IF(ISNUMBER('{NC}'!$H${rb}),'{NC}'!$H${rb},\"n/a\")", FMT_MOD),
        ("Plan LR if you BOOK this program (live history + filing + mod walk)",
         f"='{NC}'!$N${rb}", FMT_PCT),
        ("Plan LR as ASSERTED by the net selection",
         f"='{NC}'!$O${rb}", FMT_PCT),
        ("Difference (earned-shape residual, pts)",
         f"=IF(OR(NOT(ISNUMBER('{NC}'!$N${rb})),NOT(ISNUMBER('{NC}'!$O${rb}))),\"n/a\","
         f"('{NC}'!$N${rb}-'{NC}'!$O${rb})*100)", "+0.00;-0.00;0.00"),
    ]
    for lbl, f, fmt in recon:
        label(ws, f"A{r}", lbl)
        formula(ws, f"C{r}", f, fmt=fmt, align=ALIGN_C, border=BORDER_THIN,
                bold=("BOOK" in lbl or "CHANGE" in lbl))
        r += 1
    put(ws, f"A{r}",
        "The mod change is dated at the SAME date as the filing — one filing normally "
        "carries both levers. It compounds on the year-end mod already projected for "
        "the prior year, so asking for it puts the combo on the stepped path (D70) "
        "whether or not the Mod Log already holds actions.", fnt=F_SMALL_IT)
    r += 1
    put(ws, f"A{r}",
        "The solve matches the WRITTEN-basis average (the literal meaning of the "
        "target); the earned-shape residual above is the honest difference vs the "
        "asserted path. Attribution cannot track delivery under net mode.", fnt=F_SMALL_IT)
    if r >= ND_HDR:
        raise AssertionError(
            f"Net Delivery band now ends at row {r} but the summary header sits at "
            f"{ND_HDR}: raise ND_HDR / ND_FIRST (the band is fixed height, so the "
            "shift below it is a constant).")
    r += 2

    # ---- chart: log-share stacked delivery + delivered line ----
    # Staging rows go to the BOTTOM of the sheet — where the band used to sit —
    # while the chart itself is anchored beside the band at the top. A chart
    # does not care where its source cells live (D74).
    cs = g2 + 3 + nd + 4
    put(ws, f"A{cs - 1}", "Chart data (formulas — do not edit)", fnt=F_SMALL_IT)
    for j, h in enumerate(["Month", "Rate share", "Pricing share", "Delivered @ proj mods",
                           "Target"]):
        put(ws, f"{col(1 + j)}{cs}", h, fnt=font(GREY_DARK, size=9))
    for j in range(12):
        rr_ = cs + 1 + j
        src = tb + 27 + j
        formula(ws, f"A{rr_}", f'=TEXT(DATE(nr_PlanYear,{j + 1},1),"mmm")')
        formula(ws, f"B{rr_}",
                f"=IF(OR('{NC}'!$AG${src}=\"\",'{NC}'!$E${tb}=0),\"\","
                f"'{NC}'!$E${tb}*LN(1+'{NC}'!$AG${src})/LN(1+'{NC}'!$E${tb}))", fmt=PCT_S)
        formula(ws, f"C{rr_}",
                f"=IF(OR('{NC}'!$AH${src}=\"\",'{NC}'!$E${tb}=0),\"\","
                f"'{NC}'!$E${tb}*LN(1+'{NC}'!$AH${src})/LN(1+'{NC}'!$E${tb}))", fmt=PCT_S)
        formula(ws, f"D{rr_}",
                f"=IF('{NC}'!$AG${src}=\"\",\"\",(1+'{NC}'!$AG${src})"
                f"*IF('{NC}'!$F${tb}=1,'{NC}'!$O${src}/'{NC}'!$O${src - 12},1)-1)",
                fmt=PCT_S)
        formula(ws, f"E{rr_}", f"=IF('{NC}'!$D${tb}=0,\"\",'{NC}'!$E${tb})", fmt=PCT_S)
        for cL in "ABCDE":
            ws[f"{cL}{rr_}"].font = font(GREY_DARK, size=8)
    stack = BarChart()
    stack.type = "col"
    stack.grouping = "stacked"
    stack.overlap = 100
    stack.gapWidth = 45
    stack.add_data(Reference(ws, min_col=2, min_row=cs, max_col=3, max_row=cs + 12),
                   titles_from_data=True)
    stack.set_categories(Reference(ws, min_col=1, min_row=cs + 1, max_row=cs + 12))
    stack.series[0].graphicalProperties.solidFill = STEEL
    stack.series[1].graphicalProperties.solidFill = "F4B8B8"
    lines = LineChart()
    lines.add_data(Reference(ws, min_col=4, min_row=cs, max_col=5, max_row=cs + 12),
                   titles_from_data=True)
    lines.set_categories(Reference(ws, min_col=1, min_row=cs + 1, max_row=cs + 12))
    for s, rgb, dash in ((lines.series[0], NAVY, False), (lines.series[1], "C00000", True)):
        s.graphicalProperties.line.solidFill = rgb
        s.graphicalProperties.line.width = int(2.0 * 12700)
        if dash:
            s.graphicalProperties.line.dashStyle = "dash"
        s.marker = Marker(symbol="none")
        s.smooth = False
    stack += lines
    stack.title = ("How the month gets to the target: rate share + pricing share "
                   "(log-scaled to sum to x)")
    stack.style = 2
    stack.height = 8.5
    stack.width = 16
    stack.visible_cells_only = False
    stack.x_axis.delete = False
    stack.y_axis.delete = False
    stack.y_axis.number_format = "0.0%"
    # after the += merge, so it lands on the COMBINED chart legend (D69)
    chart_legend(stack)
    # right of the band's month columns (N..Y), level with its pickers, so the
    # tab opens on the picture and the numbers that produced it side by side
    ws.add_chart(stack, f"{col(ND_FLAG_COL + 2)}{ND_BAND_TOP + 1}")

    # one uniform width across both year bands: the summary's columns and the
    # month columns share them, and 10 fits every value on this tab
    set_widths(ws, {"A": 26})
    for j in range(24):
        ws.column_dimensions[col(ND_PRIOR_COL + j)].width = 10
    set_widths(ws, {col(ND_FLAG_COL): 24})
    # column A only: the tab now has three stacked sections with different
    # headers, and pinning one section's header while you read another is worse
    # than pinning none — but the row labels are wanted throughout (D74)
    presentation_setup(ws, gridlines_off=True, freeze="B1", tab_color=NAVY)
    print_setup(ws)
