"""Sheet builders: _lists (validation), Inputs (the §4 tables), Control.

Dimensioning (D37): each workbook covers ONE line of business; input rows are
BU x STATE combinations and the selector key is "BU|State". The policy term is
a single workbook-level input; tbl_Seasonality is keyed by state.
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter as col_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

from .build_workbook import Ctx, Layout as L
from .xlstyle import (
    ALIGN_C, ALIGN_L, ALIGN_WRAP, BORDER_THIN, F_HEADER, F_LABEL, F_SMALL, F_SMALL_IT,
    F_SUB, FAIL_RED, FILL_GREY, FILL_NAVY, FILL_PANEL, FILL_PANEL_2, FILL_RED,
    FILL_GREEN, FMT_DATE, FMT_EP, FMT_GEN, FMT_IDX, FMT_INT, FMT_MOD, FMT_PCT,
    FMT_PCT_SIGNED, FMT_PTS_SIGNED, GREY_DARK, NAVY, PASS_GREEN, STEEL, WARN_AMBER,
    add_dv, font, formula, header_row, input_cell, jump, label, link, nav_bar, note,
    presentation_setup, print_setup, put, rng, section, set_widths,
    status_banner_cf, title,
)

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False,
)


def _carried(ctx, attr):
    """A carried-forward setting, or None on a seeded build (D82)."""
    return getattr(ctx.carried, attr, None) if ctx.carried is not None else None


_dv = add_dv          # the helper started here; it now serves every entry sheet


# ---------------------------------------------------------------------------
# _lists
# ---------------------------------------------------------------------------


def build_lists(ctx: Ctx):
    """Validation lists. BU and state lists are DYNAMIC (D42): they extract the
    unique values actually present in tbl_LR via classic first-occurrence
    helpers, so renaming or adding a BU/state in tbl_LR flows straight into
    every dropdown, the Control selectors, and the State Summary — no
    regeneration needed. The remaining lists are fixed enumerations."""
    ws = ctx.wb["_lists"]
    n = L.LR_ROWS
    first, last = 3, 3 + n - 1
    label(ws, "A1",
          "Validation lists (hidden). Columns A/B/I are LIVE unique values from tbl_LR; "
          "helper machinery in columns M:P (D42).", bold=True)

    # ---- static enumerations ----
    # Each entry pins its OWN letter, so E and H simply stand empty now that the
    # LR-basis list and its KPI display toggle are gone (D107). Re-lettering to
    # close the gap would move four surviving lists for cosmetics nobody can
    # see — this sheet is hidden.
    static_cols = {
        "C": ("lst_status", "Rate change status", ["taken", "planned"]),
        "D": ("lst_yn", "Yes/No flags", ["Y", "N"]),
        "F": ("lst_onoff", "Toggles", ["ON", "OFF"]),
        "G": ("lst_scenario", "Scenario selector", ["Base", "S1", "S2", "S3", "S4"]),
    }
    for letter, (name, desc, values) in static_cols.items():
        label(ws, f"{letter}2", desc, bold=True)
        for i, v in enumerate(values):
            put(ws, f"{letter}{3 + i}", v, fnt=F_LABEL)
        ctx.define(name, "_lists", f"${letter}$3:${letter}${2 + len(values)}", desc)

    # ---- first-occurrence helpers (aligned with tbl_LR rows) ----
    label(ws, "M2", "BU 1st?", bold=True)
    label(ws, "N2", "BU rank", bold=True)
    label(ws, "O2", "State 1st?", bold=True)
    label(ws, "P2", "State rank", bold=True)
    for k in range(1, n + 1):
        r = first + k - 1
        ar = L.LR_FIRST + k - 1
        if k == 1:
            formula(ws, f"M{r}", f'=IF(Inputs!$A${ar}="",0,1)')
            formula(ws, f"O{r}", f'=IF(Inputs!$B${ar}="",0,1)')
        else:
            formula(ws, f"M{r}",
                    f'=IF(Inputs!$A${ar}="",0,'
                    f"IF(COUNTIF(Inputs!$A${L.LR_FIRST}:$A${ar - 1},Inputs!$A${ar})=0,1,0))")
            formula(ws, f"O{r}",
                    f'=IF(Inputs!$B${ar}="",0,'
                    f"IF(COUNTIF(Inputs!$B${L.LR_FIRST}:$B${ar - 1},Inputs!$B${ar})=0,1,0))")
        formula(ws, f"N{r}", f"=IF($M{r}=0,0,SUM($M${first}:$M{r}))")
        formula(ws, f"P{r}", f"=IF($O{r}=0,0,SUM($O${first}:$O{r}))")
        for cL in "MNOP":
            ws[f"{cL}{r}"].font = F_LABEL
    label(ws, "M1", "count:", bold=True)
    formula(ws, f"N1", f"=SUM($M${first}:$M${last})")
    formula(ws, f"P1", f"=SUM($O${first}:$O${last})")

    # ---- live unique lists ----
    label(ws, "A2", "Business units (live)", bold=True)
    label(ws, "B2", "States (live)", bold=True)
    label(ws, "I2", "BU filter (live)", bold=True)
    put(ws, f"I{first}", "All", fnt=F_LABEL)
    for k in range(1, n + 1):
        r = first + k - 1
        formula(ws, f"A{r}",
                f'=IF({k}>$N$1,"",INDEX(lr_bu,MATCH({k},$N${first}:$N${last},0)))')
        formula(ws, f"B{r}",
                f'=IF({k}>$P$1,"",INDEX(lr_state,MATCH({k},$P${first}:$P${last},0)))')
        formula(ws, f"I{r + 1}", f'=IF($A{r}="","",$A{r})')
        for cL in "ABI":
            ws[f"{cL}{r}"].font = F_LABEL

    # Self-sizing ranges via the non-volatile INDEX form (no OFFSET).
    ctx.define("lst_bu", "_lists",
               f"$A${first}:INDEX(_lists!$A${first}:$A${last},MAX(1,_lists!$N$1))",
               "LIVE unique business units from tbl_LR (self-sizing, D42)")
    ctx.define("lst_state", "_lists",
               f"$B${first}:INDEX(_lists!$B${first}:$B${last},MAX(1,_lists!$P$1))",
               "LIVE unique states from tbl_LR (self-sizing, D42)")
    ctx.define("lst_bu_all", "_lists",
               f"$I${first}:INDEX(_lists!$I${first}:$I${last + 1},MAX(1,_lists!$N$1+1))",
               "State Summary BU filter: 'All' + live business units (D42)")
    ctx.define("lst_state_cnt", "_lists", "$P$1",
               "Count of distinct states currently in tbl_LR")
    set_widths(ws, {c: 18 for c in "ABCDEFGHI"})
    set_widths(ws, {c: 9 for c in "MNOP"})


# ---------------------------------------------------------------------------
# Inputs
# ---------------------------------------------------------------------------

# Paste-friendly layout (D40): input columns are CONTIGUOUS — the key and all
# engine helper columns live to the RIGHT of each dataset, never in between.
# Labels are plain-English first with the technical symbol in parentheses
# (D43) so they still cross-reference Methodology and the named ranges.


@dataclass(frozen=True)
class LrCol:
    """One tbl_LR column, declared once.

    The header, width, number format, sample seed, required flag, data
    validation and defined name all hang off this single ordered list, so a
    column can be moved by moving one entry — nothing can drift out of step
    with the letter it lives at. (Before D72 these were six parallel
    structures keyed by position; reordering meant editing all six in lockstep
    and a mismatch showed up as a validation quietly guarding the wrong
    column.)
    """

    name: str                 # defined name, e.g. "lr_mind"
    header: str
    key: str | None           # sample-row dict key; None = not seeded
    fmt: str
    width: int
    desc: str
    required: bool = False
    dv: str = ""              # data-validation group tag (see LR_DV)


# Four blocks, left to right (D107): WHO the row is and how big it is; what the
# INDICATION said; where the MODS are going; and the handful of plan-year
# levers. Premium leads because it is the first thing anyone looks up and the
# weight behind every aggregate — it used to sit at column N, past six mods.
#
# The indication block is mostly CARRY-THROUGH: prospective trends, the expense
# ratio, ALAE/ULAE, the combined ratio and the two loads are recorded so the
# workbook holds the whole indication, and no engine formula reads any of them
# (the D96 contract, now eight columns wider). Target LR is the one they relate
# to arithmetically — target = (combined - expense) / (ALAE x ULAE) — and the
# Checks sheet says so ADVISORY, because which convention a line uses is not
# something this workbook should overrule.
#
# The mod block runs CHRONOLOGICALLY — M_prior, M_0 (with its as-of date),
# M_endPrior, M_1, M_2 — behind M_ind, which leads because it is the
# indication's assumption rather than a point on the projected path.
LR_COLS: tuple[LrCol, ...] = (
    # ---- who, and how big ------------------------------------------------
    LrCol("lr_bu", "BU", "bu", FMT_GEN, 9, "tbl_LR business unit", True, "bu"),
    LrCol("lr_state", "State", "state", FMT_GEN, 8, "tbl_LR state", True, "state"),
    LrCol("lr_ep", "Adj plan EP (000s)", "ep", FMT_EP, 12,
          "ADJUSTED plan earned premium (000s) — the weight behind Portfolio "
          "totals and State Summary aggregates", False, "ep"),
    # ---- the indication --------------------------------------------------
    LrCol("lr_lrproj", "Projected loss ratio", "lr_proj", FMT_PCT, 11,
          "Projected loss ratio from the indication, AT CURRENT RATE LEVEL "
          "(convert a proposed-level pick before entering it, D107)", True, "lr"),
    LrCol("lr_premtrend", "Prospective premium trend", "prem_trend", FMT_PCT, 12,
          "Prospective annual premium trend from the indication. Carry-through: "
          "no engine formula reads it (D107)", False, "pct"),
    LrCol("lr_losstrend", "Prospective loss trend", "loss_trend", FMT_PCT, 12,
          "Prospective annual loss trend from the indication. Carry-through: "
          "no engine formula reads it (D107)", False, "pct"),
    LrCol("lr_expense", "Expense ratio", "expense", FMT_PCT, 11,
          "Expense ratio underlying the target. Carry-through: no engine formula "
          "reads it, but the target-LR advisory on Checks uses it (D107)",
          False, "ratio"),
    LrCol("lr_alae", "ALAE factor", "alae", FMT_IDX, 10,
          "ALAE load as a factor on losses. Carry-through: no engine formula "
          "reads it, but the target-LR advisory on Checks uses it (D107)",
          False, "factor"),
    LrCol("lr_ulae", "ULAE factor", "ulae", FMT_IDX, 10,
          "ULAE load as a factor on losses. Carry-through: no engine formula "
          "reads it, but the target-LR advisory on Checks uses it (D107)",
          False, "factor"),
    LrCol("lr_combined", "Combined ratio", "combined", FMT_PCT, 11,
          "Target combined ratio from the indication. Carry-through: no engine "
          "formula reads it, but the target-LR advisory on Checks uses it (D107)",
          False, "ratio"),
    # Reference only, by design (D96): the loss ratio that carries the profit
    # provision. Nothing computes from it — it is the benchmark you read the
    # plan LR against, and the line the loss-ratio charts plot.
    LrCol("lr_target", "Target loss ratio (profit provision)", "target", FMT_PCT, 12,
          "Target loss ratio — the LR that earns the profit provision. Reference "
          "only: no engine formula reads it; the exhibits show the gap and the "
          "charts plot it. Checks compares it to (combined - expense) / "
          "(ALAE x ULAE) where all four are entered", False, "lr"),
    LrCol("lr_catload", "Cat load", "cat_load", FMT_PCT, 10,
          "Catastrophe load carried in the indication. Carry-through: no engine "
          "formula reads it (D107)", False, "pct"),
    LrCol("lr_largeload", "Large loss load", "large_load", FMT_PCT, 11,
          "Large-loss load carried in the indication. Carry-through: no engine "
          "formula reads it (D107)", False, "pct"),
    # ---- where the mods are going ----------------------------------------
    LrCol("lr_mind", "Mod assumed in indication (M_ind)", "m_ind", FMT_MOD, 13,
          "M_ind: avg schedule mod assumed in the indication", True, "mod"),
    LrCol("lr_mprior", "Mod ~1 yr before as-of (M_prior, opt)", "m_prior", FMT_MOD, 13,
          "M_prior: avg mod ~12 months before M_0 (optional)", False, "mod"),
    LrCol("lr_m0", "Current avg written mod (M_0)", "m0", FMT_MOD, 12,
          "M_0: current avg written mod", True, "mod"),
    LrCol("lr_m0asof", "Current mod as-of date", "m0_asof", FMT_DATE, 11,
          "M_0 as-of date (close of day)", True, "date"),
    LrCol("lr_mendprior", "Projected mod, end of CURRENT yr (M_endPrior)",
          "m_end_prior", FMT_MOD, 13,
          "M_endPrior: projected avg written mod at 12/31/(P-1) — the base the "
          "Mod Log's stepped actions compound on (D70)", False, "mod"),
    LrCol("lr_m1", "Projected mod, end of plan yr (M_1)", "m1", FMT_MOD, 12,
          "M_1: projected avg written mod at 12/31/P", True, "mod"),
    LrCol("lr_m2", "Projected mod, end plan yr+1 (M_2, opt)", "m2", FMT_MOD, 12,
          "M_2: projected avg written mod at 12/31/(P+1); blank = M_1", False, "mod"),
    # ---- plan-year levers ------------------------------------------------
    LrCol("lr_trend", "Net trend, plan yr+1 (opt)", "trend", FMT_PCT, 11,
          "Annual net loss-over-premium trend for the P+1 view", False, "pct"),
    LrCol("lr_aother", "Other adj factor (A_other)", "a_other", FMT_IDX, 10,
          "A_other manual adjustment factor", False, "aother"),
    LrCol("lr_modadj", "Apply mod adjustment?", "modadj", FMT_GEN, 11,
          "Per-combo mod adjustment toggle (ON/OFF)", True, "onoff"),
    LrCol("lr_netp", "Net rate selection, plan yr (opt)", "netp", FMT_PCT, 12,
          "OPTIONAL net rate selection for P: YoY combined rate x mod target "
          "from 1/1/P (blank = explicit program, D39)", False, "net"),
    LrCol("lr_netp1", "Net rate selection, plan yr+1 (opt)", "netp1", FMT_PCT, 12,
          "OPTIONAL net selection for P+1 (blank = carry the P selection)", False, "net"),
)

# data-validation groups, applied to every column carrying the tag
LR_DV: dict[str, dict] = {
    # tbl_LR defines the roster, so its own BU/State dropdowns are
    # SUGGESTION-ONLY (blocking=False): pick an existing value or type a new
    # one — every list downstream then follows automatically (D42).
    "bu": dict(kind="list", blocking=False, formula1="=lst_bu"),
    "state": dict(kind="list", blocking=False, formula1="=lst_state"),
    "onoff": dict(kind="list", formula1="=lst_onoff"),
    # Ratios are entered as decimal fractions like every other percent here, so
    # the bound is generous enough for a combined ratio well over 100% while
    # still catching the classic paste of 65 for 65%.
    "ratio": dict(kind="decimal", operator="between", formula1="0", formula2="3",
                  error="Enter a ratio as a decimal fraction (0 to 300%)."),
    "factor": dict(kind="decimal", operator="between", formula1="0.5", formula2="2",
                   error="ALAE/ULAE factors are multipliers on losses and must "
                         "lie in [0.5, 2.0] — 1.05 means a 5% load."),
    "lr": dict(kind="decimal", operator="between", formula1="0", formula2="3",
               error="Loss ratio must be between 0 and 300% (entered as a fraction)."),
    "pct": dict(kind="decimal", operator="between", formula1="-0.5", formula2="1",
                error="Enter a decimal fraction between -50% and +100%."),
    "mod": dict(kind="decimal", operator="between", formula1="0.5", formula2="1.5",
                error="Schedule mods must lie in [0.5, 1.5]."),
    "ep": dict(kind="decimal", operator="greaterThanOrEqual", formula1="0",
               error="Plan EP must be non-negative."),
    "aother": dict(kind="decimal", operator="between", formula1="0.5", formula2="2",
                   error="A_other must lie in [0.5, 2.0]."),
    "date": dict(kind="date", operator="between", formula1="DATE(1990,1,1)",
                 formula2="DATE(2100,12,31)",
                 error="Enter a real as-of date (it must precede 12/31 of the plan "
                       "year; the Checks sheet enforces that bound)."),
    "net": dict(kind="decimal", operator="between", formula1="-0.5", formula2="1",
                error="Net rate selection must lie in [-50%, +100%] (decimal "
                      "fraction; blank = off)."),
}

# All three are DERIVED from the length of LR_COLS and none is ever typed, so
# adding or removing a column moves them together. No letter is named here on
# purpose: the old comments said "S today" / "U today" / "W today" and were a
# column behind reality from the day D96 widened the block.
LR_LAST_COL = col_letter(len(LR_COLS))
# A BLANK column separates the paste block from the machinery to its right, and
# another separates the Key helper from the live results. Without the first one
# the block and the helper are contiguous: Ctrl+Shift+Right runs straight past
# the edge of what you may edit, and a paste one column too wide lands on the
# key formulas instead of in dead space. The gap is the fence.
LR_KEY_COL = col_letter(len(LR_COLS) + 2)       # BU|State helper
LR_ECHO_COL = len(LR_COLS) + 4                  # 1-based — live results


def lr_letter(name: str) -> str:
    """Column letter of a tbl_LR column, by defined name."""
    return col_letter(next(i for i, c in enumerate(LR_COLS, 1) if c.name == name))


def lr_headers() -> list[str]:
    return [c.header for c in LR_COLS]


def rl_headers() -> list[str]:
    return ["BU", "State", "Effective date", "Filed chg %", "Status (taken / planned)",
            "In indication? (Y/N)", "Achievement % (planned)", "Comment"]


RL_HELPER_HEADERS = ["Key", "r_eff", "ln(1+r)", "Eff month", "Days on/after", "Earlier cnt",
                     "Same cnt", "First?", "Dup month?", "Seq"]


def build_inputs(ctx: Ctx):
    ws = ctx.wb["Inputs"]
    p = ctx.p
    title(ws, "A1", f"Inputs — {ctx.lob.name}",
          "Loss-ratio and mod inputs (one row per BU x state), seasonality, and the policy "
          "term — blue font = enter values; yellow fill = required. Rate changes live on "
          "the Rate Log sheet. Structure is fixed; do not insert or delete rows.")
    nav_bar(ws, 3, 1, ["Control", "Rate Log", "Mod Log", "State Summary", "Checks",
                       "Read Me"], step=2)
    # A carried build holds the user's real book, so the SAMPLE banner would be
    # a lie — and a banner that lies is how a real row gets deleted (D82).
    put(ws, "A4",
        (f"CARRIED FORWARD from {ctx.carried.source.name}: these are YOUR inputs, "
         f"re-laid into a freshly built workbook. Oracle-tie rows on Checks are baked "
         f"for {ctx.we_key}." if ctx.carried is not None else
         "SAMPLE DATA: every populated row below is illustrative and should be replaced "
         f"with your book. {ctx.we_key} carries the documented worked example used by "
         "the Checks sheet."),
        fnt=font(FAIL_RED, bold=True, italic=True))
    # anchored on the key helper, not a literal letter: the note describes the
    # block from OUTSIDE it, so it has to move when the block widens
    put(ws, f"{LR_KEY_COL}4",
        f"tbl_LR A:{LR_LAST_COL} is one contiguous paste block — keys, helpers, and live "
        "results sit to the right. The rate change log lives on the Rate Log sheet.",
        fnt=F_SMALL_IT)

    # ------------------------------------------------------------------ tbl_LR
    section(ws, L.LR_HDR - 1, "A",
            "tbl_LR — one row per BU x state (projected LR and mod inputs). "
            f"Columns A:{LR_LAST_COL} are one contiguous paste block.")
    header_row(ws, L.LR_HDR, 1, lr_headers(), widths=[c.width for c in LR_COLS])
    ws.row_dimensions[L.LR_HDR].height = 44
    K = LR_KEY_COL
    put(ws, f"{K}{L.LR_HDR}", "Key", fnt=font(GREY_DARK, bold=True, size=9), fill=FILL_GREY)
    put(ws, f"{K}{L.LR_HDR - 1}", "engine helper — do not edit", fnt=F_SMALL_IT)
    for i in range(L.LR_ROWS):
        r = L.LR_FIRST + i
        row = ctx.lr_rows[i] if i < len(ctx.lr_rows) else None
        for c, spec in enumerate(LR_COLS, 1):
            v = row.get(spec.key) if (row and spec.key) else None
            input_cell(ws, ws.cell(row=r, column=c).coordinate, v,
                       fmt=spec.fmt, required=spec.required)
        formula(ws, f"{K}{r}", f'=IF(OR($A{r}="",$B{r}=""),"",$A{r}&"|"&$B{r})',
                fmt=FMT_GEN, border=BORDER_THIN)
        ws[f"{K}{r}"].font = font(GREY_DARK, size=9)
    ws[f"{lr_letter('lr_modadj')}{L.LR_HDR}"].comment = Comment(
        "Per-combo mod adjustment toggle. Set OFF if the indication's premium trend already "
        "reflects schedule mod drift — otherwise the drift is double-counted. The Control "
        "sheet master toggle must also be ON for the adjustment to apply.", "generator")
    ws[f"{lr_letter('lr_netp')}{L.LR_HDR}"].comment = Comment(
        "OPTIONAL net rate selection (DECISIONS.md D39). When set, cohorts written on/after "
        "1/1/P abandon the planned rate program and the projected mod path: each cohort's "
        "combined net price (rate x mod) renews this % above its year-ago cohort. History "
        "before 1/1/P still earns exactly as modeled. Blank = explicit program (classic). "
        "The P+1 column defaults to the P selection when blank.", "generator")
    ws[f"{lr_letter('lr_mendprior')}{L.LR_HDR}"].comment = Comment(
        "Where you expect the average written mod to land at 12/31 of the CURRENT year. "
        "Inert until the combo has actions on the Mod Log; from the first action it becomes "
        "the level those actions compound on, and the drift path lands here instead of "
        "running on to M_1 (DECISIONS.md D70).", "generator")
    tbl = Table(displayName="tbl_LR", ref=f"A{L.LR_HDR}:{LR_LAST_COL}{L.LR_LAST}")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)
    for c, spec in enumerate(LR_COLS, 1):
        ctx.define(spec.name, "Inputs",
                   f"${col_letter(c)}${L.LR_FIRST}:${col_letter(c)}${L.LR_LAST}", spec.desc)
    ctx.define("lr_key", "Inputs", f"${K}${L.LR_FIRST}:${K}${L.LR_LAST}",
               "Helper: concatenated BU|State key (right of the paste block, D40)")

    # ---- live per-row results, right of the Key helper -----------------------
    # The hidden _calc results table aligns 1:1 with tbl_LR, so editing any
    # input on row n moves that same row's plan LR on the same screen. Every
    # column here is DERIVED from LR_KEY_COL: adding a tbl_LR column has to move
    # the whole right-hand block, and a hard-coded letter is how the blank test
    # ($T) ends up pointing at a data column instead of the key.
    echo = [col_letter(LR_ECHO_COL + j) for j in range(4)]
    put(ws, f"{echo[0]}{L.LR_HDR - 1}", "live results (formulas — do not edit)",
        fnt=F_SMALL_IT)
    header_row(ws, L.LR_HDR, LR_ECHO_COL,
               ["This combo's CY plan LR", "Chg vs projected (pts)",
                "Rate earn-in (A_rate)", "Mod drift (A_mod)"],
               widths=[13, 12, 11, 11], fill=FILL_GREY,
               fnt=font(GREY_DARK, bold=True, size=9))
    for i in range(L.LR_ROWS):
        r = L.LR_FIRST + i
        n = i + 1
        blank = f'${K}{r}=""'
        for cL, f, fmt in (
                (echo[0], f"INDEX(calc_cylr_p,{n})", FMT_PCT),
                (echo[1], f"(INDEX(calc_cylr_p,{n})-INDEX(calc_lrcur,{n}))*100",
                 FMT_PTS_SIGNED),
                (echo[2], f"INDEX(calc_arate_p,{n})", FMT_IDX),
                (echo[3], f"INDEX(calc_amod_p,{n})", FMT_IDX)):
            formula(ws, f"{cL}{r}", f'=IF({blank},"",{f})', fmt=fmt)
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=9)
            ws[f"{cL}{r}"].alignment = ALIGN_C

    # tbl_RateLog moved to its own "Rate Log" sheet (sheets_ratelog.build_rate_log):
    # identical paste block A:H, its own frozen header, and same-row live results.

    # --------------------------------------------------------- tbl_Seasonality
    section(ws, L.SE_HDR - 1, "A",
            "tbl_Seasonality — optional per-STATE monthly written weights (relative; blank row = uniform)")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    header_row(ws, L.SE_HDR, 1, ["State"] + months + ["Row sum"],
               widths=[9] + [7] * 12 + [9])
    for i in range(L.SE_ROWS):
        r = L.SE_FIRST + i
        row = ctx.season_rows[i] if i < len(ctx.season_rows) else None
        input_cell(ws, f"A{r}", row and row["state"], required=False)
        for m in range(12):
            input_cell(ws, ws.cell(row=r, column=2 + m).coordinate,
                       (row["weights"][m] if row else None), fmt="0.00", required=False)
        formula(ws, f"N{r}", f'=IF($A{r}="","",SUM($B{r}:$M{r}))', fmt="0.00")
        ws[f"N{r}"].font = font(GREY_DARK, size=9)
    tbl = Table(displayName="tbl_Seasonality", ref=f"A{L.SE_HDR}:M{L.SE_LAST}")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)
    ctx.define("se_state", "Inputs", f"$A${L.SE_FIRST}:$A${L.SE_LAST}", "Seasonality state column")
    ctx.define("se_block", "Inputs", f"$B${L.SE_FIRST}:$M${L.SE_LAST}", "Seasonality 12-month weight block")
    ctx.define("se_sum", "Inputs", f"$N${L.SE_FIRST}:$N${L.SE_LAST}", "Helper: seasonality row sums")

    # ---------------------------------------------------- workbook parameters
    section(ws, L.WB_HDR - 1, "A", "Workbook parameters")
    label(ws, f"A{L.WB_HDR}", "Line of business")
    put(ws, f"B{L.WB_HDR}", ctx.lob.name, fnt=font(NAVY, bold=True))
    label(ws, f"A{L.TERM_ROW}", "Policy term (months)")
    input_cell(ws, f"B{L.TERM_ROW}", ctx.lob.term_months, fmt=FMT_INT)
    ctx.define("nr_TermMonths", "Inputs", f"$B${L.TERM_ROW}",
               "Workbook policy term in months (this workbook covers one LOB)")
    put(ws, f"C{L.TERM_ROW}",
        "Applies to every BU x state in this workbook (one workbook per LOB, D37).",
        fnt=F_SMALL_IT)

    # ------------------------------------------------------------- validations
    # DV sits ON the input cells only (a paste passes straight over it); the
    # paste blocks contain no formula columns (D40).
    fr, lr_ = L.LR_FIRST, L.LR_LAST
    # one validation per LR_DV tag, over every column carrying it — so a moved
    # column takes its own rule with it and cannot end up guarding a neighbour
    for tag, spec in LR_DV.items():
        ranges = [f"{col_letter(i)}{fr}:{col_letter(i)}{lr_}"
                  for i, c in enumerate(LR_COLS, 1) if c.dv == tag]
        if not ranges:
            continue
        kw = dict(spec)
        _dv(ws, kw.pop("kind"), ranges, **kw)
    _dv(ws, "decimal", [rng(2, L.SE_FIRST, 13, L.SE_LAST)], operator="between",
        formula1="0", formula2="100", error="Seasonality weights are non-negative.")
    _dv(ws, "list", [f"A{L.SE_FIRST}:A{L.SE_LAST}"], formula1="=lst_state")
    _dv(ws, "whole", [f"B{L.TERM_ROW}"], operator="between",
        formula1="1", formula2="12", error="Term must be 1..12 months.")

    # freeze BU + State columns as well as the header rows, so a row's identity
    # stays pinned while editing the right-hand inputs and live results
    presentation_setup(ws, gridlines_off=False, freeze=f"C{L.LR_FIRST}", tab_color=WARN_AMBER)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Control
# ---------------------------------------------------------------------------


def build_control(ctx: Ctx):
    ws = ctx.wb["Control"]
    p = ctx.p
    title(ws, "B2", f"Calendar-Year Plan Loss Ratio — {ctx.lob.name}",
          f"Plan year {p} | one workbook per LOB; combos are BU x state | generated by "
          f"src/build_workbook.py v{ctx.cfg.version} | classic formula mode")

    # One-click navigation back to the analysis sheets (D45): change the
    # selection below, then jump straight to the sheet you came from.
    label(ws, "B4", "Jump to:", bold=True)
    for i, sheet in enumerate(["Inputs", "Rate Log", "Mod Log", "Portfolio", "State Summary",
                               "Program Flow", "Net Delivery", "Bridge", "Walkthrough",
                               "One-Pager", "Flow Dashboard", "Scenarios", "Compare",
                               "Solver", "Attribution", "Rate Engine", "Mod Engine",
                               "Checks", "Methodology", "Read Me"]):
        target = f"'{sheet}'!A1" if " " in sheet else f"{sheet}!A1"
        jump(ws, ws.cell(row=4, column=3 + i).coordinate, target, sheet, size=9)

    # Selectors
    section(ws, 5, "B", "Selection")
    label(ws, "B6", "Plan year (P)")
    input_cell(ws, "C6", p, fmt="0")
    label(ws, "B7", "Business unit")
    input_cell(ws, "C7", _carried(ctx, "sel_bu") or ctx.we_row["bu"])
    label(ws, "B8", "State")
    input_cell(ws, "C8", _carried(ctx, "sel_state") or ctx.we_row["state"])
    label(ws, "B9", "Scenario in focus")
    input_cell(ws, "C9", "Base", required=False)
    label(ws, "E6", "Selected key")
    formula(ws, "F6", '=$C$7&"|"&$C$8')
    label(ws, "E7", "Combo found in tbl_LR?")
    formula(ws, "F7", '=IF($G$7,"yes","NO — add the row to tbl_LR")')
    formula(ws, "G7", "=COUNTIF(lr_key,$F$6)>0")
    ws["G7"].font = font(GREY_DARK, size=8)
    label(ws, "E8", "Policy term (months)")
    link(ws, "F8", "=nr_TermMonths", fmt="0")
    label(ws, "E9", "Line of business")
    put(ws, "F9", ctx.lob.name, fnt=font(NAVY, bold=True))
    # stale-year banner (D44): calculations always follow the plan year above;
    # this flags that sample data / prose still describe the generated year
    formula(ws, "B10",
            f'=IF(nr_PlanYear={p},"","NOTE: plan year differs from the generated {p}. '
            'All CALCULATIONS follow the plan year entered above; sample data and some '
            'printed notes still describe the generated book. Regenerate for a clean copy.")')
    ws["B10"].font = font("BF8F00", bold=True, size=9)

    ctx.define("nr_PlanYear", "Control", "$C$6", "Plan year P")
    ctx.define("nr_SelBU", "Control", "$C$7", "Selected business unit")
    ctx.define("nr_SelState", "Control", "$C$8", "Selected state")
    ctx.define("nr_SelScenario", "Control", "$C$9", "Scenario surfaced in the Control KPI row")
    ctx.define("nr_SelKey", "Control", "$F$6", "Selected BU|State key")
    ctx.define("nr_SelOK", "Control", "$G$7", "TRUE when the selected combo exists in tbl_LR")

    # Toggles
    section(ws, 11, "B", "Global toggles")
    label(ws, "B12", "Seasonality (written weights)")
    input_cell(ws, "C12", _carried(ctx, "season_on") or "OFF")
    note(ws, "D12", "ON applies tbl_Seasonality weights for the selected STATE; OFF = uniform writings (classic parallelogram).")
    label(ws, "B13", "Mod adjustment (master)")
    input_cell(ws, "C13", _carried(ctx, "mod_master") or "ON")
    put(ws, "D13",
        "Set OFF if the indication's premium trend already reflects schedule mod drift — "
        "otherwise the drift is double-counted.",
        fnt=font(FAIL_RED, size=9, italic=True), align=ALIGN_WRAP)
    # Row 14 is deliberately EMPTY. It held the "Projected-LR KPI shows"
    # toggle (D16), which chose between the LR as entered and the LR at current
    # rate level — two names for the same number since D107 removed the basis.
    # The trend default stays at C15: the harness and the carry-forward test
    # both address it literally, and sliding it up to fill a cosmetic gap would
    # move a cell two files name by address.
    label(ws, "B15", f"Default net trend for CY {p + 1}")
    _td = _carried(ctx, "trend_default")
    input_cell(ws, "C15", 0.0 if _td is None else _td, fmt=FMT_PCT)
    note(ws, "D15",
         f"Annual net loss-over-premium trend applied once for the CY {p + 1} view wherever a "
         "combo's tbl_LR trend cell is blank; a per-state entry overrides this default.")
    ctx.define("nr_SeasonOn", "Control", "$C$12", "Global seasonality toggle (ON/OFF)")
    ctx.define("nr_ModAdjMaster", "Control", "$C$13", "Master mod-adjustment toggle (ON/OFF)")
    ctx.define("nr_TrendDefault", "Control", "$C$15",
               "Default net trend for the P+1 view (per-combo blank cells inherit it)")

    put(ws, "B16",
        '=IF(nr_SelOK,"","WARNING: the selected BU | state has no row in tbl_LR - engines show neutral factors until a row is added.")',
        fnt=font(FAIL_RED, bold=True))

    # KPI cards — plain-English captions; values blank to "—" on a bad
    # selection so a missing combo can never fabricate plausible numbers
    section(ws, 17, "B", "Key results — selected BU x state")
    ok = "nr_SelOK"
    cards = [
        ('="Projected loss ratio"', f'=IF({ok},nr_LRproj,"—")', FMT_PCT,
         "from the indication, at current rate level"),
        (f'="CY "&nr_PlanYear&" plan loss ratio"', f'=IF({ok},nr_CYLR_P,"—")', FMT_PCT,
         "projected LR x rate earn-in x mod drift x other"),
        ('="CY earned rate chg vs indication"', f'=IF({ok},nr_EChgVsInd,"—")',
         FMT_PCT_SIGNED,
         "how the year's earned rate compares with the level the indication assumed"),
        (f'="Carryover into "&(nr_PlanYear+1)', f'=IF({ok},nr_YoY_P1,"—")', FMT_PCT_SIGNED,
         "earned rate change already locked in for next year"),
        # The two mod numbers are always what the user typed, so they are always
        # shown — but on a net selection, or with either mod toggle off, they
        # are NOT driving the plan LR beside them, and a KPI that doesn't say so
        # reads as though the drift is in force. Disclose, don't blank (D46).
        ('="Written mod: current / projected"', None, FMT_GEN,
         '=IF(NOT(nr_SelOK),"current written mod  /  projected at plan-year end",'
         'IF(nr_ModAdjMaster="OFF","master toggle OFF — this drift is NOT applied",'
         'IF(nr_NetMode,"net selection: merged into the net factor, A_mod = 1.000",'
         'IF(nr_ModAdjRow="OFF","mod adjustment OFF for this combo — A_mod = 1.000",'
         '"current written mod  /  projected at plan-year end"))))'),
        ('="Checks status"', "=ck_overall", FMT_GEN, "see the Checks sheet"),
        # D65: book-level assertion vs logged program — "—" when nothing in the
        # book carries a net selection, so it never invents a comparison
        ('="Book: program vs asserted"',
         '=IF(OR(SUMPRODUCT(calc_netmode)=0,SUM(calc_ep)=0),"—",'
         "(SUM(calc_w_cylr_prog)-SUM(calc_w_cylr))/SUM(calc_ep)*100)",
         FMT_PTS_SIGNED,
         "plan LR if you book the logged program instead of the net target"),
    ]
    for i, (lbl_f, val_f, fmt, sub) in enumerate(cards):
        c1 = 2 + i * 2  # B, D, F, H, J, L
        put(ws, ws.cell(row=18, column=c1).coordinate, lbl_f, fnt=F_SMALL,
            fill=FILL_PANEL_2, align=ALIGN_L)
        put(ws, ws.cell(row=18, column=c1 + 1).coordinate, None, fill=FILL_PANEL_2)
        if val_f is None:  # the mod card: TWO numeric cells, chartable and formattable
            put(ws, ws.cell(row=19, column=c1).coordinate, f'=IF({ok},nr_M0,"—")',
                fnt=font("1F3864", bold=True, size=16), fmt=FMT_MOD, fill=FILL_PANEL,
                align=ALIGN_L)
            put(ws, ws.cell(row=19, column=c1 + 1).coordinate, f'=IF({ok},nr_M1,"—")',
                fnt=font("1F3864", bold=True, size=16), fmt=FMT_MOD, fill=FILL_PANEL,
                align=ALIGN_L)
        else:
            put(ws, ws.cell(row=19, column=c1).coordinate, val_f,
                fnt=font("1F3864", bold=True, size=16), fmt=fmt, fill=FILL_PANEL,
                align=ALIGN_L)
            put(ws, ws.cell(row=19, column=c1 + 1).coordinate, None, fill=FILL_PANEL)
        put(ws, ws.cell(row=20, column=c1).coordinate, sub, fnt=F_SMALL_IT, fill=FILL_PANEL)
        put(ws, ws.cell(row=20, column=c1 + 1).coordinate, None, fill=FILL_PANEL)
    status_banner_cf(ws, "L19", size=14)

    # Summary table P vs P+1
    section(ws, 22, "B", "Rate and price flow summary")
    put(ws, "B23", "Metric", fnt=F_HEADER, fill=FILL_NAVY)
    put(ws, "C23", '="CY "&nr_PlanYear', fnt=F_HEADER, fill=FILL_NAVY, align=ALIGN_C)
    put(ws, "D23", '="CY "&(nr_PlanYear+1)&" (indicative)"', fnt=F_HEADER, fill=FILL_NAVY, align=ALIGN_C)
    rows = [
        ("Earned rate level (E_CY)", "=nr_ECY_P", "=nr_ECY_P1", FMT_IDX),
        ("Rate earn-in factor (A_rate)", "=nr_Arate_P", "=nr_Arate_P1", FMT_IDX),
        ("Earned schedule mod", "=nr_MEarned_P", "=nr_MEarned_P1", FMT_MOD),
        ("Mod drift factor (A_mod)", "=nr_Amod_P", "=nr_Amod_P1", FMT_IDX),
        ("Earned rate chg (year over year)", "=nr_YoY_P", "=nr_YoY_P1", FMT_PCT_SIGNED),
        ("CY plan loss ratio", "=nr_CYLR_P", "=nr_CYLR_P1", FMT_PCT),
    ]
    for i, (lbl_t, fp_, fp1, fmt) in enumerate(rows):
        r = 24 + i
        label(ws, f"B{r}", lbl_t)
        link(ws, f"C{r}", fp_, fmt=fmt, align=ALIGN_C, bold=(i == len(rows) - 1))
        link(ws, f"D{r}", fp1, fmt=fmt, align=ALIGN_C, bold=(i == len(rows) - 1))
    put(ws, "B30",
        "Next-year column is indicative — see the canonical caveat on the Bridge.",
        fnt=F_SMALL_IT)

    # Compare block: side-by-side with ANY other combo, served from the
    # all-combo _calc results (a data feature — no second engine, no selector
    # split; every deep-dive tab still shows one agreed selection)
    section(ws, 22, "F", "Compare with another combo")
    label(ws, "F23", "This view")
    formula(ws, "G23", "=nr_SelKey", fmt=FMT_GEN)
    ws["G23"].font = font(NAVY, bold=True)
    label(ws, "F24", "Compare to: business unit")
    input_cell(ws, "G24", "", required=False)
    label(ws, "F25", "Compare to: state")
    input_cell(ws, "G25", "", required=False)
    formula(ws, "I24", '=IF(OR($G$24="",$G$25=""),"",$G$24&"|"&$G$25)')
    ws["I24"].font = font(GREY_DARK, size=8)
    formula(ws, "I25", '=IF($I$24="",0,IF(COUNTIF(calc_key,$I$24)=0,0,'
                       "MATCH($I$24,calc_key,0)))", fmt=FMT_INT)
    ws["I25"].font = font(GREY_DARK, size=8)
    cmp_metrics = [
        ("Projected LR (current level)", "nr_LRcur", "calc_lrcur", FMT_PCT),
        ('="CY "&nr_PlanYear&" plan LR"', "nr_CYLR_P", "calc_cylr_p", FMT_PCT),
        ("Rate earn-in (A_rate)", "nr_Arate_P", "calc_arate_p", FMT_IDX),
        ("Mod drift (A_mod)", "nr_Amod_P", "calc_amod_p", FMT_IDX),
        ("Earned rate level (E_CY)", "nr_ECY_P", "calc_ecy_p", FMT_IDX),
        ("Indication rate level (CRL_ind)", "nr_CRLind", "calc_crl", FMT_IDX),
        ('="Carryover into "&(nr_PlanYear+1)', "nr_YoY_P1", "calc_carry",
         FMT_PCT_SIGNED),
        ('="CY "&(nr_PlanYear+1)&" plan LR"', "nr_CYLR_P1", "calc_cylr_p1", FMT_PCT),
    ]
    put(ws, "G26", "selected", fnt=F_SMALL)
    put(ws, "H26", "compare", fnt=F_SMALL)
    put(ws, "I26", "difference", fnt=F_SMALL)
    for i, (lbl, sel_name, calc_name, fmt) in enumerate(cmp_metrics):
        r = 27 + i
        if lbl.startswith("="):
            formula(ws, f"F{r}", lbl)
            ws[f"F{r}"].font = F_LABEL
        else:
            label(ws, f"F{r}", lbl)
        link(ws, f"G{r}", f'=IF(nr_SelOK,{sel_name},"—")', fmt=fmt, align=ALIGN_C)
        formula(ws, f"H{r}", f'=IF($I$25=0,"—",INDEX({calc_name},$I$25))', fmt=fmt,
                align=ALIGN_C)
        formula(ws, f"I{r}",
                f'=IF(OR($I$25=0,NOT(nr_SelOK)),"",IF(AND(ISNUMBER($G{r}),ISNUMBER($H{r})),'
                f'$H{r}-$G{r},""))',
                fmt="+0.0000;-0.0000;0.0000", align=ALIGN_C)
        ws[f"I{r}"].font = font(GREY_DARK, size=9)
    put(ws, "F35", "Pick a BU and state to light this up; the compare column reads the "
                   "same hidden all-combo engine every exhibit uses.", fnt=F_SMALL_IT)
    _dv(ws, "list", ["G24"], blocking=False, formula1="=lst_bu")
    _dv(ws, "list", ["G25"], blocking=False, formula1="=lst_state")

    # Scenario spotlight
    section(ws, 32, "B", "Scenario spotlight")
    label(ws, "B33", "Scenario in focus")
    link(ws, "C33", "=nr_SelScenario", align=ALIGN_C)
    label(ws, "B34", "CY plan LR under scenario")
    link(ws, "C34", "=INDEX(sc_res_cylr,MATCH(nr_SelScenario,sc_res_names,0))", fmt=FMT_PCT,
         align=ALIGN_C, bold=True)
    label(ws, "B35", "Delta vs Base")
    link(ws, "C35", "=(INDEX(sc_res_cylr,MATCH(nr_SelScenario,sc_res_names,0))-nr_CYLR_P)*100",
         fmt=FMT_PTS_SIGNED, align=ALIGN_C)
    note(ws, "B37", "Scenario definitions live on the Scenarios sheet; the Bridge and engine "
                    "sheets always show Base inputs (see Methodology).")

    _dv(ws, "list", ["C7"], formula1="=lst_bu")
    _dv(ws, "list", ["C8"], formula1="=lst_state")
    _dv(ws, "list", ["C9"], formula1="=lst_scenario")
    _dv(ws, "list", ["C12", "C13"], formula1="=lst_onoff")
    _dv(ws, "whole", ["C6"], operator="between", formula1="2000", formula2="2100",
        error="Plan year must be a 4-digit year.")
    _dv(ws, "decimal", ["C15"], operator="between", formula1="-0.5", formula2="1",
        error="Enter a decimal fraction between -50% and +100%.")

    set_widths(ws, {"A": 2, "B": 30, "C": 16, "D": 16, "E": 22, "F": 16, "G": 14,
                    "H": 16, "I": 14, "J": 16, "K": 14, "L": 16, "M": 14})
    presentation_setup(ws, gridlines_off=True, tab_color=NAVY)
    print_setup(ws)
