"""Sheet builders: Flow Dashboard, _oracle, Checks, Methodology, Read Me."""

from __future__ import annotations

import datetime as dt

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule

from .build_workbook import GENERATOR_VERSION, Ctx, Layout as L
from .xlstyle import (
    ALIGN_C, ALIGN_WRAP, BAND_FILL, BORDER_THIN, F_HEADER, F_INPUT, F_LABEL, F_LINK,
    F_SMALL_IT, FAIL_RED, FILL_AMBER, FILL_GREEN, FILL_GREY, FILL_NAVY, FILL_PANEL,
    FILL_RED, FILL_REQ, FMT_DATE, FMT_GEN, FMT_IDX, FMT_INT, FMT_MOD, FMT_PCT,
    GREY_DARK, LINK_GREEN, NAVY, PASS_GREEN, STEEL, col, font, formula, header_row,
    jump, label, link, nav_bar, note, presentation_setup, print_setup, prose, put,
    quote_sheet as _q, section, set_widths, text_height, title,
)

RE = "'Rate Engine'"


def _excel_serial(d: dt.date) -> int:
    return (d - dt.date(1899, 12, 30)).days


def _line(series, rgb, width_pt=2.25, dashed=False):
    series.graphicalProperties.line.solidFill = rgb
    series.graphicalProperties.line.width = int(width_pt * 12700)
    if dashed:
        series.graphicalProperties.line.dashStyle = "dash"
    series.marker = Marker(symbol="none")
    series.smooth = False


def _chart(c, title_text, height=8, width=15, y_title=None):
    c.title = title_text
    c.style = 2
    c.height = height
    c.width = width
    if y_title:
        c.y_axis.title = y_title
    # Explicit axis visibility: openpyxl omits <c:delete> for None, and Excel
    # interprets the omission as deleted=1 on resave (DECISIONS.md D36).
    if c.x_axis.delete is None:
        c.x_axis.delete = False
    if c.y_axis.delete is None:
        c.y_axis.delete = False
    return c


# ---------------------------------------------------------------------------
# Flow Dashboard
# ---------------------------------------------------------------------------

CD0 = 80  # chart-data header row (monthly detail table)


def build_flow_dashboard(ctx: Ctx):
    ws = ctx.wb["Flow Dashboard"]
    p = ctx.p
    title(ws, "B2", "Flow Dashboard — rate and price flow through the plan",
          "Selected BU x state. The plan year is shaded in the index chart. All series "
          "trace to the Rate Engine and Mod Engine sheets via the monthly detail table below.")
    link(ws, "J2", "=nr_SelKey", bold=True)
    jump(ws, "L2", "Control!C7", "Change BU/state selection >")

    # ---- monthly detail table (36 rows; doubles as chart data) ----
    put(ws, f"B{CD0 - 1}", "Monthly flow detail (formulas — the charts read this table)",
        fnt=F_SMALL_IT)
    heads = ["Month", "Written idx W", "Earned idx E_m", "Written mod", "Earned mod",
             "Price idx (written)", "Price idx (earned)", "Plan year",
             "YoY earned rate", "YoY earned price", "Unearned runway"]
    for j, h in enumerate(heads):
        put(ws, f"{col(2 + j)}{CD0}", h, fnt=font(GREY_DARK, size=9))
    for j in range(L.N_MONTHS):
        r = CD0 + 1 + j
        mcol = col(L.RE_MATRIX_COL + j)
        coh = L.RE_COH_FIRST + 12 + j  # cohort row for this reported month
        link(ws, f"B{r}", f"={RE}!{mcol}${L.RE_COH_HDR}", fmt="mmm-yy")
        link(ws, f"C{r}", f"={RE}!$U${coh}", fmt=FMT_IDX)  # in-force index (D39)
        link(ws, f"D{r}", f"={RE}!{mcol}${L.RE_ROW_EM}", fmt=FMT_IDX)
        link(ws, f"E{r}", f"='Mod Engine'!$E${coh}", fmt=FMT_MOD)
        link(ws, f"F{r}", f"={RE}!{mcol}${L.RE_ROW_MODM}", fmt=FMT_MOD)
        # price index: the net path already combines rate x mod — never double-count
        formula(ws, f"G{r}", f"=IF(nr_NetMode,$C{r},$C{r}*($E{r}/nr_MInd))", fmt=FMT_IDX)
        formula(ws, f"H{r}",
                f'=IF(OR($D{r}="",$F{r}=""),"",IF(nr_NetMode,$D{r},$D{r}*($F{r}/nr_MInd)))',
                fmt=FMT_IDX)
        formula(ws, f"I{r}", f"=IF(YEAR($B{r})=nr_PlanYear,1,0)", fmt=FMT_INT)
        pct_m = "+0.0%;-0.0%;0.0%"
        if j >= 12:  # first 12 reported months have no year-ago comparison
            formula(ws, f"J{r}", f'=IF(OR($D{r}="",$D{r - 12}=""),"",$D{r}/$D{r - 12}-1)',
                    fmt=pct_m)
            formula(ws, f"K{r}", f'=IF(OR($H{r}="",$H{r - 12}=""),"",$H{r}/$H{r - 12}-1)',
                    fmt=pct_m)
        formula(ws, f"L{r}", f'=IF(OR($C{r}="",$D{r}=""),"",$C{r}/$D{r}-1)', fmt=pct_m)
        for cL in "BCDEFGHIJKL":
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=8)

    cats = Reference(ws, min_col=2, min_row=CD0 + 1, max_row=CD0 + L.N_MONTHS)

    # (a) written vs earned index + plan-year band
    band = BarChart()
    band.add_data(Reference(ws, min_col=9, min_row=CD0, max_row=CD0 + L.N_MONTHS),
                  titles_from_data=True)
    band.gapWidth = 0
    band.series[0].graphicalProperties.solidFill = BAND_FILL
    band.series[0].graphicalProperties.line.noFill = True
    band.y_axis.scaling.min = 0
    band.y_axis.scaling.max = 1
    band.y_axis.delete = True
    lines = LineChart()
    lines.add_data(Reference(ws, min_col=3, min_row=CD0, max_col=4, max_row=CD0 + L.N_MONTHS),
                   titles_from_data=True)
    lines.set_categories(cats)
    _line(lines.series[0], STEEL)
    _line(lines.series[1], NAVY)
    lines.y_axis.axId = 200
    lines.y_axis.number_format = "0.00"
    lines.y_axis.crosses = "max"
    lines.y_axis.delete = False   # secondary axis must stay visible (D36)
    lines.x_axis.delete = False
    lines.y_axis.title = "Rate index"
    band.set_categories(cats)
    band += lines
    _chart(band, "Cumulative written vs earned rate index (plan year shaded)", height=8.5, width=16)
    ws.add_chart(band, "B4")

    # (c) written vs earned mod path
    mods = LineChart()
    mods.add_data(Reference(ws, min_col=5, min_row=CD0, max_col=6, max_row=CD0 + L.N_MONTHS),
                  titles_from_data=True)
    mods.set_categories(cats)
    _line(mods.series[0], STEEL)
    _line(mods.series[1], NAVY)
    mods.y_axis.number_format = "0.000"
    _chart(mods, "Written vs earned schedule mod path", height=8.5, width=16,
           y_title="Avg schedule mod")
    ws.add_chart(mods, "B22")

    # (d) combined price index
    price = LineChart()
    price.add_data(Reference(ws, min_col=7, min_row=CD0, max_col=8, max_row=CD0 + L.N_MONTHS),
                   titles_from_data=True)
    price.set_categories(cats)
    _line(price.series[0], STEEL)
    _line(price.series[1], NAVY)
    price.y_axis.number_format = "0.00"
    _chart(price, "Combined price index = rate index x (mod / M_ind)", height=8.5, width=16,
           y_title="Price index")
    ws.add_chart(price, "K22")

    # ---- (e) quarterly table + (b) YoY chart ----
    qt = 48
    section(ws, qt - 1, "B",
            "Quarterly earned rate and price change, and unearned runway — plan year and next")
    header_row(ws, qt, 2, ["Quarter", "Earned rate level (E_Q)", "Prior-year E_Q",
                           "YoY earned rate chg", "Written level at qtr end",
                           "Unearned runway", "Earned price level (P_Q)", "Prior-year P_Q",
                           "YoY earned price chg"],
               widths=[10, 13, 13, 13, 13, 13, 13, 12, 13])
    ws.row_dimensions[qt].height = 30
    rows = [(p, q) for q in range(1, 5)] + [(p + 1, q) for q in range(1, 5)]
    for i, (yr, q) in enumerate(rows):
        r = qt + 1 + i
        off = (yr - (p - 1)) * 12 + (q - 1) * 3          # first month offset in matrix
        off_prior = off - 12
        c1, c2 = col(L.RE_MATRIX_COL + off), col(L.RE_MATRIX_COL + off + 2)
        p1, p2 = col(L.RE_MATRIX_COL + off_prior), col(L.RE_MATRIX_COL + off_prior + 2)
        formula(ws, f"B{r}", f'="Q{q} "&{"nr_PlanYear" if yr == p else "(nr_PlanYear+1)"}',
                align=ALIGN_C)
        formula(ws, f"C{r}",
                f"=SUM({RE}!{c1}${L.RE_ROW_NUM}:{c2}${L.RE_ROW_NUM})"
                f"/SUM({RE}!{c1}${L.RE_ROW_DEN}:{c2}${L.RE_ROW_DEN})", fmt=FMT_IDX, align=ALIGN_C)
        formula(ws, f"D{r}",
                f"=SUM({RE}!{p1}${L.RE_ROW_NUM}:{p2}${L.RE_ROW_NUM})"
                f"/SUM({RE}!{p1}${L.RE_ROW_DEN}:{p2}${L.RE_ROW_DEN})", fmt=FMT_IDX, align=ALIGN_C)
        formula(ws, f"E{r}", f"=$C{r}/$D{r}-1", fmt="+0.0%;-0.0%;0.0%", align=ALIGN_C)
        coh_end = L.RE_COH_FIRST + 12 + off + 2
        link(ws, f"F{r}", f"={RE}!$N${coh_end}", fmt=FMT_IDX, align=ALIGN_C)
        formula(ws, f"G{r}", f"=$F{r}/$C{r}-1", fmt="+0.0%;-0.0%;0.0%", align=ALIGN_C)
        # combined price: rate x quarterly earned mod / M_ind (net path already combined)
        formula(ws, f"H{r}",
                f"=IF(nr_NetMode,$C{r},$C{r}*SUM({RE}!{c1}${L.RE_ROW_MODNUM}:{c2}${L.RE_ROW_MODNUM})"
                f"/SUM({RE}!{c1}${L.RE_ROW_DEN}:{c2}${L.RE_ROW_DEN})/nr_MInd)",
                fmt=FMT_IDX, align=ALIGN_C)
        formula(ws, f"I{r}",
                f"=IF(nr_NetMode,$D{r},$D{r}*SUM({RE}!{p1}${L.RE_ROW_MODNUM}:{p2}${L.RE_ROW_MODNUM})"
                f"/SUM({RE}!{p1}${L.RE_ROW_DEN}:{p2}${L.RE_ROW_DEN})/nr_MInd)",
                fmt=FMT_IDX, align=ALIGN_C)
        ws[f"I{r}"].font = font(GREY_DARK, size=9)
        formula(ws, f"J{r}", f"=$H{r}/$I{r}-1", fmt="+0.0%;-0.0%;0.0%", align=ALIGN_C)
    put(ws, f"B{qt + 10}",
        "Runway = written level / earned level - 1: rate locked in but not yet earned "
        "(D27). Price = rate x earned mod / M_ind (series coincide under a net selection). "
        "YoY compares the same period one year earlier.", fnt=F_SMALL_IT)

    # ---- (b) MONTHLY YoY rate vs price chart (quarterly grain hid the
    # mid-month steps the engine already computes; the quarterly table above
    # remains the runway summary) ----
    yoym = BarChart()
    yoym.type = "col"
    yoym.gapWidth = 40
    m0, m1 = CD0 + 13, CD0 + L.N_MONTHS  # Jan P .. Dec P+1 (24 valid YoY points)
    yoym.add_data(Reference(ws, min_col=10, min_row=m0, max_row=m1), titles_from_data=False)
    yoym.add_data(Reference(ws, min_col=11, min_row=m0, max_row=m1), titles_from_data=False)
    yoym.series[0].tx = SeriesLabel(v="Earned RATE chg, YoY")
    yoym.series[1].tx = SeriesLabel(v="Earned PRICE chg, YoY")
    yoym.set_categories(Reference(ws, min_col=2, min_row=m0, max_row=m1))
    yoym.series[0].graphicalProperties.solidFill = STEEL
    yoym.series[1].graphicalProperties.solidFill = NAVY
    yoym.y_axis.number_format = "0.0%"
    _chart(yoym, "YoY earned RATE vs PRICE change by MONTH — plan year and next",
           height=8.5, width=16, y_title="YoY earned change")
    ws.add_chart(yoym, "K4")

    # ---- NEW: unearned-runway line (the most plannable number in the file) ----
    run = LineChart()
    run.add_data(Reference(ws, min_col=12, min_row=CD0 + 1, max_row=CD0 + L.N_MONTHS),
                 titles_from_data=False)
    run.series[0].tx = SeriesLabel(v="Unearned runway")
    run.set_categories(Reference(ws, min_col=2, min_row=CD0 + 1, max_row=CD0 + L.N_MONTHS))
    _line(run.series[0], NAVY)
    run.legend = None
    run.y_axis.number_format = "0.0%"
    _chart(run, "Unearned runway: rate on the books not yet earned (written / earned - 1)",
           height=8.5, width=14)
    ws.add_chart(run, "T4")

    # per-chart "what to look for" captions (single overflow lines)
    captions = [
        ("B21", "Watch the gap: written (steel) runs ahead of earned (navy) — the gap is "
                "rate still to earn in."),
        ("K21", "Monthly grain shows the step when an action lands mid-month; quarterly "
                "bars hide it."),
        ("T21", "Converges to zero as filed actions finish earning; jumps when new rate "
                "goes on the books."),
        ("B39", "Mod drift reaches earned premium with the same lag as rate — a price "
                "change in slow motion."),
        ("K39", "Rate x mod combined: the net price customers actually pay."),
    ]
    for addr, text in captions:
        put(ws, addr, text, fnt=F_SMALL_IT)

    # ---- carryover ledger: why CY P+1 starts above CY P, action by action ----
    cl = qt + 12
    section(ws, cl, "B", "Carryover ledger — where the head start into the following year "
                         "comes from")
    formula(ws, f"B{cl + 1}",
            '=IF(nr_NetMode,"Net selection active: the carryover is set by the net renewal '
            'path — this ledger describes the explicit program counterfactual.",'
            '"Each action earns partly in the plan year and fully in the next; the '
            'difference is carryover.")')
    ws[f"B{cl + 1}"].font = F_SMALL_IT
    header_row(ws, cl + 2, 2,
               ["#", "Effective", "Effective %", "Share of plan yr earned",
                "Share of next yr earned", "Carryover contribution"],
               widths=None, fill=FILL_GREY, fnt=font(GREY_DARK, bold=True, size=9))
    wrng = f"{RE}!$H${L.RE_COH_FIRST}:$H${L.RE_COH_LAST}"
    ecp = f"{RE}!$Q${L.RE_COH_FIRST}:$Q${L.RE_COH_LAST}"
    ecp1 = f"{RE}!$R${L.RE_COH_FIRST}:$R${L.RE_COH_LAST}"
    absmi = f"{RE}!$G${L.RE_COH_FIRST}:$G${L.RE_COH_LAST}"

    def _share(r, ec):
        mj = f"YEAR($C{r})*12+MONTH($C{r})-1"
        frac = f"(EOMONTH($C{r},0)-$C{r}+1)/DAY(EOMONTH($C{r},0))"
        return (f'=IF($C{r}="","",(SUMPRODUCT(({absmi}>{mj})*{wrng},{ec})'
                f"+{frac}*SUMPRODUCT(({absmi}={mj})*{wrng},{ec}))"
                f"/SUMPRODUCT({wrng},{ec}))")

    for j in range(1, 9):
        r = cl + 2 + j
        cnt = f"COUNTIFS(rl_key,nr_SelKey,rl_seq,{j})"
        put(ws, f"B{r}", j, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
        formula(ws, f"C{r}", f'=IF({cnt}=0,"",SUMIFS(rl_eff,rl_key,nr_SelKey,rl_seq,{j}))',
                fmt="m/d/yy", align=ALIGN_C)
        formula(ws, f"D{r}", f'=IF({cnt}=0,"",SUMIFS(rl_reff,rl_key,nr_SelKey,rl_seq,{j}))',
                fmt="+0.0%;-0.0%;0.0%", align=ALIGN_C)
        formula(ws, f"E{r}", _share(r, ecp), fmt=FMT_PCT, align=ALIGN_C)
        formula(ws, f"F{r}", _share(r, ecp1), fmt=FMT_PCT, align=ALIGN_C)
        formula(ws, f"G{r}", f'=IF($C{r}="","",EXP(LN(1+$D{r})*($F{r}-$E{r}))-1)',
                fmt="+0.0%;-0.0%;0.0%", align=ALIGN_C)
        formula(ws, f"H{r}", f'=IF($C{r}="",0,LN(1+$D{r})*($F{r}-$E{r}))', fmt=FMT_IDX)
        ws[f"H{r}"].font = font(GREY_DARK, size=8)
    tot_r = cl + 11
    label(ws, f"C{tot_r}", "Actions combined", bold=True)
    formula(ws, f"G{tot_r}", f"=EXP(SUM($H${cl + 3}:$H${cl + 10}))-1",
            fmt="+0.0%;-0.0%;0.0%", align=ALIGN_C, bold=True, fill=FILL_PANEL)
    label(ws, f"C{tot_r + 1}", "Timing / compounding interaction (residual)")
    formula(ws, f"G{tot_r + 1}",
            f"=(1+nr_YoY_P1)/EXP(SUM($H${cl + 3}:$H${cl + 10}))-1",
            fmt="+0.0%;-0.0%;0.0%", align=ALIGN_C)
    label(ws, f"C{tot_r + 2}", "Total carryover (earned rate chg into next year)", bold=True)
    link(ws, f"G{tot_r + 2}", "=nr_YoY_P1", fmt="+0.0%;-0.0%;0.0%", align=ALIGN_C,
         bold=True, fill=FILL_PANEL)

    set_widths(ws, {"A": 2, "B": 11, "C": 12, "D": 12, "E": 12, "F": 12, "G": 13, "H": 13,
                    "I": 11, "J": 12, "K": 12, "L": 12})
    presentation_setup(ws, gridlines_off=True, tab_color=NAVY)
    print_setup(ws)


# ---------------------------------------------------------------------------
# _oracle
# ---------------------------------------------------------------------------


def build_oracle_sheet(ctx: Ctx):
    ws = ctx.wb["_oracle"]
    p = ctx.p
    m, c = ctx.oracle_m, ctx.oracle_c
    put(ws, "A1",
        "_oracle (hidden): constants computed by src/engine.py at build time for the seeded "
        "worked example, plus the input fingerprint that scopes the Checks oracle ties.",
        fnt=font(GREY_DARK, bold=True, size=9))

    consts = [
        ("orc_crl", m.crl_ind, "Oracle CRL_ind"),
        ("orc_ecy_pm1", m.e_cy[p - 1], "Oracle E_CY(P-1), monthly convention"),
        ("orc_ecy_p", m.e_cy[p], "Oracle E_CY(P), monthly convention"),
        ("orc_ecy_p1", m.e_cy[p + 1], "Oracle E_CY(P+1), monthly convention"),
        ("orc_mbar_p", m.m_earned[p], "Oracle CY P earned mod, monthly convention"),
        ("orc_arate_p", m.a_rate_p, "Oracle A_rate(P)"),
        ("orc_arate_p1", m.a_rate_p1, "Oracle A_rate(P+1)"),
        ("orc_amod_p", m.a_mod_p, "Oracle A_mod(P)"),
        ("orc_cylr_p", m.cy_lr_p, "Oracle CY plan LR(P), monthly convention"),
        ("orc_cylr_p1", m.cy_lr_p1, "Oracle CY LR(P+1), monthly convention"),
        ("orc_echg", m.earned_chg_vs_ind, "Oracle earned chg vs indication"),
        ("orc_cylr_p_cont", c.cy_lr_p, "Oracle CY plan LR(P), continuous closed form"),
        ("orc_ecy_p_cont", c.e_cy[p], "Oracle E_CY(P), continuous closed form"),
        ("orc_solver_r", ctx.oracle_solver_r, "Oracle solver required change (worked example)"),
    ]
    label(ws, "A3", "Baked oracle constants", bold=True)
    for i, (name, val, desc) in enumerate(consts):
        r = 4 + i
        put(ws, f"A{r}", desc, fnt=F_LABEL)
        put(ws, f"B{r}", val, fnt=font(GREY_DARK), fmt="0.0000000000")
        ctx.define(name, "_oracle", f"$B${r}", desc + " (baked at build time)")

    # fingerprint
    fp0 = 4 + len(consts) + 2
    label(ws, f"A{fp0 - 1}", "Worked-example input fingerprint (expected vs live)", bold=True)
    key = ctx.we_key
    we = ctx.we_row
    rate_rows = [r for r in ctx.rate_rows if f"{r['bu']}|{r['state']}" == key]
    eff_serial_sum = sum(_excel_serial(r["eff"]) for r in rate_rows)
    filed_sum = sum(r["filed"] for r in rate_rows)
    reff_sum = sum(
        r["filed"] * ((1.0 if r["achievement"] is None else r["achievement"])
                      if r["status"] == "planned" else 1.0)
        for r in rate_rows)
    checks = [
        ("Plan year", p, "=nr_PlanYear", 1e-9),
        ("Selected key", key, "=nr_SelKey", None),
        ("Projected LR", we["lr_proj"], "=nr_LRproj", 1e-9),
        ("LR basis", we["basis"], "=nr_Basis", None),
        ("M_ind", we["m_ind"], "=nr_MInd", 1e-9),
        ("M_0", we["m0"], "=nr_M0", 1e-9),
        ("M_0 as-of (serial)", _excel_serial(we["m0_asof"]), "=nr_M0Asof", 0.5),
        ("M_1", we["m1"], "=nr_M1", 1e-9),
        ("M_prior (0 = blank)", we["m_prior"] or 0, "=nr_MPrior", 1e-9),
        ("M_2 (0 = blank)", we["m2"] or 0, "=nr_M2", 1e-9),
        ("Net trend (effective)", we["trend"] or 0, "=nr_Trend", 1e-9),
        ("Trend default (Control)", 0, "=N(nr_TrendDefault)", 1e-9),
        ("Net rate selection off", "FALSE", '=IF(nr_NetMode,"TRUE","FALSE")', None),
        ("A_other", we["a_other"], "=nr_AOther", 1e-9),
        ("Mod adj (combo row)", we["modadj"], "=nr_ModAdjRow", None),
        ("Mod adj (master)", "ON", "=nr_ModAdjMaster", None),
        ("Seasonality inactive", "FALSE", '=IF(re_season_active,"TRUE","FALSE")', None),
        ("Term months", ctx.lob.term_months, "=nr_TermMonths", 1e-9),
        ("Rate rows for combo", len(rate_rows), f'=COUNTIF(rl_key,"{key}")', 1e-9),
        ("Sum of filed %", filed_sum, f'=SUMIF(rl_key,"{key}",rl_filed)', 1e-9),
        ("Sum of effective serials", eff_serial_sum, f'=SUMIF(rl_key,"{key}",rl_eff)', 0.5),
        ("Sum of effective changes", reff_sum, f'=SUMIF(rl_key,"{key}",rl_reff)', 1e-9),
        ("Planned rows", sum(1 for r in rate_rows if r["status"] == "planned"),
         f'=COUNTIFS(rl_key,"{key}",rl_status,"planned")', 1e-9),
        ("Considered rows", sum(1 for r in rate_rows if r["considered"] == "Y"),
         f'=COUNTIFS(rl_key,"{key}",rl_cons,"Y")', 1e-9),
    ]
    for j, h in enumerate(["Input", "Expected", "Live", "Match?"]):
        put(ws, f"{col(1 + j)}{fp0}", h, fnt=font(GREY_DARK, bold=True, size=9))
    for i, (lbl, exp, act_f, tol) in enumerate(checks):
        r = fp0 + 1 + i
        put(ws, f"A{r}", lbl, fnt=F_LABEL)
        put(ws, f"B{r}", exp, fnt=font(GREY_DARK))
        formula(ws, f"C{r}", act_f)
        ws[f"C{r}"].font = font(GREY_DARK)
        if tol is None:
            formula(ws, f"D{r}", f"=$C{r}=$B{r}")
        else:
            formula(ws, f"D{r}", f"=ABS($C{r}-$B{r})<={tol}")
        ws[f"D{r}"].font = font(GREY_DARK)
    fp_last = fp0 + len(checks)
    label(ws, f"A{fp_last + 2}", "Fingerprint intact?", bold=True)
    formula(ws, f"B{fp_last + 2}", f"=COUNTIF($D${fp0 + 1}:$D${fp_last},FALSE)=0")
    ctx.define("orc_fp", "_oracle", f"$B${fp_last + 2}",
               "TRUE while the seeded worked-example inputs are unchanged (scopes oracle ties)")
    set_widths(ws, {"A": 44, "B": 18, "C": 18, "D": 10})


# ---------------------------------------------------------------------------
# Checks
# ---------------------------------------------------------------------------


def build_checks(ctx: Ctx):
    ws = ctx.wb["Checks"]
    p = ctx.p
    title(ws, "A1", "Checks — validation panel",
          "Live formulas versus structure, identities, and oracle constants baked at build "
          "time. FAIL = investigate before using results. WARN = review advisory.")
    label(ws, "A3", "Overall status", bold=True)
    nav_bar(ws, 3, 9, ["Control", "Inputs", "Read Me"], step=2)

    rows: list[tuple] = []  # (category, description, expected_f, actual_f, tol, kind)
    A = rows.append
    sel = "br_selrow"

    # --- structural (always on) ---
    A(("Structure", "Earning-profile row sums = 1 for every in-window cohort",
       "=0", "=MAX(re_rowsum_chk)", 1e-9, "FAIL"))
    A(("Structure", "Monthly earned index within [min W, max W] of contributing cohorts",
       "=0", "=MAX(re_viol_row)", 1e-9, "FAIL"))
    A(("Structure", "E_CY(P): monthly-matrix aggregation = closed-form ec aggregation",
       "=nr_ECY_P", "=re_ecy_matrix", 1e-9, "FAIL"))
    A(("Structure", "E_CY(P): visible Rate Engine = hidden _calc block (selected combo)",
       "=0", f"=IF({sel}=0,0,ABS(INDEX(calc_ecy_p,{sel})-nr_ECY_P))", 1e-9, "FAIL"))
    A(("Structure", "CRL_ind: visible Rate Engine = hidden _calc block (selected combo)",
       "=0", f"=IF({sel}=0,0,ABS(INDEX(calc_crl,{sel})-nr_CRLind))", 1e-9, "FAIL"))
    A(("Structure", "Bridge chain: LR_cur x A_rate x A_mod x A_other = CY plan LR",
       "=0", "=ABS(nr_LRcur*nr_Arate_P*nr_Amod_P*nr_AOther-nr_CYLR_P)", 1e-9, "FAIL"))
    A(("Structure", "Bridge chain (P+1): with (1+trend) factor = CY LR(P+1)",
       "=0", "=ABS(nr_LRcur*(1+nr_Trend)*nr_Arate_P1*nr_Amod_P1*nr_AOther-nr_CYLR_P1)",
       1e-9, "FAIL"))
    A(("Structure", "Portfolio row = Bridge result (selected combo)",
       "=0", f"=IF({sel}=0,0,ABS(INDEX(calc_cylr_p,{sel})-nr_CYLR_P))", 1e-9, "FAIL"))
    A(("Structure", "Scenario engine reproduces Base when all levers blank",
       "=0",
       "=IF(COUNT(sc_dpts,sc_shift,sc_ach,sc_dm1,sc_dnet)>0,0,"
       "ABS(INDEX(sc_res_cylr,2)-nr_CYLR_P))",
       1e-9, "FAIL"))
    A(("Identity", "Linear-mod identity: CY P earned mod = written mod at 1/1/P (§3.3.5)",
       None, None, None, "IDENTITY"))
    A(("Structure", "Attribution factors multiply back to the actual CY LR",
       None, None, None, "ATTR"))

    # --- oracle ties (fingerprint-scoped) ---
    ties = [
        ("CRL_ind ties oracle", "=orc_crl", "=nr_CRLind", 1e-6),
        ("E_CY(P-1) ties oracle", "=orc_ecy_pm1", "=nr_ECY_Pm1", 1e-6),
        ("E_CY(P) ties oracle", "=orc_ecy_p", "=nr_ECY_P", 1e-6),
        ("E_CY(P+1) ties oracle", "=orc_ecy_p1", "=nr_ECY_P1", 1e-6),
        ("CY P earned mod ties oracle", "=orc_mbar_p", "=nr_MEarned_P", 1e-6),
        ("A_rate(P) ties oracle", "=orc_arate_p", "=nr_Arate_P", 1e-6),
        ("A_rate(P+1) ties oracle", "=orc_arate_p1", "=nr_Arate_P1", 1e-6),
        ("A_mod(P) ties oracle", "=orc_amod_p", "=nr_Amod_P", 1e-6),
        (f"CY {p} plan LR ties oracle (monthly convention)", "=orc_cylr_p", "=nr_CYLR_P", 1e-6),
        (f"CY {p + 1} LR ties oracle", "=orc_cylr_p1", "=nr_CYLR_P1", 1e-6),
        (f"CY {p} plan LR within 0.10 pts of continuous closed form", "=orc_cylr_p_cont",
         "=nr_CYLR_P", 1e-3),
    ]
    for desc, e, a, tol in ties:
        A(("Oracle tie", desc, e, a, tol, "ORACLE"))
    A(("Oracle tie", "Solver Mode A returns the seeded planned change (acceptance §11.7)",
       None, None, None, "SOLVER"))
    # The Mode B cross-check (April row with r = +5% equals the base CY LR)
    # presumes the seeded single +5% @ 4/1/P planned program; emit it only
    # when that pattern is actually present (it is absent in some per-BU files).
    we_planned = [r for r in ctx.rate_rows
                  if f"{r['bu']}|{r['state']}" == ctx.we_key and r["status"] == "planned"]
    if (len(we_planned) == 1 and we_planned[0]["eff"] == dt.date(p, 4, 1)
            and abs(ctx.oracle_solver_r - 0.05) < 1e-9):
        A(("Oracle tie", "Solver Mode B April row equals the base CY plan LR",
           None, None, None, "SOLVERB"))

    # --- input sanity ---
    A(("Input sanity", "Achievement % within [0%, 150%]", "=0",
       '=COUNTIF(rl_ach,">1.5")+COUNTIF(rl_ach,"<0")', 0, "FAIL"))
    A(("Input sanity", "Schedule mods (M_ind, M_0, M_1) within [0.5, 1.5]", "=0",
       '=COUNTIF(lr_mind,">1.5")+COUNTIF(lr_mind,"<0.5")+COUNTIF(lr_m0,">1.5")'
       '+COUNTIF(lr_m0,"<0.5")+COUNTIF(lr_m1,">1.5")+COUNTIF(lr_m1,"<0.5")', 0, "FAIL"))
    A(("Input sanity", "Optional mods (M_prior, M_2) within [0.5, 1.5]", "=0",
       '=COUNTIF(lr_mprior,">1.5")+COUNTIF(lr_mprior,"<0.5")'
       '+COUNTIF(lr_m2,">1.5")+COUNTIF(lr_m2,"<0.5")', 0, "FAIL"))
    A(("Input sanity", "Every effective change keeps 1 + r > 0", "=0",
       '=COUNTIF(rl_reff,"<=-1")', 0, "FAIL"))
    A(("Input sanity", "M_0 as-of dates precede 12/31 of the plan year", "=0",
       "=SUMPRODUCT((lr_key<>\"\")*(N(lr_m0asof)>=DATE(nr_PlanYear,12,31))*1)", 0, "FAIL"))
    A(("Input sanity", "Rate-log status is taken/planned on every populated row", "=0",
       '=SUMPRODUCT((rl_key<>"")*(rl_status<>"taken")*(rl_status<>"planned")*1)', 0, "FAIL"))
    A(("Input sanity", "LR basis is current/proposed (blank treated as current)", "=0",
       '=SUMPRODUCT((lr_basis<>"current")*(lr_basis<>"proposed")*(lr_basis<>"")*1)', 0, "FAIL"))
    A(("Input sanity", "Plan EP is non-negative", "=0", '=COUNTIF(lr_ep,"<0")', 0, "FAIL"))
    A(("Input sanity", "Net rate selections within [-50%, +100%]", "=0",
       '=COUNTIF(lr_netp,">1")+COUNTIF(lr_netp,"<-0.5")'
       '+COUNTIF(lr_netp1,">1")+COUNTIF(lr_netp1,"<-0.5")', 0, "FAIL"))
    A(("Advisory", "Considered flags appear only on taken rows (§3.2.7)", "=0",
       '=SUMPRODUCT((rl_status="planned")*(rl_cons="Y")*1)', 0, "WARN"))
    A(("Advisory", "No two changes share a BU|LOB cohort month (exact-blend support, D4)",
       "=0", "=SUM(rl_dup)", 0, "WARN"))
    A(("Advisory", "Rate-change dates within the Jan (P-2) .. Dec (P+1) engine window", "=0",
       '=COUNTIF(rl_eff,"<"&DATE(nr_PlanYear-2,1,1))+COUNTIF(rl_eff,">"&DATE(nr_PlanYear+1,12,31))',
       0, "WARN"))
    A(("Advisory", "A_other <> 1 carries a label", "=0",
       '=SUMPRODUCT((lr_key<>"")*(lr_aother<>"")*(lr_aother<>1)*(lr_aotherlbl="")*1)', 0, "WARN"))
    A(("Advisory", "Populated seasonality rows have a positive weight sum", "=0",
       '=SUMPRODUCT((se_state<>"")*(N(se_sum)=0)*1)', 0, "WARN"))
    A(("Advisory", "Considered? column is Y/N (blank counts as N)", "=0",
       '=SUMPRODUCT((rl_key<>"")*(rl_cons<>"Y")*(rl_cons<>"N")*(rl_cons<>"")*1)', 0, "WARN"))
    A(("Advisory", "No duplicate BU x state combos in tbl_LR (roster integrity)", "=0",
       '=SUMPRODUCT((lr_key<>"")*(COUNTIF(lr_key,lr_key)>1)*1)', 0, "WARN"))
    A(("Advisory", "Net selections pair with an active mod adjustment (else rate-only)", "=0",
       '=SUMPRODUCT((lr_key<>"")*(1-ISBLANK(lr_netp))*(lr_modadj="OFF")*1)', 0, "WARN"))
    A(("Advisory", "No planned rows on/after 1/1/P are superseded by a net selection", "=0",
       '=IF(NOT(nr_NetMode),0,COUNTIFS(rl_key,nr_SelKey,rl_status,"planned",'
       'rl_eff,">="&DATE(nr_PlanYear,1,1)))', 0, "WARN"))
    A(("Advisory", "Selected BU|State exists in tbl_LR", "=0", "=IF(nr_SelOK,0,1)", 0, "WARN"))
    # ---- appended v2.1 checks (append-only: Checks!G16 and earlier rows never move) ----
    A(("Structure", "Walkthrough assembly equals the Bridge result", "=0",
       "=ABS(nr_WalkLR-nr_CYLR_P)", 1e-9, "FAIL"))
    A(("Advisory", "Taken rows still carrying an achievement % (it is ignored — restate "
       "Filed % as achieved)", "=0",
       '=SUMPRODUCT((rl_status="taken")*(rl_ach<>"")*(rl_ach<>1)*1)', 0, "WARN"))
    A(("Structure", "Every distinct state fits on the State Summary display", "=0",
       f"=MAX(0,lst_state_cnt-{len(ctx.cfg.states) + 3})", 0, "FAIL"))

    header_row(ws, L.CK_HDR, 1,
               ["#", "Category", "Check", "Expected", "Actual", "Tolerance", "Status", ""],
               widths=[5, 12, 78, 13, 13, 11, 10, 4])
    fmt_num = "0.0000000000;-0.0000000000;0"
    for i, (cat, desc, exp_f, act_f, tol, kind) in enumerate(rows):
        r = L.CK_FIRST + i
        put(ws, f"A{r}", i + 1, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
        put(ws, f"B{r}", cat, fnt=font(GREY_DARK, size=9))
        put(ws, f"C{r}", desc, fnt=F_LABEL)
        if kind == "IDENTITY":
            formula(ws, f"D{r}", "=me_identity", fmt=FMT_MOD, align=ALIGN_C)
            formula(ws, f"E{r}", "=nr_MEarned_P", fmt=FMT_MOD, align=ALIGN_C)
            # the monthly-discretization residual scales with the mod path's
            # daily slope; 0.0002 is the floor from §3.3.5 (D44)
            formula(ws, f"F{r}", "=MAX(0.0002,ABS('Mod Engine'!$H$7)*2)",
                    fmt="0.00000", align=ALIGN_C)
            ws[f"F{r}"].font = font(GREY_DARK, size=9)
            formula(ws, f"G{r}",
                    f'=IF(NOT(me_identity_ok),"N/A",IF(ABS(nr_MEarned_P-me_identity)<=$F{r},'
                    '"PASS","FAIL"))', align=ALIGN_C, bold=True)
        elif kind == "ATTR":
            formula(ws, f"D{r}", "=N(att_actlr)", fmt=FMT_PCT, align=ALIGN_C)
            formula(ws, f"E{r}",
                    "=IF(N(att_actlr)=0,0,ABS('Attribution'!$D$25*'Attribution'!$C$26"
                    "*'Attribution'!$C$27*'Attribution'!$C$28*'Attribution'!$C$29-att_actlr))",
                    fmt=fmt_num, align=ALIGN_C)
            put(ws, f"F{r}", 1e-7, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
            formula(ws, f"G{r}",
                    '=IF(N(att_actlr)=0,"N/A",IF($E' + str(r) + '<=$F' + str(r) +
                    ',"PASS","FAIL"))', align=ALIGN_C, bold=True)
        elif kind == "SOLVER":
            formula(ws, f"D{r}", "=orc_solver_r", fmt="0.000%", align=ALIGN_C)
            formula(ws, f"E{r}", "=IF(ISNUMBER(nr_SolverR),nr_SolverR,-999)", fmt="0.000%",
                    align=ALIGN_C)
            put(ws, f"F{r}", 1e-6, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
            formula(ws, f"G{r}",
                    "=IF(OR(NOT(orc_fp),ABS('Solver'!$C$7-orc_cylr_p)>0.000001,"
                    "'Solver'!$C$8<>DATE(nr_PlanYear,4,1)),\"N/A\","
                    f'IF(ABS($E{r}-$D{r})<=$F{r},"PASS","FAIL"))', align=ALIGN_C, bold=True)
        elif kind == "SOLVERB":
            formula(ws, f"D{r}", "=orc_cylr_p", fmt=FMT_PCT, align=ALIGN_C)
            formula(ws, f"E{r}", "='Solver'!$G$32", fmt=FMT_PCT, align=ALIGN_C)
            put(ws, f"F{r}", 1e-6, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
            formula(ws, f"G{r}",
                    "=IF(OR(NOT(orc_fp),'Solver'!$C$26<>0.05),\"N/A\","
                    f'IF(ABS($E{r}-$D{r})<=$F{r},"PASS","FAIL"))', align=ALIGN_C, bold=True)
        else:
            formula(ws, f"D{r}", exp_f, fmt=fmt_num, align=ALIGN_C)
            formula(ws, f"E{r}", act_f, fmt=fmt_num, align=ALIGN_C)
            put(ws, f"F{r}", tol, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
            if kind == "ORACLE":
                formula(ws, f"G{r}",
                        f'=IF(NOT(orc_fp),"N/A",IF(ABS($E{r}-$D{r})<=$F{r},"PASS","FAIL"))',
                        align=ALIGN_C, bold=True)
            elif kind == "WARN":
                formula(ws, f"G{r}", f'=IF(ABS($E{r}-$D{r})<=$F{r},"PASS","WARN")',
                        align=ALIGN_C, bold=True)
            else:
                formula(ws, f"G{r}", f'=IF(ABS($E{r}-$D{r})<=$F{r},"PASS","FAIL")',
                        align=ALIGN_C, bold=True)
    last = L.CK_FIRST + len(rows) - 1
    g_rng = f"$G${L.CK_FIRST}:$G${last}"
    formula(ws, "C3",
            f'=IF(COUNTIF({g_rng},"FAIL")=0,"ALL CHECKS PASS",'
            f'"CHECKS FAILING: "&COUNTIF({g_rng},"FAIL"))', bold=True, fill=FILL_PANEL)
    formula(ws, "E3", f'=COUNTIF({g_rng},"WARN")&" warning(s), "'
                      f'&COUNTIF({g_rng},"N/A")&" not applicable"', fmt=FMT_GEN)
    ws["E3"].font = font(GREY_DARK, size=9)
    ctx.define("ck_overall", "Checks", "$C$3", "Overall PASS/FAIL status (shown on Control)")
    for value, fill, fcolor in (('"PASS"', FILL_GREEN, PASS_GREEN),
                                ('"FAIL"', FILL_RED, FAIL_RED),
                                ('"WARN"', FILL_AMBER, "7F6000"),
                                ('"N/A"', FILL_GREY, GREY_DARK)):
        ws.conditional_formatting.add(
            f"G{L.CK_FIRST}:G{last}",
            CellIsRule(operator="equal", formula=[value], fill=fill,
                       font=font(fcolor, bold=True)))
    ws.conditional_formatting.add(
        "C3", CellIsRule(operator="equal", formula=['"ALL CHECKS PASS"'], fill=FILL_GREEN,
                         font=font(PASS_GREEN, bold=True)))
    ws.conditional_formatting.add(
        "C3", CellIsRule(operator="notEqual", formula=['"ALL CHECKS PASS"'], fill=FILL_RED,
                         font=font(FAIL_RED, bold=True)))
    note(ws, f"A{last + 2}",
         "Oracle ties compare live formulas to constants computed by src/engine.py for the "
         "seeded worked example; they report N/A - INPUTS CHANGED once you replace the sample "
         "data (the structural and sanity checks above remain live). Regenerate the workbook "
         "to re-bake oracle constants for new inputs.")
    presentation_setup(ws, gridlines_off=True, freeze=f"A{L.CK_FIRST}", tab_color=PASS_GREEN)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Methodology
# ---------------------------------------------------------------------------


def _para(ws, row, text, width_cols="B", bold=False, size=10):
    c = put(ws, f"B{row}", text, fnt=font(NAVY if bold else GREY_DARK, bold=bold, size=size),
            align=ALIGN_WRAP)
    return row + 1


def build_methodology(ctx: Ctx):
    ws = ctx.wb["Methodology"]
    p = ctx.p
    title(ws, "B2", "Methodology — as implemented",
          "Regulatory-defensibility writeup: every formula, convention, and named range in "
          "this workbook. Cf. Werner & Modlin, Basic Ratemaking, Ch. 5 (on-leveling / "
          "current rate level).")
    r = 4
    blocks = [
        ("1. Purpose and framework", [
         "This workbook converts a projected loss ratio from a rate level indication "
         "(policy-year basis, rate fully earned) into a calendar-year plan loss ratio at the "
         "average earned rate level of the plan year, extended with a schedule-mod adjustment. "
         "The method is standard on-leveling / current-rate-level methodology (Werner & Modlin, "
         "Basic Ratemaking, Chapter 5), implemented as a discretized monthly written-index x "
         "earning-matrix engine that reproduces the classical parallelogram under uniform "
         "writings and generalizes it to seasonal writings, sub-annual terms, mid-month "
         "effective dates, and partially-achieved planned changes.",
         f"Dimensioning: each workbook covers ONE line of business ({ctx.lob.name} here) with "
         "a single workbook-level policy term; every input row, engine block, and Portfolio "
         "row is one business unit x state combination keyed BU|State (DECISIONS.md D37).",
         ]),
        ("2. Timeline alignment — why no additional loss trend", [
         f"The indication is assumed effective 7/1/{p - 1}. Its policy year has average loss "
         f"date 7/1/{p} — the midpoint of plan year {p}. The indication's projected loss ratio "
         "is therefore already trended to the plan year's loss center of gravity, and no "
         "additional loss trending is applied for the plan-year result. The CY (P+1) view is "
         "indicative only: it holds the indication fixed and applies the net-trend input once.",
         ]),
        ("3. Conventions", [
         "Rate-change effective dates are start-of-day: a change effective on the 1st applies "
         "to that whole day's writings. Schedule-mod as-of dates are close-of-day: 'M_0 as of "
         "9/30' anchors the written-mod path at the 10/1 boundary, so mod anchors use serial "
         "date + 1 in-sheet (DECISIONS.md D2).",
         "Writing cohorts are monthly, written at mid-month, spanning Jan (P-2) .. Dec (P+1) — "
         "48 months. The window extends 12 months earlier than the brief's 36 so that "
         "E_CY(P-1), required for the year-over-year earned rate change, is complete (D1).",
         "A cohort with term T earns 1/(2T) in its writing month, 1/T in each of the next T-1 "
         "months, and 1/(2T) in month T (mid-month convention). The closed-form year fraction "
         "used in compact blocks is ec(k,Y) = C(yEnd-k) - C(yStart-1-k) with "
         "C(o) = MIN(1, MAX(0, (o+0.5)/T)) — provably identical to summing the matrix row (D10).",
         "Mid-month effective dates split the cohort month's written weight at the FIRST "
         "in-month change date: weight p = days on/after that date / days in month at the "
         "fully-compounded end-of-month index, (1-p) at the pre-month index. Exact when at "
         "most one change lands in a BU|LOB month; a Checks advisory flags anything denser (D4).",
         "Planned rows earn filed % x achievement % (blank achievement = 100%). Taken rows "
         "always use filed %. CRL_ind is the product of (1 + effective change) over rows "
         "flagged considered-in-indication; a planned row flagged considered is included and "
         "flagged by a Checks advisory (D7).",
         ]),
        ("4. Rate engine", [
         "W_k: cumulative product of all changes effective on or before the cohort date, base "
         "1.000 before the first logged change; computed as EXP(SUMPRODUCT(conditions x "
         "ln(1+r))) — the exact log-sum identity, blank-row safe (D12).",
         "E_m = SUM_k w_k e(k,m) W_k / SUM_k w_k e(k,m);  E_CY(Y) aggregates numerator and "
         "denominator over the year's months (an aggregate ratio, never an average of monthly "
         "ratios).  A_rate(P) = CRL_ind / E_CY(P), likewise A_rate(P+1).",
         "Seasonality: optional 12-month written-weight vector per STATE (tbl_Seasonality), "
         "normalized in-engine; the global Control toggle AND a populated row are both "
         "required. Uniform writings reproduce the classical parallelogram.",
         ]),
        ("5. Schedule mod engine", [
         "The written-mod path is piecewise-linear through (as-of, M_0), (12/31/P, M_1), "
         "(12/31/P+1, M_2 with blank = M_1), plus an optional (as-of - 12 months, M_prior) "
         "anchor. With M_prior blank the backward anchor is placed ON the M_0->M_1 line, so "
         "backward extrapolation is exactly that segment extended (§3.3.1). Cohort mods are "
         "evaluated at actual mid-month dates.",
         "The CY earned mod is the exposure-weighted AVERAGE of cohort written mods through "
         "the same earning matrix (mods are levels, not changes). A_mod(P) = M_ind / "
         "earned-mod(P) when both the tbl_LR combo toggle and the Control master toggle are "
         "ON; otherwise 1.000. Set the toggle OFF if the indication's premium trend already "
         "reflects mod drift — otherwise the drift is double-counted (§3.3.4).",
         "Identity check (§3.3.5): with uniform writings, annual term, and a globally linear "
         "path, the CY earned mod equals the written mod at 1/1 of that year, within 0.0002; "
         "the check self-suspends when its preconditions fail (D29).",
         ]),
        ("6. The bridge", [
         "LR_current = LR_input x (1 + s) when the input basis is 'proposed'; unchanged when "
         "'current'.  CY_LR(P) = LR_current x A_rate(P) x A_mod(P) x A_other.  "
         "CY_LR(P+1) = LR_current x (1 + net_trend) x A_rate(P+1) x A_mod(P+1) x A_other "
         "(indicative). A_other defaults to 1.000 and requires a text label when it differs.",
         "Communication metrics: CY earned rate change vs indication = E_CY(P)/CRL_ind - 1; "
         "year-over-year earned rate changes E_CY(P)/E_CY(P-1) - 1 and E_CY(P+1)/E_CY(P) - 1 "
         "(the carryover-plus-new-actions figures planning teams quote).",
         ]),
        ("6b. Net rate selection (optional, per BU x state)", [
         "When a combo's 'Net sel P' input is set (tbl_LR; blank = off), planning switches "
         "from the explicit rate program to an OUTCOME target: define the combined net price "
         "index q(k) = W(k) x M_w(k)/M_ind (rate and schedule-mod price together, normalized "
         "to the indication's mod). Cohorts written before 1/1/P keep their fully-modeled "
         "q — actual rate history and actual mod drift, the carryover half of CY P earnings. "
         "Cohorts written on/after 1/1/P instead renew at the selection: q(k) = q(k-12) x "
         "(1 + x), superseding planned rate rows and the projected mod path for those "
         "cohorts only. The optional P+1 selection covers cohorts written in P+1 (blank = "
         "carry x).",
         "Under a net selection the bridge shows one combined factor A_net = CRL_ind / "
         "E_CY[q] on the rate line and A_mod = 1 (rate and price are inseparable in a net "
         "target). Note the compounding consequence, visible in the Rate Engine's 'Q net' "
         "column: a cohort renewing business that already carries a recent increase compounds "
         "the selection on top of it — that is the literal meaning of an x% net target on "
         "renewals. Solver and Scenarios' planned-row levers describe the explicit-program "
         "counterfactual (on-sheet notes flag this); the D-net scenario lever varies the "
         "selection itself. Attribution suspends its premium-side steps for net-mode combos.",
         ]),
        ("7. Solver (E1)", [
         "The CY earned index is linear in a single unknown change r, so the required rate "
         "solves in closed form: E(r) = (C_pre + (1+r) C_post) / D. The base index W0 uses "
         "taken rows only — the Solver replaces the ENTIRE planned program with one action at "
         "the chosen date (D13). When a taken change shares the effective month, the split "
         "follows the engine's first-change-date blend exactly (D31). Guard rails warn on "
         "thin post-effective exposure and on rates beyond the configured reasonability bound.",
         ]),
        ("8. Attribution (E2)", [
         "Sequential multiplicative decomposition of (actual - plan) CY LR, in the stated "
         "order: 1) rate magnitude (achieved % at planned dates), 2) rate timing (achieved % "
         "at actual dates), 3) mod drift (actual mod path, plan M_ind held fixed — D31), "
         "4) loss-side residual. The decomposition is order-dependent; a different ordering "
         "would allocate interaction terms differently. Taken rows are assumed booked as "
         "logged.",
         ]),
        ("9. Assumptions and limitations", [
         "Rate changes apply to all written premium — no new/renewal split in v1 (future "
         "hook: differential new vs renewal rate achievement).",
         "Exposure/audit premium effects (e.g., WC payroll audits) are out of scope.",
         "No retention response to rate is modeled (future hook).",
         "Premium-volume planning (EP x LR -> plan losses) is a future hook; plan EP here "
         "only weights the Portfolio total.",
         "Automated input refresh from Snowflake / Power Query is a future hook; in v1 all "
         "inputs are entered or pasted into the Inputs tables manually.",
         "Multiple rate changes within one BU|LOB cohort month blend at the first change "
         "date (advisory-flagged). Enter changes in distinct months for exact treatment.",
         "All ratios are currency-agnostic. Percent inputs are stored as decimal fractions.",
         "The engine window is Jan (P-2) .. Dec (P+1); changes before it fold into the base "
         "index, changes after it never earn and affect only CRL_ind if flagged considered.",
         ]),
    ]
    # widths set up-front so prose() reads the real host-column width
    set_widths(ws, {"A": 2, "B": 30, "C": 30, "D": 90})

    # ---- sections (rows 4..17 are reserved for the Contents block) ----
    toc: list[tuple[str, int]] = []
    r = 19
    for heading, paras in blocks:
        toc.append((heading, r))
        section(ws, r, "B", heading)
        r += 1
        for t in paras:
            prose(ws, f"D{r}", t, size=10)
            r += 1
        r += 1

    # ---- 9b. Engine column dictionary (the Rate Engine cohort block) ----
    toc.append(("9b. Rate Engine column dictionary", r))
    section(ws, r, "B", "9b. Rate Engine column dictionary — the cohort block, column by column")
    r += 1
    prose(ws, f"D{r}",
          "The 48-row cohort block appears on the Rate Engine and (same formulas) in every "
          "hidden _calc combo block. Each row is one monthly writing cohort, Jan (P-2) "
          "through Dec (P+1).", size=10)
    r += 1
    coldict = [
        ("A", "#", "Cohort number, 1..48."),
        ("B-D", "Written month / Month end / Days",
         "The month this cohort of policies is written, plus calendar helpers for the "
         "mid-month split."),
        ("E", "Mo#", "Calendar month number — drives the seasonality weight lookup."),
        ("F", "Mid-month", "The serial date at the middle of the writing month; the cohort "
                           "writes there and its mod is sampled there."),
        ("G", "Abs mo#", "Absolute month index (year x 12 + month - 1); the earning matrix "
                         "is keyed on it."),
        ("H", "Weight w", "Written-exposure weight: 1.00 under uniform writings, or the "
                          "state's normalized seasonality weight."),
        ("I", "Index before mo (W_pre)", "Cumulative rate index from changes effective "
                                         "strictly before the month."),
        ("J", "Index at mo end (W_post)", "Cumulative rate index including changes through "
                                          "month end."),
        ("K", "# chgs in mo", "Count of this combo's rate changes effective inside the month."),
        ("L", "Days after 1st chg", "Days in the month on/after the first in-month change "
                                    "date."),
        ("M", "Split p", "Day-weighted share of the month's writings at the post-change "
                         "level: p = L / days in month."),
        ("N", "Written rate index (W)", "The cohort's blended written index: "
                                        "(1 - p) x W_pre + p x W_post."),
        ("O", "Written mod (M_w)", "The written-mod path sampled at the cohort's mid-month "
                                   "(computed on the Mod Engine)."),
        ("P-R", "Earned in P-1 / P / P+1 (ec)",
         "The share of the cohort's premium each calendar year earns: ec = C(end) - "
         "C(start - 1) with C(o) = MIN(1, MAX(0, (o + 0.5) / T))."),
        ("S", "Hist net price (q)", "W x M_w / M_ind — the cohort's combined net price as "
                                    "modeled (used under a net rate selection)."),
        ("T", "Net path (Q)", "Equals q before 1/1/P; from 1/1/P each cohort renews at "
                              "Q(k - 12) x (1 + net selection)."),
        ("U", "Index in force", "What the aggregates actually read: the net path Q under a "
                                "net selection, the written index W otherwise."),
        ("rows 66-73", "Aggregate strip",
         "Per earned month: numerator SUM w-e-W, denominator SUM w-e, earned index E_m, "
         "min/max W bounds with a violation check, earned mod by month, and the mod "
         "numerator behind the price series."),
    ]
    for letter, hdr_txt, meaning in coldict:
        put(ws, f"B{r}", letter, fnt=font(NAVY, bold=True, size=9))
        put(ws, f"C{r}", hdr_txt, fnt=font(GREY_DARK, bold=True, size=9), align=ALIGN_WRAP)
        prose(ws, f"D{r}", meaning, size=9)
        r += 1
    r += 1

    # ---- 10. named ranges, demoted behind a technical-appendix divider ----
    toc.append(("10. Technical appendix — named ranges", r))
    section(ws, r, "B", "10. Technical appendix — named-range dictionary (auto-generated)")
    r += 1
    prose(ws, f"D{r}",
          "Every defined name in the workbook, for auditors and formula readers. Nothing "
          "below is needed to USE the workbook — the exhibits and the sections above are "
          "self-contained.", size=9)
    r += 1
    for j, h in enumerate(["Name", "Refers to", "Description"]):
        put(ws, f"{col(2 + j)}{r}", h, fnt=F_HEADER, fill=FILL_NAVY)
    r += 1
    for name, (ref, desc) in sorted(ctx.names.items()):
        put(ws, f"B{r}", name, fnt=font(NAVY, size=9))
        put(ws, f"C{r}", ref, fnt=font(GREY_DARK, size=9))
        put(ws, f"D{r}", desc or "", fnt=font(GREY_DARK, size=9))
        r += 1

    # ---- Contents (written last, into the reserved rows) ----
    section(ws, 4, "B", "Contents")
    for i, (heading, row_) in enumerate(toc):
        jump(ws, f"B{5 + i}", f"Methodology!B{row_}", heading, size=10)
    presentation_setup(ws, gridlines_off=True, tab_color="D9D9D9")
    print_setup(ws)


# ---------------------------------------------------------------------------
# Read Me
# ---------------------------------------------------------------------------


READ_ME_GUIDE = [
    ("Control", "Pick the plan year, BU and state; global toggles; the six headline KPIs."),
    ("Inputs", "Enter or paste the book: loss ratios and mods (tbl_LR), rate changes, "
               "seasonality, policy term."),
    ("Portfolio", "Every BU x state combo at once — heatmapped, with EP-weighted totals."),
    ("State Summary", "One row per state with a BU filter ('All' = EP-weighted) — the "
                      "leadership exhibit."),
    ("Bridge", "The selected combo's answer: projected LR -> CY plan LR, factor by factor."),
    ("Walkthrough", "The fully worked example — every calculation for the selected combo, "
                    "start to finish."),
    ("Rate Engine", "Audit trail: 48 writing cohorts x earning matrix behind the earned "
                    "rate level."),
    ("Mod Engine", "Audit trail: the written schedule-mod path and the earned mod."),
    ("Flow Dashboard", "When rate and price actually earn in — monthly flow charts and the "
                       "runway table."),
    ("Scenarios", "Four what-if lever sets side by side against Base."),
    ("Solver", "Inverse planning: the required change for a target CY LR, and timing "
               "sensitivity."),
    ("Attribution", "After the year closes: decompose actual vs plan into rate, timing, "
                    "mod, and loss."),
    ("Checks", "The trust panel — must show ALL CHECKS PASS."),
    ("Methodology", "The full writeup: formulas as implemented, conventions, named ranges."),
]

# Canonical glossary (D43). The State Summary column guide retired here; every
# symbol used on any exhibit gets its one plain-English home.
READ_ME_GLOSSARY = [
    ("Indication rate level (CRL_ind)",
     "The cumulative rate level the indication assumed FULLY EARNED — the product of all "
     "rate changes flagged 'in indication'. This is the 'assumed rate in the indication'."),
    ("Earned rate level (E_CY)",
     "The rate level the plan calendar year ACTUALLY earns. Policies written last year carry "
     "older rates into its early months, and new actions earn in only gradually as the book "
     "turns over — so the earned level rarely matches the indication's."),
    ("Rate earn-in (A_rate)",
     "Indication rate level / earned rate level. Above 1.000: the plan year earns less rate "
     "than the indication assumed, pushing the plan LR up. Below 1.000: extra rate the "
     "indication never counted earns in, pushing it down."),
    ("Schedule mods (M_ind / M_0 / M_1)",
     "The average schedule mod assumed in the indication; the current written average; the "
     "projected average at plan-year end. Optional anchors: M_prior (~a year before the "
     "as-of) and M_2 (end of the following year). Drift in the average mod is a price "
     "change that earns in exactly like rate."),
    ("Earned mod / mod drift (A_mod)",
     "The average schedule mod the plan year's premium actually carries, and "
     "A_mod = M_ind / earned mod — the identical earn-in logic applied to the price lever."),
    ("Written vs earned",
     "'Written' = the level at which new policies go on the books each month. 'Earned' = "
     "what the calendar year's premium actually carries, once each policy's term spreads it "
     "across months. Every factor in this workbook is an earned-basis answer."),
    ("Achievement %",
     "For PLANNED rate changes only: the share of the filed change you expect to realize "
     "(80% of a +5% filing = +4.0% effective). Once a change is TAKEN, enter the achieved "
     "rate directly in Filed % — achievement is ignored on taken rows."),
    ("In indication? (Y/N)",
     "Marks the rate changes the indication already counted in its projected loss ratio. "
     "They form CRL_ind; changes not counted there earn in as extra rate."),
    ("Carryover",
     "E_CY(P+1) / E_CY(P) - 1: the earned rate change already locked in for the following "
     "year by this year's actions, before any new ones."),
    ("Unearned runway",
     "Written level / earned level - 1 at a point in time: rate already on the books but "
     "not yet visible in earned premium — it will earn in over the coming months."),
    ("Net rate selection (optional)",
     "Instead of an explicit program, a combo can target a combined rate + price "
     "year-over-year change from 1/1 of the plan year; history still earns as modeled, and "
     "planned rows plus the mod projection are superseded from that date (Methodology 6b). "
     "The bridge then shows one combined factor with mod drift = 1.000."),
    ("Adjusted plan EP",
     "Adjusted plan earned premium (000s) — the weight behind every EP-weighted total on "
     "Portfolio and State Summary."),
    ("Other adjustment (A_other)",
     "A manual factor for anything outside rate and mod earn-in; requires a written reason "
     "when it differs from 1.000."),
]


def build_readme(ctx: Ctx):
    ws = ctx.wb["Read Me"]
    p = ctx.p
    PROSE_W = 100  # column C hosts all paragraphs
    title(ws, "B2", f"Calendar-Year Plan Loss Ratio Workbook — {ctx.lob.name}",
          f"Plan year {p}  |  one workbook per LOB; rows are BU x state  |  version "
          f"{ctx.cfg.version}  |  built "
          f"{dt.date.today():%m/%d/%Y} by src/build_workbook.py (generator v{GENERATOR_VERSION})"
          f"  |  classic formula mode (Excel-2007-era functions only)")

    def para(row, text, size=10, bold=False, color=None):
        prose(ws, f"C{row}", text, size=size, bold=bold, color=color, width=PROSE_W)

    r = 5
    section(ws, r, "B", "Where to go")
    r += 1
    for name, desc in READ_ME_GUIDE:
        jump(ws, f"B{r}", f"{_q(name)}!A1", name, bold=True)
        para(r, desc)
        r += 1
    put(ws, f"B{r}", "_lists / _calc / _oracle", fnt=font(GREY_DARK, size=9))
    para(r, "Hidden machinery: validation lists, the per-combo engine blocks, and the baked "
            "oracle constants. Nothing to edit there.", size=9)
    r += 2

    section(ws, r, "B", "What this workbook does")
    r += 1
    purpose = [
        "Converts the projected loss ratio from a rate level indication (policy-year basis, "
        "rate fully earned) into a CALENDAR-YEAR plan loss ratio at the plan year's average "
        "earned rate level — the number a plan is actually written against.",
        "Method: standard on-leveling (Werner & Modlin, Basic Ratemaking, Ch. 5) through a "
        "monthly written-index x earning-matrix engine, extended with a schedule-mod "
        "adjustment, optional seasonality, and an optional net rate selection.",
        "Every displayed figure is calculated, never typed, and ties to an independent "
        "Python oracle at build time; the Checks sheet proves the workbook's integrity live.",
    ]
    for t in purpose:
        para(r, t)
        r += 1
    r += 1

    put(ws, f"B{r}", "THE BRIDGE IN ONE LINE", fnt=font(NAVY, bold=True, size=11),
        fill=FILL_PANEL)
    prose(ws, f"C{r}", "CY plan LR  =  projected LR (current level)  x  rate earn-in  x  "
                       "mod drift  x  other adj", size=11, bold=True, color=NAVY,
          width=PROSE_W)
    ws[f"C{r}"].fill = FILL_PANEL
    r += 1
    para(r, "The following year (indicative) additionally applies one year of net trend and "
            "uses that year's earn-in factors.", size=9)
    r += 2

    section(ws, r, "B", "Quick start")
    r += 1
    steps = [
        "Replace the SAMPLE rows on the Inputs sheet with your book: tbl_LR (one row per "
        "BU x state), the rate change log (one row per change), tbl_Seasonality (optional, "
        "by state), and the policy term. Do not insert or delete rows — spare rows are "
        "provided. Rename BUs and states directly in tbl_LR: every dropdown and exhibit "
        "follows the live roster automatically.",
        "On Control, pick the plan year, BU, and state, and set the global toggles "
        "(seasonality, the mod-adjustment master switch, the default net trend).",
        "Read the KPI row on Control; audit any figure through Rate Engine -> Mod Engine "
        "-> Bridge.",
        "State Summary is the per-state exhibit (BU filter; 'All' combines on adjusted-EP "
        "weights). Portfolio shows every combo at once.",
        "Scenarios: up to four lever sets (rate points, timing shift, achievement, mod "
        "path, net selection) side by side against Base.",
        "Optional NET RATE SELECTION per combo (tbl_LR): a combined rate + price "
        "year-over-year target from 1/1 of the plan year — see Methodology 6b for exactly "
        "what it supersedes.",
        "Solver: Mode A returns the single change needed for a target CY LR at a chosen "
        "date (it replaces the whole planned program — see the note there); Mode B shows "
        "the full-year LR under each possible start month.",
        "Attribution (after the year closes): enter achieved changes, actual dates, the "
        "actual mod path, and the actual CY LR for a plan-vs-actual waterfall.",
        "Checks must show ALL CHECKS PASS (also surfaced on Control). Oracle-tie rows apply "
        "only while the seeded sample is intact and report N/A afterward — regenerate the "
        "workbook to re-bake them for your data. For a new plan year, regenerate from "
        "config/config.yaml and paste inputs into the fresh copy (table schemas are stable).",
    ]
    for i, s in enumerate(steps):
        put(ws, f"B{r}", f"{i + 1}.", fnt=font(NAVY, bold=True, size=10))
        para(r, s)
        r += 1
    r += 1

    section(ws, r, "B", "Color legend")
    r += 1
    legend = [
        ("Blue font on light yellow", "Required input — enter a value", F_INPUT, FILL_REQ),
        ("Blue font, no fill", "Optional input / lever", F_INPUT, None),
        ("Black font", "Formula calculated on the same sheet", F_LABEL, None),
        ("Green font", "Link pulled from another sheet", F_LINK, None),
        ("Grey small font", "Engine helper — do not edit", font(GREY_DARK, size=9), FILL_GREY),
    ]
    for lbl, desc, fnt_, fill_ in legend:
        put(ws, f"B{r}", lbl, fnt=fnt_, fill=fill_, border=BORDER_THIN)
        put(ws, f"C{r}", desc, fnt=font(GREY_DARK, size=10))
        r += 1
    r += 1

    section(ws, r, "B", "Glossary — terms and symbols")
    r += 1
    for term, definition in READ_ME_GLOSSARY:
        put(ws, f"B{r}", term, fnt=font(NAVY, bold=True, size=10), align=ALIGN_WRAP)
        para(r, definition)
        r += 1
    r += 1

    section(ws, r, "B", "Conventions and notes")
    r += 1
    notes_ = [
        "Percentages are stored as decimal fractions (5.0% = 0.05) and formatted as percents.",
        "No VBA or macros anywhere. No merged cells. No external links. No volatile "
        "functions. Automatic calculation. Classic formula mode: only Excel-2007-era "
        "functions, so the file recalculates in LibreOffice and any corporate Excel build.",
        "Sheets are NOT protected (an intentional choice — the Checks panel is the integrity "
        "mechanism; protection without passwords would add friction without security).",
        "First-open checklist: (1) open the file and let it calculate; (2) confirm Control "
        f"shows ALL CHECKS PASS; (3) confirm the Bridge shows {ctx.oracle_m.cy_lr_p:.2%} "
        f"for {ctx.we_key} while sample data is intact; (4) replace sample inputs with "
        "your book.",
    ]
    for s in notes_:
        para(r, s)
        r += 1
    r += 1
    section(ws, r, "B", "Version log")
    r += 1
    put(ws, f"C{r}", f"v{ctx.cfg.version} — {dt.date.today():%m/%d/%Y} — built from "
                     "config/config.yaml (see DECISIONS.md in the repo for every judgment call).",
        fnt=font(GREY_DARK, size=10))
    set_widths(ws, {"A": 2, "B": 30, "C": PROSE_W})
    presentation_setup(ws, gridlines_off=True, tab_color="D9D9D9")
    print_setup(ws)
