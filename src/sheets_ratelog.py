"""Rate Log sheet: the rate-change entry table on its OWN sheet.

Split out of Inputs so the log gets a correct frozen header and, most
importantly, a same-row feedback loop: type a date or a change and columns
U:X show that combo's plan LR move on the same line. The paste block A:H is
byte-identical to the pre-split layout (same columns, same DV, same capacity)
— only the sheet changed. Engine formulas read the rl_* names, which this
module re-registers against SHEETS.RATE_LOG.
"""

from __future__ import annotations

from openpyxl.worksheet.table import Table

from .build_workbook import Ctx, Layout as L, SHEETS
from .sheets_inputs import RL_HELPER_HEADERS, TABLE_STYLE, _dv, rl_headers
from .xlstyle import (
    ALIGN_C, ALIGN_L, BORDER_THIN, F_SMALL_IT, FAIL_RED, FILL_GREY, FMT_DATE,
    FMT_IDX, FMT_INT, FMT_PCT, GREY_DARK, WARN_AMBER, font, formula, header_row,
    input_cell, nav_bar, presentation_setup, print_setup, put, quote_sheet,
    set_widths, title,
)

RE = quote_sheet(SHEETS.RATE_ENGINE)
PCT_S = "+0.0%;-0.0%;0.0%"


def build_rate_log(ctx: Ctx):
    ws = ctx.wb[SHEETS.RATE_LOG]
    title(ws, "A1", f"Rate Change Log — {ctx.lob.name}",
          "One row per rate change per BU x state. Columns A:H are one contiguous paste "
          "block; planned rows earn filed % x achievement %. The live columns on the right "
          "show the impact of every edit on the same row.")
    nav_bar(ws, 3, 1, ["Control", "Inputs", "Bridge", "Walkthrough", "Checks"], step=2)
    put(ws, "A4",
        "SAMPLE DATA: replace the populated rows with your own log. Do not insert or "
        "delete rows — spare capacity is provided below the sample.",
        fnt=font(FAIL_RED, bold=True, italic=True))

    header_row(ws, L.RL_HDR, 1, rl_headers(), widths=[9, 8, 11, 9, 12, 12, 13, 34])
    ws.row_dimensions[L.RL_HDR].height = 30
    header_row(ws, L.RL_HDR, 10, RL_HELPER_HEADERS + ["1st planned", "1st taken"],
               widths=[14, 8, 8, 11, 12, 10, 9, 7, 11, 6, 6, 7], fill=FILL_GREY,
               fnt=font(GREY_DARK, bold=True, size=9))
    put(ws, f"J{L.RL_HDR - 1}", "engine helpers (formulas — do not edit)", fnt=F_SMALL_IT)
    header_row(ws, L.RL_HDR, 22,
               ["This combo's CY plan LR", "Chg vs projected (pts)",
                "% of plan yr this action earns (selected combo)", "Warnings"],
               widths=[13, 12, 15, 24], fill=FILL_GREY,
               fnt=font(GREY_DARK, bold=True, size=9))
    put(ws, f"V{L.RL_HDR - 1}", "live results (formulas — do not edit)", fnt=F_SMALL_IT)
    ws.column_dimensions["I"].width = 2

    # share-of-year machinery for column X (exact, via the selected combo's
    # Rate Engine block; rows of other combos show blank)
    wrng = f"{RE}!$H${L.RE_COH_FIRST}:$H${L.RE_COH_LAST}"
    ecp = f"{RE}!$Q${L.RE_COH_FIRST}:$Q${L.RE_COH_LAST}"
    absmi = f"{RE}!$G${L.RE_COH_FIRST}:$G${L.RE_COH_LAST}"

    for i in range(L.RL_ROWS):
        r = L.RL_FIRST + i
        row = ctx.rate_rows[i] if i < len(ctx.rate_rows) else None
        input_cell(ws, f"A{r}", row and row["bu"])
        input_cell(ws, f"B{r}", row and row["state"])
        input_cell(ws, f"C{r}", row and row["eff"], fmt=FMT_DATE)
        input_cell(ws, f"D{r}", row and row["filed"], fmt=FMT_PCT)
        input_cell(ws, f"E{r}", row and row["status"])
        input_cell(ws, f"F{r}", row and row["considered"])
        input_cell(ws, f"G{r}", row and row["achievement"], fmt=FMT_PCT, required=False)
        input_cell(ws, f"H{r}", (row["comment"] if row else None), required=False)
        ws[f"H{r}"].alignment = ALIGN_L
        # helpers J..S (identical formulas to the pre-split Inputs layout)
        formula(ws, f"J{r}", f'=IF(OR($A{r}="",$B{r}=""),"",$A{r}&"|"&$B{r})', border=BORDER_THIN)
        formula(ws, f"K{r}", f'=IF($C{r}="","",$D{r}*IF($E{r}="planned",IF($G{r}="",1,$G{r}),1))',
                fmt=FMT_PCT)
        formula(ws, f"L{r}", f'=IF($C{r}="",0,IF(1+$K{r}>0,LN(1+$K{r}),0))', fmt=FMT_IDX)
        formula(ws, f"M{r}", f'=IF($C{r}="","",DATE(YEAR($C{r}),MONTH($C{r}),1))', fmt=FMT_DATE)
        formula(ws, f"N{r}", f'=IF($C{r}="",0,EOMONTH($C{r},0)-$C{r}+1)', fmt=FMT_INT)
        formula(ws, f"O{r}", f'=IF($C{r}="",0,COUNTIFS(rl_key,$J{r},rl_effmonth,$M{r},rl_eff,"<"&$C{r}))',
                fmt=FMT_INT)
        formula(ws, f"P{r}", f'=IF($C{r}="",0,COUNTIFS($J${L.RL_FIRST}:$J{r},$J{r},$M${L.RL_FIRST}:$M{r},$M{r},$C${L.RL_FIRST}:$C{r},$C{r}))',
                fmt=FMT_INT)
        formula(ws, f"Q{r}", f'=IF($C{r}="",0,IF(AND($O{r}=0,$P{r}=1),1,0))', fmt=FMT_INT)
        formula(ws, f"R{r}", f'=IF($C{r}="",0,IF(COUNTIFS(rl_key,$J{r},rl_effmonth,$M{r})>1,1,0))',
                fmt=FMT_INT)
        formula(ws, f"S{r}",
                f'=IF($C{r}="",0,COUNTIFS(rl_key,$J{r},rl_eff,"<"&$C{r})'
                f"+COUNTIFS($J${L.RL_FIRST}:$J{r},$J{r},$C${L.RL_FIRST}:$C{r},$C{r}))",
                fmt=FMT_INT)
        # first PLANNED row of this key effective inside the plan year — the
        # filing Net Delivery adopts as its default change (D63). Same
        # discipline as the taken flag below: the window and the status in the
        # flag are exactly the ones its consumer filters on (D58).
        jan1, dec31 = "DATE(nr_PlanYear,1,1)", "DATE(nr_PlanYear,12,31)"
        formula(ws, f"T{r}",
                f'=IF(OR($C{r}="",$E{r}<>"planned",$C{r}<{jan1},$C{r}>{dec31}),0,'
                f'IF(AND(COUNTIFS(rl_key,$J{r},rl_status,"planned",'
                f'rl_eff,">="&{jan1},rl_eff,"<"&$C{r})=0,'
                f'COUNTIFS($J${L.RL_FIRST}:$J{r},$J{r},$E${L.RL_FIRST}:$E{r},"planned",'
                f"$C${L.RL_FIRST}:$C{r},$C{r})=1),1,0))",
                fmt=FMT_INT)
        # first change among the TAKEN rows of this key + cohort month. The
        # plain first-flag (Q) is status-blind, so a planned row earlier in
        # the month would steal it and the Solver's taken-only base would
        # silently drop the taken change from that cohort (D58).
        formula(ws, f"U{r}",
                f'=IF(OR($C{r}="",$E{r}<>"taken"),0,'
                f'IF(AND(COUNTIFS(rl_key,$J{r},rl_effmonth,$M{r},rl_eff,"<"&$C{r},'
                f'rl_status,"taken")=0,'
                f"COUNTIFS($J${L.RL_FIRST}:$J{r},$J{r},$M${L.RL_FIRST}:$M{r},$M{r},"
                f'$C${L.RL_FIRST}:$C{r},$C{r},$E${L.RL_FIRST}:$E{r},"taken")=1),1,0))',
                fmt=FMT_INT)
        for cL in "JKLMNOPQRSTU":
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=9)
        # live echo columns V:Y — the same-row feedback loop
        formula(ws, f"V{r}",
                f'=IF($J{r}="","",IF(COUNTIF(lr_key,$J{r})=0,"not in tbl_LR",'
                f"INDEX(calc_cylr_p,MATCH($J{r},lr_key,0))))", fmt=FMT_PCT, align=ALIGN_C)
        formula(ws, f"W{r}",
                f'=IF(OR($J{r}="",COUNTIF(lr_key,$J{r})=0),"",'
                f"(INDEX(calc_cylr_p,MATCH($J{r},lr_key,0))"
                f"-INDEX(calc_lrcur,MATCH($J{r},lr_key,0)))*100)",
                fmt='+0.00;-0.00;0.00', align=ALIGN_C)
        formula(ws, f"X{r}",
                f'=IF(OR($C{r}="",$J{r}<>nr_SelKey),"",'
                f"(SUMPRODUCT(({absmi}>YEAR($C{r})*12+MONTH($C{r})-1)*{wrng},{ecp})"
                f"+(EOMONTH($C{r},0)-$C{r}+1)/DAY(EOMONTH($C{r},0))"
                f"*SUMPRODUCT(({absmi}=YEAR($C{r})*12+MONTH($C{r})-1)*{wrng},{ecp}))"
                f"/SUMPRODUCT({wrng},{ecp}))", fmt="0%", align=ALIGN_C)
        formula(ws, f"Y{r}",
                f'=IF($C{r}="","",IF(AND($E{r}="taken",$G{r}<>"",$G{r}<>1),'
                f'"achievement ignored — restate Filed %",""))')
        for cL in "VWXY":
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=9)
        ws[f"Y{r}"].font = font(FAIL_RED, size=9, italic=True)

    tbl = Table(displayName="tbl_RateLog", ref=f"A{L.RL_HDR}:H{L.RL_LAST}")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)
    rl_names = {
        "rl_bu": ("A", "tbl_RateLog business unit"), "rl_state": ("B", "tbl_RateLog state"),
        "rl_eff": ("C", "Rate change effective date (start of day)"),
        "rl_filed": ("D", "Filed rate change (decimal fraction)"),
        "rl_status": ("E", "taken | planned"),
        "rl_cons": ("F", "Considered in the indication? (Y/N)"),
        "rl_ach": ("G", "Achievement % (planned rows; blank = 100%)"),
        "rl_key": ("J", "Helper: BU|State key (right of the paste block, D40)"),
        "rl_reff": ("K", "Helper: effective change = filed x achievement (planned only)"),
        "rl_ln1p": ("L", "Helper: ln(1 + effective change); 0 for blank rows"),
        "rl_effmonth": ("M", "Helper: first day of the effective month"),
        "rl_daysafter": ("N", "Helper: days in effective month on/after the change"),
        "rl_first": ("Q", "Helper: 1 for the first change in its BU|State cohort month"),
        "rl_dup": ("R", "Helper: 1 when >1 change shares a BU|State cohort month"),
        "rl_seq": ("S", "Helper: chronological rank of the change within its BU|State key"),
        "rl_firstplanned": ("T", "Helper: 1 for the key's FIRST planned change effective "
                                 "inside the plan year (Net Delivery's default filing, D63)"),
        "rl_firsttaken": ("U", "Helper: 1 for the first TAKEN change in its key + cohort "
                               "month (the Solver's status-aware split flag, D58)"),
    }
    for name, (colL, desc) in rl_names.items():
        ctx.define(name, SHEETS.RATE_LOG, f"${colL}${L.RL_FIRST}:${colL}${L.RL_LAST}", desc)

    fr, lr_ = L.RL_FIRST, L.RL_LAST
    _dv(ws, "list", [f"A{fr}:A{lr_}"], formula1="=lst_bu")
    _dv(ws, "list", [f"B{fr}:B{lr_}"], formula1="=lst_state")
    _dv(ws, "list", [f"E{fr}:E{lr_}"], formula1="=lst_status")
    _dv(ws, "list", [f"F{fr}:F{lr_}"], formula1="=lst_yn")
    _dv(ws, "decimal", [f"D{fr}:D{lr_}"], operator="between", formula1="-0.5", formula2="2",
        error="Filed change must lie in [-50%, +200%] (decimal fraction).")
    _dv(ws, "decimal", [f"G{fr}:G{lr_}"], operator="between", formula1="0", formula2="1.5",
        error="Achievement must lie in [0%, 150%].")
    _dv(ws, "date", [f"C{fr}:C{lr_}"], operator="between",
        formula1="DATE(1990,1,1)", formula2="DATE(2100,12,31)", error="Enter a real date.")

    presentation_setup(ws, gridlines_off=False, freeze=f"C{L.RL_FIRST}",
                       tab_color=WARN_AMBER)
    print_setup(ws)
