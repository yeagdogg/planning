"""The Book: one combined, filterable roll-up across every line (D66).

Six workbooks, one per LOB (D37), each publishing one row per BU x state.
This module stacks those harvested rows into a seventh workbook and lays the
familiar exhibits over them.

What "not live" means precisely: the ENGINES are frozen — every per-combo
figure is a harvested value, so changing an input in an LOB workbook does not
move this file until you regenerate, recalculate and re-harvest. Every FILTER
and AGGREGATE is a live formula, so switching the view recalculates instantly.

Filter doctrine: an exhibit keyed by a dimension applies the OTHER filters,
never its own — the per-state exhibits honour the LOB and BU pickers, the
by-LOB roll-up honours BU and state, and the all-combo grid honours all three.

Weighting: adjusted plan EP throughout, and EP x written weight for the
monthly legs. Both are valid across lines — seasonality is state-keyed and
shared across LOBs, and written-basis YoY legs are term-independent (pinned by
tests/test_program_flow.py::test_term_independent), so Inland Marine's
six-month term combines safely. Earning fractions are never aggregated.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from .sheets_main import (SS_COLS, SS_G_BRIDGE, SS_G_HIST, SS_G_P1,
                         ss_c, ss_conditional_formats, ss_groups, ss_l, ss_outline)
from .xlstyle import (ALIGN_C, ALIGN_L, BORDER_THIN, FAIL_RED, FILL_GREY, FILL_NAVY,
    FONT_NAME,
    FILL_PANEL, FMT_DATE, FMT_DATE_S, FMT_IDX, FMT_INT, FMT_MOD, FMT_MONTH, FMT_PCT,
    FMT_PTS_SIGNED, F_LABEL, F_SMALL_IT, GREY_DARK, NAVY, PASS_GREEN, STEEL_LIGHT,
    assert_formulas_balanced, col, font, formula, header_row, input_cell, jump,
    label, link, nav_bar, presentation_setup, print_setup, put, quote_sheet,
    section, set_widths, status_banner_cf, title,
)

from .xlstyle import FMT_PCT_SIGNED as PCT_S   # one definition, seven readers
from .xlstyle import FMT_PTS_COL as PTS  # points in a column already headed (pts)
ALL = "All"


def _ss_chain(r) -> str:
    """The visible bridge product, in the shared column order."""
    return "*".join(f'${ss_l(k)}{r}'
                    for k in ("lrcur", "arate", "amod", "aother"))


def _ss_chain1(r) -> str:
    return (f'${ss_l("lrcur")}{r}*(1+${ss_l("trend")}{r})'
            f'*${ss_l("arate1")}{r}*${ss_l("amod1")}{r}'
            f'*${ss_l("aother")}{r}')

BOOK = "_book"
PIVOT = "Pivot Data"
SHEET_ORDER = ["Read Me", "Control", "State Summary", "Portfolio", "Roll-ups",
               "Program Flow", PIVOT, "Checks", BOOK]

# ---- _book layout ---------------------------------------------------------
BK_HDR = 1
BK_FIRST = 2
# scalar columns, in order: (field, header, format)
SCALARS = [
    ("lob", "LOB", None), ("bu", "BU", None), ("state", "State", None),
    ("key", "Key", None), ("ukey", "LOB|BU|State", None),
    ("ep", "Adj plan EP", "#,##0"),
    ("term", "Term", FMT_INT),
    ("crl", "CRL_ind", FMT_IDX), ("ecy_p", "E_CY(P)", FMT_IDX),
    ("mbar_p", "Earned mod", FMT_MOD), ("arate_p", "A_rate", FMT_IDX),
    ("amod_p", "A_mod", FMT_IDX), ("aother", "A_other", FMT_IDX),
    ("lrcur", "Projected LR", FMT_PCT), ("cylr_p", "Plan LR", FMT_PCT),
    ("cylr_p1", "Plan LR +1", FMT_PCT), ("trend", "Net trend", PCT_S),
    ("arate_p1", "A_rate +1", FMT_IDX), ("amod_p1", "A_mod +1", FMT_IDX),
    ("echg", "Earned chg vs ind", PCT_S), ("carry", "Carryover", PCT_S),
    ("netmode", "Net?", FMT_INT), ("netx", "Net target", PCT_S),
    ("modeff", "Mod adj on", FMT_INT),
    ("cylr_prog", "Plan LR (program)", FMT_PCT),
    ("cylr1_prog", "Plan LR +1 (program)", FMT_PCT),
    ("proggap", "Program vs asserted (pts)", PTS),
    ("progident", "Non-net identity", FMT_IDX),
    ("ntaken", "# taken", FMT_INT), ("nplanned", "# planned", FMT_INT),
    ("avg_rate", "Avg rate ratio", FMT_IDX), ("avg_mod", "Avg mod ratio", FMT_IDX),
    ("avg_del", "Avg delivered ratio", FMT_IDX),
]
# EP-weighted products computed at build time (SUMIFS cannot multiply columns)
WEIGHTED = [("w_crl", "crl"), ("w_ecy", "ecy_p"), ("w_mbar", "mbar_p"),
            ("w_arate", "arate_p"), ("w_amod", "amod_p"), ("w_aother", "aother"),
            ("w_lrcur", "lrcur"), ("w_cylr", "cylr_p"), ("w_cylr1", "cylr_p1"),
            ("w_trend", "trend"), ("w_arate1", "arate_p1"),
            ("w_amod1", "amod_p1"), ("w_prog", "cylr_prog"),
            ("w_prog1", "cylr1_prog")]
# harvested EP-weighted mods (the raw per-combo mods are not published)
WEIGHTED_DIRECT = [("w_mind", "m_ind_w"), ("w_m0", "m0_w"), ("w_m1", "m1_w"),
                   ("w_target", "target_w")]                     # D96

# ---- D103: the long (tidy) pivot dataset ----------------------------------
# (Measure, Category, harvested field, the field is ALREADY x EP)
PIVOT_ANNUAL = [
    ("Projected LR (current level)", "Bridge", "lrcur", False),
    ("Rate earn-in (A_rate)", "Bridge", "arate_p", False),
    ("Mod drift (A_mod)", "Bridge", "amod_p", False),
    ("Other adjustment (A_other)", "Bridge", "aother", False),
    ("Plan LR", "Bridge", "cylr_p", False),
    ("Target LR", "Bridge", "target_w", True),
    ("Plan LR — program basis", "Bridge", "cylr_prog", False),
    ("Plan LR +1", "Next year", "cylr_p1", False),
    ("Rate earn-in +1", "Next year", "arate_p1", False),
    ("Mod drift +1", "Next year", "amod_p1", False),
    ("Net trend", "Next year", "trend", False),
    ("Plan LR +1 — program basis", "Next year", "cylr1_prog", False),
    ("Indication rate level (CRL)", "Levels", "crl", False),
    ("Earned rate level", "Levels", "ecy_p", False),
    ("Earned mod", "Levels", "mbar_p", False),
    ("Mod in indication (M_ind)", "Mods", "m_ind_w", True),
    ("Current mod (M_0)", "Mods", "m0_w", True),
    ("Projected mod (M_1)", "Mods", "m1_w", True),
]
# (Measure, harvested monthly leg, the mass is gated by the mod adjustment)
PIVOT_MONTHLY = [
    ("Delivered YoY", "delivered", False),
    ("Rate YoY", "rate", False),
    ("Mod YoY", "mod", True),
]
PIVOT_HEADERS = ["LOB", "BU", "State", "Category", "Measure", "Month",
                 "Weight", "Weighted value"]
# monthly families: EP x weight, and the same times each leg (mod legs are
# additionally gated by the combo's mod-adjustment flag)
MONTHLY = ["epw", "epw_del", "epw_rate", "epwm", "epwm_mod"]
SLOTS = 4


@dataclass
class BookCtx:
    """Minimal name registry — Ctx is LOB-specific and Layout is a configured
    global, so the book keeps its own (mirrors Ctx.define's case guard, D32)."""

    wb: Workbook
    data: object
    names: dict = field(default_factory=dict)

    def define(self, name: str, sheet: str, ref: str, description: str):
        for existing in self.names:
            if existing.lower() == name.lower() and existing != name:
                raise ValueError(f"Defined-name case collision: {name!r} vs {existing!r}")
        if name in self.names:
            raise ValueError(f"Defined name {name!r} registered twice")
        self.names[name] = (f"{quote_sheet(sheet)}!{ref}", description)

    def flush_names(self):
        for name, (ref, _d) in sorted(self.names.items()):
            self.wb.defined_names.add(DefinedName(name, attr_text=ref))


def _cols() -> dict:
    """field -> column index on _book (single source of truth for addresses)."""
    m, c = {}, 1
    for f, _h, _fmt in SCALARS:
        m[f], c = c, c + 1
    for f, _src in WEIGHTED + WEIGHTED_DIRECT:
        m[f], c = c, c + 1
    for fam in MONTHLY:
        for j in range(12):
            m[f"{fam}_{j}"], c = c, c + 1
    for j in range(SLOTS):
        for k, part in enumerate(("date", "pct", "tok")):
            m[f"slot{j}_{part}"], c = c, c + 1
    return m


COLS = _cols()
LAST_COL = max(COLS.values())


def rng(field_: str, n: int) -> str:
    """A full-height _book column range for a harvested field."""
    cl = col(COLS[field_])
    # always quote the hidden sheet: a bare _book! is ambiguous with a defined
    # name (both may start with an underscore), which Excel rejects outright
    return f"'{BOOK}'!${cl}${BK_FIRST}:${cl}${BK_FIRST + n - 1}"


# ---------------------------------------------------------------------------
# hidden data sheet
# ---------------------------------------------------------------------------


def pivot_rows(data, plan_year: int) -> list[tuple]:
    """The long dataset: one row per combo x measure, x month for the legs (D103).

    Long rather than wide for a reason that is not taste. The three monthly legs
    do NOT share a denominator — rate and delivered weight by ``epw`` (EP x the
    seasonality weight), the mod leg by ``epwm``, which is additionally gated by
    the combo's mod-adjustment flag so combos with it OFF drop out rather than
    entering as zeros. In a wide table nothing stops a reader pairing the wrong
    numerator and denominator, and the result is a plausible number that is
    quietly wrong. Here each row carries its OWN weight, so one calculated field

        Weighted = SUM([Weighted value]) / SUM([Weight])

    is correct for every measure, every slice, every subtotal — and for a single
    row too, which is why there is no raw value column: it would be redundant
    and it is the one field that could be dragged in and averaged.

    Rows with no weight are OMITTED rather than written as zeros. A combo with
    no premium, or a month whose mod adjustment is off, has no answer — not an
    answer of zero — and an absent row cannot drag an average toward it.
    """
    out = []
    for rec in data.rows:
        ep = rec["ep"] or 0.0
        dims = (rec["lob"], rec["bu"], rec["state"])
        if ep > 0:
            for measure, cat, src, pre in PIVOT_ANNUAL:
                v = rec[src]
                if v is None:
                    continue        # see the docstring: no answer, not a zero
                out.append((*dims, cat, measure, None, ep, v if pre else ep * v))
        modeff = 1.0 if (rec["modeff"] or 0) == 1 else 0.0
        for j in range(12):
            epw = rec["epw"][j] or 0.0
            if epw <= 0:
                continue
            month = dt.date(plan_year, j + 1, 1)
            # /12 so Weight means PREMIUM in every row of this table. The
            # seasonality weight averages 1 over the year (12 x se(m)/sum(se)),
            # so raw epw sums to 12 x EP — a fine relative mass, and a
            # confusing thing to call "Weight" beside annual rows that carry EP
            # itself. Rescaled, a month's Weight is that month's written
            # premium and the twelve sum to EP. Every ratio is unchanged.
            for measure, leg, gated in PIVOT_MONTHLY:
                mass = (epw * modeff if gated else epw) / 12.0
                if mass <= 0:
                    continue
                out.append((*dims, "Monthly", measure, month,
                            mass, mass * (rec[leg][j] or 0.0)))
    return out


def build_pivot_data(ctx: BookCtx, plan_year: int) -> int:
    """Write the long dataset as an Excel TABLE, so a pivot built on it grows
    with the roster instead of pointing at a stale rectangle (D103).

    Nothing else goes on this sheet — no title, no notes. A pivot source has to
    start at A1 or the table reference and the header row stop agreeing; the
    guidance lives on Read Me, where someone will actually look for it.
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ws = ctx.wb[PIVOT]
    rows = pivot_rows(ctx.data, plan_year)
    for i, h in enumerate(PIVOT_HEADERS):
        put(ws, f"{col(1 + i)}1", h, fnt=font(GREY_DARK, bold=True, size=9))
    for i, rec in enumerate(rows):
        r = 2 + i
        lob, bu, state, cat, measure, month, weight, wv = rec
        put(ws, f"A{r}", lob)
        put(ws, f"B{r}", bu)
        put(ws, f"C{r}", state)
        put(ws, f"D{r}", cat)
        put(ws, f"E{r}", measure)
        if month is not None:
            put(ws, f"F{r}", month, fmt=FMT_MONTH)
        put(ws, f"G{r}", weight, fmt="#,##0.000")
        put(ws, f"H{r}", wv, fmt="#,##0.000")
    tab = Table(displayName="tbl_Pivot", ref=f"A1:H{1 + len(rows)}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    set_widths(ws, {"A": 17, "B": 8, "C": 7, "D": 11, "E": 30, "F": 10,
                    "G": 14, "H": 15})
    ws.freeze_panes = "A2"
    print_setup(ws)
    return len(rows)


def build_book_data(ctx: BookCtx):
    ws = ctx.wb[BOOK]
    rows = ctx.data.rows
    n = len(rows)
    put(ws, "A1", None)
    heads = ([h for _f, h, _x in SCALARS]
             + [f for f, _s in WEIGHTED + WEIGHTED_DIRECT]
             + [f"{fam}_{j + 1}" for fam in MONTHLY for j in range(12)]
             + [f"chg{j + 1}_{p}" for j in range(SLOTS)
                for p in ("date", "pct", "tok")])
    for i, h in enumerate(heads):
        put(ws, f"{col(1 + i)}{BK_HDR}", h, fnt=font(GREY_DARK, bold=True, size=8))
    for i, rec in enumerate(rows):
        r = BK_FIRST + i
        ep = rec["ep"] or 0.0
        for f, _h, fmt in SCALARS:
            v = (f"{rec['lob']}|{rec['key']}" if f == "ukey" else rec[f])
            put(ws, f"{col(COLS[f])}{r}", v, fmt=fmt or "General")
        for f, src in WEIGHTED:
            put(ws, f"{col(COLS[f])}{r}", ep * (rec[src] or 0.0), fmt=FMT_IDX)
        for f, src in WEIGHTED_DIRECT:
            put(ws, f"{col(COLS[f])}{r}", rec[src] or 0.0, fmt=FMT_IDX)
        modeff = 1.0 if (rec["modeff"] or 0) == 1 else 0.0
        for j in range(12):
            epw = rec["epw"][j] or 0.0
            vals = {"epw": epw,
                    "epw_del": epw * (rec["delivered"][j] or 0.0),
                    "epw_rate": epw * (rec["rate"][j] or 0.0),
                    "epwm": epw * modeff,
                    "epwm_mod": epw * modeff * (rec["mod"][j] or 0.0)}
            for fam, v in vals.items():
                put(ws, f"{col(COLS[f'{fam}_{j}'])}{r}", v, fmt=FMT_IDX)
        for j in range(SLOTS):
            d, pct, tok = rec["slots"][j]
            put(ws, f"{col(COLS[f'slot{j}_date'])}{r}", d, fmt=FMT_DATE_S)
            put(ws, f"{col(COLS[f'slot{j}_pct'])}{r}", pct, fmt=PCT_S)
            put(ws, f"{col(COLS[f'slot{j}_tok'])}{r}", tok)
    for f in ("lob", "bu", "state", "key", "ukey", "ep", "netmode", "netx",
              "progident",
              "proggap", "ntaken", "nplanned", "cylr_p", "cylr_prog", "modeff"):
        ctx.define(f"bk_{f}", BOOK, rng(f, n).split("!")[1],
                   f"Harvested per-combo {f}")
    ws.sheet_state = "hidden"
    return n


# ---------------------------------------------------------------------------
# shared filter vocabulary
# ---------------------------------------------------------------------------


_N = 0          # harvested row count, set by build_book()


def crit(*dims) -> str:
    """SUMIFS criteria pairs for the named filter dimensions."""
    return "".join(f",{rng(d, _N)},bkc_{d}" for d in dims)


def sumifs(value_field: str, *dims, extra: str = "") -> str:
    return f"SUMIFS({rng(value_field, _N)}{extra}{crit(*dims)})"


def wtd(num_field: str, *dims, extra: str = "") -> str:
    """EP-weighted mean of a metric, blank when nothing is in view."""
    den = sumifs("ep", *dims, extra=extra)
    return f'IF({den}=0,"—",{sumifs(num_field, *dims, extra=extra)}/{den})'


# ---------------------------------------------------------------------------
# Control
# ---------------------------------------------------------------------------


def build_control(ctx: BookCtx, n: int):
    ws = ctx.wb["Control"]
    d = ctx.data
    title(ws, "A1", f"The Book — every line of business at once ({d.plan_year} plan)",
          "Harvested from the six LOB workbooks. The engines are frozen; every "
          "filter and total below is live. An exhibit keyed by a dimension applies "
          "the OTHER filters — the per-state views honour LOB and BU, the by-LOB "
          "roll-up honours BU and state, and the all-combo grid honours all three.")
    nav_bar(ws, 3, 1, ["State Summary", "Portfolio", "Roll-ups", "Program Flow",
                       "Checks", "Read Me"], step=2)
    put(ws, "A4", f"As of {d.as_of} — harvested from {len(d.sources)} workbooks, "
                  f"{n} combos. Regenerate and recalculate the LOB files, then "
                  f"rebuild this book, to refresh.", fnt=F_SMALL_IT)

    section(ws, 6, "A", "Filters")
    for i, (lbl, name, values) in enumerate((
            ("Line of business", "bkf_lob", d.lobs),
            ("Business unit", "bkf_bu", d.business_units),
            ("State", "bkf_state", d.states))):
        r = 7 + i
        label(ws, f"A{r}", lbl, bold=True)
        c = input_cell(ws, f"B{r}", ALL)
        c.alignment = ALIGN_C
        ctx.define(name, "Control", f"$B${r}", f"{lbl} filter ('{ALL}' = every one)")
        # the pick list lives beside the picker so the DV needs no helper sheet
        for j, v in enumerate([ALL] + list(values)):
            put(ws, f"{col(20 + i)}{7 + j}", v, fnt=font(GREY_DARK, size=8))
        rngref = f"$#{col(20 + i)}$7:$#{col(20 + i)}${7 + len(values)}".replace("#", "")
        dv = DataValidation(type="list", formula1=f"=Control!{rngref}",
                            allow_blank=False, showErrorMessage=True)
        ws.add_data_validation(dv)
        dv.add(f"B{r}")
        # SUMIFS wildcard criteria (the State Summary idiom)
        dim = ("lob", "bu", "state")[i]
        formula(ws, f"D{r}", f'=IF({name}="{ALL}","*",{name})')
        ws[f"D{r}"].font = font(GREY_DARK, size=8)
        ctx.define(f"bkc_{dim}", "Control", f"$D${r}",
                   f"{lbl} SUMIFS criterion ('*' when {ALL})")

    section(ws, 11, "A", "In view")
    cards = [
        ("Combos in view",
         f'=COUNTIFS({rng("key", n)},"?*"{crit("lob", "bu", "state")})',
         FMT_INT, "rows passing all three filters"),
        ("Adjusted plan EP", f'={sumifs("ep", "lob", "bu", "state")}', "#,##0",
         "the weight behind every average here"),
        (f"CY {d.plan_year} plan loss ratio",
         f'={wtd("w_cylr", "lob", "bu", "state")}', FMT_PCT,
         "EP-weighted across everything in view"),
        ("Plan LR — program basis", f'={wtd("w_prog", "lob", "bu", "state")}',
         FMT_PCT, "the logged program instead of any asserted net target"),
        ("Program vs asserted",
         f'=IF({sumifs("ep", "lob", "bu", "state")}=0,"—",'
         f'({sumifs("w_prog", "lob", "bu", "state")}-'
         f'{sumifs("w_cylr", "lob", "bu", "state")})'
         f'/{sumifs("ep", "lob", "bu", "state")}*100)',
         FMT_PTS_SIGNED, "0.00 pts when nothing asserts a net target"),
        ("Net-target combos", f'={sumifs("netmode", "lob", "bu", "state")}', FMT_INT,
         "combos carrying an asserted net selection"),
    ]
    for i, (lbl, f, fmt, sub) in enumerate(cards):
        c1 = 1 + i * 2
        put(ws, f"{col(c1)}12", lbl, fnt=F_LABEL, fill=FILL_PANEL)
        put(ws, f"{col(c1 + 1)}12", None, fill=FILL_PANEL)
        formula(ws, f"{col(c1)}13", f, fmt=fmt)
        ws[f"{col(c1)}13"].font = font(NAVY, bold=True, size=14)
        ws[f"{col(c1)}13"].fill = FILL_PANEL
        put(ws, f"{col(c1 + 1)}13", None, fill=FILL_PANEL)
        put(ws, f"{col(c1)}14", sub, fnt=F_SMALL_IT)

    section(ws, 16, "A", "Sources")
    header_row(ws, 17, 1, ["Line of business", "Workbook", "Built / modified",
                           "Version", "Combos", "Source checks"],
               widths=[20, 42, 18, 10, 9, 24])
    for i, s in enumerate(d.sources):
        r = 18 + i
        put(ws, f"A{r}", s.lob)
        put(ws, f"B{r}", s.path, fnt=font(GREY_DARK, size=9))
        put(ws, f"C{r}", s.modified, align=ALIGN_C)
        put(ws, f"D{r}", s.version, align=ALIGN_C)
        put(ws, f"E{r}", s.combos, fmt=FMT_INT, align=ALIGN_C)
        # D83: what each source's OWN trust panel said when it was harvested.
        # A red LOB used to roll up into a green book without a trace.
        put(ws, f"F{r}", s.checks or "(not read)",
            fnt=font(PASS_GREEN if s.checks_ok else FAIL_RED, bold=not s.checks_ok,
                     size=9), align=ALIGN_C)
    tot = 18 + len(d.sources)
    put(ws, f"A{tot}", "TOTAL", fnt=font(NAVY, bold=True))
    formula(ws, f"E{tot}", f"=SUM($E$18:$E${tot - 1})", fmt=FMT_INT, align=ALIGN_C,
            bold=True)
    ctx.define("bk_srccombos", "Control", f"$E${tot}",
               "Combos the sources reported, summed")
    put(ws, f"A{tot + 2}",
        "Every figure in this book is a harvested value from these files — it "
        "cannot drift from them, and it cannot follow them either. Rebuild after "
        "any change.", fnt=F_SMALL_IT)
    set_widths(ws, {"A": 20, "B": 14, "C": 16, "D": 12, "E": 12, "F": 14, "G": 12,
                    "H": 16, "I": 12, "J": 14, "K": 12, "L": 18})
    presentation_setup(ws, gridlines_off=True, tab_color=NAVY)
    print_setup(ws)


def steel_fill():
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=STEEL_LIGHT)


# ---------------------------------------------------------------------------
# State Summary — the flagship
# ---------------------------------------------------------------------------

SS_HDR, SS_FIRST = 7, 8


def build_state_summary(ctx: BookCtx, n: int):
    ws = ctx.wb["State Summary"]
    d = ctx.data
    states = d.states
    title(ws, "A1", f"State Summary — every line, one row per state ({d.plan_year} plan)",
          "Adjusted-EP weighted across the lines and business units in view (set them "
          "on Control). Harvested values; the aggregation is live. Rate-change history "
          "shows only where the filters resolve to a single combo — the same honesty "
          "rule as the per-LOB exhibit, one dimension deeper.")
    nav_bar(ws, 3, 1, ["Control", "Portfolio", "Roll-ups", "Program Flow", "Checks"],
            step=2)
    formula(ws, "A5",
            '="Showing: "&IF(bkf_lob="All","every line",bkf_lob)&"  |  "'
            '&IF(bkf_bu="All","every business unit",bkf_bu)'
            '&IF(bkf_state="All",""," | (state filter does not apply here)")')
    ws["A5"].font = font(GREY_DARK, italic=True)

    # The layout is SHARED with the per-LOB exhibit (SS_COLS): this sheet is
    # its mirror, and the two used to be two hand-maintained copies of the same
    # 35-column order — the exact arrangement that lets one drift from the
    # other. Only the captions differ, because the book has no live plan year to
    # interpolate and its honesty rule is one dimension deeper (D90).
    book_caption = {
        SS_G_BRIDGE: f"CY {d.plan_year} plan — the bridge, left to right",
        SS_G_HIST: ("Rate change history — single-combo views only. "
                    "T = taken, P = planned, * = not in the indication"),
        SS_G_P1: f"CY {d.plan_year + 1} indicative",
    }
    for c1, c2, cap in ss_groups():
        text = book_caption.get(cap, cap)
        for cc in range(c1, c2 + 1):
            cell = ws.cell(row=6, column=cc, value=text if cc == c1 else None)
            cell.font = font(NAVY, bold=True, size=9)
            cell.fill = FILL_PANEL
    book_header = {
        "ecy": f"{d.plan_year} earned rate level",
        "mbar": f"{d.plan_year} earned mod",
        "planlr": f"CY {d.plan_year} plan LR",
        "trend": "Net trend",
        "arate1": "Rate earn-in +1",
        "amod1": "Mod drift +1",
        "planlr1": f"CY {d.plan_year + 1} plan LR",
        "ep": "Adj plan EP",
    }
    header_row(ws, SS_HDR, 1,
               [book_header.get(c.key, c.header) for c in SS_COLS],
               widths=[c.width for c in SS_COLS])
    ws.row_dimensions[SS_HDR].height = 42
    # the bridge's "=" prefix has to be a FORMULA producing text: a literal
    # cell value starting with "=" is stored as a formula by openpyxl, and an
    # invalid one makes Excel refuse the entire workbook (D41, again)
    ws.cell(row=SS_HDR, column=ss_c("planlr")).value = f'="=  CY {d.plan_year} plan LR"'

    # per-state aggregates honour the LOB and BU filters (an exhibit keyed by a
    # dimension never filters on its own — see the module docstring)
    dims = ("lob", "bu")
    wide = [("mind", "w_mind", FMT_MOD), ("m0", "w_m0", FMT_MOD),
            ("m1", "w_m1", FMT_MOD), ("crl", "w_crl", FMT_IDX),
            ("ecy", "w_ecy", FMT_IDX), ("mbar", "w_mbar", FMT_MOD),
            ("lrcur", "w_lrcur", FMT_PCT), ("arate", "w_arate", FMT_IDX),
            ("target", "w_target", FMT_PCT),
            ("amod", "w_amod", FMT_IDX), ("aother", "w_aother", FMT_IDX),
            ("trend", "w_trend", PCT_S), ("arate1", "w_arate1", FMT_IDX),
            ("amod1", "w_amod1", FMT_IDX)]
    for i, st in enumerate(states):
        r = SS_FIRST + i
        band = FILL_PANEL if i % 2 else None
        put(ws, f"A{r}", st, fnt=font(NAVY, bold=True), align=ALIGN_C, fill=band,
            border=BORDER_THIN)
        st_c = f',{rng("state", n)},$A{r}'
        ep = f'SUMIFS({rng("ep", n)}{st_c}{crit(*dims)})'
        cnt = f'COUNTIFS({rng("key", n)},"?*"{st_c}{crit(*dims)})'
        nm = f'SUMIFS({rng("netmode", n)}{st_c}{crit(*dims)})'

        def w(field_):
            return (f'IF({ep}=0,"—",'
                    f'SUMIFS({rng(field_, n)}{st_c}{crit(*dims)})/{ep})')

        formula(ws, f"B{r}", f"={ep}", fmt="#,##0", align=ALIGN_C, fill=band,
                border=BORDER_THIN)
        for key, fld, fmt in wide:
            formula(ws, f"{ss_l(key)}{r}", f"={w(fld)}", fmt=fmt,
                    align=ALIGN_C, fill=band, border=BORDER_THIN)
        # rate-change slots: only when the filters resolve to exactly one combo,
        # and only where that combo HAS the change. Two guards, not one. _calc
        # publishes "" for a slot nobody filed, but a cached "" reads back as
        # None and lands on _book as a genuinely blank cell — and INDEX over a
        # blank returns 0, not "". That 0 renders 1/0/00 under the date format
        # and 0.0% under the signed percent, which is what the per-LOB exhibit
        # never shows because there the "" never round-trips through a value.
        # Presence is tested on the slot's DATE: a real one is a serial well
        # above zero, so COUNT cannot false-positive, and all three parts of a
        # slot are published together so one test governs the trio.
        uk = f'bkf_lob&"|"&bkf_bu&"|"&$A{r}'
        mt = f'MATCH({uk},{rng("ukey", n)},0)'
        for j in range(SLOTS):
            has = f'COUNT(INDEX({rng(f"slot{j}_date", n)},{mt}))'
            for part, fmt in (("date", FMT_DATE_S), ("pct", PCT_S),
                              ("tok", None)):
                formula(ws, f'{ss_l(f"chg{j + 1}_{part}")}{r}',
                        f'=IF({cnt}<>1,"—",IF({has}=0,"",'
                        f'INDEX({rng(f"slot{j}_{part}", n)},{mt})))',
                        fmt=fmt or "General", align=ALIGN_C, fill=band,
                        border=BORDER_THIN)
        for key in ("ntaken", "nplanned"):
            formula(ws, f"{ss_l(key)}{r}",
                    f'=SUMIFS({rng(key, n)}{st_c}{crit(*dims)})', fmt=FMT_INT,
                    align=ALIGN_C, fill=band, border=BORDER_THIN)
        chain = _ss_chain(r)
        chain1 = _ss_chain1(r)
        formula(ws, f'{ss_l("planlr")}{r}', f'=IF({cnt}=0,"",IF({cnt}=1,{chain},{w("w_cylr")}))',
                fmt=FMT_PCT, align=ALIGN_C, fill=band, border=BORDER_THIN, bold=True)
        formula(ws, f'{ss_l("mix")}{r}',
                f'=IF({cnt}<=1,"",(${ss_l("planlr")}{r}-{chain})*100)', fmt=PTS,
                align=ALIGN_C, fill=band, border=BORDER_THIN)
        formula(ws, f'{ss_l("planlr1")}{r}', f'=IF({cnt}=0,"",IF({cnt}=1,{chain1},{w("w_cylr1")}))',
                fmt=FMT_PCT, align=ALIGN_C, fill=band, border=BORDER_THIN, bold=True)
        formula(ws, f'{ss_l("netsel")}{r}',
                f'=IF({cnt}=0,"",IF({nm}=0,"—",'
                f'SUMIFS({rng("netx", n)}{st_c}{crit(*dims)})/{nm}))',
                fmt=PCT_S, align=ALIGN_C, fill=band, border=BORDER_THIN)
        formula(ws, f'{ss_l("progbasis")}{r}', f'=IF({nm}=0,"—",{w("w_prog")})', fmt=FMT_PCT,
                align=ALIGN_C, fill=band, border=BORDER_THIN)
        formula(ws, f'{ss_l("proggap")}{r}',
                f'=IF({nm}=0,"—",(${ss_l("progbasis")}{r}-${ss_l("planlr")}{r})*100)', fmt=PTS,
                align=ALIGN_C, fill=band, border=BORDER_THIN, bold=True)

    tot = SS_FIRST + len(states) + 1
    put(ws, f"A{tot}", "TOTAL", fnt=font(NAVY, bold=True), align=ALIGN_C,
        fill=steel_fill())
    ep_t = f'SUMIFS({rng("ep", n)}{crit(*dims)})'
    formula(ws, f"B{tot}", f"={ep_t}", fmt="#,##0", align=ALIGN_C, bold=True,
            fill=steel_fill())
    for key, fld, fmt in wide + [("planlr", "w_cylr", FMT_PCT),
                                 ("planlr1", "w_cylr1", FMT_PCT),
                                 ("progbasis", "w_prog", FMT_PCT)]:
        formula(ws, f"{ss_l(key)}{tot}",
                f'=IF({ep_t}=0,"n/a",SUMIFS({rng(fld, n)}{crit(*dims)})/{ep_t})',
                fmt=fmt, align=ALIGN_C, bold=(key in ("planlr", "planlr1")),
                fill=steel_fill())
    for key in ("ntaken", "nplanned"):
        formula(ws, f"{ss_l(key)}{tot}",
                f'=SUMIFS({rng(key, n)}{crit(*dims)})', fmt=FMT_INT, align=ALIGN_C,
                fill=steel_fill())
    # the chronology band carries no total — four dated changes do not add up —
    # but it still has to be part of the band, or the total row reads as though
    # it stops twelve columns early (the per-LOB exhibit does the same)
    for j in range(1, SLOTS + 1):
        for part in ("date", "pct", "tok"):
            put(ws, f"{ss_l(f'chg{j}_{part}')}{tot}", None, fill=steel_fill())
    formula(ws, f'{ss_l("mix")}{tot}',
            f'=IF($B${tot}=0,"",(${ss_l("planlr")}{tot}-{_ss_chain(tot)})*100)',
            fmt=PTS, align=ALIGN_C, fill=steel_fill())
    formula(ws, f'{ss_l("proggap")}{tot}',
            f'=IF($B${tot}=0,"",(${ss_l("progbasis")}{tot}-${ss_l("planlr")}{tot})*100)',
            fmt=PTS, align=ALIGN_C, bold=True, fill=steel_fill())
    ctx.define("bk_ss_ep", "State Summary", f"$B${tot}", "State Summary total EP in view")
    ctx.define("bk_ss_lr", "State Summary", f'${ss_l("planlr")}${tot}',
               "State Summary total plan LR")
    ctx.define("bk_ss_prog", "State Summary", f'${ss_l("progbasis")}${tot}',
               "State Summary total program-basis plan LR")

    notes = [
        "Every metric is adjusted-plan-EP weighted over the lines and business units "
        "in view; the state filter on Control does not apply here — this IS the "
        "per-state view.",
        "Plan LR is the product of the four factors to its left wherever the filters "
        "resolve to a single combo; where a row combines combos the exact EP-weighted "
        "value is shown and Mix carries the difference (factor averages do not "
        "compound across a mixed book).",
        "Program basis values the same rows on the logged rate program and projected "
        "mod path instead of any asserted net target — dashed where nothing asserts one.",
        "Values are harvested from the six LOB workbooks and do not move until the "
        "book is rebuilt. Sources and as-of stamp: Control.",
    ]
    for i, t in enumerate(notes):
        put(ws, f"A{tot + 2 + i}", t, fnt=F_SMALL_IT)
    ss_outline(ws)          # same collapse behaviour as the per-LOB exhibit (D101)
    # D104: and the same visual layer. The heatmaps, the EP data bar and the
    # amber planned-change tokens were per-LOB only; the Book had shared the
    # column map since D90 and still drew none of them.
    ss_conditional_formats(ws, SS_FIRST, SS_FIRST + len(states) - 1, FONT_NAME)
    presentation_setup(ws, gridlines_off=True, freeze=f"B{SS_FIRST}", tab_color=NAVY)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Portfolio — every combo, filterable
# ---------------------------------------------------------------------------

PF_HDR, PF_FIRST = 6, 7


def build_portfolio(ctx: BookCtx, n: int):
    ws = ctx.wb["Portfolio"]
    d = ctx.data
    title(ws, "A1", f"Portfolio — every combo in the book ({d.plan_year} plan)",
          "One row per line x business unit x state, straight from the harvest. "
          "Filter this grid with Excel's own filters (row 6) or narrow the whole "
          "book on Control; sorting is safe here because every cell is a value.")
    nav_bar(ws, 3, 1, ["Control", "State Summary", "Roll-ups", "Program Flow",
                       "Checks"], step=2)
    grid = [("lob", "Line of business", None, 20), ("bu", "BU", None, 9),
            ("state", "State", None, 7), ("ep", "Adj plan EP", "#,##0", 12),
            ("lrcur", "Projected LR", FMT_PCT, 11),
            ("arate_p", "Rate earn-in", FMT_IDX, 11),
            ("amod_p", "Mod drift", FMT_IDX, 10),
            ("aother", "Other adj", FMT_IDX, 10),
            ("cylr_p", f"CY {d.plan_year} plan LR", FMT_PCT, 12),
            ("cylr_p1", f"CY {d.plan_year + 1} plan LR", FMT_PCT, 12),
            ("echg", "Earned chg vs ind", PCT_S, 12),
            ("carry", "Carryover", PCT_S, 11),
            ("netx", "Net target", PCT_S, 10),
            ("cylr_prog", "Plan LR (program)", FMT_PCT, 12),
            ("proggap", "Program vs asserted (pts)", PTS, 12),
            ("ntaken", "# taken", FMT_INT, 8),
            ("nplanned", "# planned", FMT_INT, 9)]
    header_row(ws, PF_HDR, 1, [h for _f, h, _x, _w in grid],
               widths=[w for _f, _h, _x, w in grid])
    ws.row_dimensions[PF_HDR].height = 30
    for i in range(n):
        r = PF_FIRST + i
        src = BK_FIRST + i
        for j, (f, _h, fmt, _w) in enumerate(grid):
            cl = col(COLS[f])
            if f == "netx":
                fx = (f"=IF('{BOOK}'!${col(COLS['netmode'])}${src}=0,\"—\","
                      f"'{BOOK}'!${cl}${src})")
            elif f == "proggap":
                fx = (f"=IF('{BOOK}'!${col(COLS['netmode'])}${src}=0,\"—\","
                      f"'{BOOK}'!${cl}${src})")
            else:
                fx = f"='{BOOK}'!${cl}${src}"
            link(ws, ws.cell(row=r, column=1 + j).coordinate, fx,
                 fmt=fmt or "General", align=ALIGN_C if j > 2 else ALIGN_L)
    last = PF_FIRST + n - 1
    ws.auto_filter.ref = f"A{PF_HDR}:{col(len(grid))}{last}"
    put(ws, f"A{last + 2}",
        "Every cell here mirrors the harvested table — filter or sort freely. "
        "The Control filters do not narrow this grid; use the column filters.",
        fnt=F_SMALL_IT)
    presentation_setup(ws, gridlines_off=True, freeze=f"D{PF_FIRST}", tab_color=NAVY)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Roll-ups — by line, by business unit, by state
# ---------------------------------------------------------------------------


def build_rollups(ctx: BookCtx, n: int):
    ws = ctx.wb["Roll-ups"]
    d = ctx.data
    title(ws, "A1", "Roll-ups — the book cut three ways",
          "Each table applies the OTHER two Control filters, never its own, so a "
          "roll-up always shows every member of its own dimension. Totals are "
          "adjusted-EP weighted; the mix line is the honest residual, because factor "
          "averages do not compound across a mixed book.")
    nav_bar(ws, 3, 1, ["Control", "State Summary", "Portfolio", "Program Flow",
                       "Checks"], step=2)
    r = 5
    for dim, members, other in (("lob", d.lobs, ("bu", "state")),
                                ("bu", d.business_units, ("lob", "state")),
                                ("state", d.states, ("lob", "bu"))):
        label_txt = {"lob": "By line of business", "bu": "By business unit",
                     "state": "By state"}[dim]
        section(ws, r, "A", label_txt)
        header_row(ws, r + 1, 1,
                   ["", "Combos", "Adj plan EP", "Projected LR", "Rate earn-in",
                    "Mod drift", f"CY {d.plan_year} plan LR",
                    f"CY {d.plan_year + 1} plan LR", "Plan LR (program)",
                    "Program vs asserted (pts)", "Net combos"],
                   widths=[20, 9, 13, 12, 11, 10, 12, 12, 13, 13, 11])
        first = r + 2
        for i, m in enumerate(members):
            rr = first + i
            put(ws, f"A{rr}", m, fnt=font(NAVY, bold=True))
            mc = f',{rng(dim, n)},$A{rr}'
            ep = f'SUMIFS({rng("ep", n)}{mc}{crit(*other)})'
            nm = f'SUMIFS({rng("netmode", n)}{mc}{crit(*other)})'
            formula(ws, f"B{rr}",
                    f'=COUNTIFS({rng("key", n)},"?*"{mc}{crit(*other)})', fmt=FMT_INT,
                    align=ALIGN_C)
            formula(ws, f"C{rr}", f"={ep}", fmt="#,##0", align=ALIGN_C)
            for cc, fld, fmt in (("D", "w_lrcur", FMT_PCT), ("E", "w_arate", FMT_IDX),
                                 ("F", "w_amod", FMT_IDX), ("G", "w_cylr", FMT_PCT),
                                 ("H", "w_cylr1", FMT_PCT), ("I", "w_prog", FMT_PCT)):
                formula(ws, f"{cc}{rr}",
                        f'=IF({ep}=0,"—",SUMIFS({rng(fld, n)}{mc}{crit(*other)})/{ep})',
                        fmt=fmt, align=ALIGN_C, bold=(cc == "G"))
            formula(ws, f"J{rr}", f'=IF({nm}=0,"—",($I{rr}-$G{rr})*100)', fmt=PTS,
                    align=ALIGN_C)
            formula(ws, f"K{rr}", f"={nm}", fmt=FMT_INT, align=ALIGN_C)
        tr = first + len(members)
        put(ws, f"A{tr}", "TOTAL", fnt=font(NAVY, bold=True), fill=steel_fill())
        ep_t = f'SUMIFS({rng("ep", n)}{crit(*other)})'
        formula(ws, f"B{tr}", f'=COUNTIFS({rng("key", n)},"?*"{crit(*other)})',
                fmt=FMT_INT, align=ALIGN_C, fill=steel_fill())
        formula(ws, f"C{tr}", f"={ep_t}", fmt="#,##0", align=ALIGN_C, bold=True,
                fill=steel_fill())
        for cc, fld, fmt in (("D", "w_lrcur", FMT_PCT), ("E", "w_arate", FMT_IDX),
                             ("F", "w_amod", FMT_IDX), ("G", "w_cylr", FMT_PCT),
                             ("H", "w_cylr1", FMT_PCT), ("I", "w_prog", FMT_PCT)):
            formula(ws, f"{cc}{tr}",
                    f'=IF({ep_t}=0,"n/a",SUMIFS({rng(fld, n)}{crit(*other)})/{ep_t})',
                    fmt=fmt, align=ALIGN_C, bold=(cc == "G"), fill=steel_fill())
        formula(ws, f"J{tr}", f'=IF($C{tr}=0,"",($I{tr}-$G{tr})*100)', fmt=PTS,
                align=ALIGN_C, fill=steel_fill())
        formula(ws, f"K{tr}", f'=SUMIFS({rng("netmode", n)}{crit(*other)})',
                fmt=FMT_INT, align=ALIGN_C, fill=steel_fill())
        formula(ws, f"A{tr + 1}",
                f'="Mix / interaction: "&TEXT($G{tr}-$D{tr}*$E{tr}*$F{tr},'
                f'"+0.00%;-0.00%;0.00%")&" — the EP-weighted plan LR minus the '
                f'product of the EP-weighted factors above it."')
        ws[f"A{tr + 1}"].font = F_SMALL_IT
        if dim == "lob":
            ctx.define("bk_lob_ep", "Roll-ups", f"$C${tr}", "By-LOB total EP")
            ctx.define("bk_lob_first", "Roll-ups", f"$C${first}", "First LOB row EP")
        r = tr + 4
    set_widths(ws, {"A": 20})
    presentation_setup(ws, gridlines_off=True, tab_color=NAVY)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Program Flow — monthly legs across the book
# ---------------------------------------------------------------------------


def build_program_flow(ctx: BookCtx, n: int):
    ws = ctx.wb["Program Flow"]
    d = ctx.data
    states = d.states
    title(ws, "A1", f"Program Flow — what the book delivers month by month "
                    f"({d.plan_year})",
          "Year-over-year change on renewals from the programs AS LOGGED, weighted by "
          "adjusted EP x written weight across the lines and business units in view. "
          "Valid across lines: seasonality is state-keyed and the written-basis legs "
          "are term-independent, so the six-month book combines safely.")
    nav_bar(ws, 3, 1, ["Control", "State Summary", "Portfolio", "Roll-ups", "Checks"],
            step=2)
    dims = ("lob", "bu")
    r = 5
    grids = [("rate", "The rate leg — YoY written rate change on renewals",
              "epw_rate", "epw"),
             ("mod", "The pricing leg — YoY written schedule-mod change",
              "epwm_mod", "epwm"),
             ("del", "Delivered net — rate x pricing, the YoY change customers see",
              "epw_del", "epw")]
    starts = []
    for key, head, numfam, denfam in grids:
        section(ws, r, "A", head)
        put(ws, f"A{r + 1}",
            "Bottom row: every state above collapsed onto one line, EP x weight "
            "weighted across the lines and business units in view — an average, not "
            "a total.", fnt=F_SMALL_IT)
        put(ws, f"A{r + 2}", "State", fnt=font("FFFFFF", bold=True), fill=FILL_NAVY,
            align=ALIGN_C)
        for j in range(12):
            cell = ws.cell(row=r + 2, column=2 + j)
            cell.value = f'=TEXT(DATE({d.plan_year},{j + 1},1),"mmm")&" {d.plan_year}"'
            cell.font = font("FFFFFF", bold=True, size=9)
            cell.fill = FILL_NAVY
            cell.alignment = ALIGN_C
        starts.append(r + 2)
        for i, st in enumerate(states):
            rr = r + 3 + i
            put(ws, f"A{rr}", st, fnt=font(NAVY, bold=True), align=ALIGN_C)
            st_c = f',{rng("state", n)},$A{rr}'
            for j in range(12):
                den = f'SUMIFS({rng(f"{denfam}_{j}", n)}{st_c}{crit(*dims)})'
                num = f'SUMIFS({rng(f"{numfam}_{j}", n)}{st_c}{crit(*dims)})'
                formula(ws, ws.cell(row=rr, column=2 + j).coordinate,
                        f'=IF({den}=0,"—",{num}/{den})', fmt=PCT_S, align=ALIGN_C)
                ws.cell(row=rr, column=2 + j).font = font(GREY_DARK, size=9)
        rt = r + 3 + len(states)
        formula(ws, f"A{rt}", '=IF(bkf_lob="All","BOOK AVG","LINE AVG")',
                align=ALIGN_C)
        ws[f"A{rt}"].font = font(NAVY, bold=True, size=9)
        for j in range(12):
            den = f'SUMIFS({rng(f"{denfam}_{j}", n)}{crit(*dims)})'
            num = f'SUMIFS({rng(f"{numfam}_{j}", n)}{crit(*dims)})'
            cc = ws.cell(row=rt, column=2 + j)
            formula(ws, cc.coordinate, f'=IF({den}=0,"—",{num}/{den})', fmt=PCT_S,
                    align=ALIGN_C)
            cc.font = font(NAVY, bold=True, size=9)
        if key == "del":
            ws.conditional_formatting.add(
                f"B{r + 3}:M{r + 2 + len(states)}",
                ColorScaleRule(start_type="min", start_color="D6E8D5",
                               mid_type="percentile", mid_value=50,
                               mid_color="FFF2CC", end_type="max",
                               end_color="F4B8B8"))
        r = rt + 3
    ctx.define("bk_pf_first", "Program Flow", f"$B${starts[0] + 1}",
               "First cell of the rate-leg grid")
    put(ws, f"A{r}",
        "Legs are weighted means, so rate x pricing need not multiply to delivered "
        "exactly under mix — delivered is the exact statistic. Earning patterns are "
        "never aggregated across lines with different policy terms.", fnt=F_SMALL_IT)
    set_widths(ws, {"A": 10, **{col(2 + j): 10 for j in range(12)}})
    presentation_setup(ws, gridlines_off=True, freeze=f"B{starts[0] + 1}",
                       tab_color=NAVY)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Checks
# ---------------------------------------------------------------------------


def build_checks(ctx: BookCtx, n: int):
    ws = ctx.wb["Checks"]
    d = ctx.data
    title(ws, "A1", "Checks — does this book hang together?",
          "The engines are frozen, so these verify the HARVEST and the aggregation: "
          "that every source reported, that the parts sum to the whole, and that the "
          "identities carried over intact. FAIL = rebuild before using results.")
    label(ws, "A3", "Overall status", bold=True)
    nav_bar(ws, 3, 6, ["Control", "State Summary", "Read Me"], step=2)
    rows = [
        ("Harvest", "Every configured line reported its combos",
         f"={len(d.sources)}",
         f'=COUNTIF(Control!$A$18:$A${17 + len(d.sources)},"?*")', 0),
        ("Harvest", "Rows harvested = combos the sources reported",
         f"={n}", "=bk_srccombos", 0),
        ("Harvest", "Every harvested row carries a key",
         f"={n}", f'=COUNTIF({rng("key", n)},"?*")', 0),
        # D83: the sources' own status, frozen at harvest time and counted here
        # so a failing LOB cannot hide inside a green book.
        ("Harvest", "Every source reported a passing Checks panel when harvested",
         f"={len(d.sources)}",
         f'=COUNTIF(Control!$F$18:$F${17 + len(d.sources)},"ALL CHECKS PASS")'
         f'+COUNTIF(Control!$F$18:$F${17 + len(d.sources)},"PASS WITH *")', 0),
        # each exhibit applies its OWN filter set, so reconcile against the
        # matching subset — otherwise these hold only in the unfiltered view
        ("Reconciliation", "By-line roll-up total = the rows it should cover",
         f'={sumifs("ep", "bu", "state")}', "=bk_lob_ep", 1e-6),
        ("Reconciliation", "State Summary total EP = the rows it should cover",
         f'={sumifs("ep", "lob", "bu")}', "=bk_ss_ep", 1e-6),
        ("Identity", "Program basis = headline on every combo with no net selection",
         "=0", f'=SUM({rng("progident", n)})', 1e-9),
        ("Identity", "Net targets only where a net selection is flagged",
         "=0", f'=SUMPRODUCT(({rng("netmode", n)}=0)*'
               f'({rng("netx", n)}<>0)*1)', 0),
        ("Advisory", "Combos whose logged program lands 1pt+ from the asserted plan LR",
         "=0", f'=SUMPRODUCT((ABS({rng("proggap", n)})>1)*1)', 0),
    ]
    header_row(ws, 5, 1, ["#", "Category", "Check", "Expected", "Actual", "Tolerance",
                          "Status"], widths=[5, 14, 68, 14, 14, 11, 10])
    for i, (cat, desc, exp, act, tol) in enumerate(rows):
        r = 6 + i
        put(ws, f"A{r}", i + 1, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
        put(ws, f"B{r}", cat, fnt=font(GREY_DARK, size=9))
        put(ws, f"C{r}", desc, fnt=F_LABEL)
        formula(ws, f"D{r}", exp, fmt=FMT_IDX, align=ALIGN_C)
        formula(ws, f"E{r}", act, fmt=FMT_IDX, align=ALIGN_C)
        put(ws, f"F{r}", tol, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
        kind = "WARN" if cat == "Advisory" else "FAIL"
        formula(ws, f"G{r}",
                f'=IF(ABS($E{r}-$D{r})<=$F{r},"PASS","{kind}")', align=ALIGN_C,
                bold=True)
    last = 6 + len(rows) - 1
    # same three-state banner as the LOB workbooks (D80) — advisories are
    # visible here rather than rounded up into "pass"
    formula(ws, "C3",
            f'=IF(COUNTIF($G$6:$G${last},"FAIL")>0,"CHECKS FAILING: "'
            f'&COUNTIF($G$6:$G${last},"FAIL"),'
            f'IF(COUNTIF($G$6:$G${last},"WARN")>0,"PASS WITH "'
            f'&COUNTIF($G$6:$G${last},"WARN")&" WARNING(S)","ALL CHECKS PASS"))',
            bold=True, fill=FILL_PANEL)
    status_banner_cf(ws, "C3")
    ctx.define("bk_overall", "Checks", "$C$3", "Book-level status banner")
    put(ws, f"A{last + 2}",
        "These cannot re-derive the engines — that is the per-LOB harness's job "
        "(tools/verify_workbook.py). What they prove is that this book faithfully "
        "reflects the six workbooks it was built from.", fnt=F_SMALL_IT)
    presentation_setup(ws, gridlines_off=True, tab_color=PASS_GREEN)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Read Me
# ---------------------------------------------------------------------------


def build_readme(ctx: BookCtx, n: int):
    ws = ctx.wb["Read Me"]
    d = ctx.data
    title(ws, "B2", f"The Book — {d.plan_year} plan, every line of business",
          f"Harvested {d.as_of} from {len(d.sources)} workbooks | {n} combos | "
          f"classic formula mode | engines frozen, filters live")
    r = 5
    section(ws, r, "B", "What this is")
    r += 1
    for t in ("One row per line x business unit x state, harvested from the six LOB "
              "workbooks after they were recalculated. Every per-combo number here is "
              "a VALUE copied from the file that computed it — this book cannot drift "
              "from those files, and it cannot follow them either.",
              "Every filter, weighted average and total IS live, so changing the view "
              "on Control recalculates instantly.",
              # The old wording said "regenerate the LOB workbooks" as the refresh
              # step. Regeneration REPLACES a workbook with a fresh sample-seeded
              # copy — the one action that destroys a season of pasted inputs — and
              # it is not needed to refresh anything (D82/D83).
              "To refresh after editing inputs: save the LOB workbook(s) in Excel, "
              "then rebuild this book (python tools/build_book.py). That is the whole "
              "loop.",
              "Regenerate a LOB workbook only for STRUCTURAL changes (new state, new "
              "capacity, new generator version) — regeneration replaces the file's "
              "contents, so carry your inputs across with "
              "'python -m src.build_workbook --lob \"<line>\" --carry-forward'.",
              "The harvest refuses a workbook that was never recalculated, and refuses "
              "one whose own Checks panel is failing — see 'Source checks' on Control "
              "for what each file reported when it was read."):
        put(ws, f"B{r}", t, fnt=font(GREY_DARK, size=10))
        r += 1
    r += 1
    section(ws, r, "B", "Where to go")
    r += 1
    for name, desc in (
            ("Control", "Filters (line / business unit / state), the in-view KPI band, "
                        "and the source table with each file's stamp."),
            ("State Summary", "The flagship: one row per state, EP-weighted across the "
                              "lines and units in view, with the full bridge chain."),
            ("Portfolio", "Every combo in the book, with Excel filters on the grid — "
                          "sort and slice freely; the cells are values."),
            ("Roll-ups", "The book cut by line, by business unit, and by state, each "
                         "with its mix residual."),
            ("Program Flow", "Month-by-month YoY rate, pricing and delivered legs "
                             "across the book."),
            (PIVOT, "Build your own view: the whole book as one long table "
                    "(tbl_Pivot), ready for a PivotTable. Recipe below."),
            ("Checks", "Harvest and reconciliation checks — must read ALL CHECKS PASS.")):
        jump(ws, f"B{r}", f"{quote_sheet(name)}!A1", name, bold=True)
        put(ws, f"C{r}", desc, fnt=font(GREY_DARK, size=10))
        r += 1
    r += 1
    section(ws, r, "B", "How the numbers combine")
    r += 1
    for t in ("Weighting is adjusted plan EP throughout, and EP x written weight for "
              "the monthly legs — the same weights the per-LOB workbooks use.",
              "An exhibit keyed by a dimension applies the OTHER filters, never its "
              "own: the per-state views honour line and business unit, the by-line "
              "roll-up honours unit and state.",
              "Weighted factor averages do not compound: the EP-weighted plan LR is "
              "not the product of the EP-weighted factors. Where both are shown, the "
              "exact weighted value is displayed and the difference is labelled Mix.",
              "Combining lines is safe for loss ratios and written-basis legs "
              "(seasonality is state-keyed; the legs are term-independent). Earning "
              "patterns are never aggregated across different policy terms.",
              "Mixing lines with different loss economics into one headline is a "
              "judgment, not an arithmetic fact — every roll-up keeps its member lines "
              "visible beside the total for that reason."):
        put(ws, f"B{r}", t, fnt=font(GREY_DARK, size=10))
        r += 1
    r += 1
    section(ws, r, "B", f"Building your own pivot on '{PIVOT}'")
    r += 1
    for t in ("The sheet is one long table, tbl_Pivot: LOB | BU | State | Category | "
              "Measure | Month | Weight | Weighted value. One row per combo per "
              "measure, and per month for the three monthly legs.",
              "Insert > PivotTable on any cell in the table. Then add ONE calculated "
              "field — Analyze > Fields, Items & Sets > Calculated Field — named "
              "whatever you like, with the formula:      = 'Weighted value' / Weight",
              "That one field is the correct premium-weighted answer for EVERY "
              "measure, at every subtotal, under any combination of filters. Put "
              "Measure on Rows or in the Filter; put LOB, BU, State, Month wherever "
              "you want them.",
              "Do not use Sum or Average on 'Weighted value' or 'Weight' by "
              "themselves as a result — they are the numerator and denominator, not "
              "answers. (Summing Weight alone IS meaningful for one measure at a "
              "time: it is premium.)",
              "Why there is no plain 'Value' column: the calculated field is exact "
              "even for a single row, so a raw value would add nothing except the "
              "one field that could be averaged into a wrong number. Averaging loss "
              "ratios across states is the mistake this table is shaped to prevent.",
              "Each row carries its OWN weight, which is why one formula works "
              "everywhere: annual rows weight by adjusted plan EP, the rate and "
              "delivered legs by that month's written premium, and the mod leg by "
              "the same premium restricted to combos whose mod adjustment is on. "
              "Combos with no premium, and months with no applicable weight, are "
              "absent rather than zero.",
              "Monthly premium comes from tbl_Seasonality (per state, on each LOB "
              "workbook's Inputs). A blank seasonality row means UNIFORM writing, so "
              "if you have not filled it in, month-to-month shape here is coming "
              "from rate and mod anniversaries, not from volume.",
              "Weighted factors still do not compound: a pivot showing weighted "
              "A_rate, A_mod and A_other beside weighted Plan LR will not reconcile "
              "by multiplication. That gap is the mix residual — see Roll-ups."):
        put(ws, f"B{r}", t, fnt=font(GREY_DARK, size=10))
        r += 1
    r += 1
    section(ws, r, "B", "Sources")
    r += 1
    for s in d.sources:
        put(ws, f"B{r}", s.lob, fnt=font(GREY_DARK, bold=True, size=10))
        put(ws, f"C{r}", f"{s.path}   —   {s.combos} combos, v{s.version}, "
                         f"modified {s.modified}", fnt=font(GREY_DARK, size=10))
        r += 1
    set_widths(ws, {"A": 2, "B": 22, "C": 96})
    presentation_setup(ws, gridlines_off=True, tab_color="D9D9D9")
    print_setup(ws)


# ---------------------------------------------------------------------------
# assembly
# ---------------------------------------------------------------------------


def build_book(data) -> Workbook:
    """Assemble the book workbook from a harvest (tools/harvest.BookData)."""
    global _N
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEET_ORDER:
        wb.create_sheet(name)
    ctx = BookCtx(wb=wb, data=data)
    _N = len(data.rows)
    n = build_book_data(ctx)
    build_control(ctx, n)
    build_state_summary(ctx, n)
    build_portfolio(ctx, n)
    build_rollups(ctx, n)
    build_program_flow(ctx, n)
    build_pivot_data(ctx, data.plan_year)
    build_checks(ctx, n)
    build_readme(ctx, n)
    ctx.flush_names()
    assert_formulas_balanced(wb)          # D89 — same guard as the LOB builder
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    return wb
