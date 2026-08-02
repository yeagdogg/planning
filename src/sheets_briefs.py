"""One-Pager and Compare sheets — the two hand-off exhibits.

One-Pager: a print-ready portrait brief for the selected combo (the thing you
hand a VP): answer band, mini bridge, key facts, planned actions, the two
auto-composed sentences, and small waterfall + runway charts.

Compare: any two combos side by side across the full factor chain, served
entirely from the hidden all-combo _calc results — no second engine, no
selector split.
"""

from __future__ import annotations

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

from .build_workbook import Ctx, Layout as L, SHEETS
from .sheets_report import CD0
from .xlstyle import (ALIGN_C, BORDER_THIN, DOWN_BAR, FAIL_RED, FILL_GREEN, FILL_GREY,
    FILL_PANEL, FILL_RED, FMT_DATE_S, FMT_IDX, FMT_INT, FMT_MOD, FMT_PCT, F_LABEL,
    F_SMALL_IT, GREY_DARK, NAVY, PASS_GREEN, STEEL, TOTAL_BAR, UP_BAR, col, font,
    formula, header_row, input_cell, jump, label, link, nav_bar, presentation_setup,
    print_setup, put, quote_sheet, section, set_widths, status_banner_cf, title,
)

from .xlstyle import FMT_PCT_SIGNED as PCT_S   # one definition, seven readers
from .xlstyle import FMT_PTS_COL as PTS_S  # points in a column already headed (pts)


def build_one_pager(ctx: Ctx):
    ws = ctx.wb[SHEETS.ONE_PAGER]
    br = quote_sheet(SHEETS.BRIDGE)
    flow = ctx.wb[SHEETS.FLOW]
    title(ws, "B2", f"One-Pager — {ctx.lob.name}",
          "Print-ready brief for the selected BU x state (synced to Control). "
          "Everything here is live; nothing is typed.")
    label(ws, "B4", "Combo:")
    link(ws, "C4", "=nr_SelKey", bold=True)
    jump(ws, "E4", "Control!C7", "Change selection >")
    formula(ws, "B5",
            '=IF(nr_SelOK,"Projected "&TEXT(nr_LRproj,"0.0%")&"  ->  CY "&nr_PlanYear&'
            '" plan "&TEXT(nr_CYLR_P,"0.0%")&"   ("&TEXT((nr_CYLR_P-nr_LRcur)*100,'
            '"+0.0;-0.0;0.0")&" pts vs current level)",'
            '"Select a combo on Control first.")')
    ws["B5"].font = font(NAVY, bold=True, size=14)
    ws["B5"].fill = FILL_PANEL
    for cc in range(3, 8):
        put(ws, ws.cell(row=5, column=cc).coordinate, None, fill=FILL_PANEL)

    # ---- mini bridge ----
    section(ws, 7, "B", "The bridge")
    steps = [
        ("Projected LR at current rate level", None, "=nr_LRcur"),
        ('=IF(nr_NetMode,"x  net earn-in (rate + price)","x  rate earn-in")',
         "=nr_Arate_P", None),
        ('=IF(nr_NetMode,"x  mod drift (merged into net)","x  schedule-mod drift")',
         "=nr_Amod_P", None),
        ("x  other adjustment", "=nr_AOther", None),
    ]
    r = 8
    for lbl, fac, seed in steps:
        if lbl.startswith("="):
            formula(ws, f"B{r}", lbl)
            ws[f"B{r}"].font = F_LABEL
        else:
            label(ws, f"B{r}", lbl)
        if seed:
            link(ws, f"D{r}", seed, fmt=FMT_PCT, align=ALIGN_C)
        else:
            link(ws, f"C{r}", fac, fmt=FMT_IDX, align=ALIGN_C)
            formula(ws, f"D{r}", f"=$D{r - 1}*$C{r}", fmt=FMT_PCT, align=ALIGN_C)
        r += 1
    formula(ws, f"B{r}", '="CY "&nr_PlanYear&" PLAN LOSS RATIO"')
    ws[f"B{r}"].font = font(NAVY, bold=True)
    link(ws, f"D{r}", "=nr_CYLR_P", fmt=FMT_PCT, align=ALIGN_C, bold=True, fill=FILL_PANEL)
    r += 1
    formula(ws, f"B{r}", '="CY "&(nr_PlanYear+1)&" (indicative, + net trend)"')
    ws[f"B{r}"].font = F_LABEL
    link(ws, f"D{r}", "=nr_CYLR_P1", fmt=FMT_PCT, align=ALIGN_C)

    # ---- key facts ----
    section(ws, 15, "B", "Key facts")
    facts = [
        ("Adjusted plan EP (000s)",
         '=IF(br_selrow=0,"",N(INDEX(lr_ep,br_selrow)))', "#,##0"),
        ("Earned rate chg vs indication", "=IF(nr_SelOK,nr_EChgVsInd,\"\")", PCT_S),
        ('="Carryover into "&(nr_PlanYear+1)', "=IF(nr_SelOK,nr_YoY_P1,\"\")", PCT_S),
        ("Pricing program",
         '=IF(NOT(nr_SelOK),"",IF(nr_NetMode,"NET RATE SELECTION: "&TEXT(nr_NetSelP,'
         '"+0.0%;-0.0%")&" combined from 1/1","explicit rate program (see actions below)"))',
         None),
        ("Workbook integrity", "=ck_overall", None),
    ]
    r = 16
    for lbl, f, fmt in facts:
        if lbl.startswith("="):
            formula(ws, f"B{r}", lbl)
            ws[f"B{r}"].font = F_LABEL
        else:
            label(ws, f"B{r}", lbl)
        link(ws, f"D{r}", f, fmt=fmt or "General", align=ALIGN_C)
        r += 1
    status_banner_cf(ws, "D20")

    # ---- planned actions ----
    section(ws, 22, "B", "Rate actions on the books (chronological)")
    header_row(ws, 23, 2, ["#", "Effective", "Effective %", "Status"],
               widths=None, fill=FILL_GREY, fnt=font(GREY_DARK, bold=True, size=9))
    for j in range(1, 6):
        r = 23 + j
        cnt = f"COUNTIFS(rl_key,nr_SelKey,rl_seq,{j})"
        put(ws, f"B{r}", j, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
        formula(ws, f"C{r}", f'=IF({cnt}=0,"",SUMIFS(rl_eff,rl_key,nr_SelKey,rl_seq,{j}))',
                fmt=FMT_DATE_S, align=ALIGN_C)
        formula(ws, f"D{r}", f'=IF({cnt}=0,"",SUMIFS(rl_reff,rl_key,nr_SelKey,rl_seq,{j}))',
                fmt=PCT_S, align=ALIGN_C)
        formula(ws, f"E{r}",
                f'=IF({cnt}=0,"",IF(COUNTIFS(rl_key,nr_SelKey,rl_seq,{j},rl_status,'
                f'"planned")>0,"planned","taken"))', align=ALIGN_C)
    formula(ws, "B29", '=IF(COUNTIFS(rl_key,nr_SelKey)>5,"+ "&(COUNTIFS(rl_key,nr_SelKey)'
                       '-5)&" more on the Rate Log","")')
    ws["B29"].font = F_SMALL_IT

    # ---- say it upward ----
    section(ws, 31, "B", "In two sentences")
    formula(ws, "B32",
            '="CY "&nr_PlanYear&" plan loss ratio for "&nr_SelKey&" is "'
            '&TEXT(nr_CYLR_P,"0.0%")&", versus "&TEXT(nr_LRcur,"0.0%")&" projected at '
            'current rate level ("&TEXT((nr_CYLR_P-nr_LRcur)*100,"+0.0;-0.0;0.0")&" pts)."')
    ws["B32"].font = font(NAVY, size=11)
    formula(ws, "B33",
            '=IF(nr_NetMode,"Driver: the "&TEXT(nr_NetSelP,"+0.0%;-0.0%")&" net rate '
            'selection from 1/1 — rate and price together renew at that pace.",'
            '"Drivers: rate earn-in "&TEXT(nr_LRcur*(nr_Arate_P-1)*100,"+0.0;-0.0;0.0")'
            '&" pts, schedule-mod drift "&TEXT(nr_LRcur*nr_Arate_P*(nr_Amod_P-1)*100,'
            '"+0.0;-0.0;0.0")&" pts; the year earns "&TEXT(nr_EChgVsInd,'
            "\"+0.0%;-0.0%;0.0%\")&\" vs the indication's assumed level.\")")
    ws["B33"].font = font(NAVY, size=11)

    # ---- charts: the bridge waterfall (reads Bridge staging) + runway ----
    wf = BarChart()
    wf.type = "col"
    wf.grouping = "stacked"
    wf.overlap = 100
    wf.gapWidth = 60
    bws = ctx.wb[SHEETS.BRIDGE]
    cd = L.BR_CHART_DATA
    wf.add_data(Reference(bws, min_col=3, min_row=cd, max_col=6, max_row=cd + 6),
                titles_from_data=True)
    wf.set_categories(Reference(bws, min_col=2, min_row=cd + 1, max_row=cd + 6))
    wf.series[0].graphicalProperties.noFill = True
    wf.series[1].graphicalProperties.solidFill = UP_BAR
    wf.series[2].graphicalProperties.solidFill = DOWN_BAR
    wf.series[3].graphicalProperties.solidFill = TOTAL_BAR
    wf.legend = None
    wf.y_axis.number_format = "0%"
    # source staging rows on the Bridge are outline-hidden — without this the
    # chart plots nothing (openpyxl attr is visible_cells_only, NOT plotVisOnly)
    wf.visible_cells_only = False
    wf.title = "Projected -> plan loss ratio"
    wf.style = 2
    wf.height = 7
    wf.width = 13
    wf.x_axis.delete = False
    wf.y_axis.delete = False
    ws.add_chart(wf, "B36")

    run = LineChart()
    run.add_data(Reference(flow, min_col=12, min_row=CD0, max_row=CD0 + L.N_MONTHS),
                 titles_from_data=True)
    run.set_categories(Reference(flow, min_col=2, min_row=CD0 + 1,
                                 max_row=CD0 + L.N_MONTHS))
    run.series[0].graphicalProperties.line.solidFill = NAVY
    run.series[0].graphicalProperties.line.width = int(2.25 * 12700)
    run.series[0].marker = Marker(symbol="none")
    run.series[0].smooth = False
    run.legend = None
    run.y_axis.number_format = "0.0%"
    run.title = "Rate still to earn in (unearned runway)"
    run.style = 2
    run.height = 7
    run.width = 13
    run.visible_cells_only = False
    run.x_axis.delete = False
    run.y_axis.delete = False
    ws.add_chart(run, "B51")

    set_widths(ws, {"A": 2, "B": 34, "C": 12, "D": 14, "E": 12, "F": 12, "G": 12})
    presentation_setup(ws, gridlines_off=True, tab_color=NAVY)
    print_setup(ws, landscape=False)
    ws.page_setup.fitToHeight = 1
    ws.print_area = "A1:G66"


def build_compare(ctx: Ctx):
    ws = ctx.wb[SHEETS.COMPARE]
    title(ws, "B2", "Compare — any two combos, factor by factor",
          "Served from the same hidden all-combo engine every exhibit reads. The deep-dive "
          "tabs stay on the Control selection; this page is free to look anywhere.")
    nav_bar(ws, 3, 2, ["Control", "Inputs", "Rate Log", "Portfolio", "Checks"], step=2)

    for label_, c_bu, c_st, seed_bu, seed_st in (
            ("Combo A", "C5", "C6", ctx.we_row["bu"], ctx.we_row["state"]),
            ("Combo B", "E5", "E6", None, None)):
        put(ws, f"{c_bu[0]}4", label_, fnt=font(NAVY, bold=True))
        input_cell(ws, c_bu, seed_bu, required=False)
        input_cell(ws, c_st, seed_st, required=False)
        for rng_, f_ in ((c_bu, "=lst_bu"), (c_st, "=lst_state")):
            dv = DataValidation(type="list", formula1=f_, allow_blank=True,
                                showErrorMessage=False)
            ws.add_data_validation(dv)
            dv.add(rng_)
    label(ws, "B5", "Business unit")
    label(ws, "B6", "State")
    # resolved keys + _calc rows (grey machinery)
    formula(ws, "G5", '=IF(OR($C$5="",$C$6=""),"",$C$5&"|"&$C$6)')
    formula(ws, "G6", '=IF(OR($E$5="",$E$6=""),"",$E$5&"|"&$E$6)')
    formula(ws, "H5", '=IF($G$5="",0,IF(COUNTIF(calc_key,$G$5)=0,0,MATCH($G$5,calc_key,0)))',
            fmt=FMT_INT)
    formula(ws, "H6", '=IF($G$6="",0,IF(COUNTIF(calc_key,$G$6)=0,0,MATCH($G$6,calc_key,0)))',
            fmt=FMT_INT)
    for a in ("G5", "G6", "H5", "H6"):
        ws[a].font = font(GREY_DARK, size=8)

    hdr = 8
    header_row(ws, hdr, 2, ["Metric", "Combo A", "Combo B", "B - A"],
               widths=[34, 13, 13, 13])
    formula(ws, f"C{hdr}", '=IF($G$5="","Combo A",$G$5)')
    formula(ws, f"D{hdr}", '=IF($G$6="","Combo B",$G$6)')
    for a in (f"C{hdr}", f"D{hdr}"):
        c = ws[a]
        c.font = font("FFFFFF", bold=True)

    metrics = [
        ("Projected LR (as input)", "lr", "lr_lrproj", FMT_PCT),
        ("Projected LR (current level)", "calc", "calc_lrcur", FMT_PCT),
        ("Mod assumed in indication (M_ind)", "lr", "lr_mind", FMT_MOD),
        ("Current written mod (M_0)", "lr", "lr_m0", FMT_MOD),
        ("Projected mod, plan-yr end (M_1)", "lr", "lr_m1", FMT_MOD),
        ("Indication rate level (CRL_ind)", "calc", "calc_crl", FMT_IDX),
        ("Earned rate level (E_CY)", "calc", "calc_ecy_p", FMT_IDX),
        ("Rate earn-in (A_rate)", "calc", "calc_arate_p", FMT_IDX),
        ("Earned mod (plan year)", "calc", "calc_mbar_p", FMT_MOD),
        ("Mod drift (A_mod)", "calc", "calc_amod_p", FMT_IDX),
        ("Other adjustment (A_other)", "calc", "calc_aother", FMT_IDX),
        ("PLAN_LR_ROW", "calc", "calc_cylr_p", FMT_PCT),
        ("Chg vs projected (pts)", "delta", None, PTS_S),
        ("Net trend (next year)", "calc", "calc_trend", PCT_S),
        ("Rate earn-in, next yr", "calc", "calc_arate_p1", FMT_IDX),
        ("Mod drift, next yr", "calc", "calc_amod_p1", FMT_IDX),
        ("NEXT_LR_ROW", "calc", "calc_cylr_p1", FMT_PCT),
        ("Carryover into next yr", "calc", "calc_carry", PCT_S),
        ("Adjusted plan EP (000s)", "calc", "calc_ep", "#,##0"),
        ("Net rate selection", "net", None, PCT_S),
    ]
    r = hdr + 1
    for lbl, kind, name, fmt in metrics:
        bold_ = lbl in ("PLAN_LR_ROW", "NEXT_LR_ROW")
        if lbl == "PLAN_LR_ROW":
            formula(ws, f"B{r}", '="CY "&nr_PlanYear&" PLAN LOSS RATIO"')
            ws[f"B{r}"].font = font(NAVY, bold=True)
        elif lbl == "NEXT_LR_ROW":
            formula(ws, f"B{r}", '="CY "&(nr_PlanYear+1)&" plan LR (indicative)"')
            ws[f"B{r}"].font = F_LABEL
        else:
            label(ws, f"B{r}", lbl, bold=bold_)
        for cc, hrow in (("C", "$H$5"), ("D", "$H$6")):
            if kind == "lr":
                f = f'=IF({hrow}=0,"",N(INDEX({name},{hrow})))'
            elif kind == "delta":
                f = (f'=IF({hrow}=0,"",(INDEX(calc_cylr_p,{hrow})'
                     f"-INDEX(calc_lrcur,{hrow}))*100)")
            elif kind == "net":
                f = (f'=IF({hrow}=0,"",IF(INDEX(calc_netmode,{hrow})=0,"—",'
                     f"INDEX(calc_netx,{hrow})))")
            else:
                f = f'=IF({hrow}=0,"",INDEX({name},{hrow}))'
            formula(ws, f"{cc}{r}", f, fmt=fmt, align=ALIGN_C, bold=bold_,
                    fill=FILL_PANEL if bold_ else None, border=BORDER_THIN)
        formula(ws, f"E{r}",
                f'=IF(OR($H$5=0,$H$6=0),"",IF(AND(ISNUMBER($C{r}),ISNUMBER($D{r})),'
                f'$D{r}-$C{r},""))', fmt="+0.0000;-0.0000;0.0000", align=ALIGN_C)
        ws[f"E{r}"].font = font(GREY_DARK, size=9)
        r += 1
    put(ws, f"B{r + 1}",
        "Blank column = pick a BU and state above (or the combo is not in tbl_LR). "
        "Differences show only when both combos resolve.", fnt=F_SMALL_IT)

    set_widths(ws, {"A": 2, "B": 34, "C": 13, "D": 13, "E": 13, "F": 3, "G": 12, "H": 6})
    presentation_setup(ws, gridlines_off=True, tab_color=NAVY)
    print_setup(ws)
