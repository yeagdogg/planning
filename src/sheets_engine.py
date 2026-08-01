"""Rate Engine and Mod Engine sheets, plus the shared cohort-block writer.

The cohort block (48 monthly writing cohorts, Jan (P-2) .. Dec (P+1)) is the
heart of the workbook. ONE writer emits it everywhere it appears — the visible
Rate Engine, every hidden `_calc` combo block, and the scenario / attribution /
solver variants — so the audited formula pattern is identical throughout.

Generic block columns (relative to column A of the block):
  A idx | B SOM | C EOM | D days | E month# | F mid-month serial | G abs month
  H w_k | I W_pre | J W_post | K #chg | L first-chg days | M p | N W_k
  O M_w  | P ec(P-1) | Q ec(P) | R ec(P+1)
"""

from __future__ import annotations

from dataclasses import dataclass

from .build_workbook import Ctx, Layout as L
from .xlstyle import (
    ALIGN_C, BORDER_THIN, F_HEADER, F_LABEL, F_SMALL_IT, FILL_GREY, FILL_NAVY, FILL_PANEL,
    FMT_DATE, FMT_GEN, FMT_IDX, FMT_INT, FMT_MOD, FMT_PCT, GREY_DARK, STEEL,
    col, chart_legend, font, formula, header_row, jump, label, link, note, presentation_setup,
    print_setup, put, section, set_widths, title,
)

COH_HEADERS = ["#", "Written month", "Month end", "Days", "Mo#", "Mid-month",
               "Abs mo#", "Weight w", "Index before mo (W_pre)", "Index at mo end (W_post)",
               "# chgs in mo", "Days after 1st chg", "Split p", "Written rate index (W)",
               "Written mod (M_w)", "Earned in P-1 (ec)", "Earned in P (ec)",
               "Earned in P+1 (ec)", "Hist net price (q)", "Net path (Q)", "Index in force"]

# D70 stepped-mod columns, written at ``mod_col`` — the mod leg's I..N in
# miniature, so the two legs can be audited side by side
MOD_HEADERS = ["Mod idx before mo", "Mod idx at mo end", "# mod acts in mo",
               "Days after 1st act", "Mod step index"]


@dataclass(frozen=True)
class LogRefs:
    """Where a cohort block reads its rate-change log columns."""

    key_cond: str        # formula-ready selected-key reference, e.g. "nr_SelKey" or "$A$25"
    eff: str = "rl_eff"
    ln: str = "rl_ln1p"
    effmonth: str = "rl_effmonth"
    first: str = "rl_first"
    daysafter: str = "rl_daysafter"
    taken_only: bool = False   # solver base: taken rows only (DECISIONS.md D13)
    status: str = "rl_status"  # which log's status column taken_only filters on

    @property
    def sp_status(self) -> str:   # extra SUMPRODUCT factor
        return f'*({self.status}="taken")' if self.taken_only else ""

    @property
    def ifs_status(self) -> str:  # extra SUMIFS/COUNTIFS condition pair
        return f',{self.status},"taken"' if self.taken_only else ""


@dataclass(frozen=True)
class ModAnchors:
    """Absolute cell refs of the four mod-path anchors and three slopes."""

    x: tuple  # (x0, x1, x2, x3) close-of-day serials
    m: tuple  # (m0, m1, m2, m3) mod values
    s: tuple  # (s0, s1, s2) segment slopes


def write_mod_anchor_cells(ws, ctx: Ctx, row: int, first_col: int,
                           mind_ref, m0_ref, asof_ref, m1_ref, mprior_ref, m2_ref,
                           dm1_expr: str = "", steps_cond: str = "",
                           mend_ref: str = "") -> ModAnchors:
    """Write the 4 anchors (x, M) + 3 slopes into one row of cells.

    Anchor x-coordinates are close-of-day serials: date + 1 (DECISIONS.md D2).
    ``dm1_expr`` shifts M_1 and M_2 (scenario lever), e.g. "+N(INDEX(sc_dm1,2))".
    Blank M_prior -> the backward anchor is placed ON the M_0->M_1 line, which
    makes backward extrapolation exactly the extended M_0->M_1 segment (§3.3.1).

    Under ``steps_cond`` the third anchor moves back to 12/31/(P-1) (D70), and
    an as-of date on or after that lands it on top of the second anchor: the
    slope denominator goes to zero and the whole block reads #DIV/0!. "Mods as
    of 12/31 of the prior year" is an entirely ordinary thing to type, so the
    degenerate case is handled rather than validated away. There is simply no
    history segment left to run: the engine drops the switch anchor
    (``MonthlyEngine`` skips it unless ``x_switch`` is strictly later), leaving
    the M_prior -> M_0 line extrapolated, or a flat M_0 when M_prior is blank.
    Nudging x2 one day past x1 and putting m2 on that same line reproduces
    both, because a one-day segment at slope s is the line at slope s.
    """
    c = lambda i: f"${col(first_col + i)}${row}"
    x0, x1, x2, x3 = (c(i) for i in range(4))
    m0, m1, m2, m3 = (c(i) for i in range(4, 8))
    s0, s1, s2 = (c(i) for i in range(8, 11))
    sw = "DATE(nr_PlanYear-1,12,31)+1"           # the D70 switch coordinate
    deg = f"{sw}<={x1}"                          # no history segment survives
    fmls = {
        x0: f"=EDATE({asof_ref},-12)+1",
        x1: f"={asof_ref}+1",
        x2: ("=DATE(nr_PlanYear,12,31)+1" if not steps_cond else
             f"=IF({steps_cond},IF({deg},{x1}+1,{sw}),DATE(nr_PlanYear,12,31)+1)"),
        x3: "=DATE(nr_PlanYear+1,12,31)+1",
        m1: f"={m0_ref}",
        m2: (f"={m1_ref}{dm1_expr}" if not steps_cond else
             f"=IF({steps_cond},IF({deg},IF(N({mprior_ref})=0,{m1},"
             f"{m1}+({m1}-{mprior_ref})/({x1}-{x0})),{mend_ref}),"
             f"{m1_ref}{dm1_expr})"),
        m3: f"=IF(N({m2_ref})=0,{m1_ref},{m2_ref}){dm1_expr}",
        m0: f"=IF(N({mprior_ref})=0,{m1}-({m2}-{m1})/({x2}-{x1})*({x1}-{x0}),{mprior_ref})",
        s0: f"=({m1}-{m0})/({x1}-{x0})",
        s1: f"=({m2}-{m1})/({x2}-{x1})",
        s2: f"=({m3}-{m2})/({x3}-{x2})",
    }
    for addr, f in fmls.items():
        formula(ws, addr.replace("$", ""), f, fmt=FMT_GEN)
        ws[addr.replace("$", "")].font = font(GREY_DARK, size=9)
    return ModAnchors(x=(x0, x1, x2, x3), m=(m0, m1, m2, m3), s=(s0, s1, s2))


def mod_drift_at_switch(m0_ref: str, asof_ref: str, m1_ref: str,
                        mprior_ref: str) -> str:
    """Where the drift path reaches 12/31/(P-1) — the fallback for a blank
    M_endPrior (D70).

    Blank does not mean "the mod stops at M_0". It means the planner has not
    restated the level, so the path they DID describe runs on to the switch
    date, and that is what the oracle uses (``MonthlyEngine`` evaluates the
    UNSWITCHED anchor path at 12/31/(P-1) when ``m_end_prior`` is None). M_0
    is a cruder guess that silently discards the drift already entered — on
    the sample book it is off by about a tenth of a mod point, which is small,
    wrong, and invisible.

    Two segments, because an as-of date at or after the switch puts it on the
    M_prior -> M_0 leg instead. With M_prior blank the two legs are the same
    line by construction, so the one branch covers it. Neither denominator can
    vanish: the as-of date is validated before 12/31/P, and the M_prior anchor
    is a fixed twelve months back.
    """
    sw = "(DATE(nr_PlanYear-1,12,31)+1)"
    xa, xe = f"({asof_ref}+1)", "(DATE(nr_PlanYear,12,31)+1)"
    xp = f"(EDATE({asof_ref},-12)+1)"
    seg1 = f"({m0_ref}+({m1_ref}-{m0_ref})/({xe}-{xa})*({sw}-{xa}))"
    seg0 = f"({mprior_ref}+({m0_ref}-{mprior_ref})/({xa}-{xp})*({sw}-{xp}))"
    return f"IF(AND({sw}<{xa},N({mprior_ref})<>0),{seg0},{seg1})"


def mod_value_formula(x_expr: str, a: ModAnchors) -> str:
    """Piecewise-linear mod path at ``x_expr`` (extrapolates both ends)."""
    return (
        f"=IF({x_expr}<{a.x[1]},{a.m[0]}+{a.s[0]}*({x_expr}-{a.x[0]}),"
        f"IF({x_expr}<{a.x[2]},{a.m[1]}+{a.s[1]}*({x_expr}-{a.x[1]}),"
        f"{a.m[2]}+{a.s[2]}*({x_expr}-{a.x[2]})))"
    )


def write_cohort_block(
    ws, ctx: Ctx, first_row: int, refs: LogRefs, t_ref: str,
    srow_ref: str | None, ssum_ref: str | None,
    anchors: ModAnchors | None, include_rate=True, include_mod=True,
    header=True, header_row_at: int | None = None, net: dict | None = None,
    mod_refs: LogRefs | None = None, mod_col: int | None = None,
    mod_base_ref: str = "", mod_count_ref: str = "",
):
    """Emit the 48-row cohort block starting at ``first_row``, columns A..R.

    ``net`` (D39) adds the net-rate-selection columns S/T/U:
      S  q_hist = W x M_w/M_ind (mod leg only while the mod adjustment is on)
      T  Q      = q_hist before 1/1/P; Q(k-12) x (1 + x) from 1/1/P
      U  index in force = Q in net mode, W otherwise (aggregates read U)
    Expected keys: mode (boolean condition), x, x1, modeff (condition), mind.
    Requires an M_w column O (written here or linked in by the caller).
    """
    if header:
        hr = header_row_at if header_row_at is not None else first_row - 1
        header_row(ws, hr, 1, COH_HEADERS,
                   fill=FILL_GREY, fnt=font(GREY_DARK, bold=True, size=9))
        if mod_col is not None:
            header_row(ws, hr, mod_col, MOD_HEADERS,
                       fill=FILL_GREY, fnt=font(GREY_DARK, bold=True, size=9))
    last = first_row + L.N_COH - 1
    for i in range(L.N_COH):
        r = first_row + i
        put(ws, f"A{r}", i + 1, fnt=font(GREY_DARK, size=9), align=ALIGN_C)
        if i == 0:
            formula(ws, f"B{r}", "=DATE(nr_PlanYear-2,1,1)", fmt="mmm yyyy")
        else:
            formula(ws, f"B{r}", f"=EDATE($B{r - 1},1)", fmt="mmm yyyy")
        formula(ws, f"C{r}", f"=EOMONTH($B{r},0)", fmt=FMT_DATE)
        formula(ws, f"D{r}", f"=$C{r}-$B{r}+1", fmt=FMT_INT)
        formula(ws, f"E{r}", f"=MONTH($B{r})", fmt=FMT_INT)
        formula(ws, f"F{r}", f"=($B{r}+$C{r}+1)/2", fmt="0.0")
        formula(ws, f"G{r}", f"=YEAR($B{r})*12+MONTH($B{r})-1", fmt=FMT_INT)
        if srow_ref is None:
            formula(ws, f"H{r}", "=1", fmt="0.00")
        else:
            formula(ws, f"H{r}",
                    f'=IF(OR(nr_SeasonOn="OFF",{srow_ref}=0,{ssum_ref}=0),1,'
                    f"12*INDEX(se_block,{srow_ref},$E{r})/{ssum_ref})", fmt="0.00")
        if include_rate:
            k = refs.key_cond
            formula(ws, f"I{r}",
                    f"=EXP(SUMPRODUCT(({refs.eff}<$B{r})*(rl_key={k}){refs.sp_status}*{refs.ln}))",
                    fmt=FMT_IDX)
            formula(ws, f"J{r}",
                    f"=EXP(SUMPRODUCT(({refs.eff}<=$C{r})*(rl_key={k}){refs.sp_status}*{refs.ln}))",
                    fmt=FMT_IDX)
            formula(ws, f"K{r}",
                    f'=COUNTIFS(rl_key,{k},{refs.eff},">="&$B{r},{refs.eff},"<="&$C{r}{refs.ifs_status})',
                    fmt=FMT_INT)
            formula(ws, f"L{r}",
                    f"=SUMIFS({refs.daysafter},rl_key,{k},{refs.effmonth},$B{r},{refs.first},1{refs.ifs_status})",
                    fmt=FMT_INT)
            formula(ws, f"M{r}", f"=IF($K{r}=0,0,MIN(1,$L{r}/$D{r}))", fmt="0.000")
            formula(ws, f"N{r}", f"=IF($K{r}=0,$J{r},(1-$M{r})*$I{r}+$M{r}*$J{r})", fmt=FMT_IDX)
        if include_mod and anchors is not None:
            drift = mod_value_formula(f"$F{r}", anchors)[1:]
            if mod_refs is None or mod_col is None:
                formula(ws, f"O{r}", f"={drift}", fmt=FMT_MOD)
            else:
                # D70: five columns mirroring the rate leg's I..M, then the
                # same blend. Drift owns cohorts before 1/1/P; the step log
                # owns the rest, compounding on the current-year-end mod.
                mk = mod_refs.key_cond
                c = [col(mod_col + i) for i in range(5)]
                formula(ws, f"{c[0]}{r}",
                        f"=EXP(SUMPRODUCT(({mod_refs.eff}<$B{r})*(ml_key={mk})"
                        f"{mod_refs.sp_status}*{mod_refs.ln}))", fmt=FMT_IDX)
                formula(ws, f"{c[1]}{r}",
                        f"=EXP(SUMPRODUCT(({mod_refs.eff}<=$C{r})*(ml_key={mk})"
                        f"{mod_refs.sp_status}*{mod_refs.ln}))", fmt=FMT_IDX)
                formula(ws, f"{c[2]}{r}",
                        f"=COUNTIFS(ml_key,{mk},{mod_refs.eff},\">=\"&$B{r},"
                        f"{mod_refs.eff},\"<=\"&$C{r}{mod_refs.ifs_status})", fmt=FMT_INT)
                formula(ws, f"{c[3]}{r}",
                        f"=SUMIFS({mod_refs.daysafter},ml_key,{mk},"
                        f"{mod_refs.effmonth},$B{r},{mod_refs.first},1"
                        f"{mod_refs.ifs_status})", fmt=FMT_INT)
                formula(ws, f"{c[4]}{r}",
                        f"=IF(${c[2]}{r}=0,${c[1]}{r},"
                        f"(1-MIN(1,${c[3]}{r}/$D{r}))*${c[0]}{r}"
                        f"+MIN(1,${c[3]}{r}/$D{r})*${c[1]}{r})", fmt=FMT_IDX)
                for i in range(5):
                    ws[f"{c[i]}{r}"].font = font(GREY_DARK, size=9)
                formula(ws, f"O{r}",
                        f"=IF(AND({mod_count_ref}>0,$G{r}>=nr_MStartP),"
                        f"{mod_base_ref}*${c[4]}{r},{drift})", fmt=FMT_MOD)
        for cL, (ms, me) in (("P", ("nr_MStartPm1", "nr_MEndPm1")),
                             ("Q", ("nr_MStartP", "nr_MEndP")),
                             ("R", ("nr_MStartP1", "nr_MEndP1"))):
            formula(ws, f"{cL}{r}",
                    f"=MIN(1,MAX(0,({me}-$G{r}+0.5)/{t_ref}))"
                    f"-MIN(1,MAX(0,({ms}-1-$G{r}+0.5)/{t_ref}))", fmt="0.0000")
        if net is not None:
            # blank (not zero) when the combo runs the explicit program, so the
            # net columns stop reading as live machinery when the mode is off
            formula(ws, f"S{r}",
                    f'=IF({net["mode"]},$N{r}*IF({net["modeff"]},$O{r}/{net["mind"]},1),"")',
                    fmt=FMT_IDX)
            formula(ws, f"T{r}",
                    f'=IF(NOT({net["mode"]}),"",IF($G{r}<nr_MStartP,$S{r},$T{r - 12}*'
                    f"(1+IF($G{r}<nr_MStartP1,{net['x']},{net['x1']}))))", fmt=FMT_IDX)
            formula(ws, f"U{r}", f"=IF({net['mode']},$T{r},$N{r})", fmt=FMT_IDX)
    return dict(
        first=first_row, last=last,
        w=f"$H${first_row}:$H${last}", W=f"$N${first_row}:$N${last}",
        Mw=f"$O${first_row}:$O${last}", ec_pm1=f"$P${first_row}:$P${last}",
        ec_p=f"$Q${first_row}:$Q${last}", ec_p1=f"$R${first_row}:$R${last}",
        absmi=f"$G${first_row}:$G${last}", mid=f"$F${first_row}:$F${last}",
        som=f"$B${first_row}:$B${last}",
        U=(f"$U${first_row}:$U${last}" if net is not None
           else f"$N${first_row}:$N${last}"),
    )


def cy_aggregate(w_rng: str, ec_rng: str, val_rng: str | None) -> str:
    """E_CY / Mbar_CY aggregate-ratio formula over a cohort block."""
    if val_rng is None:
        return f"=SUMPRODUCT({w_rng},{ec_rng})"
    return f"=SUMPRODUCT({w_rng},{ec_rng},{val_rng})/SUMPRODUCT({w_rng},{ec_rng})"


# ---------------------------------------------------------------------------
# Rate Engine (visible, selected combo)
# ---------------------------------------------------------------------------


def build_rate_engine(ctx: Ctx):
    ws = ctx.wb["Rate Engine"]
    p = ctx.p
    title(ws, "A1", "Rate Engine — monthly written index x earning matrix",
          "Selected BU x LOB. 48 writing cohorts Jan (P-2) .. Dec (P+1); earned months Jan (P-1) .. Dec (P+1). "
          "Helper columns preferred over mega-formulas; every column defined on Methodology.")

    # Selected combo banner
    label(ws, "A3", "Selected:")
    link(ws, "B3", "=nr_SelKey", bold=True)
    jump(ws, "D3", "Control!C7", "Change BU/state selection >")

    # Header scalars (left: outputs; right: parameters)
    section(ws, 4, "A", "Calendar-year aggregates")
    hdr = [
        ("Indication rate level, fully earned (CRL_ind)",
         '=EXP(SUMPRODUCT((rl_key=nr_SelKey)*(rl_cons="Y")*rl_ln1p))', FMT_IDX,
         "nr_CRLind", "Cumulative rate level in the indication: product over considered rows"),
        ('="CY "&(nr_PlanYear-1)&" earned rate level  E_CY(P-1)"', None, FMT_IDX,
         "nr_ECY_Pm1", "Calendar-year earned rate index, plan year minus 1"),
        ('="CY "&nr_PlanYear&" earned rate level  E_CY(P)"', None, FMT_IDX,
         "nr_ECY_P", "Calendar-year earned rate index, plan year"),
        ('="CY "&(nr_PlanYear+1)&" earned rate level  E_CY(P+1)"', None, FMT_IDX,
         "nr_ECY_P1", "Calendar-year earned rate index, plan year + 1"),
        ('="Rate earn-in factor  A_rate("&nr_PlanYear&") = CRL / E_CY"', "=nr_CRLind/nr_ECY_P", FMT_IDX,
         "nr_Arate_P", "Rate adjustment factor for the plan year"),
        ('="Rate earn-in factor  A_rate("&(nr_PlanYear+1)&")"', "=nr_CRLind/nr_ECY_P1", FMT_IDX,
         "nr_Arate_P1", "Rate adjustment factor for plan year + 1"),
        ("CY earned rate chg vs indication", "=nr_ECY_P/nr_CRLind-1", "+0.0%;-0.0%;0.0%",
         "nr_EChgVsInd", "E_CY(P)/CRL_ind - 1"),
        ('="Earned rate chg "&(nr_PlanYear-1)&" -> "&nr_PlanYear', "=nr_ECY_P/nr_ECY_Pm1-1", "+0.0%;-0.0%;0.0%",
         "nr_YoY_P", "Year-over-year earned rate change into P"),
        ('="Earned rate chg "&nr_PlanYear&" -> "&(nr_PlanYear+1)&" (carryover + new)"', "=nr_ECY_P1/nr_ECY_P-1",
         "+0.0%;-0.0%;0.0%", "nr_YoY_P1", "Year-over-year earned rate change into P+1"),
    ]
    cf, cl = L.RE_COH_FIRST, L.RE_COH_LAST
    # Aggregates read the IN-FORCE index column U: the combined net path when a
    # net rate selection is active, the explicit written index otherwise (D39).
    ecy = {
        "nr_ECY_Pm1": cy_aggregate(f"$H${cf}:$H${cl}", f"$P${cf}:$P${cl}", f"$U${cf}:$U${cl}"),
        "nr_ECY_P": cy_aggregate(f"$H${cf}:$H${cl}", f"$Q${cf}:$Q${cl}", f"$U${cf}:$U${cl}"),
        "nr_ECY_P1": cy_aggregate(f"$H${cf}:$H${cl}", f"$R${cf}:$R${cl}", f"$U${cf}:$U${cl}"),
    }
    for i, (lbl, f, fmt, name, desc) in enumerate(hdr):
        r = 5 + i
        if lbl.startswith("="):
            formula(ws, f"A{r}", lbl)
            ws[f"A{r}"].font = F_LABEL
        else:
            label(ws, f"A{r}", lbl)
        formula(ws, f"C{r}", f if f else ecy[name], fmt=fmt, border=BORDER_THIN,
                bold=name in ("nr_ECY_P", "nr_Arate_P"))
        ctx.define(name, "Rate Engine", f"$C${r}", desc)

    prm = [
        ("Policy term T (months, workbook-level)", "=nr_TermMonths",
         FMT_INT, None, None),  # green link; the name lives on Inputs (D37)
        ("Seasonality row (0 = none)",
         "=IF(COUNTIF(se_state,nr_SelState)=0,0,MATCH(nr_SelState,se_state,0))",
         FMT_INT, "re_srow", "tbl_Seasonality row for the selected state, 0 if absent"),
        ("Seasonality row sum", "=IF($F$6=0,0,INDEX(se_sum,$F$6))", "0.00",
         "re_ssum", "Sum of the selected state's 12 seasonality weights"),
        ("Seasonality active?", '=AND(nr_SeasonOn="ON",$F$6>0,$F$7>0)', FMT_GEN,
         "re_season_active", "TRUE when non-uniform written weights are in force"),
        (f"Months Jan-Dec {p - 1} (abs index)", "=(nr_PlanYear-1)*12", FMT_INT,
         "nr_MStartPm1", "Absolute month index of Jan (P-1); year*12+month-1 convention"),
        (f"Months Jan-Dec {p} (abs index)", "=nr_PlanYear*12", FMT_INT,
         "nr_MStartP", "Absolute month index of Jan P"),
        (f"Months Jan-Dec {p + 1} (abs index)", "=(nr_PlanYear+1)*12", FMT_INT,
         "nr_MStartP1", "Absolute month index of Jan (P+1)"),
        (f"E_CY({p}) recomputed from matrix", None, FMT_IDX,
         "re_ecy_matrix", "Cross-check: E_CY(P) summed from the monthly matrix columns"),
    ]
    for i, (lbl, f, fmt, name, desc) in enumerate(prm):
        r = 5 + i
        label(ws, f"E{r}", lbl)
        if f:
            if name is None:
                link(ws, f"F{r}", f, fmt=fmt, border=BORDER_THIN)
            else:
                formula(ws, f"F{r}", f, fmt=fmt, border=BORDER_THIN)
        if name:
            ctx.define(name, "Rate Engine", f"$F${r}", desc)
    for i, (name, desc) in enumerate([("nr_MEndPm1", "Abs month index of Dec (P-1)"),
                                      ("nr_MEndP", "Abs month index of Dec P"),
                                      ("nr_MEndP1", "Abs month index of Dec (P+1)")]):
        r = 9 + i
        formula(ws, f"G{r}", f"=$F${r}+11", fmt=FMT_INT)
        ctx.define(name, "Rate Engine", f"$G${r}", desc)
    # matrix cross-check E_CY(P): year-P columns of the matrix
    mc0 = L.RE_MATRIX_COL + 12
    mc1 = L.RE_MATRIX_COL + 23
    formula(ws, "F12",
            f"=SUM({col(mc0)}{L.RE_ROW_NUM}:{col(mc1)}{L.RE_ROW_NUM})"
            f"/SUM({col(mc0)}{L.RE_ROW_DEN}:{col(mc1)}{L.RE_ROW_DEN})", fmt=FMT_IDX,
            border=BORDER_THIN)

    # Cohort block (net columns S/T/U driven by the Bridge's nr_Net* cells)
    section(ws, L.RE_COH_HDR - 1, "A",
            "Writing cohorts (48 months) — index construction and CY earning fractions")
    refs = LogRefs(key_cond="nr_SelKey")
    write_cohort_block(ws, ctx, L.RE_COH_FIRST, refs, "$F$5", "$F$6", "$F$7",
                       anchors=None, include_mod=False, header=True,
                       header_row_at=L.RE_COH_HDR,
                       net=dict(mode="nr_NetMode", x="nr_NetSelP", x1="nr_NetSelP1",
                                modeff='nr_ModAdjEff="ON"', mind="nr_MInd"))
    ws.row_dimensions[L.RE_COH_HDR].height = 42
    for i in range(L.N_COH):
        r = L.RE_COH_FIRST + i
        link(ws, f"O{r}", f"='Mod Engine'!$E{r}", fmt=FMT_MOD)
    formula(ws, "H3",
            '=IF(nr_NetMode,"NET RATE SELECTION ACTIVE: index, E_CY, and A_rate are the '
            'COMBINED rate x mod path from column U (A_mod = 1).","")')
    ws["H3"].font = font(STEEL, bold=True, size=10)

    # Earning matrix
    mcol0, mcol1 = L.RE_MATRIX_COL, L.RE_MATRIX_LAST_COL
    put(ws, f"{col(mcol0)}{L.RE_COH_HDR - 2}",
        "Earning matrix e(k, m): share of cohort k earned in month m (mid-month convention)",
        fnt=font(STEEL, bold=True, size=10))
    for j in range(L.N_MONTHS):
        cL = col(mcol0 + j)
        if j == 0:
            formula(ws, f"{cL}{L.RE_ROW_MIDX}", "=(nr_PlanYear-1)*12", fmt=FMT_INT)
            formula(ws, f"{cL}{L.RE_COH_HDR}", "=DATE(nr_PlanYear-1,1,1)", fmt="mmm-yy")
        else:
            prev = col(mcol0 + j - 1)
            formula(ws, f"{cL}{L.RE_ROW_MIDX}", f"={prev}{L.RE_ROW_MIDX}+1", fmt=FMT_INT)
            formula(ws, f"{cL}{L.RE_COH_HDR}", f"=EDATE({prev}{L.RE_COH_HDR},1)", fmt="mmm-yy")
        c = ws[f"{cL}{L.RE_COH_HDR}"]
        c.font = F_HEADER
        c.fill = FILL_NAVY
        c.alignment = ALIGN_C
        ws[f"{cL}{L.RE_ROW_MIDX}"].font = font(GREY_DARK, size=8)
        for i in range(L.N_COH):
            r = L.RE_COH_FIRST + i
            formula(ws, f"{cL}{r}",
                    f"=IF(OR({cL}$15<$G{r},{cL}$15>$G{r}+$F$5),0,"
                    f"IF(OR({cL}$15=$G{r},{cL}$15=$G{r}+$F$5),0.5,1)/$F$5)", fmt="0.000")
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=8)
        wrng = f"$H${L.RE_COH_FIRST}:$H${L.RE_COH_LAST}"
        Wrng = f"$U${L.RE_COH_FIRST}:$U${L.RE_COH_LAST}"  # in-force index (D39)
        erng = f"{cL}${L.RE_COH_FIRST}:{cL}${L.RE_COH_LAST}"
        formula(ws, f"{cL}{L.RE_ROW_NUM}", f"=SUMPRODUCT({wrng},{erng},{Wrng})", fmt=FMT_IDX)
        formula(ws, f"{cL}{L.RE_ROW_DEN}", f"=SUMPRODUCT({wrng},{erng})", fmt=FMT_IDX)
        formula(ws, f"{cL}{L.RE_ROW_EM}",
                f'=IF({cL}{L.RE_ROW_DEN}=0,"",{cL}{L.RE_ROW_NUM}/{cL}{L.RE_ROW_DEN})',
                fmt=FMT_IDX, bold=True)
        formula(ws, f"{cL}{L.RE_ROW_MINW}",
                f"=MIN(INDEX({Wrng},MAX(1,{cL}$15-$G${L.RE_COH_FIRST}-$F$5+1)):"
                f"INDEX({Wrng},{cL}$15-$G${L.RE_COH_FIRST}+1))", fmt=FMT_IDX)
        formula(ws, f"{cL}{L.RE_ROW_MAXW}",
                f"=MAX(INDEX({Wrng},MAX(1,{cL}$15-$G${L.RE_COH_FIRST}-$F$5+1)):"
                f"INDEX({Wrng},{cL}$15-$G${L.RE_COH_FIRST}+1))", fmt=FMT_IDX)
        formula(ws, f"{cL}{L.RE_ROW_VIOL}",
                f'=IF({cL}{L.RE_ROW_EM}="",0,MAX(0,{cL}{L.RE_ROW_MINW}-{cL}{L.RE_ROW_EM},'
                f"{cL}{L.RE_ROW_EM}-{cL}{L.RE_ROW_MAXW}))", fmt="0.000000")
        formula(ws, f"{cL}{L.RE_ROW_MODM}",
                f'=IF({cL}{L.RE_ROW_DEN}=0,"",SUMPRODUCT({cL}${L.RE_COH_FIRST}:{cL}${L.RE_COH_LAST},'
                f"$H${L.RE_COH_FIRST}:$H${L.RE_COH_LAST},'Mod Engine'!$E${L.RE_COH_FIRST}:$E${L.RE_COH_LAST})"
                f"/{cL}{L.RE_ROW_DEN})", fmt="0.000")
        formula(ws, f"{cL}{L.RE_ROW_MODNUM}",
                f"=SUMPRODUCT({cL}${L.RE_COH_FIRST}:{cL}${L.RE_COH_LAST},"
                f"$H${L.RE_COH_FIRST}:$H${L.RE_COH_LAST},'Mod Engine'!$E${L.RE_COH_FIRST}:$E${L.RE_COH_LAST})",
                fmt="0.0000")
        for rr in (L.RE_ROW_MINW, L.RE_ROW_MAXW, L.RE_ROW_VIOL, L.RE_ROW_MODM,
                   L.RE_ROW_MODNUM):
            ws[f"{cL}{rr}"].font = font(GREY_DARK, size=8)
    for rr, lbl in ((L.RE_ROW_NUM, "Monthly numerator  SUM w-e-W"),
                    (L.RE_ROW_DEN, "Monthly denominator  SUM w-e"),
                    (L.RE_ROW_EM, "Earned index E_m"),
                    (L.RE_ROW_MINW, "min W over contributing cohorts"),
                    (L.RE_ROW_MAXW, "max W over contributing cohorts"),
                    (L.RE_ROW_VIOL, "bound violation (Checks)"),
                    (L.RE_ROW_MODM, "Earned mod by month (M_w from Mod Engine)"),
                    (L.RE_ROW_MODNUM, "Mod numerator SUM w-e-M (quarterly price)")):
        put(ws, f"{col(mcol0 - 1)}{rr}", lbl, fnt=font(GREY_DARK, size=9))
        # repeat the strip labels at the RIGHT edge too — a reader scrolled 36
        # columns into the matrix should not lose track of what each row is
        put(ws, f"{col(mcol1 + 2)}{rr}", lbl, fnt=font(GREY_DARK, size=9))
    ctx.define("re_em_row", "Rate Engine",
               f"${col(mcol0)}${L.RE_ROW_EM}:${col(mcol1)}${L.RE_ROW_EM}",
               "Monthly earned rate index E_m, Jan (P-1) .. Dec (P+1)")
    ctx.define("re_viol_row", "Rate Engine",
               f"${col(mcol0)}${L.RE_ROW_VIOL}:${col(mcol1)}${L.RE_ROW_VIOL}",
               "Per-month earned-index bound violations (should be all 0)")
    ctx.define("re_den_row", "Rate Engine",
               f"${col(mcol0)}${L.RE_ROW_DEN}:${col(mcol1)}${L.RE_ROW_DEN}",
               "Monthly earned-exposure denominators")

    # Per-cohort matrix row-sum check (column BC)
    ckL = col(mcol1 + 1)
    put(ws, f"{ckL}{L.RE_COH_HDR}", "row-sum chk", fnt=font(GREY_DARK, bold=True, size=9),
        fill=FILL_GREY)
    for i in range(L.N_COH):
        r = L.RE_COH_FIRST + i
        formula(ws, f"{ckL}{r}",
                f"=IF(OR($G{r}<{col(mcol0)}$15,$G{r}+$F$5>{col(mcol1)}$15),0,"
                f"ABS(SUM({col(mcol0)}{r}:{col(mcol1)}{r})-1))", fmt="0.000000")
        ws[f"{ckL}{r}"].font = font(GREY_DARK, size=8)
    ctx.define("re_rowsum_chk", "Rate Engine",
               f"${ckL}${L.RE_COH_FIRST}:${ckL}${L.RE_COH_LAST}",
               "Per-cohort earning-profile row sums minus 1 (in-window cohorts; should be 0)")

    set_widths(ws, {"A": 5, "B": 11, "C": 11, "D": 6, "E": 5, "F": 10, "G": 8, "H": 6,
                    "I": 9, "J": 9, "K": 6, "L": 10, "M": 7, "N": 9, "O": 9, "P": 8,
                    "Q": 8, "R": 8, "S": 9, "T": 9, "U": 9})
    for j in range(L.N_MONTHS):
        ws.column_dimensions[col(L.RE_MATRIX_COL + j)].width = 7.5
    # mid-month sampling date reads as a date, not a raw serial
    for i in range(L.N_COH):
        ws[f"F{L.RE_COH_FIRST + i}"].number_format = "m/d/yy h:mm"
    # index-construction detail columns are collapsible (audit on demand)
    for cL in "IJKLM":
        ws.column_dimensions[cL].outlineLevel = 1
    presentation_setup(ws, gridlines_off=False, zoom=85, freeze=f"I{L.RE_COH_FIRST}",
                       tab_color=STEEL)
    print_setup(ws)


# ---------------------------------------------------------------------------
# Mod Engine (visible, selected combo)
# ---------------------------------------------------------------------------


def build_mod_engine(ctx: Ctx):
    ws = ctx.wb["Mod Engine"]
    p = ctx.p
    title(ws, "A1", "Mod Engine — written schedule-mod path and earned mod",
          "Piecewise-linear written mod path through the tbl_LR anchors; earned mod is the "
          "exposure-weighted AVERAGE through the same earning matrix (mods are levels, not changes).")
    label(ws, "A3", "Selected:")
    link(ws, "B3", "=nr_SelKey", bold=True)
    jump(ws, "D3", "Control!C7", "Change BU/state selection >")

    # Anchor table
    section(ws, 4, "E", "Mod path anchors (x = close-of-day boundary: date + 1; see Methodology)")
    header_row(ws, 5, 5, ["Anchor", "x (boundary date)", "M value", "Slope / day", "pts / month"],
               widths=[26, 12, 10, 12, 10], fill=FILL_NAVY)
    # D70: with actions in the Mod Log the drift path stops a year early, at
    # 12/31/(P-1) on M_endPrior, and the step log carries the plan year onward.
    # Anchor A2 moves rather than being overridden, so the two mechanisms meet
    # end to end and no movement is ever counted twice.
    anchor_rows = [
        ("A0: backward (as-of - 12 mo)", "=EDATE(nr_M0Asof,-12)+1",
         "=IF(N(nr_MPrior)=0,$G$7-($G$8-$G$7)/($F$8-$F$7)*($F$7-$F$6),nr_MPrior)"),
        ("A1: M_0 as-of", "=nr_M0Asof+1", "=nr_M0"),
        ("A2: M_1 at plan-yr end (12/31 of the CURRENT year when the log drives)",
         "=IF($C$11>0,DATE(nr_PlanYear-1,12,31)+1,DATE(nr_PlanYear,12,31)+1)",
         "=IF($C$11>0,$C$12,nr_M1)"),
        ("A3: M_2 at plan-yr+1 end", "=DATE(nr_PlanYear+1,12,31)+1",
         "=IF(N(nr_M2)=0,nr_M1,nr_M2)"),
    ]
    for i, (lbl, fx, fm) in enumerate(anchor_rows):
        r = 6 + i
        label(ws, f"E{r}", lbl)
        formula(ws, f"F{r}", fx, fmt="m/d/yy", border=BORDER_THIN)
        formula(ws, f"G{r}", fm, fmt=FMT_MOD, border=BORDER_THIN)
        if i < 3:
            formula(ws, f"H{r}", f"=($G${r + 1}-$G${r})/($F${r + 1}-$F${r})", fmt="0.000000",
                    border=BORDER_THIN)
            # the human-readable drift speed (mod points per average month)
            formula(ws, f"I{r}", f"=$H${r}*30.4375*100", fmt="0.00", border=BORDER_THIN)
            ws[f"I{r}"].font = font(GREY_DARK, size=9)
    note(ws, "E10", "Blank M_prior places A0 on the M_0->M_1 line (backward extrapolation of that "
                    "segment, §3.3.1). Blank M_2 = M_1 (flat beyond 12/31/P). With actions in "
                    "the Mod Log, A2 lands a year earlier — at 12/31 of the current year on "
                    "M_endPrior — and the log carries the path from there (D70).")
    anchors = ModAnchors(
        x=("$F$6", "$F$7", "$F$8", "$F$9"), m=("$G$6", "$G$7", "$G$8", "$G$9"),
        s=("$H$6", "$H$7", "$H$8"))

    # Header echo + outputs
    hdr = [("Mod assumed in indication (M_ind)", "=nr_MInd", FMT_MOD),
           ("Current avg written mod (M_0)", "=nr_M0", FMT_MOD),
           ("Current mod as-of date", "=nr_M0Asof", FMT_DATE),
           ("Projected mod, end of plan yr (M_1)", "=nr_M1", FMT_MOD),
           ("Mod ~1 yr before as-of (M_prior, opt)", "=nr_MPrior", FMT_MOD),
           ("Projected mod, end plan yr+1 (M_2, opt)", "=nr_M2", FMT_MOD)]
    section(ws, 4, "A", "Inputs (from tbl_LR via Bridge)")
    for i, (lbl, f, fmt) in enumerate(hdr):
        r = 5 + i
        label(ws, f"A{r}", lbl)
        link(ws, f"C{r}", f, fmt=fmt, border=BORDER_THIN)
    # D70: the two cells that decide whether this sheet runs on drift or on the
    # Mod Log. Derived here rather than on the Bridge — the resolver block is
    # already at its layout budget, and these are only ever read locally.
    label(ws, "A11", "Mod actions logged for this combo")
    formula(ws, "C11", "=COUNTIF(ml_key,nr_SelKey)", fmt=FMT_INT, border=BORDER_THIN)
    label(ws, "A12", "Projected mod, end of CURRENT yr (M_endPrior)")
    formula(ws, "C12",
            "=IF(br_selrow=0,nr_M0,IF(N(INDEX(lr_mendprior,br_selrow))=0,"
            + mod_drift_at_switch("nr_M0", "nr_M0Asof", "nr_M1", "N(nr_MPrior)")
            + ",INDEX(lr_mendprior,br_selrow)))", fmt=FMT_MOD, border=BORDER_THIN)
    for rr in ("C11", "C12"):
        ws[rr].font = font(GREY_DARK, size=10)
    # one flat line, no wrap: column A is 6 wide and a wrapped note here would
    # stretch the row down the sheet
    put(ws, "A13",
        '=IF($C$11=0,"No mod actions logged — the written path is pure drift '
        '(M_0 -> M_1 -> M_2), exactly as before.","The Mod Log drives from 1/1: '
        '"&$C$11&" action(s) compound on "&TEXT($C$12,"0.000")&", and drift is '
        'confined to history.")', fnt=F_SMALL_IT)

    out = [
        ('="Earned mod CY "&(nr_PlanYear-1)', "F", FMT_MOD, "nr_MEarned_Pm1",
         "Exposure-weighted CY P-1 earned schedule mod"),
        ('="Earned mod CY "&nr_PlanYear', "G", FMT_MOD, "nr_MEarned_P",
         "Exposure-weighted CY P earned schedule mod"),
        ('="Earned mod CY "&(nr_PlanYear+1)', "H", FMT_MOD, "nr_MEarned_P1",
         "Exposure-weighted CY P+1 earned schedule mod"),
    ]
    section(ws, 4, "J", "Earned mod and adjustment factors")
    cf, cl = L.ME_COH_FIRST, L.ME_COH_LAST
    for i, (lbl, ecL, fmt, name, desc) in enumerate(out):
        r = 5 + i
        formula(ws, f"J{r}", lbl)
        ws[f"J{r}"].font = F_LABEL
        formula(ws, f"L{r}",
                f"=SUMPRODUCT($D${cf}:$D${cl},${ecL}${cf}:${ecL}${cl},$E${cf}:$E${cl})"
                f"/SUMPRODUCT($D${cf}:$D${cl},${ecL}${cf}:${ecL}${cl})",
                fmt=fmt, border=BORDER_THIN, bold=(name == "nr_MEarned_P"))
        ctx.define(name, "Mod Engine", f"$L${r}", desc)
    label(ws, "J8", f"Mod drift factor  A_mod({p}) = M_ind / earned mod")
    formula(ws, "L8",
            '=IF(nr_NetMode,1,IF(AND(nr_ModAdjMaster="ON",nr_ModAdjRow="ON"),'
            "nr_MInd/nr_MEarned_P,1))",
            fmt=FMT_IDX, border=BORDER_THIN, bold=True)
    ctx.define("nr_Amod_P", "Mod Engine", "$L$8",
               "Mod adjustment factor for the plan year (1 under a net selection, D39)")
    label(ws, "J9", f"Mod drift factor  A_mod({p + 1})")
    formula(ws, "L9",
            '=IF(nr_NetMode,1,IF(AND(nr_ModAdjMaster="ON",nr_ModAdjRow="ON"),'
            "nr_MInd/nr_MEarned_P1,1))",
            fmt=FMT_IDX, border=BORDER_THIN)
    ctx.define("nr_Amod_P1", "Mod Engine", "$L$9",
               "Mod adjustment factor for plan year + 1 (1 under a net selection)")
    label(ws, "J11", "Identity: written mod at 1/1/P")
    formula(ws, "K11", "=DATE(nr_PlanYear,1,1)", fmt="0")
    ws["K11"].font = font(GREY_DARK, size=9)
    formula(ws, "L11", mod_value_formula("$K$11", anchors), fmt=FMT_MOD, border=BORDER_THIN)
    ctx.define("me_identity", "Mod Engine", "$L$11",
               "Written mod at 1/1/P — equals CY P earned mod under the §3.3.5 identity")
    label(ws, "J12", "Identity applicable?")
    # a logged action puts a STEP in the path, and the Jan-1 identity is a
    # property of straight lines — so the Mod Log suspends it (D70)
    formula(ws, "L12",
            "=AND(N(nr_MPrior)=0,nr_TermMonths=12,NOT(re_season_active),$C$11=0)",
            border=BORDER_THIN)
    ctx.define("me_identity_ok", "Mod Engine", "$L$12",
               "TRUE when the linear-mod identity preconditions hold (no M_prior kink, "
               "annual term, uniform writings, no logged mod actions)")

    # Cohort table (links to Rate Engine + local M_w)
    section(ws, 15, "A", "Writing cohorts — written mod path (columns A-D, F-H link to the "
                         "Rate Engine; T-Y build the Mod Log's step index)")
    header_row(ws, 16, 1, ["#", "Written month", "Mid-month", "Weight w", "Written mod (M_w)",
                           "Earned in P-1", "Earned in P", "Earned in P+1"],
               widths=[5, 12, 10, 8, 10, 9, 9, 9], fill=FILL_GREY,
               fnt=font(GREY_DARK, bold=True, size=9))
    ws.row_dimensions[16].height = 30
    # D70 step helpers at T..Y — the mod leg's own index construction, the same
    # five shapes the rate leg builds in I..M. Everything derives from the
    # cohort month in B, so the strip is self-contained and auditable in place.
    header_row(ws, 16, 20, ["Month end", "Mod idx before mo", "Mod idx at mo end",
                            "# mod acts in mo", "Days after 1st act", "Mod step index"],
               widths=[10, 11, 11, 9, 11, 10], fill=FILL_GREY,
               fnt=font(GREY_DARK, bold=True, size=9))
    re = "'Rate Engine'"
    for i in range(L.N_COH):
        r = L.ME_COH_FIRST + i
        link(ws, f"A{r}", f"={re}!$A{r}", fmt=FMT_INT, align=ALIGN_C)
        link(ws, f"B{r}", f"={re}!$B{r}", fmt="mmm yyyy")
        link(ws, f"C{r}", f"={re}!$F{r}", fmt="0.0")
        link(ws, f"D{r}", f"={re}!$H{r}", fmt="0.00")
        formula(ws, f"T{r}", f"=EOMONTH($B{r},0)", fmt=FMT_DATE)
        formula(ws, f"U{r}",
                f"=EXP(SUMPRODUCT((ml_eff<$B{r})*(ml_key=nr_SelKey)*ml_ln1p))", fmt=FMT_IDX)
        formula(ws, f"V{r}",
                f"=EXP(SUMPRODUCT((ml_eff<=$T{r})*(ml_key=nr_SelKey)*ml_ln1p))", fmt=FMT_IDX)
        formula(ws, f"W{r}",
                f'=COUNTIFS(ml_key,nr_SelKey,ml_eff,">="&$B{r},ml_eff,"<="&$T{r})',
                fmt=FMT_INT)
        formula(ws, f"X{r}",
                f"=SUMIFS(ml_daysafter,ml_key,nr_SelKey,ml_effmonth,$B{r},ml_first,1)",
                fmt=FMT_INT)
        formula(ws, f"Y{r}",
                f"=IF($W{r}=0,$V{r},(1-MIN(1,$X{r}/($T{r}-$B{r}+1)))*$U{r}"
                f"+MIN(1,$X{r}/($T{r}-$B{r}+1))*$V{r})", fmt=FMT_IDX)
        for cL in "TUVWXY":
            ws[f"{cL}{r}"].font = font(GREY_DARK, size=9)
        # B is the first of the cohort month, so B >= 1/1/P is exactly the
        # block writer's absolute-month test $G >= nr_MStartP
        formula(ws, f"E{r}",
                f"=IF(AND($C$11>0,$B{r}>=DATE(nr_PlanYear,1,1)),$C$12*$Y{r},"
                f"{mod_value_formula(f'$C{r}', anchors)[1:]})", fmt=FMT_MOD)
        link(ws, f"F{r}", f"={re}!$P{r}", fmt="0.0000")
        link(ws, f"G{r}", f"={re}!$Q{r}", fmt="0.0000")
        link(ws, f"H{r}", f"={re}!$R{r}", fmt="0.0000")
    ctx.define("me_mw_col", "Mod Engine", f"$E${L.ME_COH_FIRST}:$E${L.ME_COH_LAST}",
               "Written mod M_w(k) by cohort")
    # 48-row echo is audit detail — grouped, collapsible, still calculating
    for rr in range(L.ME_COH_FIRST, L.ME_COH_LAST + 1):
        ws.row_dimensions[rr].outlineLevel = 1

    # the path chart the tab always deserved (reads the Flow monthly table)
    from openpyxl.chart import LineChart, Reference
    from .sheets_report import CD0
    flow = ctx.wb["Flow Dashboard"]
    pathc = LineChart()
    pathc.add_data(Reference(flow, min_col=5, min_row=CD0, max_row=CD0 + L.N_MONTHS),
                   titles_from_data=True)
    pathc.add_data(Reference(flow, min_col=6, min_row=CD0, max_row=CD0 + L.N_MONTHS),
                   titles_from_data=True)
    pathc.set_categories(Reference(flow, min_col=2, min_row=CD0 + 1,
                                   max_row=CD0 + L.N_MONTHS))
    from openpyxl.chart.marker import Marker as _Marker
    for s, rgb in zip(pathc.series, (STEEL, "1F3864")):
        s.graphicalProperties.line.solidFill = rgb
        s.graphicalProperties.line.width = int(2.25 * 12700)
        s.marker = _Marker(symbol="none")
        s.smooth = False
    pathc.y_axis.number_format = "0.000"
    pathc.title = "Written vs earned schedule mod path"
    pathc.style = 2
    pathc.height = 7.5
    pathc.width = 13
    pathc.visible_cells_only = False
    pathc.x_axis.delete = False
    pathc.y_axis.delete = False
    chart_legend(pathc)          # legend beside the plot, not on it (D69)
    ws.add_chart(pathc, "J14")

    set_widths(ws, {"A": 6, "B": 12, "C": 10, "D": 7, "E": 9, "F": 9, "G": 9, "H": 9,
                    "I": 3, "J": 26, "K": 10, "L": 11})
    # step-index construction is audit detail, like the rate leg's I..M
    for cL in "TUVWXY":
        ws.column_dimensions[cL].outlineLevel = 1
    presentation_setup(ws, gridlines_off=False, zoom=90, freeze="A17", tab_color=STEEL)
    print_setup(ws)
