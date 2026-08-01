"""Workbook verification harness (brief §9 "Workbook verification").

Dimensioning (D37): one workbook per LOB; combos are BU x state. The harness
verifies one LOB workbook per invocation (--lob, default Property):

  A. Static scans of the formula layer: no merged cells, no volatile /
     dynamic-array / prohibited functions, no external references, defined
     names intact.
  B. Recalculate (Excel COM / LibreOffice) and prove zero formula errors,
     plus the chart-axis regression guard (D36).
  C. Oracle ties in the default state: worked-example Bridge cells, monthly
     series, EVERY BU x state combo, solver, scenario-blank identity.
  D. Toggle exercises: mutate inputs on scratch copies, recalculate, re-read,
     tie each state to a fresh oracle run, rescan for errors. (Runs a
     representative selector subsample; phase C already ties all combos.)

Usage:
    python tools/verify_workbook.py [--lob "Property"] [--workbook PATH] [--quick]

Exit code 0 = every assertion passed.
"""

from __future__ import annotations

import argparse
import datetime as dt
import math
import re
import shutil
import sys
from dataclasses import replace
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src import engine  # noqa: E402
from src.build_workbook import (  # noqa: E402
    Layout as L, SHEETS, _bu, degenerate_combo, load_config, output_path,
    sample_lr_rows, sample_mod_rows, sample_rate_rows, sample_seasonality_rows,
    sample_to_combo,
)
from src.sheets_inputs import lr_letter  # noqa: E402
from tools.recalc import recalc  # noqa: E402

ERR_STRINGS = ("#DIV/0!", "#REF!", "#N/A", "#NAME?", "#VALUE!", "#NUM!", "#NULL!",
               "#SPILL!", "#CALC!", "#GETTING_DATA")
BANNED_FUNCS = [
    "OFFSET", "INDIRECT", "TODAY", "NOW", "RAND", "RANDBETWEEN", "RANDARRAY",
    "XLOOKUP", "XMATCH", "FILTER", "SORTBY", "UNIQUE", "SEQUENCE", "LET",
    "LAMBDA", "IFERROR", "IFNA", "TEXTJOIN", "IFS", "SWITCH", "MAXIFS",
    "MINIFS", "CONCAT", "TOCOL", "TOROW", "VSTACK", "HSTACK", "TAKE", "DROP",
]
# word-boundary match so COUNTIFS/SUMIFS don't false-positive on IFS, etc.
BANNED_RE = re.compile(r"(?<![A-Z0-9_.])(" + "|".join(BANNED_FUNCS) + r")\(")

PASS = 0
FAIL = 0
FAILURES: list[str] = []


def _states(cfg):
    """Configured states in order, de-duplicated — the workbook's live
    roster extracts UNIQUES, so a repeated state in the config would
    otherwise walk the exhibit past its last live row (D62)."""
    return list(dict.fromkeys(cfg.states))


def check(desc: str, ok: bool, detail: str = ""):
    global PASS, FAIL
    if ok:
        PASS += 1
    else:
        FAIL += 1
        FAILURES.append(f"{desc}  {detail}")
        print(f"  FAIL  {desc}  {detail}")


def approx(a, b, tol=1e-9):
    if a is None or b is None:
        return False
    if isinstance(a, str) or isinstance(b, str):
        return False
    return abs(float(a) - float(b)) <= tol


def name_cell(wb, name: str):
    dn = wb.defined_names[name]
    for sheet, ref in dn.destinations:
        return wb[sheet][ref.replace("$", "")]
    raise KeyError(name)


def nval(wb, name: str):
    return name_cell(wb, name).value


def scan_errors(wb) -> list[str]:
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERR_STRINGS:
                    out.append(f"{ws.title}!{c.coordinate}={c.value}")
    return out


# ---------------------------------------------------------------------------
# Phase A: static scans
# ---------------------------------------------------------------------------


def phase_a(path: Path):
    print("Phase A: static formula-layer scans")
    wb = openpyxl.load_workbook(path, data_only=False)
    merged = {ws.title: len(ws.merged_cells.ranges) for ws in wb.worksheets
              if len(ws.merged_cells.ranges)}
    check("no merged cells anywhere", not merged, str(merged))

    banned_hits, ext_hits, longest = [], [], 0
    n_formulas = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    n_formulas += 1
                    longest = max(longest, len(v))
                    for hit in BANNED_RE.findall(v.upper()):
                        banned_hits.append(f"{ws.title}!{c.coordinate}: {hit}(")
                    if "[" in v and "]" in v:
                        ext_hits.append(f"{ws.title}!{c.coordinate}")
    check("no volatile / dynamic-array / prohibited functions", not banned_hits,
          "; ".join(banned_hits[:5]))
    check("no external-workbook references", not ext_hits, "; ".join(ext_hits[:5]))
    check("defined names all present and unbroken",
          all("#REF" not in (wb.defined_names[n].attr_text or "")
              for n in wb.defined_names) and len(list(wb.defined_names)) > 80,
          f"{len(list(wb.defined_names))} names")
    print(f"  formulas: {n_formulas:,}; longest: {longest} chars")
    check("automatic calculation mode",
          wb.calculation.calcMode in (None, "auto"), str(wb.calculation.calcMode))
    return wb


# ---------------------------------------------------------------------------
# Phase B/C: recalc + oracle ties in the default state
# ---------------------------------------------------------------------------


def tie_default_state(path: Path, cfg, lob, do_recalc=True):
    print("Phase B: recalculate and scan for errors")
    if do_recalc:
        used = recalc(path)
        print(f"  recalculated with {used}")
    wb = openpyxl.load_workbook(path, data_only=True)
    errs = scan_errors(wb)
    check("zero formula-error cells in every sheet (incl. hidden)", not errs,
          "; ".join(errs[:8]))
    check("Checks sheet: ALL CHECKS PASS", nval(wb, "ck_overall") == "ALL CHECKS PASS",
          str(nval(wb, "ck_overall")))

    # Chart axes must stay visible after the Excel resave (D36), and every
    # chart must plot hidden cells: "plot visible cells only" would blank any
    # chart whose staging rows are outline-hidden (openpyxl attr is
    # visible_cells_only — plotVisOnly is silently ignored).
    import zipfile
    deleted = 0
    total_axes = 0
    vis_only = []
    legend_over = []
    n_legends = 0
    title_over = []
    n_titles = 0
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if re.match(r"xl/charts/chart\d+\.xml$", name):
                xml = z.read(name).decode("utf-8", errors="replace")
                total_axes += len(re.findall(r"<c:(catAx|valAx|dateAx)>", xml))
                deleted += len(re.findall(r'<c:delete val="1"\s*/>', xml))
                # openpyxl writes an unprefixed default namespace; Excel's
                # resave rewrites with the c: prefix — accept either
                if not re.search(r'<(c:)?plotVisOnly val="0"', xml):
                    vis_only.append(name)
                # D69: a legend with no explicit <c:overlay val="0"> is drawn ON
                # TOP of the series. openpyxl omits the element and ECMA-376
                # defaults a missing overlay to TRUE, so Excel materialises
                # val="1" on resave — the same omitted-element trap as D36.
                leg = re.search(r"<(c:)?legend>.*?</(c:)?legend>", xml, re.S)
                if leg:
                    n_legends += 1
                    if not re.search(r'<(c:)?overlay val="0"', leg.group(0)):
                        legend_over.append(name.split("/")[-1])
                # D71: chart and AXIS titles carry the same <c:overlay>, and an
                # overlaid axis title is drawn behind the plot area. Guarding
                # every title (not just the legend) is what keeps a newly added
                # chart from silently reintroducing it.
                for t in re.finditer(r"<(c:)?title>.*?</(c:)?title>", xml, re.S):
                    n_titles += 1
                    if not re.search(r'<(c:)?overlay val="0"', t.group(0)):
                        title_over.append(name.split("/")[-1])
    check("chart axes visible after Excel resave (1 intentional hide only)",
          total_axes >= 16 and deleted == 1,
          f"{deleted} deleted of {total_axes} axes")
    check("every chart plots hidden cells (plotVisOnly=0)", not vis_only,
          "; ".join(vis_only[:5]))
    check(f"no legend overlaps its plot area ({n_legends} legends, D69)",
          not legend_over, "; ".join(legend_over[:6]))
    check(f"no chart or axis title overlaps its plot area ({n_titles} titles, D71)",
          not title_over, "; ".join(title_over[:6]))

    print("Phase C: oracle ties (default state)")
    p = cfg.plan_year
    lr_rows = sample_lr_rows(cfg, lob)
    rate_rows = sample_rate_rows(cfg, lob)
    mod_rows = sample_mod_rows(cfg, lob)   # D70: the oracle must see the log
    we = next(r for r in lr_rows if r["bu"] == _bu(cfg, 0) and r["state"] == cfg.states[0])
    combo = sample_to_combo(cfg, lob, we, rate_rows, mod_rows)
    m = engine.run_bridge(p, combo, "monthly")
    c = engine.run_bridge(p, combo, "continuous")

    ties = [
        ("nr_CRLind", m.crl_ind), ("nr_ECY_Pm1", m.e_cy[p - 1]), ("nr_ECY_P", m.e_cy[p]),
        ("nr_ECY_P1", m.e_cy[p + 1]), ("nr_Arate_P", m.a_rate_p),
        ("nr_Arate_P1", m.a_rate_p1), ("nr_MEarned_P", m.m_earned[p]),
        ("nr_MEarned_P1", m.m_earned[p + 1]), ("nr_Amod_P", m.a_mod_p),
        ("nr_Amod_P1", m.a_mod_p1), ("nr_LRcur", m.lr_current), ("nr_CYLR_P", m.cy_lr_p),
        ("nr_CYLR_P1", m.cy_lr_p1), ("nr_EChgVsInd", m.earned_chg_vs_ind),
        ("nr_YoY_P", m.yoy_earned_p), ("nr_YoY_P1", m.yoy_earned_p1),
        ("nr_WalkLR", m.cy_lr_p),   # Walkthrough assembly must tie the Bridge
    ]
    for name, expected in ties:
        check(f"{name} ties oracle monthly", approx(nval(wb, name), expected, 1e-9),
              f"wb={nval(wb, name)} oracle={expected}")
    check("CY LR within 0.10 pts of continuous closed form",
          approx(nval(wb, "nr_CYLR_P"), c.cy_lr_p, 1e-3))
    if lob.term_months == 12 and p == 2027:
        check("worked example CY LR rounds to 63.36%",
              round(float(nval(wb, "nr_CYLR_P")), 4) == 0.6336, str(nval(wb, "nr_CYLR_P")))

    # D76: the current-process reconciliation, on the default (drift-only)
    # selection. The waterfall legs are exercised on a stepped combo in 9j.
    me_top = L.ME_COH_LAST + 2
    rec = engine.process_reconciliation(p, combo)
    me0 = wb["Mod Engine"]
    for cl, want, nm in ((f"K{me_top + 3}", rec.m_avg, "current-process M_avg"),
                         (f"L{me_top + 3}", rec.lr_process, "current-process plan LR"),
                         (f"K{me_top + 7}", rec.m_earned, "model earned mod"),
                         (f"L{me_top + 7}", rec.lr_model, "model plan LR"),
                         (f"K{me_top + 8}", rec.m_end_prior, "start level"),
                         (f"K{me_top + 9}", rec.m_end_plan, "end level"),
                         (f"M{me_top + 12}", rec.gap_pts, "process-minus-model gap")):
        check(f"[reconciliation] {nm} ties oracle",
              approx(me0[cl].value, want, 1e-9),
              f"wb={me0[cl].value} oracle={want}")
    check("[reconciliation] a drift-only combo shows no step split",
          not rec.stepped and me0[f"M{me_top + 4}"].value == "n/a",
          f"stepped={rec.stepped} wb={me0[f'M{me_top + 4}'].value!r}")

    # monthly series: written index, written mod, earned index by month
    re_ws, me_ws = wb["Rate Engine"], wb["Mod Engine"]
    ok_w = ok_m = ok_e = True
    for i in range(L.N_COH):
        r = L.RE_COH_FIRST + i
        ok_w &= approx(re_ws[f"N{r}"].value, m.written_index[i], 1e-9)
        ok_m &= approx(me_ws[f"E{r}"].value, m.written_mod[i], 1e-9)
    for j in range(L.N_MONTHS):
        cell = re_ws.cell(row=L.RE_ROW_EM, column=L.RE_MATRIX_COL + j).value
        exp = m.earned_index_m[j]
        ok_e &= (approx(cell, exp, 1e-9) or (cell in (None, "") and math.isnan(exp)))
    check("48 cohort written-index values tie oracle", ok_w)
    check("48 cohort written-mod values tie oracle", ok_m)
    check("36 monthly earned-index values tie oracle", ok_e)

    # EVERY BU x state combo vs oracle (via the _calc results table)
    calc = wb["_calc"]
    bad = 0
    for i, lrr in enumerate(lr_rows):
        cmb = sample_to_combo(cfg, lob, lrr, rate_rows, mod_rows)
        om = engine.run_bridge(p, cmb, "monthly")
        r = L.CALC_RES_FIRST + i
        for colL, exp in (("C", om.crl_ind), ("E", om.e_cy[p]), ("F", om.e_cy[p + 1]),
                          ("G", om.m_earned[p]), ("I", om.a_rate_p), ("K", om.a_mod_p),
                          ("M", om.lr_current), ("N", om.cy_lr_p), ("O", om.cy_lr_p1)):
            if not approx(calc[f"{colL}{r}"].value, exp, 1e-9):
                bad += 1
                if bad <= 5:
                    key = f"{lrr['bu']}|{lrr['state']}"
                    print(f"  FAIL detail: {key} _calc!{colL}{r} wb="
                          f"{calc[f'{colL}{r}'].value} oracle={exp}")
    check(f"all {len(lr_rows)} BU x state combos tie oracle (9 metrics each)", bad == 0,
          f"{bad} mismatches")

    # solver seed + Mode B
    sres = engine.solve_rate_for_target(p, combo, m.cy_lr_p, dt.date(p, 4, 1))
    check("Solver Mode A returns the seeded +5.0% (acceptance §11.7)",
          approx(nval(wb, "nr_SolverR"), 0.05, 1e-9), str(nval(wb, "nr_SolverR")))
    check("Solver Mode A ties oracle solver", approx(nval(wb, "nr_SolverR"),
          sres.required_change, 1e-9))
    sv = wb["Solver"]
    tt = engine.solver_timing_table(p, combo, r=0.05, target_cy_lr=m.cy_lr_p)
    ok_b = all(approx(sv[f"G{28 + row['month']}"].value, row["cy_lr"], 1e-9) for row in tt)
    check("Solver Mode B 12-month table ties oracle", ok_b)
    if lob.term_months == 12:
        check("Solver Mode B latest qualifying month is April",
              sv["D42"].value == f"Apr {p}", str(sv["D42"].value))

    # S1 ships seeded (+2 pts on planned rows) as a self-teaching demo; the
    # remaining blank-lever scenarios must reproduce Base exactly
    sc = wb["Scenarios"]
    s1_combo = replace(combo, rate_changes=tuple(
        replace(rc, filed_pct=rc.filed_pct + 0.02) if rc.status == "planned" else rc
        for rc in combo.rate_changes))
    s1_m = engine.run_bridge(p, s1_combo, "monthly")
    check("seeded S1 (+2 pts on planned rows) ties oracle",
          approx(sc["D14"].value, s1_m.cy_lr_p, 1e-9),
          f"wb={sc['D14'].value} oracle={s1_m.cy_lr_p}")
    ok_s = all(approx(sc[f"{cL}14"].value, m.cy_lr_p, 1e-9) for cL in "EFG")
    check("blank-lever scenarios reproduce Base CY LR", ok_s)

    # State Summary in 'All' mode: EP-weighted aggregates per state (D38)
    ss = wb["State Summary"]
    results = {}
    for lrr in lr_rows:
        om = engine.run_bridge(p, sample_to_combo(cfg, lob, lrr, rate_rows, mod_rows), "monthly")
        results[(lrr["bu"], lrr["state"])] = (lrr["ep"], om)
    bad = 0
    tot_ep = tot_l27 = tot_l28 = 0.0
    for si, state in enumerate(_states(cfg)):
        rows_ = [(ep, om) for (bu, s), (ep, om) in results.items() if s == state]
        eps = sum(ep for ep, _ in rows_)
        w27 = sum(ep * om.cy_lr_p for ep, om in rows_) / eps
        w28 = sum(ep * om.cy_lr_p1 for ep, om in rows_) / eps
        wcrl = sum(ep * om.crl_ind for ep, om in rows_) / eps
        wecy = sum(ep * om.e_cy[p] for ep, om in rows_) / eps
        r = 8 + si
        for colL, exp in (("B", eps), ("T", wcrl), ("U", wecy), ("AA", w27), ("AF", w28)):
            if not approx(ss[f"{colL}{r}"].value, exp, 1e-6):
                bad += 1
                if bad <= 4:
                    print(f"  FAIL detail: StateSummary {state} {colL}{r} "
                          f"wb={ss[f'{colL}{r}'].value} exp={exp}")
        tot_ep += eps
        tot_l27 += sum(ep * om.cy_lr_p for ep, om in rows_)
        tot_l28 += sum(ep * om.cy_lr_p1 for ep, om in rows_)
    check("State Summary (All): every state row EP-weighted correctly", bad == 0,
          f"{bad} mismatches")
    tot_row = 8 + len(cfg.states) + 3 + 1   # 3 live-roster spare rows (D42)
    check("State Summary (All): total row EP-weighted correctly",
          approx(ss[f"AA{tot_row}"].value, tot_l27 / tot_ep, 1e-9)
          and approx(ss[f"AF{tot_row}"].value, tot_l28 / tot_ep, 1e-9)
          and approx(ss[f"B{tot_row}"].value, tot_ep, 1e-6))
    # D61: the visible bridge chain reproduces the engine on single-combo rows
    bad = 0
    for si, state in enumerate(_states(cfg)):
        r = 8 + si
        rows_ = [(ep, om) for (bu, s), (ep, om) in results.items() if s == state]
        prod = (ss[f"W{r}"].value * ss[f"X{r}"].value * ss[f"Y{r}"].value
                * ss[f"Z{r}"].value)
        exact = (sum(ep * om.cy_lr_p for ep, om in rows_)
                 / sum(ep for ep, _ in rows_))
        # more than one BU per state in the All view -> mix is real and shown
        mix = ss[f"AB{r}"].value
        if len(rows_) > 1:
            if not approx(mix, (exact - prod) * 100.0, 1e-9):
                bad += 1
        elif mix not in (None, ""):
            bad += 1
    check("State Summary: Mix column is the exact residual of the visible product",
          bad == 0, f"{bad} mismatches")
    # Net Delivery (D57): the showcase net combo's closed forms, monthly grid,
    # and reconciliation must tie the oracle at the default 4/1/P date
    if len(cfg.states) > 8:
        from src.sheets_netdelivery import block_top
        net_state = cfg.states[8]
        net_row = next((x for x in lr_rows
                        if x["bu"] == _bu(cfg, 2) and x["state"] == net_state), None)
        if net_row is not None and net_row.get("netp") is not None:
            ncmb = sample_to_combo(cfg, lob, net_row, rate_rows, mod_rows)
            # D63: with no typed override the tab adopts the PLANNED plan-year
            # filing (date and effective %); the suggestion is still computed
            planned = engine.planned_change_in_plan_year(p, ncmb)
            d0 = planned[0] if planned else dt.date(p, 4, 1)
            r_star, comp = engine.suggest_net_rate(p, ncmb, d0)
            r_force = planned[1] if planned else r_star
            m1p = engine.required_m1(p, ncmb, d0, r_force)
            months = engine.net_delivery_by_month(p, ncmb, d0, r_force)
            prog = engine.net_program_plan_lr(p, ncmb, d0, r_force, m1p)
            ncs = wb["_netcalc"]
            si = list(cfg.states).index(net_state)
            t = block_top(si)
            rr = t + 51
            check("[net delivery] suggested r* ties oracle",
                  approx(ncs[f"E{rr}"].value, r_star, 1e-9),
                  f"wb={ncs[f'E{rr}'].value} oracle={r_star}")
            # D63: with both inputs blank the change in force is the planned
            # plan-year filing from the rate log, at its EFFECTIVE percent
            check("[net delivery] default change in force = the planned filing (D63)",
                  planned is not None and approx(ncs[f"F{rr}"].value, r_force, 1e-9),
                  f"wb={ncs[f'F{rr}'].value} oracle={r_force} planned={planned}")
            check("[net delivery] default date = the planned filing's date (D63)",
                  ncs[f"O{t}"].value is not None
                  and getattr(ncs[f"O{t}"].value, "date", lambda: ncs[f"O{t}"].value)() == d0,
                  f"wb={ncs[f'O{t}'].value} oracle={d0}")
            check("[net delivery] implied year-end mod M_1' ties oracle",
                  approx(ncs[f"H{rr}"].value, m1p, 1e-9))
            # D75: the same ask as a dated ACTION — the answer that survives
            # once the Mod Log drives the path and M_1' goes n/a
            ach = wb[SHEETS.NET_DELIVERY]["I5"].value
            ms = engine.required_mod_step(p, ncmb, d0, r_force, achievement=ach)
            for cl, want, nm in (("S", ms.required_step, "required mod step"),
                                 ("T", ms.directed_equivalent, "directed equivalent"),
                                 ("U", ms.m_end, "year-end mod it produces"),
                                 ("V", ms.post_share, "reachable share")):
                check(f"[net delivery] D75 {nm} ties oracle",
                      approx(ncs[f"{cl}{rr}"].value, want, 1e-9),
                      f"wb={ncs[f'{cl}{rr}'].value} oracle={want}")
            # and the tab's own summary row must read the same cells
            from src.sheets_netdelivery import ND_FIRST as _NDF
            nd_row = _NDF + list(cfg.states).index(net_state)
            check("[net delivery] D75 the TAB's summary columns show the action",
                  approx(wb[SHEETS.NET_DELIVERY][f"O{nd_row}"].value,
                         ms.required_step, 1e-9)
                  and approx(wb[SHEETS.NET_DELIVERY][f"P{nd_row}"].value,
                             ms.m_end, 1e-9),
                  f"O={wb[SHEETS.NET_DELIVERY][f'O{nd_row}'].value} "
                  f"P={wb[SHEETS.NET_DELIVERY][f'P{nd_row}'].value}")
            check("[net delivery] booked-program plan LR ties oracle",
                  approx(ncs[f"N{rr}"].value, prog, 1e-9))
            bad_m = sum(
                0 if (approx(ncs[f"AG{t + 27 + j}"].value, mr["rate_leg"], 1e-9)
                      and approx(ncs[f"AH{t + 27 + j}"].value,
                                 mr["price_leg_required"], 1e-9)
                      and approx(ncs[f"AI{t + 27 + j}"].value, mr["m_required"], 1e-9))
                else 1
                for j, mr in enumerate(months))
            check("[net delivery] all 12 grid months tie oracle (3 measures each)",
                  bad_m == 0, f"{bad_m} mismatched months")
            check("[net delivery] solve residual is zero",
                  approx(ncs[f"P{rr}"].value, 0.0, 1e-9))
            # D68: the prior-year rate leg is pure program flow — the tab's new
            # change is date-validated into the plan year, so it can never
            # reach a P-1 month. This is the tie that would catch a wrong
            # cohort offset (the band that broke silently in D64).
            npf = engine.program_flow_by_month(p, ncmb)
            bad_y = sum(0 if approx(ncs[f"AG{t + 15 + j}"].value,
                                    npf.prior_rows[j]["rate_leg"], 1e-9) else 1
                        for j in range(12))
            check("[net delivery] D68 prior-year rate leg ties the program-flow oracle",
                  bad_y == 0, f"{bad_y} mismatched months")
            from src.sheets_netdelivery import (ND_FLAG_COL as NDFC, ND_HDR as NDH,
                                                ND_PRIOR_COL as NDYC)
            nd_ws = wb[SHEETS.NET_DELIVERY]
            nd_nd = len(cfg.states) + 3
            # rate grid: section row g1, header g1+2, first data row g1+3;
            # pricing grid follows at g2 = g1 + 3 + nd + 3
            nd_g1 = next(rw[0].row for rw in
                         nd_ws.iter_rows(min_row=1, max_row=nd_ws.max_row, max_col=1)
                         if isinstance(rw[0].value, str)
                         and rw[0].value.startswith("How the target is delivered"))
            nd_g2 = nd_g1 + 3 + nd_nd + 3
            bad_y = sum(0 if approx(nd_ws.cell(row=nd_g1 + 3 + si,
                                               column=NDYC + j).value,
                                    npf.prior_rows[j]["rate_leg"], 1e-9) else 1
                        for j in range(12))
            check("[net delivery] D68 the TAB's prior-year band ties the oracle",
                  bad_y == 0, f"{bad_y} mismatched months")
            empty = all(nd_ws.cell(row=nd_g2 + 3 + si, column=NDYC + j).value
                        in (None, "") for j in range(12))
            check("[net delivery] D68 the pricing grid has no prior-year block "
                  "(no target in P-1, so nothing to solve)", empty,
                  "prior-year pricing cells must stay empty")
            check("[net delivery] D68 flags column sits right of both year bands",
                  nd_ws.cell(row=NDH, column=NDFC).value == "Flags",
                  repr(nd_ws.cell(row=NDH, column=NDFC).value))
            # the deep-dive band resolves the SAME combo through its own
            # pickers, so it must reproduce that state's block cell for cell
            # (the band had no tie until a resolver typo broke it silently)
            tb = block_top(len(cfg.states) + 3)   # band sits after the roster spares
            band_state = nval(wb, "nd_BandState")
            if band_state == net_state:
                bad_b = sum(
                    0 if approx(ncs[f"AG{tb + 27 + j}"].value,
                                ncs[f"AG{t + 27 + j}"].value, 1e-12)
                    and approx(ncs[f"AH{tb + 27 + j}"].value,
                               ncs[f"AH{t + 27 + j}"].value, 1e-12) else 1
                    for j in range(12))
                check("[net delivery] deep-dive band reproduces the state block",
                      bad_b == 0 and approx(ncs[f"F{tb + 51}"].value,
                                            ncs[f"F{rr}"].value, 1e-12),
                      f"{bad_b} month mismatches; band r={ncs[f'F{tb + 51}'].value} "
                      f"state r={ncs[f'F{rr}'].value}")
                check("[net delivery] band effective date resolves inside the plan year",
                      ncs[f"P{tb}"].value == 1,
                      f"in-window flag={ncs[f'P{tb}'].value} date={ncs[f'O{tb}'].value}")

    # Program Flow (D59): dashboard written legs, the locked block, and the
    # tab's summary averages + worked-example grid rows tie program_flow_by_month
    from src.sheets_programflow import (PF_PLAN_COL, PF_PRIOR_COL, PF_SUM_FIRST,
                                        grid_starts)
    pfr = engine.program_flow_by_month(p, combo)
    fd = wb["Flow Dashboard"]
    bad = 0
    for j, prow in enumerate(pfr.rows):
        r = 93 + j
        ok = (approx(fd[f"M{r}"].value, prow["rate_leg"], 1e-9)
              and approx(fd[f"O{r}"].value, prow["delivered"], 1e-9)
              and approx(fd[f"P{r}"].value, prow["locked_leg"], 1e-9)
              and approx(fd[f"Q{r}"].value, prow["planned_residual"], 1e-9))
        ml = prow["mod_leg"]
        ok = ok and (fd[f"N{r}"].value == "—" if ml is None
                     else approx(fd[f"N{r}"].value, ml, 1e-9))
        # D77: the pricing and delivered legs carry the same locked/planned pair
        for cl, key in (("R", "locked_mod_leg"), ("S", "planned_mod_residual"),
                        ("T", "locked_delivered"), ("U", "planned_delivered_residual")):
            want = prow[key]
            ok = ok and (fd[f"{cl}{r}"].value == "—" if want is None
                         else approx(fd[f"{cl}{r}"].value, want, 1e-9))
        if not ok:
            bad += 1
    check("[program flow] dashboard cols M..U tie oracle for all 24 YoY months",
          bad == 0, f"{bad} mismatched months")
    check("[program flow] delta-vs-net column blank for the non-net WE combo",
          fd["V93"].value in (None, ""), f"V93={fd['V93'].value!r}")
    check("[program flow] D77 delivered = locked x planned residual",
          approx(nval(wb, "fd_splitchk"), 0.0, 1e-9),
          f"fd_splitchk={nval(wb, 'fd_splitchk')}")
    check("[program flow] D77 no planned mod actions => locked pricing IS program",
          approx(nval(wb, "fd_lockedmodchk"), 0.0, 1e-9),
          f"fd_lockedmodchk={nval(wb, 'fd_lockedmodchk')}")
    check("[program flow] fd_avgdel ties the oracle plan-year average",
          approx(nval(wb, "fd_avgdel"), pfr.avg_delivered_ratio - 1.0, 1e-9),
          f"wb={nval(wb, 'fd_avgdel')} oracle={pfr.avg_delivered_ratio - 1.0}")
    # locked block: 48 cohorts tie a taken-only engine (range from the name)
    tk_ref = next(iter(wb.defined_names["fd_tkw"].destinations))[1]
    tk_first = int(tk_ref.split(":")[0].rsplit("$", 1)[1])
    tk_eng = engine.MonthlyEngine(p, combo, changes_override=tuple(
        rc for rc in combo.rate_changes if rc.status == "taken"))
    cs = wb["_calc"]
    bad = sum(0 if approx(cs[f"N{tk_first + i}"].value,
                          tk_eng.written_index(tk_eng.cohorts[i]), 1e-9) else 1
              for i in range(48))
    check("[program flow] locked block ties the taken-only engine (48 cohorts)",
          bad == 0, f"{bad} mismatched cohorts")
    pfs = wb[SHEETS.PROGRAM_FLOW]
    # summary metrics live in two column bands (D68): the prior year at
    # PF_PRIOR_COL (rate/mod/delivered) and the plan year at PF_PLAN_COL
    # (EP, net target, rate, mod, delivered, delta, gap)
    SUM_P = {k: PF_PLAN_COL + o for k, o in
             (("ep", 0), ("net", 1), ("rate", 2), ("mod", 3), ("del", 4),
              ("dlt", 5), ("gap", 6))}
    SUM_Y = {k: PF_PRIOR_COL + o for k, o in (("rate", 0), ("mod", 1), ("del", 2))}
    bad = bad_y = 0
    for i, stname in enumerate(_states(cfg)):
        rowdef = next(x for x in lr_rows if x["bu"] == _bu(cfg, 0) and x["state"] == stname)
        cpf = engine.program_flow_by_month(p, sample_to_combo(cfg, lob, rowdef, rate_rows, mod_rows))
        r = PF_SUM_FIRST + i

        def _cell(cmap, key, row_=r):
            return pfs.cell(row=row_, column=cmap[key]).value

        def _ok(cmap, key, ratio, row_=r):
            v = _cell(cmap, key, row_)
            return (v == "—") if ratio is None else approx(v, ratio - 1.0, 1e-9)

        if not (_ok(SUM_P, "rate", cpf.avg_rate_ratio)
                and _ok(SUM_P, "del", cpf.avg_delivered_ratio)
                and _ok(SUM_P, "mod", cpf.avg_mod_ratio)):
            bad += 1
        if not (_ok(SUM_Y, "rate", cpf.avg_rate_ratio_prior)
                and _ok(SUM_Y, "del", cpf.avg_delivered_ratio_prior)
                and _ok(SUM_Y, "mod", cpf.avg_mod_ratio_prior)):
            bad_y += 1
    check("[program flow] summary averages tie oracle for every live state",
          bad == 0, f"{bad} mismatched states")
    check("[program flow] D68 prior-year summary averages tie oracle for every state",
          bad_y == 0, f"{bad_y} mismatched states")
    g1, g2, g3 = grid_starts(cfg)
    bad = bad_y = 0
    for j in range(12):
        for c0, rows_, tally in ((PF_PLAN_COL, pfr.rows, "p"),
                                 (PF_PRIOR_COL, pfr.prior_rows, "y")):
            prow = rows_[j]
            ok = (approx(pfs.cell(row=g1 + 3, column=c0 + j).value,
                         prow["rate_leg"], 1e-9)
                  and approx(pfs.cell(row=g3 + 3, column=c0 + j).value,
                             prow["delivered"], 1e-9))
            ml = prow["mod_leg"]
            v2 = pfs.cell(row=g2 + 3, column=c0 + j).value
            ok = ok and (v2 == "—" if ml is None else approx(v2, ml, 1e-9))
            if not ok:
                if tally == "p":
                    bad += 1
                else:
                    bad_y += 1
    check("[program flow] worked-example grid rows tie oracle (3 grids x 12 months)",
          bad == 0, f"{bad} mismatched months")
    check("[program flow] D68 prior-year grid rows tie oracle (3 grids x 12 months)",
          bad_y == 0, f"{bad_y} mismatched months")
    # the two bands must not be the same numbers: P-1 is measured against P-2,
    # so a real program moves between them (a copy-paste of the plan-year
    # offsets would silently produce identical blocks)
    same = sum(1 for j in range(12)
               if approx(pfs.cell(row=g3 + 3, column=PF_PRIOR_COL + j).value,
                         pfs.cell(row=g3 + 3, column=PF_PLAN_COL + j).value, 1e-12))
    check("[program flow] the two year bands are genuinely different series",
          same < 12, f"{same}/12 months identical — prior band may be reading plan rows")

    # published flow columns (D60): the flat projection ties per-combo oracle
    from src.sheets_calc import FLOW_PUB, PRIOR_PUB
    from src.xlstyle import col as xcol
    bad = bad_w = bad_a = 0
    bad_y = bad_ya = bad_epw = 0
    for i, rowdef in enumerate(lr_rows):
        cmb2 = sample_to_combo(cfg, lob, rowdef, rate_rows, mod_rows)
        cpf = engine.program_flow_by_month(p, cmb2)
        eng2 = engine.MonthlyEngine(p, cmb2)
        rr2 = L.CALC_RES_FIRST + i
        for j in range(12):
            if not approx(cs[f"{xcol(FLOW_PUB['delivered'] + j)}{rr2}"].value,
                          cpf.rows[j]["delivered"], 1e-9):
                bad += 1
            epw_o = rowdef["ep"] * eng2.w(engine.month_index(p, 1) + j)
            wv = cs[f"{xcol(FLOW_PUB['epw'] + j)}{rr2}"].value
            if wv is None or epw_o == 0 or not approx(wv / epw_o, 1.0, 1e-9):
                bad_w += 1
            # D68: the prior-year legs, and the assumption that lets ONE epw
            # family serve both years — w is a function of calendar month, so
            # w(m in P-1) must equal the published w(m in P)
            pr = cpf.prior_rows[j]
            for key, val in (("delivered", pr["delivered"]), ("rate", pr["rate_leg"]),
                             ("mod", pr["mod_leg"])):
                if val is None:
                    continue
                if not approx(cs[f"{xcol(PRIOR_PUB[key] + j)}{rr2}"].value, val, 1e-9):
                    bad_y += 1
            epw_y = rowdef["ep"] * eng2.w(engine.month_index(p - 1, 1) + j)
            if wv is None or epw_y == 0 or not approx(wv / epw_y, 1.0, 1e-12):
                bad_epw += 1
        if not approx(cs[f"{xcol(FLOW_PUB['avg_del'])}{rr2}"].value,
                      cpf.avg_delivered_ratio, 1e-9):
            bad_a += 1
        for key, val in (("avg_rate", cpf.avg_rate_ratio_prior),
                         ("avg_mod", cpf.avg_mod_ratio_prior),
                         ("avg_del", cpf.avg_delivered_ratio_prior)):
            if val is None:
                continue
            if not approx(cs[f"{xcol(PRIOR_PUB[key])}{rr2}"].value, val, 1e-9):
                bad_ya += 1
    check("[program flow] published delivered legs tie oracle (all combos x 12)",
          bad == 0, f"{bad} mismatched cells")
    check("[program flow] published EP x w products tie oracle", bad_w == 0,
          f"{bad_w} mismatched cells")
    check("[program flow] published avg-ratio echoes tie the block results",
          bad_a == 0, f"{bad_a} mismatched combos")
    check("[program flow] D68 published prior-year legs tie oracle "
          "(all combos x 12 x 3)", bad_y == 0, f"{bad_y} mismatched cells")
    check("[program flow] D68 published prior-year avg echoes tie the block results",
          bad_ya == 0, f"{bad_ya} mismatched combos")
    check("[program flow] D68 the single epw family weights BOTH years "
          "(seasonality is a function of calendar month)", bad_epw == 0,
          f"{bad_epw} months where w(P-1) != published w(P)")

    # program-basis plan LR (D65): the explicit rate x mod counterfactual
    from src.sheets_calc import PROG_LR
    bad = bad_id = 0
    for i, rowdef in enumerate(lr_rows):
        cmb3 = sample_to_combo(cfg, lob, rowdef, rate_rows, mod_rows)
        lp, lp1 = engine.program_basis_plan_lr(p, cmb3)
        rr3 = L.CALC_RES_FIRST + i
        if not (approx(cs[f"{xcol(PROG_LR['cylr'])}{rr3}"].value, lp, 1e-9)
                and approx(cs[f"{xcol(PROG_LR['cylr1'])}{rr3}"].value, lp1, 1e-9)):
            bad += 1
        # the invariant: with no net selection there is only ONE path
        if not cmb3.net_mode:
            om3 = engine.run_bridge(p, cmb3, "monthly")
            if not (approx(lp, om3.cy_lr_p, 1e-12)
                    and approx(cs[f"{xcol(PROG_LR['ident'])}{rr3}"].value, 0.0, 1e-12)):
                bad_id += 1
    check("[program basis] plan LR ties the explicit-path oracle (all combos, both years)",
          bad == 0, f"{bad} mismatched combos")
    check("[program basis] identical to the headline on every non-net combo",
          bad_id == 0, f"{bad_id} mismatches")
    # the surfaces that display it
    net_combo = next((x for x in lr_rows if x.get("netp") is not None), None)
    if net_combo is not None:
        ni = lr_rows.index(net_combo)
        ncmb3 = sample_to_combo(cfg, lob, net_combo, rate_rows, mod_rows)
        lp, _ = engine.program_basis_plan_lr(p, ncmb3)
        gap = (lp - engine.run_bridge(p, ncmb3, "monthly").cy_lr_p) * 100.0
        check("[program basis] Portfolio row shows the combo's program LR and gap",
              approx(wb["Portfolio"][f"P{L.PF_FIRST + ni}"].value, lp, 1e-9)
              and approx(wb["Portfolio"][f"Q{L.PF_FIRST + ni}"].value, gap, 1e-9),
              f"wb={wb['Portfolio'][f'P{L.PF_FIRST + ni}'].value} oracle={lp}")
        # State Summary 'All' view: EP-weighted across the state's BUs
        sr = 8 + list(_states(cfg)).index(net_combo["state"])
        rows_s = [(x["ep"], sample_to_combo(cfg, lob, x, rate_rows, mod_rows))
                  for x in lr_rows if x["state"] == net_combo["state"]]
        exp_prog = (sum(ep * engine.program_basis_plan_lr(p, c)[0] for ep, c in rows_s)
                    / sum(ep for ep, _ in rows_s))
        check("[program basis] State Summary shows the EP-weighted program LR",
              approx(ss[f"AH{sr}"].value, exp_prog, 1e-9),
              f"wb={ss[f'AH{sr}'].value} oracle={exp_prog}")

    # live dimension lists (D42): _lists uniques match the seeded roster
    ls = wb["_lists"]
    got_bus = [ls[f"A{3 + i}"].value for i in range(len(cfg.business_units))]
    got_states = [ls[f"B{3 + i}"].value for i in range(len(cfg.states))]
    check("_lists live BU uniques match tbl_LR", got_bus == list(cfg.business_units),
          str(got_bus))
    check("_lists live state uniques match tbl_LR", got_states == list(cfg.states),
          str(got_states[:5]))
    return wb


# ---------------------------------------------------------------------------
# Phase D: toggle exercises
# ---------------------------------------------------------------------------


def _mutate(src: Path, scratch: Path, mutations: dict[tuple[str, str], object]) -> Path:
    shutil.copy2(src, scratch)
    wb = openpyxl.load_workbook(scratch, data_only=False)
    for (sheet, addr), value in mutations.items():
        wb[sheet][addr] = value
    wb.save(scratch)
    recalc(scratch)
    return scratch


def _read(scratch: Path):
    return openpyxl.load_workbook(scratch, data_only=True)


def phase_d(path: Path, cfg, lob, scratch_dir: Path):
    print("Phase D: toggle exercises (mutate -> recalc -> tie to oracle)")
    p = cfg.plan_year
    st = cfg.states
    lr_rows = sample_lr_rows(cfg, lob)
    rate_rows = sample_rate_rows(cfg, lob)
    mod_rows = sample_mod_rows(cfg, lob)   # D70: the oracle must see the log
    season_rows = sample_seasonality_rows(cfg)
    we = next(r for r in lr_rows if r["bu"] == _bu(cfg, 0) and r["state"] == st[0])
    base_combo = sample_to_combo(cfg, lob, we, rate_rows, mod_rows)
    base_m = engine.run_bridge(p, base_combo, "monthly")
    scratch = scratch_dir / "exercise.xlsx"
    we_row_xl = L.LR_FIRST + lr_rows.index(we)

    def combo_of(bu, state, **kw):
        row = next(r for r in lr_rows if r["bu"] == bu and r["state"] == state)
        return sample_to_combo(cfg, lob, row, rate_rows, mod_rows, **kw)

    def run(label_, mutations, assertions):
        _mutate(path, scratch, mutations)
        wb = _read(scratch)
        errs = scan_errors(wb)
        check(f"[{label_}] zero formula errors", not errs, "; ".join(errs[:5]))
        assertions(wb)

    # 1. master mod toggle OFF
    off_m = engine.run_bridge(p, replace(base_combo, mod_adjustment_enabled=False), "monthly")

    def a1(wb):
        check("[mod master OFF] A_mod = 1", approx(nval(wb, "nr_Amod_P"), 1.0, 1e-12))
        check("[mod master OFF] CY LR ties oracle", approx(nval(wb, "nr_CYLR_P"),
              off_m.cy_lr_p, 1e-9))
        check("[mod master OFF] CY LR moved UP (A_mod was < 1)",
              float(nval(wb, "nr_CYLR_P")) > base_m.cy_lr_p)
    run("mod master OFF", {("Control", "C13"): "OFF"}, a1)

    # 2. per-combo mod toggle OFF
    def a2(wb):
        check("[combo mod OFF] CY LR ties oracle", approx(nval(wb, "nr_CYLR_P"),
              off_m.cy_lr_p, 1e-9))
    run("combo mod OFF", {("Inputs", f"{lr_letter('lr_modadj')}{we_row_xl}"): "OFF"}, a2)

    # 3. basis proposed on the worked-example combo
    prop_m = engine.run_bridge(p, replace(base_combo, lr_basis="proposed", sel_change=0.05),
                               "monthly")

    def a3(wb):
        check("[basis proposed] LR_current = LR x (1+s)",
              approx(nval(wb, "nr_LRcur"), 0.65 * 1.05, 1e-12))
        check("[basis proposed] CY LR ties oracle", approx(nval(wb, "nr_CYLR_P"),
              prop_m.cy_lr_p, 1e-9))
    run("basis proposed", {("Inputs", f"{lr_letter('lr_basis')}{we_row_xl}"): "proposed"}, a3)

    # 4. seasonality ON for a state with a seeded profile
    se_state = season_rows[0]["state"] if season_rows else None
    if se_state:
        se_on = combo_of(_bu(cfg, 0), se_state, seasonality_on=True, season_rows=season_rows)
        se_off = combo_of(_bu(cfg, 0), se_state)
        se_on_m = engine.run_bridge(p, se_on, "monthly")
        se_off_m = engine.run_bridge(p, se_off, "monthly")

        def a4(wb):
            check(f"[seasonality ON {se_state}] CY LR ties oracle (seasonal weights)",
                  approx(nval(wb, "nr_CYLR_P"), se_on_m.cy_lr_p, 1e-9),
                  f"wb={nval(wb, 'nr_CYLR_P')} oracle={se_on_m.cy_lr_p}")
            check(f"[seasonality ON {se_state}] seasonality changes some CY index",
                  any(not approx(se_on_m.e_cy[y], se_off_m.e_cy[y], 1e-12)
                      for y in (p - 1, p, p + 1)),
                  f"on={se_on_m.e_cy} off={se_off_m.e_cy}")
        run(f"seasonality ON {se_state}",
            {("Control", "C8"): se_state, ("Control", "C12"): "ON"}, a4)

        def a4b(wb):
            check(f"[{se_state} seasonality OFF] CY LR ties oracle (uniform)",
                  approx(nval(wb, "nr_CYLR_P"), se_off_m.cy_lr_p, 1e-9))
        run(f"{se_state} seasonality OFF", {("Control", "C8"): se_state}, a4b)

    # 5. selector subsample: every BU x first 3 states + every showcase combo
    picks = {(bu, s) for bu in cfg.business_units for s in st[:3]}
    for special in [(_bu(cfg, 1), st[1] if len(st) > 1 else st[0]),
                    (_bu(cfg, 1), st[2] if len(st) > 2 else st[0]),
                    (_bu(cfg, 0), st[3] if len(st) > 3 else st[0]),
                    (_bu(cfg, 0), st[5] if len(st) > 5 else st[0]),
                    (_bu(cfg, 2), st[6] if len(st) > 6 else st[0]),
                    (_bu(cfg, 2), st[8] if len(st) > 8 else st[0]),  # net-selection showcase
                    degenerate_combo(cfg)]:
        picks.add(special)
    for bu, state in sorted(picks):
        om = engine.run_bridge(p, combo_of(bu, state), "monthly")

        def a5(wb, om=om, bu=bu, state=state):
            check(f"[selector {bu}|{state}] Bridge ties oracle",
                  approx(nval(wb, "nr_CYLR_P"), om.cy_lr_p, 1e-9),
                  f"wb={nval(wb, 'nr_CYLR_P')} oracle={om.cy_lr_p}")
        run(f"selector {bu}|{state}",
            {("Control", "C7"): bu, ("Control", "C8"): state}, a5)

    # 6. degenerate combo (seeded with no rate rows): factors = 1 exactly
    deg_bu, deg_state = degenerate_combo(cfg)
    deg_m = engine.run_bridge(p, combo_of(deg_bu, deg_state), "monthly")

    def a6(wb):
        check("[degenerate combo] A_rate = 1 exactly", approx(nval(wb, "nr_Arate_P"), 1.0, 1e-12))
        check("[degenerate combo] CY LR ties oracle", approx(nval(wb, "nr_CYLR_P"),
              deg_m.cy_lr_p, 1e-9))
    run("degenerate combo", {("Control", "C7"): deg_bu, ("Control", "C8"): deg_state}, a6)

    # 6b. blank out an existing combo's rate log -> neutral factors
    tgt = next((r for r in lr_rows
                if any(rr["bu"] == r["bu"] and rr["state"] == r["state"] for rr in rate_rows)
                and (r["bu"], r["state"]) != (_bu(cfg, 0), st[0])), None)
    if tgt:
        rows_xl = [L.RL_FIRST + i for i, rr in enumerate(rate_rows)
                   if rr["bu"] == tgt["bu"] and rr["state"] == tgt["state"]]
        muts = {("Control", "C7"): tgt["bu"], ("Control", "C8"): tgt["state"]}
        for r in rows_xl:
            for colL in "ABCDEFGH":
                muts[(SHEETS.RATE_LOG, f"{colL}{r}")] = None
        blank_m = engine.run_bridge(
            p, replace(combo_of(tgt["bu"], tgt["state"]), rate_changes=()), "monthly")

        def a6b(wb):
            check("[blanked rate log] A_rate = 1 exactly",
                  approx(nval(wb, "nr_Arate_P"), 1.0, 1e-12))
            check("[blanked rate log] CY LR ties oracle",
                  approx(nval(wb, "nr_CYLR_P"), blank_m.cy_lr_p, 1e-9))
        run("blanked rate log", muts, a6b)

    # 7. scenario levers, one at a time
    def transformed(dpts=0.0, shift=0, ach=None, dm1=0.0):
        changes = tuple(
            replace(rc,
                    filed_pct=rc.filed_pct + dpts,
                    effective=engine.add_months(rc.effective, shift),
                    achievement=(ach if ach is not None else rc.achievement))
            if rc.status == "planned" else rc
            for rc in base_combo.rate_changes)
        mods = base_combo.mods
        m2 = mods.m2 if mods.m2 is not None else mods.m1
        mods = replace(mods, m1=mods.m1 + dm1, m2=m2 + dm1)
        return replace(base_combo, rate_changes=changes, mods=mods)

    sc_expected = [
        engine.run_bridge(p, transformed(dpts=0.02), "monthly"),
        engine.run_bridge(p, transformed(shift=2), "monthly"),
        engine.run_bridge(p, transformed(ach=0.5), "monthly"),
        engine.run_bridge(p, transformed(dm1=0.02), "monthly"),
    ]
    muts = {("Scenarios", "C6"): 0.02, ("Scenarios", "D7"): 2,
            ("Scenarios", "E8"): 0.5, ("Scenarios", "F9"): 0.02}

    def a7(wb):
        sc = wb["Scenarios"]
        for s, om in enumerate(sc_expected, start=1):
            got = sc[f"{'CDEFG'[s]}14"].value
            check(f"[scenario S{s}] CY LR ties oracle", approx(got, om.cy_lr_p, 1e-9),
                  f"wb={got} oracle={om.cy_lr_p}")
        check("[scenario S1] rate increase lowers CY LR",
              sc_expected[0].cy_lr_p < base_m.cy_lr_p)
        check("[scenario S2] later timing raises CY LR",
              sc_expected[1].cy_lr_p > base_m.cy_lr_p)
        check("[scenario S3] under-achievement raises CY LR",
              sc_expected[2].cy_lr_p > base_m.cy_lr_p)
        check("[scenario S4] rising mods lower CY LR",
              sc_expected[3].cy_lr_p < base_m.cy_lr_p)
    run("scenario levers", muts, a7)

    # 8. solver custom target / mid-month date / achievement
    s_custom = engine.solve_rate_for_target(p, base_combo, 0.61, dt.date(p, 5, 15),
                                            achievement=0.8)

    def a8(wb):
        check("[solver custom] required r ties oracle",
              approx(nval(wb, "nr_SolverR"), s_custom.required_change, 1e-9),
              f"wb={nval(wb, 'nr_SolverR')} oracle={s_custom.required_change}")
        check("[solver custom] filed equivalent = r / 80%",
              approx(wb["Solver"]["C18"].value, s_custom.required_change / 0.8, 1e-9))
    run("solver custom", {("Solver", "C7"): 0.61, ("Solver", "C8"): dt.date(p, 5, 15),
                          ("Solver", "C9"): 0.8}, a8)

    # 8b. mixed-status cohort month (D58): a PLANNED row earlier in the same
    # month as a TAKEN row must not steal the solver base's first-change split
    mix = (engine.RateChange(dt.date(p - 1, 6, 5), 0.03, "planned", False),
           engine.RateChange(dt.date(p - 1, 6, 20), 0.02, "taken", False))
    mix_combo = replace(base_combo, rate_changes=base_combo.rate_changes + mix)
    mix_r = engine.solve_rate_for_target(
        p, mix_combo, base_m.cy_lr_p, dt.date(p, 4, 1)).required_change
    mix_r0 = L.RL_FIRST + len(rate_rows)
    muts_mix = {}
    for i, rc in enumerate(mix):
        rr_ = mix_r0 + i
        muts_mix.update({("Rate Log", f"A{rr_}"): we["bu"],
                         ("Rate Log", f"B{rr_}"): we["state"],
                         ("Rate Log", f"C{rr_}"): rc.effective,
                         ("Rate Log", f"D{rr_}"): rc.filed_pct,
                         ("Rate Log", f"E{rr_}"): rc.status,
                         ("Rate Log", f"F{rr_}"): "N"})

    def a8b(wb):
        check("[mixed-status month] solver required r ties oracle (D58)",
              approx(nval(wb, "nr_SolverR"), mix_r, 1e-9),
              f"wb={nval(wb, 'nr_SolverR')} oracle={mix_r}")
    run("mixed-status cohort month (solver base)", muts_mix, a8b)

    # 9. attribution actuals
    actuals = [engine.PlannedActual(dt.date(p, 4, 1), 0.04, dt.date(p, 6, 1))]
    mods_act = replace(base_combo.mods, m1=0.90)
    att = engine.attribution(p, base_combo, actuals, mods_act, actual_cy_lr=0.655)

    def a9(wb):
        aw = wb["Attribution"]
        for row, (lbl, f) in zip((26, 27, 28, 29), att.factors):
            check(f"[attribution] factor '{lbl}' ties oracle",
                  approx(aw[f"C{row}"].value, f, 1e-9),
                  f"wb={aw[f'C{row}'].value} oracle={f}")
        check("[attribution] final LR = actual entered",
              approx(aw["D30"].value, 0.655, 1e-12))
        ck = wb["Checks"]
        check("[attribution] reconciliation check row is PASS",
              ck["G16"].value == "PASS", str(ck["G16"].value))
    run("attribution actuals",
        {("Attribution", "E14"): 0.04, ("Attribution", "F14"): dt.date(p, 6, 1),
         ("Attribution", "C8"): 0.90, ("Attribution", "C10"): 0.655}, a9)

    # 9b. State Summary filtered to one BU: rows equal that BU's combos exactly,
    # and the rate-change slots replay the chronological log
    fbu = _bu(cfg, 1)

    def a9b(wb):
        ss = wb["State Summary"]
        bad = 0
        for si, state in enumerate(_states(cfg)):
            r = 8 + si
            om = engine.run_bridge(p, combo_of(fbu, state), "monthly")
            row_lr = next(x for x in lr_rows if x["bu"] == fbu and x["state"] == state)
            for colL, exp in (("B", row_lr["ep"]), ("AA", om.cy_lr_p),
                              ("AF", om.cy_lr_p1), ("T", om.crl_ind),
                              ("U", om.e_cy[p])):
                if not approx(ss[f"{colL}{r}"].value, exp, 1e-6):
                    bad += 1
            log = sorted([rr for rr in rate_rows
                          if rr["bu"] == fbu and rr["state"] == state],
                         key=lambda rr: rr["eff"])
            n_tk = sum(1 for rr in log if rr["status"] == "taken")
            if (ss[f"R{r}"].value or 0) != n_tk:
                bad += 1
            if (ss[f"S{r}"].value or 0) != len(log) - n_tk:
                bad += 1
            for j, rr in enumerate(log[:4]):
                d = ss.cell(row=r, column=6 + j * 3).value
                v = ss.cell(row=r, column=7 + j * 3).value
                tok = ss.cell(row=r, column=8 + j * 3).value
                eff_pct = rr["filed"] * ((1.0 if rr["achievement"] is None
                                          else rr["achievement"])
                                         if rr["status"] == "planned" else 1.0)
                if not (hasattr(d, "date") and d.date() == rr["eff"]) and d != rr["eff"]:
                    bad += 1
                if not approx(v, eff_pct, 1e-12):
                    bad += 1
                # D61 status token: T/P plus * when outside the indication
                exp_tok = ("P" if rr["status"] == "planned" else "T") + (
                    "" if rr["considered"] in ("Y", True) else "*")
                if tok != exp_tok:
                    bad += 1
            # single-BU rows are single-combo: the plan LR must BE the product
            prod = (ss[f"W{r}"].value * ss[f"X{r}"].value * ss[f"Y{r}"].value
                    * ss[f"Z{r}"].value)
            if not approx(prod, om.cy_lr_p, 1e-9):
                bad += 1
        check(f"[State Summary {fbu}] rows, totals, rate slots, and status tokens tie",
              bad == 0, f"{bad} mismatches")
    run(f"State Summary filter {fbu}", {("State Summary", "B4"): fbu}, a9b)

    # 9c. trend default inheritance: blank combos inherit 2%, explicit entries win
    trend_m = engine.run_bridge(
        p, sample_to_combo(cfg, lob, we, rate_rows, mod_rows, trend_default=0.02), "monthly")
    ovr_state = st[1] if len(st) > 1 else st[0]
    ovr_row = next(r for r in lr_rows if r["bu"] == _bu(cfg, 1) and r["state"] == ovr_state)
    ovr_m = engine.run_bridge(
        p, sample_to_combo(cfg, lob, ovr_row, rate_rows, mod_rows, trend_default=0.02), "monthly")
    ovr_calc_row = L.CALC_RES_FIRST + lr_rows.index(ovr_row)

    def a9c(wb):
        check("[trend default 2%] blank-trend combo inherits it for CY P+1",
              approx(nval(wb, "nr_CYLR_P1"), trend_m.cy_lr_p1, 1e-9),
              f"wb={nval(wb, 'nr_CYLR_P1')} oracle={trend_m.cy_lr_p1}")
        check("[trend default 2%] CY P untouched by trend",
              approx(nval(wb, "nr_CYLR_P"), base_m.cy_lr_p, 1e-9))
        got = wb["_calc"][f"O{ovr_calc_row}"].value
        check("[trend default 2%] explicit per-state trend still wins",
              approx(got, ovr_m.cy_lr_p1, 1e-9), f"wb={got} oracle={ovr_m.cy_lr_p1}")
    run("trend default 2%", {("Control", "C15"): 0.02}, a9c)

    # 9d. net rate selection on the worked-example combo (D39)
    net_m = engine.run_bridge(p, replace(base_combo, net_sel_p=0.08), "monthly")

    def a9d(wb):
        check("[net 8%] CY LR ties oracle net path",
              approx(nval(wb, "nr_CYLR_P"), net_m.cy_lr_p, 1e-9),
              f"wb={nval(wb, 'nr_CYLR_P')} oracle={net_m.cy_lr_p}")
        check("[net 8%] CY LR(P+1) ties oracle (selection carried)",
              approx(nval(wb, "nr_CYLR_P1"), net_m.cy_lr_p1, 1e-9))
        check("[net 8%] A_mod collapses to 1", approx(nval(wb, "nr_Amod_P"), 1.0, 1e-12))
        check("[net 8%] net mode flag TRUE", nval(wb, "nr_NetMode") is True)
    run("net selection 8%", {("Inputs", f"{lr_letter('lr_netp')}{we_row_xl}"): 0.08}, a9d)

    # 9e. scenario D-net lever on top of a net selection
    net_s1 = engine.run_bridge(p, replace(base_combo, net_sel_p=0.10), "monthly")

    def a9e(wb):
        got = wb["Scenarios"]["D14"].value
        check("[net + D-net 2%] scenario S1 ties oracle at 10% net",
              approx(got, net_s1.cy_lr_p, 1e-9), f"wb={got} oracle={net_s1.cy_lr_p}")
        check("[net + D-net 2%] base column still ties 8% net",
              approx(wb["Scenarios"]["C14"].value, net_m.cy_lr_p, 1e-9))
    run("net + scenario D-net", {("Inputs", f"{lr_letter('lr_netp')}{we_row_xl}"): 0.08,
                                 ("Scenarios", "C10"): 0.02}, a9e)

    # 9f. in-book rename of a BU and a state (D42): every dropdown list, the
    # selectors, and the State Summary must follow without regeneration
    ren_state = st[1] if len(st) > 1 else st[0]
    ren_row = next(x for x in lr_rows if x["bu"] == _bu(cfg, 1) and x["state"] == ren_state)
    ren_lr_xl = L.LR_FIRST + lr_rows.index(ren_row)
    muts = {("Inputs", f"{lr_letter('lr_bu')}{ren_lr_xl}"): "BU-X",
            ("Inputs", f"{lr_letter('lr_state')}{ren_lr_xl}"): "ZZ",
            ("Control", "C7"): "BU-X", ("Control", "C8"): "ZZ"}
    for i, rr in enumerate(rate_rows):
        if rr["bu"] == _bu(cfg, 1) and rr["state"] == ren_state:
            muts[(SHEETS.RATE_LOG, f"A{L.RL_FIRST + i}")] = "BU-X"
            muts[(SHEETS.RATE_LOG, f"B{L.RL_FIRST + i}")] = "ZZ"
    ren_m = engine.run_bridge(p, combo_of(_bu(cfg, 1), ren_state), "monthly")  # values unchanged

    def a9f(wb):
        ls = wb["_lists"]
        bus = {ls[f"A{3 + i}"].value for i in range(10)}
        sts = {ls[f"B{3 + i}"].value for i in range(len(st) + 3)}
        check("[rename] new BU appears in the live BU list", "BU-X" in bus, str(sorted(
            b for b in bus if b))[:80])
        check("[rename] new state appears in the live state list", "ZZ" in sts)
        check("[rename] Bridge ties oracle for the renamed combo",
              approx(nval(wb, "nr_CYLR_P"), ren_m.cy_lr_p, 1e-9),
              f"wb={nval(wb, 'nr_CYLR_P')} oracle={ren_m.cy_lr_p}")
        ss = wb["State Summary"]
        zz_row = next((r for r in range(8, 8 + len(st) + 3)
                       if ss[f"A{r}"].value == "ZZ"), None)
        check("[rename] State Summary grew a live ZZ row", zz_row is not None)
        if zz_row:
            check("[rename] ZZ row EP and CY LR tie the renamed combo",
                  approx(ss[f"B{zz_row}"].value, ren_row["ep"], 1e-6)
                  and approx(ss[f"AA{zz_row}"].value, ren_m.cy_lr_p, 1e-9),
                  f"EP={ss[f'B{zz_row}'].value} LR={ss[f'AA{zz_row}'].value}")
    run("in-book rename BU/state", muts, a9f)

    # 9g. Net Delivery exercises (D57)
    if len(st) > 8:
        from src.sheets_netdelivery import ND_FIRST, block_top
        net_state = st[8]
        net_row = next((x for x in lr_rows
                        if x["bu"] == _bu(cfg, 2) and x["state"] == net_state), None)
        if net_row is not None and net_row.get("netp") is not None:
            ncmb = sample_to_combo(cfg, lob, net_row, rate_rows, mod_rows)
            si8 = list(st).index(net_state)
            t8 = block_top(si8)
            rr8 = t8 + 51
            row_si8 = ND_FIRST + si8
            d0 = dt.date(p, 4, 1)

            # (i) net selection switched ON for the worked example combo
            we_net = replace(base_combo, net_sel_p=0.08)
            r_we, _ = engine.suggest_net_rate(p, we_net, d0)
            t0 = block_top(0)

            def a9g1(wb):
                got = wb["_netcalc"][f"E{t0 + 51}"].value
                check("[net delivery: WE net 8%] suggested r* ties oracle",
                      approx(got, r_we, 1e-9), f"wb={got} oracle={r_we}")
            run("net delivery: worked example net-on",
                {("Inputs", f"{lr_letter('lr_netp')}{we_row_xl}"): 0.08, ("Net Delivery", "B5"): _bu(cfg, 0)},
                a9g1)

            # (ii) a mid-month state date exercises the day-blend p
            d_mid = dt.date(p, 5, 15)
            r_mid, _ = engine.suggest_net_rate(p, ncmb, d_mid)

            def a9g2(wb):
                got = wb["_netcalc"][f"E{rr8}"].value
                check("[net delivery: 5/15 date] r* ties oracle (day-blend)",
                      approx(got, r_mid, 1e-9), f"wb={got} oracle={r_mid}")
            run("net delivery: mid-month date",
                {("Net Delivery", f"D{row_si8}"): d_mid}, a9g2)

            # (iii) D's month co-occupied by a NEW taken row (the affine-form
            # regression case D31/D57) + an r override driving the dual solve
            inj = engine.RateChange(dt.date(p, 6, 10), 0.02, "taken", False)
            ncmb3 = replace(ncmb, rate_changes=ncmb.rate_changes + (inj,))
            d_co = dt.date(p, 6, 20)
            r_co, _ = engine.suggest_net_rate(p, ncmb3, d_co)
            m1_ovr = engine.required_m1(p, ncmb3, d_co, 0.03)
            empty_rl = L.RL_FIRST + len(rate_rows)
            muts3 = {("Net Delivery", f"D{row_si8}"): d_co,
                     ("Net Delivery", f"E{row_si8}"): 0.03,
                     ("Rate Log", f"A{empty_rl}"): _bu(cfg, 2),
                     ("Rate Log", f"B{empty_rl}"): net_state,
                     ("Rate Log", f"C{empty_rl}"): inj.effective,
                     ("Rate Log", f"D{empty_rl}"): 0.02,
                     ("Rate Log", f"E{empty_rl}"): "taken",
                     ("Rate Log", f"F{empty_rl}"): "N"}

            def a9g3(wb):
                ncs = wb["_netcalc"]
                check("[net delivery: co-occupied month] r* ties oracle (affine A/B)",
                      approx(ncs[f"E{rr8}"].value, r_co, 1e-9),
                      f"wb={ncs[f'E{rr8}'].value} oracle={r_co}")
                check("[net delivery: r override] M_1' ties oracle dual solve",
                      approx(ncs[f"H{rr8}"].value, m1_ovr, 1e-9),
                      f"wb={ncs[f'H{rr8}'].value} oracle={m1_ovr}")
                check("[net delivery] override is the change in force",
                      approx(ncs[f"F{rr8}"].value, 0.03, 1e-12))
            run("net delivery: co-occupied month + override", muts3, a9g3)

            # (iv) master mod toggle OFF -> rate-only target, pricing dashed
            r_off, _ = engine.suggest_net_rate(
                p, replace(ncmb, mod_adjustment_enabled=False), d0)

            def a9g4(wb):
                ncs = wb["_netcalc"]
                check("[net delivery: mod master OFF] rate-only r* ties oracle",
                      approx(ncs[f"E{rr8}"].value, r_off, 1e-9),
                      f"wb={ncs[f'E{rr8}'].value} oracle={r_off}")
                check("[net delivery: mod master OFF] pricing grid goes dark",
                      ncs[f"AH{t8 + 27}"].value in (None, ""))
            run("net delivery: mod master OFF", {("Control", "C13"): "OFF"}, a9g4)

            # (v) fabrication guard: the All view shows nothing computable
            def a9g5(wb):
                nd_ws = wb["Net Delivery"]
                check("[net delivery: All view] suggestions blank, never fabricated",
                      nd_ws[f"H{row_si8}"].value in (None, ""))
                check("[net delivery: All view] Checks still ALL PASS",
                      nval(wb, "ck_overall") == "ALL CHECKS PASS")
            run("net delivery: All view", {("Net Delivery", "B5"): "All"}, a9g5)

            # (vi) D63: the default filing tracks the RATE LOG. Move the
            # showcase combo's planned plan-year row and the tab must follow;
            # delete it and the tab must fall back to the suggestion.
            pl_i = next((k for k, rr_ in enumerate(rate_rows)
                         if rr_["bu"] == _bu(cfg, 2) and rr_["state"] == net_state
                         and rr_["status"] == "planned"
                         and rr_["eff"].year == p), None)
            if pl_i is not None:
                pl_xl = L.RL_FIRST + pl_i
                d_new = dt.date(p, 8, 15)
                moved = replace(ncmb, rate_changes=tuple(
                    engine.RateChange(d_new, 0.07, "planned", False, 1.0)
                    if (rc.status == "planned" and rc.effective.year == p) else rc
                    for rc in ncmb.rate_changes))
                m1_moved = engine.required_m1(p, moved, d_new, 0.07)

                def a9g6(wb):
                    ncs = wb["_netcalc"]
                    check("[net delivery: moved planned row] change in force follows "
                          "the rate log", approx(ncs[f"F{rr8}"].value, 0.07, 1e-12),
                          f"wb={ncs[f'F{rr8}'].value}")
                    check("[net delivery: moved planned row] M_1' ties the oracle at "
                          "the new date", approx(ncs[f"H{rr8}"].value, m1_moved, 1e-9),
                          f"wb={ncs[f'H{rr8}'].value} oracle={m1_moved}")
                run("net delivery: planned filing moved",
                    {("Rate Log", f"C{pl_xl}"): d_new,
                     ("Rate Log", f"D{pl_xl}"): 0.07}, a9g6)

                # deleting it falls back to the default date + suggestion
                no_plan = replace(ncmb, rate_changes=tuple(
                    rc for rc in ncmb.rate_changes
                    if not (rc.status == "planned" and rc.effective.year == p)))
                r_sugg, _ = engine.suggest_net_rate(p, no_plan, dt.date(p, 4, 1))

                def a9g7(wb):
                    ncs = wb["_netcalc"]
                    check("[net delivery: no planned row] falls back to the suggestion",
                          approx(ncs[f"F{rr8}"].value, r_sugg, 1e-9)
                          and approx(ncs[f"E{rr8}"].value, r_sugg, 1e-9),
                          f"wb={ncs[f'F{rr8}'].value} oracle={r_sugg}")
                    check("[net delivery: no planned row] the tab reports 'none'",
                          wb["Net Delivery"][f"F{row_si8}"].value == "none",
                          repr(wb["Net Delivery"][f"F{row_si8}"].value))
                run("net delivery: planned filing removed",
                    {("Rate Log", f"A{pl_xl}"): None, ("Rate Log", f"B{pl_xl}"): None,
                     ("Rate Log", f"C{pl_xl}"): None, ("Rate Log", f"D{pl_xl}"): None,
                     ("Rate Log", f"E{pl_xl}"): None, ("Rate Log", f"F{pl_xl}"): None},
                    a9g7)

    # 9h. D75: the mod-step solve where it actually bites — a combo that
    # ALREADY carries mod actions, put onto a net selection. The default
    # sample never lands here (the net combo has an empty mod log), so the
    # stepped regime, the whole-log base, and the day-blend collision in D's
    # own month are only reachable by mutation.
    ms_row = next((x for x in lr_rows if any(
        mr["bu"] == x["bu"] and mr["state"] == x["state"] for mr in mod_rows)
        and x["bu"] == _bu(cfg, 2)), None)
    if ms_row is not None:
        from src.sheets_netdelivery import ND_FIRST as _NDF, block_top as _bt
        ms_xl = L.LR_FIRST + lr_rows.index(ms_row)
        ms_si = list(st).index(ms_row["state"])
        ms_rr = _bt(ms_si) + 51
        ms_tab = _NDF + ms_si
        ms_combo = replace(sample_to_combo(cfg, lob, ms_row, rate_rows, mod_rows),
                           net_sel_p=0.09)
        # 4/15 lands ON the logged action; 4/1 lands BEFORE it, which moves the
        # two-point split to D and changes the blend even at a zero step (D31)
        for tag, d_ms in (("on the logged action", dt.date(p, 4, 15)),
                          ("before the logged action", dt.date(p, 4, 1))):
            r_ms, _ = engine.suggest_net_rate(p, ms_combo, d_ms)
            want = engine.required_mod_step(p, ms_combo, d_ms, r_ms)

            def a9h(wb, want=want, tag=tag, r_ms=r_ms):
                ncs = wb["_netcalc"]
                check(f"[net delivery: stepped combo, {tag}] r* ties oracle",
                      approx(ncs[f"E{ms_rr}"].value, r_ms, 1e-9),
                      f"wb={ncs[f'E{ms_rr}'].value} oracle={r_ms}")
                for cl, w, nm in (("S", want.required_step, "required mod step"),
                                  ("U", want.m_end, "year-end mod"),
                                  ("V", want.post_share, "reachable share")):
                    check(f"[net delivery: stepped combo, {tag}] D75 {nm} ties oracle",
                          approx(ncs[f"{cl}{ms_rr}"].value, w, 1e-9),
                          f"wb={ncs[f'{cl}{ms_rr}'].value} oracle={w}")
                check(f"[net delivery: stepped combo, {tag}] MOD-STEPS flagged",
                      "MOD-STEPS" in str(ncs[f"M{ms_rr}"].value or ""),
                      repr(ncs[f"M{ms_rr}"].value))
                check(f"[net delivery: stepped combo, {tag}] mod columns still "
                      "reproduce the engine path", approx(nval(wb, "nd_MaxStepDiff"),
                                                          0.0, 1e-9),
                      f"nd_MaxStepDiff={nval(wb, 'nd_MaxStepDiff')}")
            run(f"net delivery: stepped combo on a net selection, {tag}",
                {("Inputs", f"{lr_letter('lr_netp')}{ms_xl}"): 0.09,
                 ("Net Delivery", "B5"): _bu(cfg, 2),
                 ("Net Delivery", f"D{ms_tab}"): d_ms}, a9h)

        # the degenerate as-of: "mods as of 12/31 of the prior year" puts the
        # D70 switch anchor on top of the M_0 anchor. The engine drops it and
        # keeps the M_prior -> M_0 line; the workbook must not divide by zero.
        deg_asof = dt.date(p - 1, 12, 31)
        deg_combo = replace(ms_combo, mods=replace(ms_combo.mods, m0_asof=deg_asof))
        deg_eng = engine.MonthlyEngine(p, deg_combo)
        r_deg, _ = engine.suggest_net_rate(p, deg_combo, dt.date(p, 4, 15))
        deg_step = engine.required_mod_step(p, deg_combo, dt.date(p, 4, 15), r_deg)

        def a9h_deg(wb):
            ncs = wb["_netcalc"]
            t_ms = _bt(ms_si)
            # match Excel error VALUES only — the cohort block's own headers
            # ("#", "# chgs in mo", "# mod acts in mo") also start with a hash
            errs = [c.coordinate for row in
                    ncs.iter_rows(min_row=t_ms, max_row=t_ms + 51, max_col=64)
                    for c in row
                    if isinstance(c.value, str) and c.value in ERR_STRINGS]
            check("[net delivery: as-of 12/31 prior year] block computes at all",
                  not errs, f"{len(errs)} error cells, e.g. {errs[:4]}")
            bad = sum(0 if approx(ncs[f"O{t_ms + 15 + j}"].value,
                                  deg_eng.written_mod(engine.month_index(p - 1, 1) + j),
                                  1e-9) else 1 for j in range(12))
            check("[net delivery: as-of 12/31 prior year] history mod path ties oracle",
                  bad == 0, f"{bad} mismatched months")
            check("[net delivery: as-of 12/31 prior year] mod step still ties oracle",
                  approx(ncs[f"S{ms_rr}"].value, deg_step.required_step, 1e-9),
                  f"wb={ncs[f'S{ms_rr}'].value} oracle={deg_step.required_step}")
            check("[net delivery: as-of 12/31 prior year] Checks still ALL PASS",
                  nval(wb, "ck_overall") == "ALL CHECKS PASS",
                  repr(nval(wb, "ck_overall")))
        run("net delivery: mod as-of on the D70 switch date",
            {("Inputs", f"{lr_letter('lr_netp')}{ms_xl}"): 0.09,
             ("Inputs", f"{lr_letter('lr_m0asof')}{ms_xl}"): deg_asof,
             ("Net Delivery", "B5"): _bu(cfg, 2),
             ("Net Delivery", f"D{ms_tab}"): dt.date(p, 4, 15)}, a9h_deg)

    # 9j. D76: the reconciliation waterfall where it has three legs to show.
    # Only a combo with logged mod actions splits into term / timing / history,
    # and the default selection has none — select one and tie every leg.
    if mod_rows:
        me_top = L.ME_COH_LAST + 2
        ms_bu, ms_st = mod_rows[0]["bu"], mod_rows[0]["state"]
        ms_lr = next((x for x in lr_rows
                      if x["bu"] == ms_bu and x["state"] == ms_st), None)
        if ms_lr is not None:
            want = engine.process_reconciliation(
                p, sample_to_combo(cfg, lob, ms_lr, rate_rows, mod_rows))

            def a9j(wb):
                me = wb["Mod Engine"]
                check("[reconciliation: stepped combo] selection followed",
                      me["B3"].value == f"{ms_bu}|{ms_st}", repr(me["B3"].value))
                for cl, w_, nm in (
                        (f"K{me_top + 3}", want.m_avg, "M_avg"),
                        (f"L{me_top + 3}", want.lr_process, "current-process plan LR"),
                        (f"M{me_top + 4}", want.term_pts, "term leg"),
                        (f"M{me_top + 5}", want.timing_pts, "timing leg"),
                        (f"M{me_top + 6}", want.history_pts, "history leg"),
                        (f"K{me_top + 7}", want.m_earned, "model earned mod"),
                        (f"K{me_top + 9}", want.m_end_plan, "end level (log-derived)"),
                        (f"K{me_top + 10}", want.theta_actual, "share of the step earned"),
                        (f"M{me_top + 12}", want.gap_pts, "gap")):
                    check(f"[reconciliation: stepped combo] {nm} ties oracle",
                          approx(me[cl].value, w_, 1e-9),
                          f"wb={me[cl].value} oracle={w_}")
                check("[reconciliation: stepped combo] waterfall closes",
                      approx(nval(wb, "me_ProcWaterfall"), 0.0, 1e-9),
                      f"wb={nval(wb, 'me_ProcWaterfall')}")
                check("[reconciliation: stepped combo] model leg IS the plan LR",
                      approx(me[f"L{me_top + 7}"].value, nval(wb, "nr_CYLR_P"), 1e-9),
                      f"wb={me[f'L{me_top + 7}'].value} bridge={nval(wb, 'nr_CYLR_P')}")
                # D77: the same selection is the only one whose Mod Log has
                # PLANNED actions, so it is where the pricing split has
                # anything to report — a locked leg that merely echoed the
                # program leg would pass every other exercise in this file
                pfr_s = engine.program_flow_by_month(
                    p, sample_to_combo(cfg, lob, ms_lr, rate_rows, mod_rows))
                fd_s = wb["Flow Dashboard"]
                bad_s = sum(
                    0 if all(approx(fd_s[f"{cl}{93 + j}"].value, prow[k], 1e-9)
                             for cl, k in (("R", "locked_mod_leg"),
                                           ("S", "planned_mod_residual"),
                                           ("T", "locked_delivered"),
                                           ("U", "planned_delivered_residual")))
                    else 1 for j, prow in enumerate(pfr_s.rows))
                check("[reconciliation: stepped combo] D77 pricing split ties oracle "
                      "on all 24 YoY months", bad_s == 0, f"{bad_s} mismatched months")
                outstanding = max(abs(r_["planned_mod_residual"]) for r_ in pfr_s.rows)
                check("[reconciliation: stepped combo] the planned pricing residual is "
                      "actually non-zero (the split has something to say)",
                      outstanding > 1e-6, f"max |planned mod residual| = {outstanding}")
                for nm in ("fd_splitchk", "fd_lockedmodchk"):
                    check(f"[reconciliation: stepped combo] {nm} holds",
                          approx(nval(wb, nm), 0.0, 1e-9), f"{nm}={nval(wb, nm)}")
            run("reconciliation: stepped combo selected",
                {("Control", "C7"): ms_bu, ("Control", "C8"): ms_st}, a9j)

    # 9i. program-basis plan LR (D65): switch the worked example ONTO a net
    # selection and the gap must open exactly as the oracle says
    we_net65 = replace(base_combo, net_sel_p=0.08)
    lp65, _ = engine.program_basis_plan_lr(p, we_net65)
    gap65 = (lp65 - engine.run_bridge(p, we_net65, "monthly").cy_lr_p) * 100.0

    def a9i(wb):
        from src.sheets_calc import PROG_LR
        from src.xlstyle import col as xcol2
        cs_ = wb["_calc"]
        rr_ = L.CALC_RES_FIRST + lr_rows.index(we)
        got_lr = cs_[f"{xcol2(PROG_LR['cylr'])}{rr_}"].value
        got_gap = cs_[f"{xcol2(PROG_LR['gap'])}{rr_}"].value
        check("[program basis] gap opens when a net selection is switched on",
              approx(got_lr, lp65, 1e-9) and approx(got_gap, gap65, 1e-9),
              f"wb lr={got_lr} gap={got_gap} oracle lr={lp65} gap={gap65}")
        brg = wb["Bridge"]
        note = next((brg[f"B{r_}"].value for r_ in range(1, 120)
                     if isinstance(brg[f"B{r_}"].value, str)
                     and "booked the logged program" in brg[f"B{r_}"].value), None)
        check("[program basis] Bridge surfaces the counterfactual under net mode",
              note is not None, "counterfactual sentence absent")
        check("[program basis] switching a combo to net keeps every check passing",
              nval(wb, "ck_overall") == "ALL CHECKS PASS",
              str(nval(wb, "ck_overall")))
    run("program basis: worked example switched to net",
        {("Inputs", f"{lr_letter('lr_netp')}{we_row_xl}"): 0.08}, a9i)

    # 9h. Program Flow / dashboard written-legs exercises (D59)
    from src.sheets_programflow import (PF_PLAN_COL as PFPC, PF_PRIOR_COL as PFYC,
                                        PF_SUM_FIRST,
                                        grid_starts as pf_grid_starts,
                                        n_disp as pf_n_disp)
    pf_g1, pf_g2, pf_g3 = pf_grid_starts(cfg)
    pf_nd = pf_n_disp(cfg)
    pf_rl0 = L.RL_FIRST + len(rate_rows)

    # (i) a NEW planned row on the WE combo moves the dashboard legs and the
    # grid; the locked leg follows the (unchanged) taken program
    add = engine.RateChange(dt.date(p, 8, 10), 0.04, "planned", False)
    pf_add = engine.program_flow_by_month(
        p, replace(base_combo, rate_changes=base_combo.rate_changes + (add,)))
    muts_add = {("Rate Log", f"A{pf_rl0}"): we["bu"],
                ("Rate Log", f"B{pf_rl0}"): we["state"],
                ("Rate Log", f"C{pf_rl0}"): add.effective,
                ("Rate Log", f"D{pf_rl0}"): add.filed_pct,
                ("Rate Log", f"E{pf_rl0}"): "planned",
                ("Rate Log", f"F{pf_rl0}"): "N"}

    def a9h1(wb):
        fd = wb["Flow Dashboard"]
        bad = sum(0 if (approx(fd[f"M{93 + j}"].value, pr["rate_leg"], 1e-9)
                        and approx(fd[f"O{93 + j}"].value, pr["delivered"], 1e-9)
                        and approx(fd[f"P{93 + j}"].value, pr["locked_leg"], 1e-9))
                  else 1 for j, pr in enumerate(pf_add.rows))
        check("[program flow: appended planned row] dashboard M/O/P tie fresh oracle",
              bad == 0, f"{bad} mismatched months")
        pfs = wb[SHEETS.PROGRAM_FLOW]
        bad = sum(0 if approx(pfs.cell(row=pf_g1 + 3, column=PFPC + j).value,
                              pf_add.rows[j]["rate_leg"], 1e-9) else 1
                  for j in range(12))
        check("[program flow: appended planned row] grid A row ties fresh oracle",
              bad == 0, f"{bad} mismatched months")
        # D68: a change dated INSIDE the plan year cannot move the prior year
        bad = sum(0 if approx(pfs.cell(row=pf_g1 + 3, column=PFYC + j).value,
                              pf_add.prior_rows[j]["rate_leg"], 1e-9) else 1
                  for j in range(12))
        check("[program flow: appended planned row] prior-year band is unmoved "
              "by a plan-year filing", bad == 0, f"{bad} mismatched months")
    run("program flow: appended planned row", muts_add, a9h1)

    # (ii) master mod OFF: mod legs dash, delivered collapses to the rate leg
    pf_off = engine.program_flow_by_month(
        p, replace(base_combo, mod_adjustment_enabled=False))

    def a9h2(wb):
        fd = wb["Flow Dashboard"]
        check("[program flow: mod master OFF] dashboard mod leg dashes",
              fd["N93"].value == "—", f"N93={fd['N93'].value!r}")
        bad = sum(0 if approx(fd[f"O{93 + j}"].value, pr["delivered"], 1e-9) else 1
                  for j, pr in enumerate(pf_off.rows))
        check("[program flow: mod master OFF] delivered = rate-only, ties oracle",
              bad == 0, f"{bad} mismatched months")
        pfs_ = wb[SHEETS.PROGRAM_FLOW]
        check("[program flow: mod master OFF] grid B dashes in BOTH year bands",
              pfs_.cell(row=pf_g2 + 3, column=PFPC).value == "—"
              and pfs_.cell(row=pf_g2 + 3, column=PFYC).value == "—")
    run("program flow: mod master OFF", {("Control", "C13"): "OFF"}, a9h2)

    # (iii) mixed-status cohort month on the LOCKED block: a planned row
    # EARLIER in the same month as a taken row must not disturb the taken-only
    # day-blend (D58 exercised on the D59 block through a real recalc)
    mixtk = (engine.RateChange(dt.date(p - 1, 6, 5), 0.03, "planned", False),
             engine.RateChange(dt.date(p - 1, 6, 20), 0.02, "taken", False))
    pf_mix = engine.program_flow_by_month(
        p, replace(base_combo, rate_changes=base_combo.rate_changes + mixtk))
    muts_mixtk = {}
    for i, rc in enumerate(mixtk):
        rr_ = pf_rl0 + i
        muts_mixtk.update({("Rate Log", f"A{rr_}"): we["bu"],
                           ("Rate Log", f"B{rr_}"): we["state"],
                           ("Rate Log", f"C{rr_}"): rc.effective,
                           ("Rate Log", f"D{rr_}"): rc.filed_pct,
                           ("Rate Log", f"E{rr_}"): rc.status,
                           ("Rate Log", f"F{rr_}"): "N"})

    def a9h3(wb):
        fd = wb["Flow Dashboard"]
        bad = sum(0 if (approx(fd[f"P{93 + j}"].value, pr["locked_leg"], 1e-9)
                        and approx(fd[f"M{93 + j}"].value, pr["rate_leg"], 1e-9))
                  else 1 for j, pr in enumerate(pf_mix.rows))
        check("[program flow: mixed-status month] locked leg ties taken-only oracle "
              "(D58 on the locked block)", bad == 0, f"{bad} mismatched months")
    run("program flow: mixed-status cohort month (locked block)", muts_mixtk, a9h3)

    # (iv) All view (D60): the EP-weighted book view ties the combined oracle
    def _state_comb(stname):
        return engine.combined_flow_by_month(
            p, [sample_to_combo(cfg, lob, x, rate_rows, mod_rows)
                for x in lr_rows if x["state"] == stname])
    comb0, comb6 = _state_comb(st[0]), _state_comb(st[6])
    book = engine.combined_flow_by_month(
        p, [sample_to_combo(cfg, lob, x, rate_rows, mod_rows) for x in lr_rows])

    def a9h4(wb):
        pfs = wb[SHEETS.PROGRAM_FLOW]
        r9 = PF_SUM_FIRST

        def _c(cidx):
            return pfs.cell(row=r9, column=cidx).value

        ok = (approx(_c(PFPC + 2), comb0.avg_rate_ratio - 1.0, 1e-9)
              and approx(_c(PFPC + 4), comb0.avg_delivered_ratio - 1.0, 1e-9)
              and (_c(PFPC + 3) == "—" if comb0.avg_mod_ratio is None
                   else approx(_c(PFPC + 3), comb0.avg_mod_ratio - 1.0, 1e-9)))
        check("[program flow: All view] summary averages tie the EP-weighted oracle",
              ok, str([_c(PFPC + k) for k in (2, 3, 4)]))
        check("[program flow: All view] D68 prior-year summary delivered ties the "
              "EP-weighted oracle",
              approx(_c(PFYC + 2), comb0.avg_delivered_ratio_prior - 1.0, 1e-9),
              f"wb={_c(PFYC + 2)} oracle={comb0.avg_delivered_ratio_prior - 1.0}")
        check("[program flow: All view] delta column stays per-BU (dash)",
              _c(PFPC + 5) == "—", repr(_c(PFPC + 5)))
        bad = bad_y = 0
        for si_, comb_ in ((0, comb0), (6, comb6)):
            for g0, key_ in ((pf_g1, "rate_leg"), (pf_g2, "mod_leg"),
                             (pf_g3, "delivered")):
                for c0, rows_ in ((PFPC, comb_.rows), (PFYC, comb_.prior_rows)):
                    for j in range(12):
                        wv = pfs.cell(row=g0 + 3 + si_, column=c0 + j).value
                        ov = rows_[j][key_]
                        ok_ = (wv == "—") if ov is None else approx(wv, ov, 1e-9)
                        if not ok_:
                            if c0 == PFPC:
                                bad += 1
                            else:
                                bad_y += 1
        check("[program flow: All view] grids tie the combined oracle "
              "(2 states x 3 grids x 12 months)", bad == 0, f"{bad} mismatches")
        check("[program flow: All view] D68 prior-year grids tie the combined oracle "
              "(2 states x 3 grids x 12 months)", bad_y == 0, f"{bad_y} mismatches")
        bad = bad_y = 0
        for c0, rows_ in ((PFPC, book.rows), (PFYC, book.prior_rows)):
            for j in range(12):
                wv = pfs.cell(row=pf_g3 + 3 + pf_nd, column=c0 + j).value
                if not approx(wv, rows_[j]["delivered"], 1e-9):
                    if c0 == PFPC:
                        bad += 1
                    else:
                        bad_y += 1
        check("[program flow: All view] IN VIEW delivered row = the whole book "
              "(cross-state EP x w weights)", bad == 0, f"{bad} mismatches")
        check("[program flow: All view] D68 IN VIEW prior-year row = the whole book",
              bad_y == 0, f"{bad_y} mismatches")
        check("[program flow: All view] Checks still ALL PASS",
              nval(wb, "ck_overall") == "ALL CHECKS PASS")
    run("program flow: All view", {("Program Flow", "B5"): "All"}, a9h4)

    # (v) net combo in view: delta columns populate on program basis
    if len(st) > 8:
        net_state5 = st[8]
        net_row5 = next((x for x in lr_rows
                         if x["bu"] == _bu(cfg, 2) and x["state"] == net_state5), None)
        if net_row5 is not None and net_row5.get("netp") is not None:
            ncmb5 = sample_to_combo(cfg, lob, net_row5, rate_rows, mod_rows)
            pf_net = engine.program_flow_by_month(p, ncmb5)
            x5, x15 = ncmb5.net_sel_p, ncmb5.net_x1
            from src.sheets_programflow import PF_SUM_FIRST as pf_first5
            si5 = list(st).index(net_state5)

            def a9h5(wb):
                fd = wb["Flow Dashboard"]
                bad = sum(0 if approx(fd[f"V{93 + j}"].value,
                                      pr["delivered"] - (x5 if j < 12 else x15), 1e-9)
                          else 1 for j, pr in enumerate(pf_net.rows))
                check("[program flow: net combo] dashboard delta-vs-assertion ties "
                      "oracle (x for P, x1 for P+1)", bad == 0,
                      f"{bad} mismatched months")
                gv = wb[SHEETS.PROGRAM_FLOW].cell(row=pf_first5 + si5,
                                                  column=PFPC + 5).value
                ov = pf_net.avg_delivered_ratio - 1.0 - x5
                check("[program flow: net combo] summary delta ties oracle",
                      approx(gv, ov, 1e-9), f"wb={gv} oracle={ov}")
                check("[program flow: net combo] zero FAILs (WARN advisories allowed)",
                      nval(wb, "ck_overall") == "ALL CHECKS PASS")
            run("program flow: net combo in view",
                {("Control", "C7"): _bu(cfg, 2), ("Control", "C8"): net_state5,
                 ("Program Flow", "B5"): _bu(cfg, 2)}, a9h5)

    # 10. plan-year change (fingerprint must go N/A, engines recompute cleanly)
    p2 = p + 1
    m_p2 = engine.run_bridge(p2, base_combo, "monthly")

    def a10(wb):
        check("[plan year P+1] CY LR ties oracle for the new year",
              approx(nval(wb, "nr_CYLR_P"), m_p2.cy_lr_p, 1e-9),
              f"wb={nval(wb, 'nr_CYLR_P')} oracle={m_p2.cy_lr_p}")
        check("[plan year P+1] fingerprint FALSE", nval(wb, "orc_fp") is False)
        check("[plan year P+1] Checks still ALL PASS (oracle rows N/A)",
              nval(wb, "ck_overall") == "ALL CHECKS PASS",
              str(nval(wb, "ck_overall")))
    run("plan year change", {("Control", "C6"): p2}, a10)

    # 10b. plan year set BACK a year (the "reproduce last year" test, D44):
    # values AND the live year labels must both follow the Control input
    pm1 = p - 1
    m_pm1 = engine.run_bridge(pm1, base_combo, "monthly")

    def a10b(wb):
        check(f"[plan year {pm1}] CY LR ties oracle computed for {pm1}",
              approx(nval(wb, "nr_CYLR_P"), m_pm1.cy_lr_p, 1e-9),
              f"wb={nval(wb, 'nr_CYLR_P')} oracle={m_pm1.cy_lr_p}")
        check(f"[plan year {pm1}] E_CY ties oracle",
              approx(nval(wb, "nr_ECY_P"), m_pm1.e_cy[pm1], 1e-9))
        ss_hdr = wb["State Summary"].cell(row=7, column=27).value
        check(f"[plan year {pm1}] State Summary header relabels live",
              ss_hdr == f"=  CY {pm1} plan LR", str(ss_hdr))
        pf_hdr = wb["Portfolio"].cell(row=L.PF_HDR, column=7).value
        check(f"[plan year {pm1}] Portfolio header relabels live",
              pf_hdr == f"CY {pm1} plan LR", str(pf_hdr))
        banner = wb["Control"]["B10"].value
        check(f"[plan year {pm1}] Control stale-year banner appears",
              isinstance(banner, str) and banner.startswith("NOTE"), str(banner)[:60])
        check(f"[plan year {pm1}] Checks still ALL PASS",
              nval(wb, "ck_overall") == "ALL CHECKS PASS")
    run(f"plan year back to {pm1}", {("Control", "C6"): pm1}, a10b)


# ---------------------------------------------------------------------------


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--lob", default="Property")
    ap.add_argument("--workbook", default=None,
                    help="override the workbook path (default derived from config + --lob)")
    ap.add_argument("--config", default="config/config.yaml")
    ap.add_argument("--quick", action="store_true", help="skip phase D exercises")
    ap.add_argument("--skip-recalc", action="store_true",
                    help="assume the workbook already has cached values")
    args = ap.parse_args(argv)

    cfg = load_config(args.config)
    L.configure(cfg)
    lob = cfg.lob(args.lob)
    path = Path(args.workbook) if args.workbook else output_path(cfg, Path(cfg.output_dir), args.lob)
    print(f"Verifying {path} (LOB: {lob.name}, term {lob.term_months} months)")
    scratch_dir = path.parent / "_verify_scratch"
    scratch_dir.mkdir(exist_ok=True)

    phase_a(path)
    tie_default_state(path, cfg, lob, do_recalc=not args.skip_recalc)
    if not args.quick:
        phase_d(path, cfg, lob, scratch_dir)
    shutil.rmtree(scratch_dir, ignore_errors=True)

    print(f"\n{'=' * 60}\nRESULT: {PASS} passed, {FAIL} failed")
    for f in FAILURES:
        print(f"  FAIL: {f}")
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())
