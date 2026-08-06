"""Verification harness for the combined book workbook (DECISIONS.md D66).

The book's engines are frozen, so this proves something different from the
per-LOB harness: that every harvested row still equals what the oracle says
for that combo, that the aggregation over it is right under every filter
state, and that the file is structurally sound.

    python tools/verify_book.py                 # phases A-D
    python tools/verify_book.py --quick         # skip the filter exercises
"""

from __future__ import annotations

import argparse
import datetime as dt
import shutil
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # noqa: E402

from src import engine  # noqa: E402
from src.build_workbook import (load_config, sample_lr_rows, sample_mod_rows,  # noqa: E402
                                sample_rate_rows,
                                sample_to_combo)
from src.sheets_book import COLS, PF_FIRST, PIVOT, SS_FIRST  # noqa: E402
from src.sheets_main import ss_l  # noqa: E402  (shared column order, D90)
from tools.build_book import book_path  # noqa: E402
from tools.harvest import harvest  # noqa: E402
from tools.recalc import recalc  # noqa: E402
from tools.verify_workbook import (BANNED_RE, approx, scan_errors)  # noqa: E402

PASS = 0
FAIL = 0
FAILURES: list[str] = []


def _banner_ok(v) -> bool:
    """The status banner reports no FAILing check (D80 three-state banner)."""
    return isinstance(v, str) and (v == "ALL CHECKS PASS"
                                   or v.startswith("PASS WITH "))


def _anchor_row(wb, name: str) -> int:
    """Row of a single-cell anchor the builder exported.

    Grids are addressed by NAME rather than by counting rows from a caption, so
    an exhibit can grow a note or a header line without every tie below it
    quietly starting to check the wrong cells.
    """
    import re as _re
    ref = str(wb.defined_names[name].attr_text)
    return int(_re.search(r"\$(\d+)$", ref).group(1))


def check(desc: str, ok: bool, detail: str = ""):
    global PASS, FAIL
    if ok:
        PASS += 1
    else:
        FAIL += 1
        FAILURES.append(f"{desc}  {detail}")
        print(f"  FAIL  {desc}  {detail}")


def phase_a(path: Path):
    """Static scans — the same house rules the per-LOB harness enforces."""
    print("Phase A: static scans")
    wb = openpyxl.load_workbook(path, data_only=False)
    banned, ext, texted, merged = [], [], [], []
    for ws in wb.worksheets:
        if ws.merged_cells.ranges:
            merged.append(ws.title)
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str):
                    continue
                if v.startswith("="):
                    if BANNED_RE.search(v):
                        banned.append(f"{ws.title}!{c.coordinate}")
                    if "[" in v and "]" in v:
                        ext.append(f"{ws.title}!{c.coordinate}")
                    # D41/D66: a LITERAL that begins with "=" is stored as a
                    # formula; an invalid one makes Excel refuse the file, so
                    # every "=" cell must parse as a formula, not prose
                    body = v[1:].lstrip()
                    if (" " in body and "(" not in v and "&" not in v
                            and '"' not in v):
                        texted.append(f"{ws.title}!{c.coordinate}: {v[:40]}")
    check("no banned functions", not banned, "; ".join(banned[:5]))
    check("no external-workbook references", not ext, "; ".join(ext[:5]))
    check("no prose stored as a formula (the D41 trap)", not texted,
          "; ".join(texted[:3]))
    check("no merged cells", not merged, "; ".join(merged[:3]))
    check("hidden data sheet stays hidden", wb["_book"].sheet_state == "hidden")
    return wb


def phase_bc(path: Path, cfg, do_recalc=True):
    print("Phase B: recalculate and scan for errors")
    if do_recalc:
        recalc(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    errs = scan_errors(wb)
    check("zero formula errors in every sheet", not errs, "; ".join(errs[:5]))
    check("book Checks panel reports no FAILing check",
          _banner_ok(wb["Checks"]["C3"].value),
          str(wb["Checks"]["C3"].value))

    p = cfg.plan_year
    book = harvest(cfg)

    # D83: the book is a snapshot, and external links are banned, so nothing in
    # the file itself can notice a source that moved on after the harvest. The
    # data for the comparison is already on both sides — the book stamps each
    # source's mtime on Control — so the harness does what the workbook cannot.
    print("Phase B2: source freshness")
    ctl = wb["Control"]
    stale = []
    for i, s in enumerate(book.sources):
        stamped = ctl[f"C{18 + i}"].value
        src_path = Path(cfg.output_dir) / s.path
        live = (dt.datetime.fromtimestamp(src_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
                if src_path.exists() else None)
        check(f"[{s.lob}] source file still present", live is not None, str(src_path))
        if live is not None and stamped != live:
            stale.append(f"{s.lob}: harvested {stamped}, file now {live}")
        banner = ctl[f"F{18 + i}"].value
        check(f"[{s.lob}] source reported a passing Checks panel at harvest",
              _banner_ok(banner), str(banner))
    check("no source changed since the harvest (rebuild the book if it did)",
          not stale, "; ".join(stale))

    print("Phase C: every harvested row vs the oracle")
    bk = wb["_book"]
    # index the oracle by (lob, bu, state) so row order never matters
    oracle = {}
    for lob in cfg.lobs:
        lr_rows = sample_lr_rows(cfg, lob)
        rate_rows = sample_rate_rows(cfg, lob)
        mod_rows = sample_mod_rows(cfg, lob)   # D70: the oracle must see the log
        for rowdef in lr_rows:
            cmb = sample_to_combo(cfg, lob, rowdef, rate_rows, mod_rows)
            om = engine.run_bridge(p, cmb, "monthly")
            prog = engine.program_basis_plan_lr(p, cmb)
            oracle[(lob.name, rowdef["bu"], rowdef["state"])] = (rowdef, om, prog)
    bad = missing = 0
    for i, rec in enumerate(book.rows):
        key = (rec["lob"], rec["bu"], rec["state"])
        if key not in oracle:
            missing += 1
            continue
        rowdef, om, prog = oracle[key]
        r = 2 + i
        pairs = ((COLS["ep"], rowdef["ep"]), (COLS["crl"], om.crl_ind),
                 (COLS["ecy_p"], om.e_cy[p]), (COLS["arate_p"], om.a_rate_p),
                 (COLS["amod_p"], om.a_mod_p), (COLS["lrcur"], om.lr_current),
                 (COLS["cylr_p"], om.cy_lr_p), (COLS["cylr_p1"], om.cy_lr_p1),
                 (COLS["cylr_prog"], prog[0]), (COLS["cylr1_prog"], prog[1]))
        if not all(approx(bk.cell(row=r, column=c).value, v, 1e-9) for c, v in pairs):
            bad += 1
    check("every harvested row ties its per-combo oracle run (10 metrics)",
          bad == 0 and missing == 0, f"{bad} mismatched, {missing} unmatched rows")

    # the aggregates, recomputed in Python over the same rows
    rows = book.rows
    ep_all = sum(r["ep"] for r in rows)
    lr_all = sum(r["ep"] * r["cylr_p"] for r in rows) / ep_all
    prog_all = sum(r["ep"] * r["cylr_prog"] for r in rows) / ep_all
    ss = wb["State Summary"]
    tot = SS_FIRST + len(book.states) + 1
    check("State Summary total EP ties the harvest",
          approx(ss[f"B{tot}"].value, ep_all, 1e-6),
          f"wb={ss[f'B{tot}'].value} py={ep_all}")
    check("State Summary total plan LR ties the EP-weighted harvest",
          approx(ss[f"{ss_l('planlr')}{tot}"].value, lr_all, 1e-9),
          f"wb={ss[f'AA{tot}'].value} py={lr_all}")
    check("State Summary total program basis ties the harvest",
          approx(ss[f"{ss_l('progbasis')}{tot}"].value, prog_all, 1e-9))
    # per-state rows, unfiltered
    bad = 0
    for i, st in enumerate(book.states):
        sub = [r for r in rows if r["state"] == st]
        ep_s = sum(r["ep"] for r in sub)
        lr_s = sum(r["ep"] * r["cylr_p"] for r in sub) / ep_s
        r = SS_FIRST + i
        if not (approx(ss[f"B{r}"].value, ep_s, 1e-6)
                and approx(ss[f"{ss_l('planlr')}{r}"].value, lr_s, 1e-9)):
            bad += 1
    check("every State Summary row ties an EP-weighted recomputation", bad == 0,
          f"{bad} mismatched states")
    # Portfolio mirrors the harvest row for row
    pf = wb["Portfolio"]
    bad = sum(0 if (pf[f"A{PF_FIRST + i}"].value == rec["lob"]
                    and approx(pf[f"I{PF_FIRST + i}"].value, rec["cylr_p"], 1e-9))
              else 1 for i, rec in enumerate(rows))
    check("Portfolio grid mirrors every harvested row", bad == 0,
          f"{bad} mismatched rows")
    # ---- D103: the long pivot dataset ------------------------------------
    # The whole design rests on ONE property: SUM(weighted)/SUM(weight) over
    # any slice is the premium-weighted answer for that slice. Check it the way
    # a pivot would — group the sheet's own cells, then recompute in Python.
    pv = wb[PIVOT]
    agg: dict = {}
    for r in range(2, pv.max_row + 1):
        lob = pv.cell(row=r, column=1).value
        if lob is None:
            break
        measure = pv.cell(row=r, column=5).value
        w = pv.cell(row=r, column=7).value or 0.0
        wv = pv.cell(row=r, column=8).value or 0.0
        for key in ((measure, None), (measure, lob)):     # book-wide and by line
            a = agg.setdefault(key, [0.0, 0.0])
            a[0] += wv
            a[1] += w
    check("Pivot Data: no row carries a zero weight (a slice could go 0/0)",
          all(a[1] > 0 for a in agg.values()))
    bad = 0
    for measure, src in (("Plan LR", "cylr_p"), ("Plan LR +1", "cylr_p1"),
                         ("Rate earn-in (A_rate)", "arate_p"),
                         ("Mod drift (A_mod)", "amod_p"),
                         ("Projected LR (current level)", "lrcur"),
                         ("Plan LR — program basis", "cylr_prog")):
        for lob in (None, *sorted({r["lob"] for r in rows})):
            sub = [r for r in rows if lob is None or r["lob"] == lob]
            d = sum(r["ep"] or 0.0 for r in sub)
            want = sum((r["ep"] or 0.0) * (r[src] or 0.0) for r in sub) / d
            got = agg[(measure, lob)]
            if not approx(got[0] / got[1], want, 1e-9):
                bad += 1
    check("Pivot Data: weighted/weight reproduces the EP-weighted book and each "
          "line, for six measures", bad == 0, f"{bad} mismatched slices")
    # premium must survive the monthly split — the /12 rescale, checked whole
    ep_month = agg[("Rate YoY", None)][1]
    check("Pivot Data: monthly weights sum back to book premium",
          approx(ep_month, ep_all, 1e-6), f"monthly={ep_month} annual={ep_all}")
    check("Pivot Data is an Excel Table (a pivot source that grows with the roster)",
          "tbl_Pivot" in pv.tables, str(list(pv.tables)))

    # ---- D110: Net Delivery -----------------------------------------------
    # The two 24-month grids are premium-weighted means of harvested legs, and
    # the P+1 half is the part that could be silently wrong — it publishes rows
    # of the engine block that nothing read before, so an off-by-twelve would
    # look entirely plausible. Recompute both years in Python from the harvest.
    nd = wb["Net Delivery"]
    sum_r = _anchor_row(wb, "bk_nd_sum")
    bad = 0
    for key, leg in (("rate", "rate"), ("del", "delivered")):
        gr = _anchor_row(wb, f"bk_nd_{key}")
        for i, st in enumerate(book.states):
            sub = [r for r in rows if r["state"] == st]
            for yi, fam in enumerate((leg, leg + "1")):
                for j in range(12):
                    den = sum((r["epw"][j] or 0.0) for r in sub)
                    got = nd.cell(row=gr + i, column=2 + yi * 12 + j).value
                    if den <= 0:
                        continue
                    want = sum((r["epw"][j] or 0.0) * (r[fam][j] or 0.0)
                               for r in sub) / den
                    if not approx(got, want, 1e-9):
                        bad += 1
    check("Net Delivery: both 24-month grids reproduce an EP x weight recomputation "
          "over the harvest, for the plan year AND the following year", bad == 0,
          f"{bad} mismatched cells")

    # The following year must not merely echo the plan year — that is exactly
    # what an off-by-twelve in the published columns would look like.
    moved = sum(1 for r in rows for j in range(12)
                if not approx(r["rate"][j], r["rate1"][j], 1e-12))
    check("Net Delivery: the following-year legs are their own year, not a copy "
          "of the plan year", moved > 0,
          f"{moved} of {len(rows) * 12} combo-months differ")

    # Targets average over the combos that ASSERT one; a combo asserting
    # nothing is not a target of zero.
    bad = 0
    for i, st in enumerate(book.states):
        sub = [r for r in rows if r["state"] == st]
        nmode = [r for r in sub if (r["netmode"] or 0) == 1]
        for cL, fld in (("D", "netx"), ("E", "netx1")):
            got = nd[f"{cL}{sum_r + i}"].value
            if not nmode:
                if got != "—":
                    bad += 1
            elif not approx(got, sum(r[fld] for r in nmode) / len(nmode), 1e-9):
                bad += 1
    check("Net Delivery: net targets average over the net-mode combos only "
          "(dashed where none asserts one)", bad == 0, f"{bad} mismatched states")

    # The required pricing leg is shown only where the filters leave ONE combo
    # that asserts a target — and where it IS shown it must be the definition.
    price_r = _anchor_row(wb, "bk_nd_price")
    shown = wrong = 0
    for i, st in enumerate(book.states):
        sub = [r for r in rows if r["state"] == st]
        nmode = [r for r in sub if (r["netmode"] or 0) == 1]
        single = len(sub) == 1 and len(nmode) == 1
        for j in range(12):
            got = nd.cell(row=price_r + i, column=2 + j).value
            if not single:
                if got != "—":
                    wrong += 1
                continue
            den = sub[0]["epw"][j] or 0.0
            if den <= 0:
                continue
            rate = sub[0]["rate"][j] or 0.0
            want = (1.0 + nmode[0]["netx"]) / (1.0 + rate) - 1.0
            shown += 1
            if not approx(got, want, 1e-9):
                wrong += 1
    check("Net Delivery: the required pricing leg is dashed unless the filters "
          "resolve to one net-mode combo, and equals (1+x)/(1+rate)-1 where it is "
          "shown", wrong == 0, f"{wrong} wrong, {shown} shown")
    return wb, book


def phase_d(path: Path, cfg, book, scratch_dir: Path):
    print("Phase D: filter exercises")
    scratch_dir.mkdir(parents=True, exist_ok=True)
    scratch = scratch_dir / "book_exercise.xlsx"
    rows = book.rows
    tot = SS_FIRST + len(book.states) + 1

    def run(label_, lob, bu, state):
        shutil.copy2(path, scratch)
        wb = openpyxl.load_workbook(scratch, data_only=False)
        ctl = wb["Control"]
        ctl["B7"], ctl["B8"], ctl["B9"] = lob, bu, state
        wb.save(scratch)
        recalc(scratch, keep_calc_settings=False)   # throwaway copy (D88)
        got = openpyxl.load_workbook(scratch, data_only=True)
        errs = scan_errors(got)
        check(f"[{label_}] zero formula errors", not errs, "; ".join(errs[:3]))
        # State Summary honours LOB + BU (never its own dimension)
        sub = [r for r in rows
               if (lob == "All" or r["lob"] == lob)
               and (bu == "All" or r["bu"] == bu)]
        ep = sum(r["ep"] for r in sub)
        ss = got["State Summary"]
        check(f"[{label_}] State Summary total EP matches the subset",
              approx(ss[f"B{tot}"].value, ep, 1e-6),
              f"wb={ss[f'B{tot}'].value} py={ep}")
        if ep:
            lr = sum(r["ep"] * r["cylr_p"] for r in sub) / ep
            check(f"[{label_}] State Summary total plan LR matches the subset",
                  approx(ss[f"{ss_l('planlr')}{tot}"].value, lr, 1e-9),
                  f"wb={ss[f'AA{tot}'].value} py={lr}")
        # The chronology band, where the v3.7.2 bug lived. Where the filters
        # resolve a state row to ONE combo the slots populate, and a slot that
        # combo never filed has to come back BLANK. It used to come back as the
        # zero INDEX returns over an empty _book cell, which the date format
        # renders 1/0/00 and the signed percent 0.0% — a rate change that never
        # happened, shown as a real one. Only reachable single-combo, which is
        # why four versions of All/All views never saw it.
        if lob != "All" and bu != "All":
            bad = []
            for i, st in enumerate(book.states):
                one = [r for r in sub if r["state"] == st]
                if len(one) != 1:
                    continue
                for j in range(4):
                    filed = one[0]["slots"][j][0] is not None
                    for part in ("date", "pct", "tok"):
                        v = ss[f"{ss_l(f'chg{j + 1}_{part}')}{SS_FIRST + i}"].value
                        if filed == (v is None):
                            bad.append(f"{st} slot{j + 1} {part}={v!r} "
                                       f"(filed={filed})")
            check(f"[{label_}] unfiled rate-change slots read blank, not zero",
                  not bad, "; ".join(bad[:4]))

        # the Control KPI band honours all three
        sub3 = [r for r in sub if state == "All" or r["state"] == state]
        ep3 = sum(r["ep"] for r in sub3)
        check(f"[{label_}] Control EP-in-view matches all three filters",
              approx(got["Control"]["C13"].value, ep3, 1e-6),
              f"wb={got['Control']['C13'].value} py={ep3}")
        check(f"[{label_}] combos in view", got["Control"]["A13"].value == len(sub3),
              f"wb={got['Control']['A13'].value} py={len(sub3)}")
        check(f"[{label_}] book Checks still report no FAIL",
              _banner_ok(got["Checks"]["C3"].value))

    run("all", "All", "All", "All")
    run(f"line={book.lobs[0]}", book.lobs[0], "All", "All")
    run(f"bu={book.business_units[0]}", "All", book.business_units[0], "All")
    run(f"state={book.states[0]}", "All", "All", book.states[0])
    run("one combo", book.lobs[-1], book.business_units[-1], book.states[-1])
    shutil.rmtree(scratch_dir, ignore_errors=True)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Verify the combined book workbook.")
    ap.add_argument("--config", default="config/config.yaml")
    ap.add_argument("--workbook", default=None)
    ap.add_argument("--quick", action="store_true", help="skip the filter exercises")
    ap.add_argument("--skip-recalc", action="store_true")
    args = ap.parse_args(argv)

    cfg = load_config(args.config)
    path = Path(args.workbook) if args.workbook else book_path(cfg, Path(cfg.output_dir))
    print(f"Verifying {path}")
    phase_a(path)
    _wb, book = phase_bc(path, cfg, do_recalc=not args.skip_recalc)
    if not args.quick:
        phase_d(path, cfg, book, path.parent / "_book_scratch")
    print(f"\n{'=' * 60}\nRESULT: {PASS} passed, {FAIL} failed")
    for f in FAILURES:
        print(f"  FAIL: {f}")
    return 1 if FAIL else 0


if __name__ == "__main__":
    raise SystemExit(main())
