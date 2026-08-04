"""A phase-D exercise without the file I/O (D105).

Phase D used to cost 3.49s per exercise, of which the actual recalculation was
0.02s. The other 99% was ceremony: copy the 2.7MB workbook, reopen it in
openpyxl, write the mutation, save, hand it to Excel, let Excel open and save it
again, then parse the result back in openpyxl -- twice. Six lines x ~50
exercises put that toll on the wall clock ~300 times a release.

This module pays it once per LINE instead. One workbook is opened in Excel and
stays open; an exercise writes its cells, calls ``Calculate``, reads the answers
straight out of the live grid, and puts the cells back. Measured on Property:

    Workbooks.Open              0.92s  ONCE
    CalculateFullRebuild        0.38s  ONCE (see _rebuild)
    ---- per exercise ----
    write a cell                0.014s
    Calculate + wait            0.013s
    read a sheet in bulk        0.03-0.18s
    SpecialCells error sweep    0.15s   (whole workbook, 26 sheets)

``Application.Calculate`` is enough after the first rebuild because this
workbook family bans every volatile and indirect function (OFFSET, INDIRECT,
TODAY, NOW), so the dependency graph is static and complete -- a changed input
dirties exactly the cells that depend on it. That is also why plain Calculate is
17x cheaper than CalculateFullRebuild, which only has to run once, on open.

WHY THIS IS SAFE THIS TIME (D86a said it was not)
-------------------------------------------------
Mutating through COM was tried before and reverted, for a good reason: under the
parallel release driver, with six Excel instances live at once, a write could
FAIL SILENTLY. One line came back with a full set of assertion failures whose
"actual" values were the workbook's default state -- the mutation had never
landed, and the harness never noticed. A verifier that can quietly fail to set
up its own exercise could just as quietly report a false PASS.

Two things close that hole, and neither existed then:

1. ``_assert_applied`` runs AFTER Calculate, over COM, on every mutated cell,
   and RAISES. A mutation that does not land stops the run instead of grading
   the wrong workbook.
2. Every exercise restores its cells and then re-reads a canary set of headline
   values, which must return to the baseline captured at open. That catches the
   opposite failure -- a restore that did not land, leaking one exercise's state
   into the next.

The third thing is that this is fast enough not to need the parallelism that
caused the original instability (D100: those COM errors were RPC_E_CALL_REJECTED
-- "busy" -- from six instances contending, not broken workbooks).

READING
-------
The shims present the openpyxl surface phase D already uses -- ``wb[sheet]``,
``ws[addr].value``, ``ws.cell(row=, column=)``, ``ws.iter_rows(...)``,
``wb.defined_names`` -- so the ~1,000 lines of assertions are untouched.

Reads are BULK. A single-cell COM read costs ~3.5ms, so the 63-cell state loops
in phase D would cost more than the recalculation they are checking; one
``UsedRange.Value`` for a whole sheet costs 0.03-0.18s and serves every cell on
it. The snapshot is taken on first touch and dropped whenever the grid changes.

``.Value`` rather than ``.Value2``: it returns dates as datetimes, which is what
openpyxl does and what the assertions expect. ``_py`` finishes the job --
stripping pywin32's local timezone stamp (Excel serials have no timezone), and
mapping Excel's error variants back to the "#VALUE!" strings the error scan
matches on. All of this is measured, not assumed; see ``tests/test_live.py``,
which diffs every cell of a real workbook against openpyxl.
"""

from __future__ import annotations

import datetime as dt
import shutil
import time
from pathlib import Path

from openpyxl.utils.cell import coordinate_to_tuple

from tools.recalc import _excel, busy_retry, com_value

# Excel enumerations, spelled out so a reader does not have to look them up
XL_CALC_MANUAL = -4135
XL_CALC_DONE = 0
XL_CELLTYPE_FORMULAS = -4123
XL_CELLTYPE_CONSTANTS = 2
XL_ERRORS = 16

# An error cell marshals through pywin32 as an INT whose low 16 bits are the
# Excel error code; every real number arrives as a float, so the type alone
# identifies an error. Verified against a live instance for all six common
# codes before this module was written.
ERR_BY_CODE = {
    2000: "#NULL!", 2007: "#DIV/0!", 2015: "#VALUE!", 2023: "#REF!",
    2029: "#NAME?", 2036: "#NUM!", 2042: "#N/A", 2043: "#GETTING_DATA",
    2045: "#SPILL!", 2046: "#CONNECT!", 2047: "#BLOCKED!", 2048: "#UNKNOWN!",
    2049: "#FIELD!", 2050: "#CALC!",
}

# A whole sheet of errors should not turn the scan into an hour of COM calls.
ERROR_SCAN_LIMIT = 200


def _py(v):
    """One COM value, as openpyxl's ``data_only=True`` would have reported it.

    Three deliberate conversions, each of which was a real difference:

    * an error variant (an ``int``) becomes its "#DIV/0!" string, because that
      is what the error scan matches on;
    * ``""`` becomes ``None`` -- a formula returning blank has no cached value
      in the file, so openpyxl reports None and a comparison against None is
      what the assertions are written against;
    * a ``pywintypes.datetime`` loses its timezone. pywin32 stamps the machine's
      local zone onto a value that has none; keeping it would make a tz-aware
      datetime compare unequal (or raise) against the oracle's naive one, and
      east of UTC it could shift the date.
    """
    if v is None:
        return None
    if isinstance(v, bool):                 # bool before int: it IS an int
        return v
    if isinstance(v, int):
        code = v & 0xFFFF
        return ERR_BY_CODE.get(code, f"#ERR:{code}")
    if isinstance(v, str):
        return v or None
    if isinstance(v, dt.datetime):
        return dt.datetime(v.year, v.month, v.day, v.hour, v.minute, v.second,
                           v.microsecond)
    return v


def _split_ref(refers_to: str) -> tuple[str, str]:
    """``"='State Summary'!$A$1"`` -> ``("State Summary", "$A$1")``."""
    s = refers_to.lstrip("=")
    sheet, sep, ref = s.rpartition("!")
    # "=#REF!" is what a name whose target was deleted looks like; it splits
    # without complaint into ("#REF", "") and would then read as a blank cell
    if not sep or not ref or not sheet:
        raise ValueError(f"defined name does not point at a cell: {refers_to!r}")
    if sheet.startswith("'") and sheet.endswith("'"):
        sheet = sheet[1:-1].replace("''", "'")
    return sheet, ref


class LiveCell:
    """Enough of an openpyxl cell for the assertions: a value and an address."""

    __slots__ = ("value", "coordinate")

    def __init__(self, value, coordinate):
        self.value = value
        self.coordinate = coordinate

    def __repr__(self):
        return f"<LiveCell {self.coordinate}={self.value!r}>"


class LiveSheet:
    """One worksheet, read in bulk and served from memory until it changes."""

    def __init__(self, book: "LiveBook", title: str):
        self.book = book
        self.title = title
        self._grid = None          # (rows, r0, c0)

    # -- the bulk snapshot ---------------------------------------------------
    def _load(self):
        if self._grid is not None:
            return self._grid
        ws = self.book.sheet(self.title)
        ur = busy_retry(lambda: ws.UsedRange)
        r0, c0 = ur.Row, ur.Column
        raw = busy_retry(lambda: ur.Value)
        if not isinstance(raw, tuple):                 # a one-cell UsedRange
            rows = ((raw,),)
        elif raw and not isinstance(raw[0], tuple):    # a single row
            rows = (raw,)
        else:
            rows = raw
        self._grid = (rows, r0, c0)
        return self._grid

    def _invalidate(self):
        self._grid = None

    def _at(self, row: int, column: int):
        rows, r0, c0 = self._load()
        i, j = row - r0, column - c0
        if 0 <= i < len(rows) and 0 <= j < len(rows[i]):
            return _py(rows[i][j])
        return None                                    # outside the used range

    # -- the openpyxl surface ------------------------------------------------
    def cell(self, row: int, column: int) -> LiveCell:
        from openpyxl.utils import get_column_letter
        return LiveCell(self._at(row, column), f"{get_column_letter(column)}{row}")

    def __getitem__(self, addr: str) -> LiveCell:
        if ":" in addr:
            raise NotImplementedError(
                f"{self.title}!{addr}: the live session reads single cells; use "
                f"iter_rows() for a block")
        row, column = coordinate_to_tuple(addr)
        return LiveCell(self._at(row, column), addr.replace("$", ""))

    def iter_rows(self, min_row=None, max_row=None, min_col=None, max_col=None):
        rows, r0, c0 = self._load()
        n_rows = len(rows)
        n_cols = max((len(r) for r in rows), default=0)
        lo_r = r0 if min_row is None else min_row
        hi_r = r0 + n_rows - 1 if max_row is None else max_row
        lo_c = c0 if min_col is None else min_col
        hi_c = c0 + n_cols - 1 if max_col is None else max_col
        from openpyxl.utils import get_column_letter
        for r in range(lo_r, hi_r + 1):
            yield tuple(
                LiveCell(self._at(r, c), f"{get_column_letter(c)}{r}")
                for c in range(lo_c, hi_c + 1))

    def __repr__(self):
        return f"<LiveSheet {self.title!r}>"


class _LiveName:
    __slots__ = ("destinations",)

    def __init__(self, sheet: str, ref: str):
        self.destinations = ((sheet, ref),)


class _LiveNames:
    """``wb.defined_names[name].destinations``, resolved lazily and cached."""

    def __init__(self, book: "LiveBook"):
        self.book = book
        self._cache: dict[str, _LiveName] = {}

    def __getitem__(self, name: str) -> _LiveName:
        got = self._cache.get(name)
        if got is None:
            try:
                refers = busy_retry(lambda: self.book._com.Names(name).RefersTo)
            except Exception as e:  # noqa: BLE001 - absent name, as openpyxl
                raise KeyError(name) from e
            got = self._cache[name] = _LiveName(*_split_ref(refers))
        return got

    def __contains__(self, name: str) -> bool:
        try:
            self[name]
            return True
        except KeyError:
            return False


class LiveBook:
    """One workbook, open in Excel for the length of a phase-D run.

    Use as a context manager; ``apply`` puts the workbook into a mutated state
    and returns an object the assertions can read like an openpyxl workbook.
    """

    def __init__(self, src: Path, scratch: Path, canary: tuple[str, ...] = ()):
        shutil.copy2(src, scratch)
        self.path = scratch
        self._xl = _excel()
        self._com = busy_retry(
            lambda: self._xl.Workbooks.Open(str(scratch.resolve()), UpdateLinks=0))
        # Calculation is only settable while a workbook is open, so it is set
        # here and put back in close() BEFORE the workbook goes away.
        self._prev_calc = busy_retry(lambda: self._xl.Calculation)
        busy_retry(lambda: setattr(self._xl, "Calculation", XL_CALC_MANUAL))
        self._sheets: dict[str, LiveSheet] = {}
        self._com_sheets: dict[str, object] = {}
        self._saved: dict[tuple[str, str], tuple[bool, object]] = {}
        self.defined_names = _LiveNames(self)
        self._rebuild()
        self.canary = {n: self.nval(n) for n in canary}
        self.exercises = 0

    # -- Excel plumbing ------------------------------------------------------
    def sheet(self, title: str):
        got = self._com_sheets.get(title)
        if got is None:
            got = self._com_sheets[title] = busy_retry(
                lambda: self._com.Worksheets(title))
        return got

    def _wait(self, what: str):
        import pythoncom
        deadline = time.monotonic() + 300.0
        while busy_retry(lambda: self._xl.CalculationState) != XL_CALC_DONE:
            if time.monotonic() > deadline:
                raise RuntimeError(f"Excel still calculating after 300s ({what})")
            pythoncom.PumpWaitingMessages()
            time.sleep(0.005)

    def _rebuild(self):
        """The one full rebuild, on open. Everything after it is incremental."""
        busy_retry(self._xl.CalculateFullRebuild)
        self._wait("full rebuild")
        self._invalidate()

    def calculate(self):
        busy_retry(self._xl.Calculate)
        self._wait("calculate")
        self._invalidate()

    def _invalidate(self):
        for sh in self._sheets.values():
            sh._invalidate()

    # -- the openpyxl surface ------------------------------------------------
    def __getitem__(self, title: str) -> LiveSheet:
        got = self._sheets.get(title)
        if got is None:
            got = self._sheets[title] = LiveSheet(self, title)
        return got

    @property
    def worksheets(self) -> list[LiveSheet]:
        return [self[ws.Name] for ws in self._com.Worksheets]

    @property
    def sheetnames(self) -> list[str]:
        return [ws.Name for ws in self._com.Worksheets]

    def nval(self, name: str):
        sheet, ref = self.defined_names[name].destinations[0]
        return self[sheet][ref.replace("$", "")].value

    # -- mutation ------------------------------------------------------------
    def _snapshot(self, sheet: str, addr: str):
        key = (sheet, addr)
        if key in self._saved:
            return
        rng = busy_retry(lambda: self.sheet(sheet).Range(addr))
        has_f = bool(busy_retry(lambda: rng.HasFormula))
        # Formula for a formula cell, Value2 for a constant: between them they
        # round-trip exactly, number format included (measured).
        self._saved[key] = (has_f, busy_retry(lambda: rng.Formula if has_f
                                              else rng.Value2))

    def _write(self, sheet: str, addr: str, value):
        # com_value, not a raw datetime: pywin32 marshals a naive datetime
        # through the local timezone and can land it on the previous day, which
        # computes a correct answer for the wrong date (D86, and 18 harness
        # failures). Every date target already carries a date number format, so
        # the serial renders exactly as the openpyxl path wrote it.
        rng = busy_retry(lambda: self.sheet(sheet).Range(addr))
        busy_retry(lambda: setattr(rng, "Value2", com_value(value)))

    def apply(self, mutations: dict[tuple[str, str], object]) -> "LiveBook":
        """Put the workbook into exactly this state and hand it back to read."""
        self.restore()
        for (sheet, addr), value in mutations.items():
            self._snapshot(sheet, addr)
            self._write(sheet, addr, value)
        self.calculate()
        self._assert_applied(mutations)
        self.exercises += 1
        return self

    def _assert_applied(self, mutations) -> None:
        """The D86a guard: prove the setup landed before grading anything.

        A silent write failure used to produce a full set of assertion failures
        whose "actual" values were the DEFAULT workbook -- which on a different
        day is a false PASS. This raises instead.
        """
        for (sheet, addr), want in mutations.items():
            got = self[sheet][addr].value
            if isinstance(want, dt.date) and not isinstance(want, dt.datetime):
                if isinstance(got, dt.datetime):
                    got = got.date()
            if isinstance(want, (int, float)) and not isinstance(want, bool) \
                    and isinstance(got, (int, float)) and not isinstance(got, bool):
                ok = abs(float(got) - float(want)) <= 1e-9
            else:
                ok = got == want
            if not ok:
                raise RuntimeError(
                    f"mutation did not take: {sheet}!{addr} is {got!r}, expected "
                    f"{want!r} — the exercise would have asserted against the "
                    f"wrong workbook state")

    def restore(self) -> None:
        """Undo every cell this session has touched, and prove it took."""
        if not self._saved:
            return
        for (sheet, addr), (has_f, saved) in self._saved.items():
            rng = busy_retry(lambda: self.sheet(sheet).Range(addr))
            if has_f:
                busy_retry(lambda: setattr(rng, "Formula", saved))
            else:
                busy_retry(lambda: setattr(rng, "Value2", saved))
        self._saved.clear()
        self.calculate()
        self._check_canary()

    def _check_canary(self) -> None:
        """State leakage is the failure mode this design has to earn its way
        past, so it is checked every time rather than sampled."""
        for name, want in self.canary.items():
            got = self.nval(name)
            same = (abs(got - want) <= 1e-9
                    if isinstance(want, (int, float)) and isinstance(got, (int, float))
                    else got == want)
            if not same:
                raise RuntimeError(
                    f"live workbook did not return to its baseline after "
                    f"{self.exercises} exercises: {name} is {got!r}, opened as "
                    f"{want!r} — a restore did not land and later exercises "
                    f"would be graded against a contaminated workbook")

    # -- lifecycle -----------------------------------------------------------
    def close(self) -> None:
        try:
            self.restore()
        finally:
            try:
                # while the workbook is still open: Excel refuses to set
                # Calculation with no workbook to own the mode
                busy_retry(lambda: setattr(self._xl, "Calculation", self._prev_calc))
            except Exception:  # noqa: BLE001
                pass
            try:
                self._com.Saved = True          # never write the scratch back
                busy_retry(self._com.Close, SaveChanges=False)
            except Exception:  # noqa: BLE001
                pass
            # Let go of every proxy now, while the apartment that created them
            # is still alive. A COM object outliving its apartment is released
            # into a torn-down one, which faults at interpreter exit rather than
            # anywhere you could catch it (the same rule recalc._com_init keeps).
            self._com = None
            self._xl = None
            self._com_sheets.clear()
            self._sheets.clear()

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        self.close()
        return False


def scan_errors_live(book: LiveBook, limit: int = ERROR_SCAN_LIMIT) -> list[str]:
    """Every error-valued cell in the workbook, via Excel's own index.

    ``SpecialCells`` asks Excel which cells are errors instead of reading 1.2M
    values and testing each one: 0.15s against 0.37s, and it yields addresses
    directly. It raises when a sheet has none, which is the common case.
    """
    import pywintypes

    out: list[str] = []
    for ws in book._com.Worksheets:
        ur = busy_retry(lambda: ws.UsedRange)
        if busy_retry(lambda: ur.Count) == 1:
            # SpecialCells on a single-cell range searches the WHOLE sheet, so
            # this degenerate case is answered directly instead
            v = _py(busy_retry(lambda: ur.Value))
            if isinstance(v, str) and v.startswith("#"):
                out.append(f"{ws.Name}!{ur.Address.replace('$', '')}={v}")
            continue
        for ctype in (XL_CELLTYPE_FORMULAS, XL_CELLTYPE_CONSTANTS):
            try:
                found = ur.SpecialCells(ctype, XL_ERRORS)
            except pywintypes.com_error:
                continue                     # no cells of that kind: the norm
            for cell in found:
                out.append(f"{ws.Name}!{cell.Address.replace('$', '')}="
                           f"{_py(cell.Value)}")
                if len(out) >= limit:
                    out.append(f"...truncated at {limit} error cells")
                    return out
    return out
