"""Harvest per-combo results from the generated LOB workbooks (DECISIONS.md D66).

The six LOB workbooks each publish one row per BU x state on their hidden
``_calc`` results table — every plan metric, the program-basis counterfactual,
the monthly rate/mod/delivered legs, and the rate-change history (D60/D61/D65/
D66). Stacking those six tables IS the book, so this module reads them
positionally and hands back plain dicts; nothing downstream needs openpyxl.

The one hazard the design has to defend against: ``data_only=True`` returns
None for a formula cell with no cached result, so a workbook that was never
recalculated would harvest as a page of zeros and roll up into a plausible,
wrong book. Every source is checked and the harvest REFUSES rather than
silently under-report.

Usage:
    from tools.harvest import harvest
    book = harvest(cfg)                 # reads output/ per the config template
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # noqa: E402

import datetime as dt
from dataclasses import dataclass, field

import openpyxl

from src.build_workbook import (GENERATOR_VERSION, Layout as L, load_config,
                                output_path)
from src.sheets_calc import BOOK_PUB, BOOK_SLOTS, FLOW_PUB, PROG_LR, W_AOTHER_COL

# scalar per-combo fields: field name -> column index on the results table
SCALARS = {
    "key": 1, "term": 2, "crl": 3, "ecy_pm1": 4, "ecy_p": 5, "ecy_p1": 6,
    "mbar_p": 7, "mbar_p1": 8, "arate_p": 9, "arate_p1": 10, "amod_p": 11,
    "amod_p1": 12, "lrcur": 13, "cylr_p": 14, "cylr_p1": 15, "echg": 16,
    "carry": 17, "aother": 18, "trend": 19, "state": 20, "bu": 21, "ep": 22,
    "m_ind_w": 23, "m0_w": 24, "m1_w": 25,
    "netmode": 37, "netx": 38,
    "modeff": FLOW_PUB["modeff"], "avg_rate": FLOW_PUB["avg_rate"],
    "avg_mod": FLOW_PUB["avg_mod"], "avg_del": FLOW_PUB["avg_del"],
    "flow_gap": FLOW_PUB["gap"], "aother_w": W_AOTHER_COL,
    "cylr_prog": PROG_LR["cylr"], "cylr1_prog": PROG_LR["cylr1"],
    "progident": PROG_LR["ident"], "proggap": PROG_LR["gap"],
    "ntaken": BOOK_PUB["ntaken"], "nplanned": BOOK_PUB["nplanned"],
}
# 12-month families: field name -> first column index
MONTHLY = {"delivered": FLOW_PUB["delivered"], "rate": FLOW_PUB["rate"],
           "mod": FLOW_PUB["mod"], "epw": FLOW_PUB["epw"]}
LAST_COL = BOOK_PUB["slot"] + BOOK_SLOTS * 3 - 1

# A row is only usable if these carry cached values. None means one of two
# things, and both must stop the harvest: the workbook was never recalculated,
# or it was built by an older generator that has no such column (ntaken is a
# COUNTIFS — always numeric in a compatible, calculated file, so it doubles as
# the schema probe).
REQUIRED = ("crl", "ecy_p", "cylr_p", "cylr_prog", "ep", "ntaken")


class HarvestError(RuntimeError):
    """Raised when a source workbook is missing, stale, or unreadable."""


@dataclass
class Source:
    lob: str
    path: str
    modified: str          # source file mtime, "YYYY-MM-DD HH:MM"
    version: str           # the generator version stamped on its Read Me
    combos: int
    checks: str = ""       # the source's own Checks banner at harvest time (D83)

    @property
    def checks_ok(self) -> bool:
        return self.checks.startswith("ALL CHECKS PASS") or self.checks.startswith("PASS WITH ")


@dataclass
class BookData:
    plan_year: int
    rows: list = field(default_factory=list)      # one dict per combo
    sources: list = field(default_factory=list)   # one Source per LOB
    as_of: str = ""                               # harvest timestamp

    @property
    def lobs(self) -> list:
        return [s.lob for s in self.sources]

    @property
    def business_units(self) -> list:
        return sorted({r["bu"] for r in self.rows})

    @property
    def states(self) -> list:
        """States in first-appearance order — the roster the book displays."""
        seen, out = set(), []
        for r in self.rows:
            if r["state"] not in seen:
                seen.add(r["state"])
                out.append(r["state"])
        return out


def _version_of(wb) -> str:
    """The version stamped in the Read Me subtitle (prose, not a name)."""
    try:
        sub = wb["Read Me"]["B3"].value or ""
    except KeyError:
        return "?"
    for part in str(sub).split("|"):
        part = part.strip()
        if part.startswith("version "):
            return part.split(" ", 1)[1].strip()
    return "?"


def _checks_banner(wb) -> str:
    """The source's own Checks!C3 status string, "" when unreadable (D83)."""
    try:
        v = wb["Checks"]["C3"].value
    except KeyError:
        return ""
    return v if isinstance(v, str) else ""


def read_lob(path: Path, lob: str, require_checks: bool = True) -> tuple[list, Source]:
    """Read one workbook's published results rows. Raises HarvestError."""
    if not path.exists():
        raise HarvestError(
            f"{lob}: {path} not found — generate the LOB workbooks first "
            f"(python -m src.build_workbook)")
    wb = openpyxl.load_workbook(path, data_only=True)
    if "_calc" not in wb.sheetnames:
        raise HarvestError(f"{lob}: {path.name} has no _calc sheet")
    ws = wb["_calc"]
    rows = []
    for r in range(L.CALC_RES_FIRST, L.CALC_RES_LAST + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, LAST_COL + 1)]
        key = vals[0]
        if key in (None, ""):
            continue                                   # spare tbl_LR row
        rec = {"lob": lob}
        for name, cidx in SCALARS.items():
            rec[name] = vals[cidx - 1]
        for name, c0 in MONTHLY.items():
            rec[name] = [vals[c0 - 1 + j] for j in range(12)]
        rec["slots"] = [
            tuple(vals[BOOK_PUB["slot"] - 1 + j * 3 + k] for k in range(3))
            for j in range(BOOK_SLOTS)]
        missing = [f for f in REQUIRED if rec[f] is None]
        if missing:
            raise HarvestError(
                f"{lob}: {path.name} row {r} ({key}) has no cached value for "
                f"{', '.join(missing)} — the workbook was either never "
                f"recalculated or was built by an older generator (this book "
                f"needs v{GENERATOR_VERSION}, the file says v{_version_of(wb)}). "
                f"Rebuild and recalculate it, then harvest again.")
        rows.append(rec)
    if not rows:
        raise HarvestError(f"{lob}: {path.name} published no combo rows")
    dupes = sorted({r["key"] for r in rows
                    if [x["key"] for x in rows].count(r["key"]) > 1})
    if dupes:
        raise HarvestError(
            f"{lob}: duplicate combo key(s) {', '.join(dupes)} — BU x state "
            f"must be unique within a workbook")
    # D83: a source whose own trust panel is red must not roll up green. The
    # gate comes AFTER the cached-value gate so a never-recalculated file keeps
    # the more specific message (its banner would be blank either way).
    banner = _checks_banner(wb)
    mtime = dt.datetime.fromtimestamp(path.stat().st_mtime)
    src = Source(lob=lob, path=path.name, modified=mtime.strftime("%Y-%m-%d %H:%M"),
                 version=_version_of(wb), combos=len(rows), checks=banner)
    if require_checks and not src.checks_ok:
        raise HarvestError(
            f"{lob}: {path.name} reports {banner or 'no Checks status'} — a workbook "
            f"whose own Checks panel is failing must not roll into the book. Fix it "
            f"(open the file, read the FAILing row), or harvest with --allow-failing-"
            f"sources to build anyway; the book will show which sources were bad.")
    return rows, src


def harvest(cfg, out_dir: str | Path | None = None,
            now: dt.datetime | None = None,
            allow_failing_sources: bool = False) -> BookData:
    """Stack every configured LOB workbook into one BookData."""
    book = BookData(plan_year=cfg.plan_year,
                    as_of=(now or dt.datetime.now()).strftime("%Y-%m-%d %H:%M"))
    seen = set()
    for lob_name in cfg.output_lobs:
        if lob_name in seen:
            raise HarvestError(f"LOB {lob_name!r} configured twice")
        seen.add(lob_name)
        path = output_path(cfg, Path(out_dir or cfg.output_dir), lob_name)
        rows, src = read_lob(path, lob_name,
                             require_checks=not allow_failing_sources)
        book.rows.extend(rows)
        book.sources.append(src)
    if not book.sources:
        raise HarvestError("no LOBs configured for output")
    return book


def main(argv=None) -> int:
    import argparse
    ap = argparse.ArgumentParser(description="Summarise a harvest of the LOB workbooks.")
    ap.add_argument("--config", default="config/config.yaml")
    ap.add_argument("--out", default=None)
    args = ap.parse_args(argv)
    book = harvest(load_config(args.config), args.out)
    print(f"harvested {len(book.rows)} combos from {len(book.sources)} workbooks "
          f"as of {book.as_of}")
    for s in book.sources:
        print(f"  {s.lob:<20} {s.combos:>4} combos  v{s.version}  {s.modified}  {s.path}")
    print(f"  business units: {', '.join(book.business_units)}")
    print(f"  states: {len(book.states)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
